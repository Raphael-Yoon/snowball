from flask import Blueprint, request, render_template, session, url_for
from datetime import datetime
import os
from io import BytesIO
from openpyxl import load_workbook
from snowball_mail import send_gmail_with_attachment
from auth import log_user_activity

# Blueprint 생성
bp_link1 = Blueprint('link1', __name__)

def get_user_info():
    """현재 로그인한 사용자 정보 반환 (세션 우선)"""
    from snowball import is_logged_in
    from auth import get_current_user
    if is_logged_in():
        # 세션에 저장된 user_info를 우선 사용
        if 'user_info' in session:
            return session['user_info']
        # 없으면 데이터베이스에서 조회
        return get_current_user()
    return None

def is_logged_in():
    """로그인 상태 확인"""
    from snowball import is_logged_in as main_is_logged_in
    return main_is_logged_in()

def generate_and_send_rcm_excel(form_data, save_local_path=None):
    """
    form_data: dict (Flask의 request.form.to_dict())
    RCM 엑셀 파일을 생성하고 이메일로 전송

    save_local_path: 테스트용 - 이 경로가 지정되면 파일을 로컬에도 저장
    """
    
    # 파일명 생성: 입력받은 파일명(param2)_RCM_YYMMDD.xlsx
    base_name = form_data.get('param2', 'output')
    today = datetime.today().strftime('%Y%m%d')
    file_name = f"{base_name}_ITGC_RCM_{today}.xlsx"
    
    # 엑셀 파일 생성 (템플릿 사용)
    excel_stream = BytesIO()
    template_path = os.path.join("static", "RCM_Generate.xlsx")
    wb = load_workbook(template_path)
    
    # RCM 시트에서 파라미터에 따라 행 필터링
    if 'RCM' in wb.sheetnames:
        ws = wb['RCM']
        rows_to_delete = []

        # Tool 사용 여부 확인 (체크되지 않으면 'N', 체크되면 'Y')
        use_os_tool = form_data.get('use_os_tool', 'N')
        use_db_tool = form_data.get('use_db_tool', 'N')

        # 2행부터 시작 (헤더 제외)
        for row_num in range(2, ws.max_row + 1):
            section_col = ws[f'BR{row_num}'].value  # B컬럼 (Section)
            value_col = ws[f'BS{row_num}'].value    # C컬럼 (Value)

            # B컬럼이 Common인 경우는 무조건 유지
            if section_col == 'Common':
                continue

            # B컬럼이 Cloud인 경우 param_cloud와 비교
            elif section_col == 'Cloud':
                if value_col == '공통':
                    continue
                elif value_col != form_data.get('param_cloud'):
                    rows_to_delete.append(row_num)

            # B컬럼이 APP인 경우
            elif section_col == 'APP':
                if value_col == '공통':
                    continue
                elif value_col != form_data.get('param3'):
                    rows_to_delete.append(row_num)

            # B컬럼이 OS인 경우 param4와 비교
            elif section_col == 'OS':
                if value_col != form_data.get('param4'):
                    rows_to_delete.append(row_num)

            # B컬럼이 DB인 경우 param5와 비교
            elif section_col == 'DB':
                if value_col != form_data.get('param5'):
                    rows_to_delete.append(row_num)

            # B컬럼이 OS_Tool인 경우 - use_os_tool이 'Y'가 아니면 삭제
            elif section_col == 'OS_Tool':
                if use_os_tool != 'Y':
                    rows_to_delete.append(row_num)

            # B컬럼이 DB_Tool인 경우 - use_db_tool이 'Y'가 아니면 삭제
            elif section_col == 'DB_Tool':
                if use_db_tool != 'Y':
                    rows_to_delete.append(row_num)
        
        # 역순으로 행 삭제 (인덱스 변화 방지)
        for row_num in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_num)
        # B, C 열 삭제
        ws.delete_cols(70, 2)
        # 커서를 B2 셀로 이동
        ws.sheet_view.selection[0].activeCell = 'B2'
        ws.sheet_view.selection[0].sqref = 'B2'
            
    # 엑셀 파일 저장 (한글 처리 개선)
    wb.save(excel_stream)
    wb.close()
    excel_stream.seek(0)

    # 테스트용: 로컬 파일 저장
    if save_local_path:
        os.makedirs(os.path.dirname(save_local_path), exist_ok=True)
        with open(save_local_path, 'wb') as f:
            f.write(excel_stream.read())
        excel_stream.seek(0)

    # 이메일 전송
    user_email = form_data.get('param1')
    if not user_email:
        return False, None, "이메일 주소가 없습니다."

    try:
        subject = 'RCM 자동생성 결과 파일'
        body = '요청하신 RCM 자동생성 엑셀 파일을 첨부합니다.'
        send_gmail_with_attachment(
            to=user_email,
            subject=subject,
            body=body,
            file_stream=excel_stream,
            file_name=file_name
        )
        return True, user_email, None
    except Exception as e:
        return False, user_email, str(e)


# ===== Blueprint 라우트 정의 =====

@bp_link1.route('/link1')
def link1():
    """RCM 자동생성 페이지"""
    user_info = get_user_info()
    users = user_info['user_name'] if user_info else "Guest"
    # 로그인된 사용자의 이메일 주소 자동 입력
    user_email = user_info.get('user_email', '') if user_info else ''

    # 로그인한 사용자만 활동 로그 기록
    if is_logged_in():
        log_user_activity(user_info, 'PAGE_ACCESS', 'RCM 페이지', '/link1',
                         request.remote_addr, request.headers.get('User-Agent'))

    return render_template('link1.jsp',
                         return_code=0,
                         users=users,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         user_email=user_email,
                         remote_addr=request.remote_addr)


@bp_link1.route('/rcm_generate', methods=['POST'])
def rcm_generate():
    """RCM 엑셀 생성 및 전송 (인증 필요)"""
    form_data = request.form.to_dict()
    success, user_email, error = generate_and_send_rcm_excel(form_data)

    if success:
        return render_template('mail_sent.jsp', user_email=user_email)
    else:
        # 에러 메시지 설정
        if not user_email:
            error_message = '이메일 주소가 입력되지 않았습니다. 담당자 정보를 확인해 주세요.'
        else:
            error_message = f'메일 전송에 실패했습니다: {error}'

        return render_template('error.jsp',
                             error_title='RCM 생성 실패',
                             error_message=error_message,
                             return_url=url_for('link1.link1'))