from flask import Flask, render_template, request, send_file
import openpyxl.utils
from werkzeug.utils import secure_filename
import os
import openpyxl
import snowball_db

def paper_request(form_data):
    print("Link2 Paper Request called")

    param1 = form_data.get('param1')
    param2 = form_data.get('param2')
    param3 = form_data.get('param3')
    param4 = form_data.get('param4')

    print("Param1 = ", param1)
    print("Param2 = ", param2)
    print("Param3 = ", param3)

    uploaded_file = request.files['param3']
    filename = secure_filename(uploaded_file.filename)
    file_path = os.path.join('uploads', uploaded_file.filename)
    uploaded_file.save(file_path)
    print('upload complete: ', param3)

    snowball_db.set_paper_request(pi_client_name=param1, pi_email=param2, pi_request_file=param4, pi_request_content=param3)

def design_generate(form_data):
    print("Link2 Design Generate called")

    param1 = form_data.get('param1')
    param2 = form_data.get('param2')
    param3 = form_data.get('param3')

    print("Param1 = ", param1)
    print("Param2 = ", param2)
    print("Param3 = ", param3)

    output_path = './downloads/' + param2 + '_설계평가.xlsx'

    if param3:
        print("Upload, Custom")
        uploaded_file = request.files['param3']
        filename = secure_filename(uploaded_file.filename)
        file_path = os.path.join('uploads', uploaded_file.filename)
        uploaded_file.save(file_path)
        test_workbook = openpyxl.load_workbook(file_path)
    else:
        print("No Upload, Standard")
        test_workbook = openpyxl.load_workbook('./paper_templates/Design_Template.xlsx')
        
    test_sheet = test_workbook['RCM']

    paper_workbook = openpyxl.load_workbook('./paper_templates/Design_Paper.xlsx')
    paper_sheet = paper_workbook['RCM']

    for row in test_sheet.iter_rows(min_row=2, max_col=5):
        for cell in row:
            paper_sheet[cell.coordinate].value = cell.value

    rcm_sheet = paper_workbook["RCM"]
    template_sheet = paper_workbook["Template"]

    for row in rcm_sheet.iter_rows(min_row=2, values_only=True):
        sheet_name = row[0]
        if not sheet_name:
            break
        copied_sheet = paper_workbook.copy_worksheet(template_sheet)
        copied_sheet.title = sheet_name
        copied_sheet["A3"].value = param1
        copied_sheet["C7"].value = row[0] #통제코드
        copied_sheet["C8"].value = row[1] #통제명
        copied_sheet["C9"].value = row[2]  #주기
        copied_sheet["C10"].value = row[3] #구분
        copied_sheet["C11"].value = row[4] #테스트절차

    paper_workbook.remove(template_sheet)

    i=2
    for row in rcm_sheet.iter_rows(min_row=2, max_row=rcm_sheet.max_row, min_col=1, max_col=1):
        cell_value = row[0].value
        if cell_value:
            rcm_sheet["F"+str(i)].value = '=HYPERLINK("{}", "{}")'.format("#" + cell_value + "!A1", cell_value)
            i = i+1

    paper_workbook.save(output_path)
    test_workbook.close()
    paper_workbook.close()

    return output_path

def design_template_download(form_data):
    print("Design Generate called")

    param1 = form_data.get('param1')
    param2 = form_data.get('param2')
    param3 = form_data.get('param3')

    print("Param1 = ", param1)
    print("Param2 = ", param2)
    print("Param3 = ", param3)

    output_path = './paper_templates/Design_Download_Template.xlsx'

    return output_path
