import sys
import os
project_root = os.path.join(os.path.dirname(__file__), 'snowball')
sys.path.insert(0, project_root)
os.chdir(project_root)

from openpyxl import load_workbook

wb = load_workbook('paper_templates/Template_Manual.xlsx')
ws = wb['Testing Table']

print("=== Testing Table 구조 확인 ===")
for row in range(1, 8):
    b_val = ws.cell(row, 2).value
    c_val = ws.cell(row, 3).value
    d_val = ws.cell(row, 4).value
    print(f"Row {row}: B=[{b_val}], C=[{c_val}], D=[{d_val}]")

print(f"\nRow 3 merged? {any('B3' in str(merged) for merged in ws.merged_cells)}")
print(f"Merged cells containing row 3: {[str(merged) for merged in ws.merged_cells if 'B3' in str(merged) or 'C3' in str(merged)]}")
