"""
파일 업로드/다운로드 관리 모듈

평가 증빙 이미지, 모집단 엑셀 등 파일 관리 기능 제공
"""

import os
import time
from werkzeug.utils import secure_filename

# ===================================================================
# 파일 업로드 공통 함수
# ===================================================================

def save_evaluation_files(uploaded_images, excel_file, rcm_id, header_id, control_code, evaluation_type='design'):
    """
    평가 파일 저장 공통 함수
    
    Args:
        uploaded_images: 이미지 파일 리스트
        excel_file: 엑셀 파일 (모집단 데이터, None 가능)
        rcm_id: RCM ID
        header_id: 평가 헤더 ID
        control_code: 통제 코드
        evaluation_type: 'design' 또는 'operation'
    
    Returns:
        dict: {'images': [...], 'excel': None or {...}}
    """
    saved_files = {
        'images': [],
        'excel': None
    }
    
    # 이미지 파일 저장
    if uploaded_images:
        upload_dir = os.path.join('static', 'uploads', f'{evaluation_type}_evaluations', 
                                  str(rcm_id), str(header_id), control_code, 'images')
        os.makedirs(upload_dir, exist_ok=True)
        
        for i, image_file in enumerate(uploaded_images):
            if image_file and image_file.filename:
                # 확장자 분리
                original_name, original_ext = os.path.splitext(image_file.filename)
                
                # 안전한 파일명 생성
                safe_name = secure_filename(original_name)
                if not safe_name or safe_name.strip() == '' or safe_name == '_':
                    safe_name = 'evidence'
                
                # 타임스탬프 추가
                timestamp = str(int(time.time()))
                safe_filename = f"{safe_name}_{timestamp}_{i}{original_ext.lower()}"
                
                # 파일 저장
                file_path = os.path.join(upload_dir, safe_filename)
                image_file.save(file_path)
                
                # 상대 경로 저장
                relative_path = f"uploads/{evaluation_type}_evaluations/{rcm_id}/{header_id}/{control_code}/images/{safe_filename}"
                saved_files['images'].append({
                    'filename': safe_filename,
                    'original_name': image_file.filename,
                    'path': relative_path,
                    'url': f"/static/{relative_path}"
                })
    
    # 엑셀 파일 저장 (모집단 데이터)
    if excel_file and excel_file.filename:
        upload_dir = os.path.join('static', 'uploads', f'{evaluation_type}_evaluations', 
                                  str(rcm_id), str(header_id), control_code, 'excel')
        os.makedirs(upload_dir, exist_ok=True)
        
        # 확장자 분리
        original_name, original_ext = os.path.splitext(excel_file.filename)
        
        # 안전한 파일명 생성
        safe_name = secure_filename(original_name)
        if not safe_name or safe_name.strip() == '' or safe_name == '_':
            safe_name = 'population_data'
        
        # 타임스탬프 추가
        timestamp = str(int(time.time()))
        safe_filename = f"{safe_name}_{timestamp}{original_ext.lower()}"
        
        # 파일 저장
        file_path = os.path.join(upload_dir, safe_filename)
        excel_file.save(file_path)
        
        # 상대 경로 저장
        relative_path = f"uploads/{evaluation_type}_evaluations/{rcm_id}/{header_id}/{control_code}/excel/{safe_filename}"
        saved_files['excel'] = {
            'filename': safe_filename,
            'original_name': excel_file.filename,
            'path': relative_path,
            'url': f"/static/{relative_path}"
        }
    
    return saved_files


def get_evaluation_files(rcm_id, header_id, control_code, evaluation_type='design'):
    """
    저장된 평가 파일 목록 조회
    
    Args:
        rcm_id: RCM ID
        header_id: 평가 헤더 ID
        control_code: 통제 코드
        evaluation_type: 'design' 또는 'operation'
    
    Returns:
        dict: {'images': [...], 'excel': None or {...}}
    """
    files = {
        'images': [],
        'excel': None
    }
    
    # 이미지 파일 조회
    image_dir = os.path.join('static', 'uploads', f'{evaluation_type}_evaluations', 
                             str(rcm_id), str(header_id), control_code, 'images')
    
    if os.path.exists(image_dir):
        for filename in sorted(os.listdir(image_dir)):
            if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp')):
                relative_path = f"uploads/{evaluation_type}_evaluations/{rcm_id}/{header_id}/{control_code}/images/{filename}"
                files['images'].append({
                    'filename': filename,
                    'path': relative_path,
                    'url': f"/static/{relative_path}"
                })
    
    # 엑셀 파일 조회 (최신 파일 1개만)
    excel_dir = os.path.join('static', 'uploads', f'{evaluation_type}_evaluations', 
                             str(rcm_id), str(header_id), control_code, 'excel')
    
    if os.path.exists(excel_dir):
        excel_files = [f for f in os.listdir(excel_dir) 
                      if f.lower().endswith(('.xlsx', '.xls', '.csv'))]
        if excel_files:
            # 최신 파일 선택
            latest_file = sorted(excel_files, reverse=True)[0]
            relative_path = f"uploads/{evaluation_type}_evaluations/{rcm_id}/{header_id}/{control_code}/excel/{latest_file}"
            files['excel'] = {
                'filename': latest_file,
                'path': relative_path,
                'url': f"/static/{relative_path}"
            }
    
    return files


def delete_evaluation_file(file_path):
    """
    평가 파일 삭제
    
    Args:
        file_path: 삭제할 파일의 상대 경로
    
    Returns:
        bool: 성공 여부
    """
    try:
        full_path = os.path.join('static', file_path.replace('uploads/', ''))
        if os.path.exists(full_path):
            os.remove(full_path)
            return True
        return False
    except Exception as e:
        print(f"파일 삭제 실패: {e}")
        return False


def validate_image_file(file):
    """
    이미지 파일 검증
    
    Args:
        file: 업로드된 파일 객체
    
    Returns:
        tuple: (bool, str) - (유효성, 에러메시지)
    """
    if not file or not file.filename:
        return False, "파일이 선택되지 않았습니다."
    
    allowed_extensions = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'}
    _, ext = os.path.splitext(file.filename)
    
    if ext.lower() not in allowed_extensions:
        return False, f"지원하지 않는 이미지 형식입니다. ({', '.join(allowed_extensions)})"
    
    # 파일 크기 제한 (10MB)
    file.seek(0, os.SEEK_END)
    file_size = file.tell()
    file.seek(0)
    
    if file_size > 10 * 1024 * 1024:
        return False, "이미지 파일은 10MB 이하만 업로드 가능합니다."
    
    return True, ""


def validate_excel_file(file):
    """
    엑셀 파일 검증
    
    Args:
        file: 업로드된 파일 객체
    
    Returns:
        tuple: (bool, str) - (유효성, 에러메시지)
    """
    if not file or not file.filename:
        return False, "파일이 선택되지 않았습니다."
    
    allowed_extensions = {'.xlsx', '.xls', '.csv'}
    _, ext = os.path.splitext(file.filename)
    
    if ext.lower() not in allowed_extensions:
        return False, f"지원하지 않는 파일 형식입니다. ({', '.join(allowed_extensions)})"
    
    # 파일 크기 제한 (50MB)
    file.seek(0, os.SEEK_END)
    file_size = file.tell()
    file.seek(0)
    
    if file_size > 50 * 1024 * 1024:
        return False, "엑셀 파일은 50MB 이하만 업로드 가능합니다."
    
    return True, ""


# ===================================================================
# 표본 추출 관련 함수
# ===================================================================

def calculate_sample_size(population_count):
    """
    모집단 크기에 따른 표본 수 계산
    
    Args:
        population_count: 모집단 개수
    
    Returns:
        int: 표본 수
    
    Examples:
        >>> calculate_sample_size(0)
        0
        >>> calculate_sample_size(1)
        1
        >>> calculate_sample_size(3)
        2
        >>> calculate_sample_size(10)
        2
        >>> calculate_sample_size(30)
        5
        >>> calculate_sample_size(100)
        20
        >>> calculate_sample_size(300)
        25
    """
    if population_count == 0:
        return 0
    elif population_count == 1:
        return 1
    elif 2 <= population_count <= 4:
        return 2
    elif 5 <= population_count <= 12:
        return 2
    elif 13 <= population_count <= 52:
        return 5
    elif 53 <= population_count <= 250:
        return 20
    else:  # > 250
        return 25


def select_random_samples(population_list, sample_size):
    """
    모집단에서 무작위 표본 선택
    
    Args:
        population_list: 모집단 리스트
        sample_size: 선택할 표본 수
    
    Returns:
        list: 선택된 표본 리스트 (인덱스 포함)
    """
    import random
    
    if not population_list or sample_size <= 0:
        return []
    
    population_count = len(population_list)
    actual_sample_size = min(sample_size, population_count)
    
    # 랜덤 인덱스 선택
    selected_indices = random.sample(range(population_count), actual_sample_size)
    
    # 인덱스와 함께 반환
    samples = []
    for idx in sorted(selected_indices):
        samples.append({
            'index': idx,
            'data': population_list[idx]
        })
    
    return samples


# ===================================================================
# 엑셀 파싱 관련 함수 (표준통제 테스트용)
# ===================================================================

def parse_population_excel(file_path, field_mapping):
    """
    모집단 엑셀 파일 파싱
    
    Args:
        file_path: 엑셀 파일 경로
        field_mapping: 필드 매핑 정보
            {
                'user_id': 'A',  # 또는 컬럼명 '사용자ID'
                'user_name': 'B',
                'department': 'C',
                'permission': 'D',
                'grant_date': 'E'
            }
    
    Returns:
        list: 파싱된 데이터 리스트
        [
            {
                'user_id': 'U001',
                'user_name': '홍길동',
                'department': '재무팀',
                'permission': '관리자',
                'grant_date': '2024-01-15'
            },
            ...
        ]
    """
    from openpyxl import load_workbook
    from datetime import datetime
    
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active

        population_data = []

        # 첫 번째 행은 헤더로 가정, 2번째 행부터 읽기
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # 빈 행 스킵
                continue
            
            # 필드 매핑에 따라 데이터 추출
            record = {}
            for field_name, column_ref in field_mapping.items():
                # 컬럼 참조가 문자열 숫자인 경우 (예: "0", "1")
                if isinstance(column_ref, str) and column_ref.isdigit():
                    col_idx = int(column_ref)
                    value = row[col_idx] if col_idx < len(row) else None
                # 컬럼 참조가 문자(A, B, C)인 경우
                elif isinstance(column_ref, str) and len(column_ref) == 1 and column_ref.isalpha():
                    col_idx = ord(column_ref.upper()) - ord('A')
                    value = row[col_idx] if col_idx < len(row) else None
                # 컬럼 인덱스(0, 1, 2)인 경우
                elif isinstance(column_ref, int):
                    value = row[column_ref] if column_ref < len(row) else None
                else:
                    value = None

                # 날짜 타입 처리
                if value and isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d')

                record[field_name] = value if value is not None else ''

            population_data.append(record)
        
        workbook.close()
        return population_data
        
    except Exception as e:
        raise Exception(f"엑셀 파일 파싱 오류: {str(e)}")


def parse_apd01_population(file_path, field_mapping, recommended_sample_size=None):
    """
    APD01 모집단 엑셀 파싱 (사용자 권한 부여)

    Args:
        file_path: 엑셀 파일 경로
        field_mapping: {
            'user_id': 컬럼_참조,
            'user_name': 컬럼_참조,
            'department': 컬럼_참조,
            'permission': 컬럼_참조,
            'grant_date': 컬럼_참조
        }
        recommended_sample_size: 권장 표본 수 (None이면 자동 계산)

    Returns:
        dict: {
            'population': [...],  # 모집단 데이터
            'count': 100,  # 모집단 수
            'sample_size': 20  # 계산된 표본 수
        }
    """
    population = parse_population_excel(file_path, field_mapping)
    count = len(population)

    # recommended_sample_size가 있으면 사용, 없으면 자동 계산
    if recommended_sample_size is not None and recommended_sample_size > 0:
        sample_size = min(recommended_sample_size, count)  # 모집단보다 클 수 없음
    else:
        sample_size = calculate_sample_size(count)

    return {
        'population': population,
        'count': count,
        'sample_size': sample_size
    }


def parse_apd07_population(file_path, field_mapping, recommended_sample_size=None):
    """
    APD07 모집단 엑셀 파싱 (데이터 직접변경 승인)

    Args:
        file_path: 엑셀 파일 경로
        field_mapping: {
            'change_id': 컬럼_참조,  # 쿼리(변경내역) - 필수
            'change_date': 컬럼_참조,  # 변경 일자 - 필수
            'change_type': 컬럼_참조,  # 변경 유형 (선택)
            'table_name': 컬럼_참조,  # 테이블명 (선택)
            'changed_by': 컬럼_참조,  # 변경자 (선택)
            'approval_date': 컬럼_참조  # 승인 일자 (선택)
        }
        recommended_sample_size: 권장 표본 수 (None이면 자동 계산)

    Returns:
        dict: {
            'population': [...],  # 모집단 데이터
            'count': 100,  # 모집단 수
            'sample_size': 20  # 계산된 표본 수
        }
    """
    population = parse_population_excel(file_path, field_mapping)
    count = len(population)

    # recommended_sample_size가 있으면 사용, 없으면 자동 계산
    if recommended_sample_size is not None and recommended_sample_size > 0:
        sample_size = min(recommended_sample_size, count)
    else:
        sample_size = calculate_sample_size(count)

    return {
        'population': population,
        'count': count,
        'sample_size': sample_size
    }


def parse_apd09_population(file_path, field_mapping, recommended_sample_size=None):
    """
    APD09 모집단 엑셀 파싱 (OS 접근권한 부여 승인)

    Args:
        file_path: 엑셀 파일 경로
        field_mapping: {
            'account': 컬럼_참조,  # 접근권한 부여 계정 - 필수
            'grant_date': 컬럼_참조  # 권한부여일 - 필수
        }
        recommended_sample_size: 권장 표본 수 (None이면 자동 계산)

    Returns:
        dict: {
            'population': [...],  # 모집단 데이터
            'count': 100,  # 모집단 수
            'sample_size': 20  # 계산된 표본 수
        }
    """
    population = parse_population_excel(file_path, field_mapping)
    count = len(population)

    # recommended_sample_size가 있으면 사용, 없으면 자동 계산
    if recommended_sample_size is not None and recommended_sample_size > 0:
        sample_size = min(recommended_sample_size, count)
    else:
        sample_size = calculate_sample_size(count)

    return {
        'population': population,
        'count': count,
        'sample_size': sample_size
    }


def parse_apd12_population(file_path, field_mapping, recommended_sample_size=None):
    """
    APD12 모집단 엑셀 파싱 (DB 접근권한 부여 승인)

    Args:
        file_path: 엑셀 파일 경로
        field_mapping: {
            'account': 컬럼_참조,  # 접근권한 부여 계정 - 필수
            'grant_date': 컬럼_참조  # 권한부여일 - 필수
        }
        recommended_sample_size: 권장 표본 수 (None이면 자동 계산)

    Returns:
        dict: {
            'population': [...],  # 모집단 데이터
            'count': 100,  # 모집단 수
            'sample_size': 20  # 계산된 표본 수
        }
    """
    population = parse_population_excel(file_path, field_mapping)
    count = len(population)

    # recommended_sample_size가 있으면 사용, 없으면 자동 계산
    if recommended_sample_size is not None and recommended_sample_size > 0:
        sample_size = min(recommended_sample_size, count)
    else:
        sample_size = calculate_sample_size(count)

    return {
        'population': population,
        'count': count,
        'sample_size': sample_size
    }


# ===================================================================
# 운영평가 결과 엑셀 다운로드 함수
# ===================================================================

def generate_operation_evaluation_excel(control_code, control_info, test_data, output_path=None):
    """
    운영평가 결과를 템플릿 기반 엑셀로 생성

    Args:
        control_code: 통제 코드 (예: APD01, APD07)
        control_info: 통제 정보 dict
            {
                'control_name': '...',
                'control_description': '...',
                'rcm_name': '...',
                etc.
            }
        test_data: 테스트 데이터 dict
            {
                'population_file': {...},
                'samples_data': {...},
                'test_results_data': {...}
            }
        output_path: 저장 경로 (None이면 임시 파일)

    Returns:
        str: 생성된 엑셀 파일 경로
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import tempfile
    from datetime import datetime
    from shutil import copyfile

    # 템플릿 파일 경로
    template_path = os.path.join('static', 'paper', f'{control_code}_paper.xlsx')

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: {template_path}")

    # 출력 경로 설정
    if output_path is None:
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        output_path = temp_file.name
        temp_file.close()

    # 템플릿 복사
    copyfile(template_path, output_path)

    # 엑셀 파일 열기
    wb = load_workbook(output_path)

    # Testing Table 시트
    if 'Testing Table' in wb.sheetnames:
        sheet = wb['Testing Table']

        # 표본 데이터 채우기
        samples = test_data['samples_data']['samples']
        test_results = test_data['test_results_data']['test_results']

        # 데이터 시작 행 (4번째 행이 헤더, 5번째 행부터 데이터)
        start_row = 5

        for idx, sample in enumerate(samples):
            row_num = start_row + idx
            sample_data = sample['data']

            # APD01 필드 매핑
            if control_code == 'APD01':
                sheet.cell(row_num, 2).value = idx + 1  # No.
                sheet.cell(row_num, 3).value = sample_data.get('user_id', '')  # 사용자ID
                sheet.cell(row_num, 4).value = sample_data.get('user_name', '')  # 사용자명
                sheet.cell(row_num, 5).value = sample_data.get('department', '')  # 부서명
                sheet.cell(row_num, 6).value = sample_data.get('permission', '')  # 권한명
                sheet.cell(row_num, 7).value = sample_data.get('grant_date', '')  # 권한부여일

                # 테스트 결과 데이터
                if idx < len(test_results):
                    result = test_results[idx]
                    sheet.cell(row_num, 8).value = result.get('request_number', '')  # 권한 요청서 번호
                    sheet.cell(row_num, 9).value = result.get('requester_id', '')  # 요청자ID
                    sheet.cell(row_num, 10).value = result.get('requester_name', '')  # 요청자명
                    sheet.cell(row_num, 11).value = result.get('requester_department', '')  # 요청자 부서명
                    sheet.cell(row_num, 12).value = result.get('approver_id', '')  # 승인자ID
                    sheet.cell(row_num, 13).value = result.get('approver_name', '')  # 승인자명
                    sheet.cell(row_num, 14).value = result.get('approval_date', '')  # 승인일자
                    sheet.cell(row_num, 18).value = result.get('notes', '')  # 참고사항/비고

            # APD07 필드 매핑
            elif control_code == 'APD07':
                sheet.cell(row_num, 2).value = idx + 1  # No.
                sheet.cell(row_num, 3).value = sample_data.get('change_id', '')  # Data 변경
                sheet.cell(row_num, 4).value = sample_data.get('changed_by', '')  # 실행자
                sheet.cell(row_num, 5).value = sample_data.get('change_date', '')  # 실행일자

                # 테스트 결과 데이터
                if idx < len(test_results):
                    result = test_results[idx]
                    sheet.cell(row_num, 6).value = result.get('request_number', '')  # 변경 요청서 번호
                    sheet.cell(row_num, 7).value = result.get('requester_id', '')  # 요청자ID
                    sheet.cell(row_num, 8).value = result.get('requester_name', '')  # 요청자명
                    sheet.cell(row_num, 9).value = result.get('requester_department', '')  # 요청자 부서명
                    sheet.cell(row_num, 10).value = result.get('approver_id', '')  # 승인자ID
                    sheet.cell(row_num, 11).value = result.get('approval_date', '')  # 승인일자
                    sheet.cell(row_num, 15).value = result.get('notes', '')  # 참고사항/비고

    # Population 시트에 모집단 데이터 넣기 (선택사항)
    # 필요시 구현

    # 저장
    wb.save(output_path)
    wb.close()

    return output_path


# ===================================================================
# 운영평가 데이터 파일 기반 저장/로드 함수
# ===================================================================

def save_operation_test_data(rcm_id, operation_header_id, control_code, population_data, field_mapping, samples, test_results_data=None):
    """
    운영평가 데이터를 템플릿 기반 엑셀 파일로 저장

    Args:
        rcm_id: RCM ID
        operation_header_id: 운영평가 헤더 ID
        control_code: 통제 코드 (예: APD01, APD07)
        population_data: 파싱된 모집단 데이터 리스트
        field_mapping: 필드 매핑 정보 dict
        samples: 선정된 표본 리스트
        test_results_data: 테스트 결과 데이터 (dict, 선택)

    Returns:
        dict: {
            'excel_path': '...',  # 생성된 엑셀 파일 경로
            'samples_path': '...' # 표본 데이터 JSON 경로 (백업용)
        }
    """
    from openpyxl import load_workbook
    from shutil import copyfile
    import json

    # 저장 디렉토리 생성 (통제코드 폴더 제거)
    base_dir = os.path.join('static', 'uploads', 'operation_evaluations',
                            str(rcm_id), str(operation_header_id))
    os.makedirs(base_dir, exist_ok=True)

    result = {}

    # 1. 템플릿 파일 복사
    if control_code in ['APD01', 'APD07', 'APD09', 'APD12', 'PC01', 'PC02', 'PC03', 'CO01']:
        # 수동통제는 공통 템플릿 사용
        template_path = os.path.join('static', 'paper', 'Template_manual.xlsx')
    else:
        template_path = os.path.join('static', 'paper', f'{control_code}_paper.xlsx')

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: {template_path}")

    excel_filename = f"{control_code}_evaluation.xlsx"
    excel_path = os.path.join(base_dir, excel_filename)

    # 기존 파일이 있으면 삭제 (Windows에서 파일 핸들 문제로 실패할 수 있으므로 무시)
    if os.path.exists(excel_path):
        try:
            os.remove(excel_path)
        except Exception as e:
            print(f"기존 파일 삭제 실패 (무시됨): {e}")

    # 템플릿 복사
    copyfile(template_path, excel_path)

    # 2. 엑셀 파일 열기 및 데이터 채우기
    wb = load_workbook(excel_path)

    # Population 시트에 모집단 데이터 추가
    if 'Population' in wb.sheetnames:
        pop_sheet = wb['Population']

        # 헤더 작성 (1행)
        if control_code == 'APD01':
            headers = ['사용자ID', '사용자명', '부서명', '권한명', '권한부여일']
            for col_idx, header in enumerate(headers, start=1):
                pop_sheet.cell(1, col_idx).value = header

            # 데이터 작성 (2행부터)
            for row_idx, record in enumerate(population_data, start=2):
                pop_sheet.cell(row_idx, 1).value = record.get('user_id', '')
                pop_sheet.cell(row_idx, 2).value = record.get('user_name', '')
                pop_sheet.cell(row_idx, 3).value = record.get('department', '')
                pop_sheet.cell(row_idx, 4).value = record.get('permission', '')
                pop_sheet.cell(row_idx, 5).value = record.get('grant_date', '')

        elif control_code == 'APD07':
            headers = ['쿼리(변경내역)', '변경일자', '변경유형', '테이블명', '변경자', '승인일자']
            for col_idx, header in enumerate(headers, start=1):
                pop_sheet.cell(1, col_idx).value = header

            # 데이터 작성
            for row_idx, record in enumerate(population_data, start=2):
                pop_sheet.cell(row_idx, 1).value = record.get('change_id', '')
                pop_sheet.cell(row_idx, 2).value = record.get('change_date', '')
                pop_sheet.cell(row_idx, 3).value = record.get('change_type', '')
                pop_sheet.cell(row_idx, 4).value = record.get('table_name', '')
                pop_sheet.cell(row_idx, 5).value = record.get('changed_by', '')
                pop_sheet.cell(row_idx, 6).value = record.get('approval_date', '')

        elif control_code == 'APD09':
            headers = ['OS 접근권한명', '접근권한부여일']
            for col_idx, header in enumerate(headers, start=1):
                pop_sheet.cell(1, col_idx).value = header

            # 데이터 작성
            for row_idx, record in enumerate(population_data, start=2):
                pop_sheet.cell(row_idx, 1).value = record.get('account', '')
                pop_sheet.cell(row_idx, 2).value = record.get('grant_date', '')

        elif control_code == 'APD12':
            headers = ['DB 접근권한명', '접근권한부여일']
            for col_idx, header in enumerate(headers, start=1):
                pop_sheet.cell(1, col_idx).value = header

            # 데이터 작성
            for row_idx, record in enumerate(population_data, start=2):
                pop_sheet.cell(row_idx, 1).value = record.get('account', '')
                pop_sheet.cell(row_idx, 2).value = record.get('grant_date', '')

        elif control_code in ['PC01', 'PC02', 'PC03']:
            headers = ['프로그램명', '반영일자']
            for col_idx, header in enumerate(headers, start=1):
                pop_sheet.cell(1, col_idx).value = header

            # 데이터 작성
            for row_idx, record in enumerate(population_data, start=2):
                pop_sheet.cell(row_idx, 1).value = record.get('program_name', '')
                pop_sheet.cell(row_idx, 2).value = record.get('deploy_date', '')

        elif control_code == 'CO01':
            headers = ['배치스케줄이름', '등록일자']
            for col_idx, header in enumerate(headers, start=1):
                pop_sheet.cell(1, col_idx).value = header

            # 데이터 작성
            for row_idx, record in enumerate(population_data, start=2):
                pop_sheet.cell(row_idx, 1).value = record.get('batch_schedule_name', '')
                pop_sheet.cell(row_idx, 2).value = record.get('register_date', '')

    # Testing Table 시트에 표본 데이터 추가
    if 'Testing Table' in wb.sheetnames:
        test_sheet = wb['Testing Table']
        start_row = 5  # 5행부터 데이터 시작
        template_rows = 25  # 템플릿 기본 행 수
        sample_count = len(samples)

        # APD07의 경우 컬럼 조정 (모집단 항목이 2개만 필요: 쿼리, 실행일자)
        if control_code == 'APD07':
            # C4셀에 '쿼리', D4셀에 '실행일자' 입력
            test_sheet.cell(4, 3).value = '쿼리'
            test_sheet.cell(4, 4).value = '실행일자'

            # E, F, G 컬럼 삭제 (3개 컬럼) - 삭제 후 H열이 E열로 이동
            test_sheet.delete_cols(5, 3)

            # 삭제 후 헤더 조정
            test_sheet.cell(4, 5).value = '변경 요청서 번호'  # E열 (원래 H)
            test_sheet.cell(4, 6).value = '요청자명'  # F열 (원래 I)
            test_sheet.cell(4, 7).value = '요청부서'  # G열 (원래 J)
            test_sheet.cell(4, 8).value = '승인자명'  # H열 (원래 K)
            test_sheet.cell(4, 9).value = '승인자부서'  # I열 (원래 L)
            test_sheet.cell(4, 10).value = '승인일자'  # J열 (원래 M)
            test_sheet.cell(4, 11).value = '승인여부'  # K열 (원래 N)
            test_sheet.cell(4, 12).value = '사전승인여부'  # L열 (원래 O)
            test_sheet.cell(4, 13).value = '결론'  # M열 (원래 P)
            test_sheet.cell(4, 14).value = '비고'  # N열 (원래 Q)

            # 템플릿 행(5~29행)의 K, L, M열 수식 초기화 및 재설정
            for row in range(5, 30):
                test_sheet.cell(row, 11).value = f'=IF(H{row}<>"","Y","N")'  # K열: 승인여부
                test_sheet.cell(row, 12).value = f'=IF(J{row}<=D{row},"Y","N")'  # L열: 사전승인여부
                test_sheet.cell(row, 13).value = None  # M열: 결론 (저장 시 입력)

        # APD09/APD12의 경우 컬럼 조정 (모집단 항목이 2개만 필요: 접근권한명, 접근권한부여일)
        elif control_code in ['APD09', 'APD12']:
            # C4셀에 '접근권한명', D4셀에 '접근권한부여일' 입력
            control_type = 'OS' if control_code == 'APD09' else 'DB'
            test_sheet.cell(4, 3).value = f'{control_type} 접근권한명'
            test_sheet.cell(4, 4).value = '접근권한부여일'

            # E, F, G 컬럼 삭제 (3개 컬럼) - 삭제 후 H열이 E열로 이동
            test_sheet.delete_cols(5, 3)

            # 삭제 후 헤더 조정
            test_sheet.cell(4, 5).value = '권한 요청서 번호'  # E열 (원래 H)
            test_sheet.cell(4, 6).value = '요청자명'  # F열 (원래 I)
            test_sheet.cell(4, 7).value = '요청부서'  # G열 (원래 J)
            test_sheet.cell(4, 8).value = '승인자명'  # H열 (원래 K)
            test_sheet.cell(4, 9).value = '승인자부서'  # I열 (원래 L)
            test_sheet.cell(4, 10).value = '승인일자'  # J열 (원래 M)
            test_sheet.cell(4, 11).value = '승인여부'  # K열 (원래 N)
            test_sheet.cell(4, 12).value = '사전승인여부'  # L열 (원래 O)
            test_sheet.cell(4, 13).value = '결론'  # M열 (원래 P)
            test_sheet.cell(4, 14).value = '비고'  # N열 (원래 Q)

            # 템플릿 행(5~29행)의 K, L, M열 수식 초기화 및 재설정
            for row in range(5, 30):
                test_sheet.cell(row, 11).value = f'=IF(H{row}<>"","Y","N")'  # K열: 승인여부
                test_sheet.cell(row, 12).value = f'=IF(J{row}<=D{row},"Y","N")'  # L열: 사전승인여부
                test_sheet.cell(row, 13).value = None  # M열: 결론 (저장 시 입력)

        elif control_code == 'PC01':
            # C4셀에 '프로그램명', D4셀에 '반영일자' 입력
            test_sheet.cell(4, 3).value = '프로그램명'
            test_sheet.cell(4, 4).value = '반영일자'

            # E, F, G 컬럼 삭제 (3개 컬럼) - APD09/APD12와 동일
            test_sheet.delete_cols(5, 3)

            # 삭제 후 헤더 조정
            test_sheet.cell(4, 5).value = '변경 요청서 번호'  # E열
            test_sheet.cell(4, 6).value = '요청자'  # F열
            test_sheet.cell(4, 7).value = '요청자부서'  # G열
            test_sheet.cell(4, 8).value = '승인자'  # H열
            test_sheet.cell(4, 9).value = '승인자부서'  # I열
            test_sheet.cell(4, 10).value = '승인일자'  # J열
            test_sheet.cell(4, 11).value = '개발담당자'  # K열
            test_sheet.cell(4, 12).value = '배포담당자'  # L열
            test_sheet.cell(4, 13).value = '예외'  # M열
            test_sheet.cell(4, 14).value = '비고'  # N열

        elif control_code == 'PC02':
            # C4셀에 '프로그램명', D4셀에 '반영일자' 입력
            test_sheet.cell(4, 3).value = '프로그램명'
            test_sheet.cell(4, 4).value = '반영일자'

            # E, F, G 컬럼 삭제 (사용자ID, 사용자명, 부서명)
            test_sheet.delete_cols(5, 3)

            # 삭제 후 E열은 원래 H열(권한 요청서 번호)을 '변경 요청서 번호'로 변경
            test_sheet.cell(4, 5).value = '변경 요청서 번호'  # E열

            # F~I열 삭제 (요청자명, 요청자부서명, 승인자명, 승인자부서명)
            test_sheet.delete_cols(6, 4)

            # 삭제 후 F열은 원래 M열(승인일자)이므로 삭제
            test_sheet.delete_cols(6, 1)

            # 삭제 후 F~H열 삭제 (승인여부, 사전승인여부, 결론)
            test_sheet.delete_cols(6, 3)

            # PC02 전용 컬럼 설정 (삭제 후 F~J열)
            test_sheet.cell(4, 6).value = '사용자테스트 유무'  # F열
            test_sheet.cell(4, 7).value = '사용자테스트담당자'  # G열
            test_sheet.cell(4, 8).value = '사용자테스트일자'  # H열
            test_sheet.cell(4, 9).value = '예외'  # I열
            test_sheet.cell(4, 10).value = '비고'  # J열 (원래 Q열)

        elif control_code == 'PC03':
            # C4셀에 '프로그램명', D4셀에 '반영일자' 입력
            test_sheet.cell(4, 3).value = '프로그램명'
            test_sheet.cell(4, 4).value = '반영일자'

            # E, F, G 컬럼 삭제 (사용자ID, 사용자명, 부서명)
            test_sheet.delete_cols(5, 3)

            # 삭제 후 E열은 원래 H열(권한 요청서 번호)을 '변경 요청서 번호'로 변경
            test_sheet.cell(4, 5).value = '변경 요청서 번호'  # E열

            # F~I열 삭제 (요청자명, 요청자부서명, 승인자명, 승인자부서명)
            test_sheet.delete_cols(6, 4)

            # 삭제 후 F열은 원래 M열(승인일자)이므로 삭제
            test_sheet.delete_cols(6, 1)

            # 삭제 후 F~H열 삭제 (승인여부, 사전승인여부, 결론)
            test_sheet.delete_cols(6, 3)

            # PC03 전용 컬럼 설정 (삭제 후 F~K열)
            test_sheet.cell(4, 6).value = '배포요청자'  # F열
            test_sheet.cell(4, 7).value = '배포요청자부서'  # G열
            test_sheet.cell(4, 8).value = '배포승인자'  # H열
            test_sheet.cell(4, 9).value = '배포승인자부서'  # I열
            test_sheet.cell(4, 10).value = '배포승인일자'  # J열
            test_sheet.cell(4, 11).value = '예외'  # K열
            test_sheet.cell(4, 12).value = '비고'  # L열

        elif control_code == 'CO01':
            # C4셀에 '배치스케줄이름', D4셀에 '등록일자' 입력
            test_sheet.cell(4, 3).value = '배치스케줄이름'
            test_sheet.cell(4, 4).value = '등록일자'

            # E, F, G 컬럼 삭제 (사용자ID, 사용자명, 부서명)
            test_sheet.delete_cols(5, 3)

            # 삭제 후 헤더 조정
            test_sheet.cell(4, 5).value = '요청번호'  # E열
            test_sheet.cell(4, 6).value = '요청자'  # F열
            test_sheet.cell(4, 7).value = '요청자부서'  # G열
            test_sheet.cell(4, 8).value = '승인자'  # H열
            test_sheet.cell(4, 9).value = '승인자부서'  # I열
            test_sheet.cell(4, 10).value = '승인일자'  # J열
            test_sheet.cell(4, 11).value = '예외'  # K열
            test_sheet.cell(4, 12).value = '비고'  # L열

        # 표본 수가 템플릿 행 수보다 많으면 행 추가
        if sample_count > template_rows:
            # 마지막 템플릿 행을 복사하여 추가
            for i in range(template_rows, sample_count):
                # 템플릿의 마지막 행(29행) 복사
                source_row = start_row + template_rows - 1
                target_row = start_row + i

                # 행 삽입
                test_sheet.insert_rows(target_row)

                # 스타일 및 병합 정보 복사
                for col in range(1, test_sheet.max_column + 1):
                    source_cell = test_sheet.cell(source_row, col)
                    target_cell = test_sheet.cell(target_row, col)

                    # 스타일 복사
                    if source_cell.has_style:
                        target_cell.font = source_cell.font.copy()
                        target_cell.border = source_cell.border.copy()
                        target_cell.fill = source_cell.fill.copy()
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = source_cell.protection.copy()
                        target_cell.alignment = source_cell.alignment.copy()

                # APD07/APD09/APD12의 경우 추가 행에도 K, L열 수식 적용
                if control_code in ['APD07', 'APD09', 'APD12']:
                    test_sheet.cell(target_row, 11).value = f'=IF(H{target_row}<>"","Y","N")'  # K열: 승인여부
                    test_sheet.cell(target_row, 12).value = f'=IF(J{target_row}<=D{target_row},"Y","N")'  # L열: 사전승인여부
                    test_sheet.cell(target_row, 13).value = None  # M열: 결론

        # 표본 수가 템플릿 행 수보다 적으면 남는 행 내용만 지우기 (행은 유지)
        elif sample_count < template_rows:
            for row_idx in range(start_row + sample_count, start_row + template_rows):
                for col in range(1, test_sheet.max_column + 1):
                    cell = test_sheet.cell(row_idx, col)
                    cell.value = None  # 값만 지우고 스타일은 유지

        for idx, sample in enumerate(samples):
            row_num = start_row + idx
            sample_data = sample['data']

            if control_code == 'APD01':
                test_sheet.cell(row_num, 2).value = idx + 1  # No.
                test_sheet.cell(row_num, 3).value = sample_data.get('user_id', '')
                test_sheet.cell(row_num, 4).value = sample_data.get('user_name', '')
                test_sheet.cell(row_num, 5).value = sample_data.get('department', '')
                test_sheet.cell(row_num, 6).value = sample_data.get('permission', '')
                test_sheet.cell(row_num, 7).value = sample_data.get('grant_date', '')

            elif control_code == 'APD07':
                test_sheet.cell(row_num, 2).value = idx + 1  # No.
                test_sheet.cell(row_num, 3).value = sample_data.get('change_id', '')  # 쿼리
                test_sheet.cell(row_num, 4).value = sample_data.get('change_date', '')  # 실행일자

            elif control_code in ['APD09', 'APD12']:
                test_sheet.cell(row_num, 2).value = idx + 1  # No.
                test_sheet.cell(row_num, 3).value = sample_data.get('account', '')  # OS/DB 접근권한명
                test_sheet.cell(row_num, 4).value = sample_data.get('grant_date', '')  # 접근권한부여일

            elif control_code in ['PC01', 'PC02', 'PC03']:
                test_sheet.cell(row_num, 2).value = idx + 1  # No.
                test_sheet.cell(row_num, 3).value = sample_data.get('program_name', '')  # 프로그램명
                test_sheet.cell(row_num, 4).value = sample_data.get('deploy_date', '')  # 반영일자

            elif control_code == 'CO01':
                test_sheet.cell(row_num, 2).value = idx + 1  # No.
                test_sheet.cell(row_num, 3).value = sample_data.get('batch_schedule_name', '')  # 배치스케줄이름
                test_sheet.cell(row_num, 4).value = sample_data.get('register_date', '')  # 등록일자

        # 테스트 결과가 있으면 채우기
        if test_results_data and 'test_results' in test_results_data:
            test_results = test_results_data['test_results']
            print(f"[DEBUG file_manager] test_results 데이터: {test_results}")

            for idx, result in enumerate(test_results):
                if idx >= len(samples):
                    break

                row_num = start_row + idx

                if control_code == 'APD01':
                    print(f"[DEBUG] APD01 row {row_num} 저장: approver_name={result.get('approver_name')}, requester_department={result.get('requester_department')}, approval_date={result.get('approval_date')}")
                    test_sheet.cell(row_num, 8).value = result.get('request_number', '')  # H열: 권한 요청서 번호
                    test_sheet.cell(row_num, 9).value = result.get('requester_name', '')  # I열: 요청자명
                    test_sheet.cell(row_num, 10).value = result.get('requester_department', '')  # J열: 요청자 부서명
                    test_sheet.cell(row_num, 11).value = result.get('approver_name', '')  # K열: 승인자명
                    test_sheet.cell(row_num, 12).value = result.get('approver_department', '')  # L열: 승인자 부서명
                    test_sheet.cell(row_num, 13).value = result.get('approval_date', '')  # M열: 승인일자
                    test_sheet.cell(row_num, 17).value = result.get('notes', '')  # Q열: 참고사항/비고

                elif control_code == 'APD07':
                    # E,F,G 컬럼 삭제 후 템플릿 구조: E(요청서번호), F(요청자명), G(요청부서), H(승인자명), I(승인자부서), J(승인일자), K(승인여부), L(사전승인여부), M(결론), N(비고)
                    test_sheet.cell(row_num, 5).value = result.get('request_number', '')  # E열: 변경 요청서 번호
                    test_sheet.cell(row_num, 6).value = result.get('requester_name', '')  # F열: 요청자명
                    test_sheet.cell(row_num, 7).value = result.get('requester_department', '')  # G열: 요청부서
                    test_sheet.cell(row_num, 8).value = result.get('approver_name', '')  # H열: 승인자명
                    test_sheet.cell(row_num, 9).value = result.get('approver_department', '')  # I열: 승인자부서
                    test_sheet.cell(row_num, 10).value = result.get('approval_date', '')  # J열: 승인일자
                    test_sheet.cell(row_num, 11).value = f'=IF(H{row_num}<>"","Y","N")'  # K열: 승인여부 (승인자명 있으면 Y)
                    test_sheet.cell(row_num, 12).value = f'=IF(J{row_num}<=D{row_num},"Y","N")'  # L열: 사전승인여부 (승인일자 <= 실행일자)
                    test_sheet.cell(row_num, 13).value = 'Y' if result.get('has_exception') else 'N'  # M열: 결론 (예외여부)
                    test_sheet.cell(row_num, 14).value = result.get('notes', '')  # N열: 비고

                elif control_code in ['APD09', 'APD12']:
                    # APD09/APD12는 E,F,G 컬럼 삭제 후 APD07과 동일한 템플릿 구조 사용
                    test_sheet.cell(row_num, 5).value = result.get('request_number', '')  # E열: 권한 요청서 번호
                    test_sheet.cell(row_num, 6).value = result.get('requester_name', '')  # F열: 요청자명
                    test_sheet.cell(row_num, 7).value = result.get('requester_department', '')  # G열: 요청부서
                    test_sheet.cell(row_num, 8).value = result.get('approver_name', '')  # H열: 승인자명
                    test_sheet.cell(row_num, 9).value = result.get('approver_department', '')  # I열: 승인자부서
                    test_sheet.cell(row_num, 10).value = result.get('approval_date', '')  # J열: 승인일자
                    test_sheet.cell(row_num, 11).value = f'=IF(H{row_num}<>"","Y","N")'  # K열: 승인여부 (승인자명 있으면 Y)
                    test_sheet.cell(row_num, 12).value = f'=IF(J{row_num}<=D{row_num},"Y","N")'  # L열: 사전승인여부 (승인일자 <= 부여일자)
                    test_sheet.cell(row_num, 13).value = 'Y' if result.get('has_exception') else 'N'  # M열: 결론 (예외여부)
                    test_sheet.cell(row_num, 14).value = result.get('notes', '')  # N열: 비고

                elif control_code == 'PC01':
                    # PC01: 프로그램 변경 승인
                    test_sheet.cell(row_num, 5).value = result.get('request_number', '')  # E열: 요청번호
                    test_sheet.cell(row_num, 6).value = result.get('requester_name', '')  # F열: 요청자
                    test_sheet.cell(row_num, 7).value = result.get('requester_department', '')  # G열: 요청자부서
                    test_sheet.cell(row_num, 8).value = result.get('approver_name', '')  # H열: 승인자
                    test_sheet.cell(row_num, 9).value = result.get('approver_department', '')  # I열: 승인자부서
                    test_sheet.cell(row_num, 10).value = result.get('approval_date', '')  # J열: 승인일자
                    test_sheet.cell(row_num, 11).value = result.get('developer', '')  # K열: 개발담당자
                    test_sheet.cell(row_num, 12).value = result.get('deployer', '')  # L열: 배포담당자
                    test_sheet.cell(row_num, 13).value = 'Y' if result.get('has_exception') else 'N'  # M열: 예외
                    test_sheet.cell(row_num, 14).value = result.get('notes', '')  # N열: 비고

                elif control_code == 'PC02':
                    # PC02: 사용자 테스트
                    test_sheet.cell(row_num, 5).value = result.get('request_number', '')  # E열: 변경 요청서 번호
                    test_sheet.cell(row_num, 6).value = result.get('user_test_yn', '')  # F열: 사용자테스트 유무
                    test_sheet.cell(row_num, 7).value = result.get('user_test_person', '')  # G열: 사용자테스트담당자
                    test_sheet.cell(row_num, 8).value = result.get('user_test_date', '')  # H열: 사용자테스트일자
                    test_sheet.cell(row_num, 9).value = 'Y' if result.get('has_exception') else 'N'  # I열: 예외
                    test_sheet.cell(row_num, 10).value = result.get('notes', '')  # J열: 비고

                elif control_code == 'PC03':
                    # PC03: 배포 승인
                    test_sheet.cell(row_num, 5).value = result.get('request_number', '')  # E열: 변경 요청서 번호
                    test_sheet.cell(row_num, 6).value = result.get('deploy_requester', '')  # F열: 배포요청자
                    test_sheet.cell(row_num, 7).value = result.get('deploy_requester_dept', '')  # G열: 배포요청자부서
                    test_sheet.cell(row_num, 8).value = result.get('deploy_approver', '')  # H열: 배포승인자
                    test_sheet.cell(row_num, 9).value = result.get('deploy_approver_dept', '')  # I열: 배포승인자부서
                    test_sheet.cell(row_num, 10).value = result.get('deploy_approval_date', '')  # J열: 배포승인일자
                    test_sheet.cell(row_num, 11).value = 'Y' if result.get('has_exception') else 'N'  # K열: 예외
                    test_sheet.cell(row_num, 12).value = result.get('notes', '')  # L열: 비고

                elif control_code == 'CO01':
                    # CO01: 배치 스케줄 승인
                    test_sheet.cell(row_num, 5).value = result.get('request_number', '')  # E열: 요청번호
                    test_sheet.cell(row_num, 6).value = result.get('requester_name', '')  # F열: 요청자
                    test_sheet.cell(row_num, 7).value = result.get('requester_department', '')  # G열: 요청자부서
                    test_sheet.cell(row_num, 8).value = result.get('approver_name', '')  # H열: 승인자
                    test_sheet.cell(row_num, 9).value = result.get('approver_department', '')  # I열: 승인자부서
                    test_sheet.cell(row_num, 10).value = result.get('approval_date', '')  # J열: 승인일자
                    test_sheet.cell(row_num, 11).value = 'Y' if result.get('has_exception') else 'N'  # K열: 예외
                    test_sheet.cell(row_num, 12).value = result.get('notes', '')  # L열: 비고

    # 저장
    try:
        wb.save(excel_path)
    except Exception as save_error:
        print(f"[ERROR] 엑셀 파일 저장 실패: {save_error}")
        raise
    finally:
        wb.close()

    result['excel_path'] = f"uploads/operation_evaluations/{rcm_id}/{operation_header_id}/{excel_filename}"

    return result


def load_operation_test_data(rcm_id, operation_header_id, control_code):
    """
    저장된 운영평가 데이터 로드 (엑셀 파일에서 직접)

    Args:
        rcm_id: RCM ID
        operation_header_id: 운영평가 헤더 ID
        control_code: 통제 코드 (예: APD01, APD07)

    Returns:
        dict: {
            'excel_file': {...},           # 엑셀 파일 정보
            'samples_data': {...},         # 표본 및 테스트 결과 데이터 (엑셀에서 읽기)
            'population_data': [...]       # 모집단 데이터 (엑셀에서 읽기)
        }
        없는 경우 None 반환
    """
    from openpyxl import load_workbook

    base_dir = os.path.join('static', 'uploads', 'operation_evaluations',
                            str(rcm_id), str(operation_header_id))

    if not os.path.exists(base_dir):
        return None

    result = {
        'excel_file': None,
        'samples_data': None,
        'population_data': []
    }

    # 엑셀 파일 확인
    excel_filename = f"{control_code}_evaluation.xlsx"
    excel_path = os.path.join(base_dir, excel_filename)

    if not os.path.exists(excel_path):
        return None

    relative_path = f"uploads/operation_evaluations/{rcm_id}/{operation_header_id}/{excel_filename}"
    result['excel_file'] = {
        'filename': excel_filename,
        'path': relative_path,
        'url': f"/static/{relative_path}",
        'full_path': excel_path
    }

    # 엑셀 파일에서 데이터 읽기
    wb = None
    try:
        # data_only=False로 변경하여 방금 저장된 값도 제대로 읽을 수 있도록 함
        wb = load_workbook(excel_path, data_only=False)

        # 1. Population 시트에서 모집단 데이터 읽기
        if 'Population' in wb.sheetnames:
            pop_sheet = wb['Population']
            population_data = []

            for row in pop_sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):  # 빈 행 스킵
                    continue

                if control_code == 'APD01':
                    population_data.append({
                        'user_id': row[0] if len(row) > 0 else '',
                        'user_name': row[1] if len(row) > 1 else '',
                        'department': row[2] if len(row) > 2 else '',
                        'permission': row[3] if len(row) > 3 else '',
                        'grant_date': row[4] if len(row) > 4 else ''
                    })
                elif control_code == 'APD07':
                    population_data.append({
                        'change_id': row[0] if len(row) > 0 else '',
                        'change_date': row[1] if len(row) > 1 else '',
                        'change_type': row[2] if len(row) > 2 else '',
                        'table_name': row[3] if len(row) > 3 else '',
                        'changed_by': row[4] if len(row) > 4 else '',
                        'approval_date': row[5] if len(row) > 5 else ''
                    })
                elif control_code in ['APD09', 'APD12']:
                    population_data.append({
                        'account': row[0] if len(row) > 0 else '',
                        'grant_date': row[1] if len(row) > 1 else ''
                    })
                elif control_code in ['PC01', 'PC02', 'PC03']:
                    population_data.append({
                        'program_name': row[0] if len(row) > 0 else '',
                        'deploy_date': row[1] if len(row) > 1 else ''
                    })
                elif control_code == 'CO01':
                    population_data.append({
                        'batch_schedule_name': row[0] if len(row) > 0 else '',
                        'register_date': row[1] if len(row) > 1 else ''
                    })

            result['population_data'] = population_data

        # 2. Testing Table 시트에서 표본 및 테스트 결과 읽기
        if 'Testing Table' in wb.sheetnames:
            test_sheet = wb['Testing Table']
            samples = []
            test_results = []
            start_row = 5

            for row_num in range(start_row, test_sheet.max_row + 1):
                # No. 컬럼이 비어있으면 종료
                no_value = test_sheet.cell(row_num, 2).value
                if not no_value:
                    break

                if control_code == 'APD01':
                    # 표본 데이터
                    sample = {
                        'index': row_num - start_row,
                        'data': {
                            'user_id': test_sheet.cell(row_num, 3).value or '',
                            'user_name': test_sheet.cell(row_num, 4).value or '',
                            'department': test_sheet.cell(row_num, 5).value or '',
                            'permission': test_sheet.cell(row_num, 6).value or '',
                            'grant_date': test_sheet.cell(row_num, 7).value or ''
                        }
                    }
                    samples.append(sample)

                    # 테스트 결과
                    test_result = {
                        'sample_index': row_num - start_row,
                        'population_data': sample['data'],
                        'request_number': test_sheet.cell(row_num, 8).value or '',  # H열: 권한 요청서 번호
                        'requester_id': '',  # 템플릿에 없음
                        'requester_name': test_sheet.cell(row_num, 9).value or '',  # I열: 요청자명
                        'requester_department': test_sheet.cell(row_num, 10).value or '',  # J열: 요청자 부서명
                        'approver_id': '',  # 템플릿에 없음
                        'approver_name': test_sheet.cell(row_num, 11).value or '',  # K열: 승인자명
                        'approver_department': test_sheet.cell(row_num, 12).value or '',  # L열: 승인자 부서명
                        'approval_date': test_sheet.cell(row_num, 13).value or '',  # M열: 승인일자
                        'notes': test_sheet.cell(row_num, 17).value or '',  # Q열: 참고사항/비고
                        'has_exception': False  # 예외 여부는 별도 로직으로 판단
                    }
                    test_results.append(test_result)

                elif control_code == 'APD07':
                    # 표본 데이터
                    sample = {
                        'index': row_num - start_row,
                        'data': {
                            'change_id': test_sheet.cell(row_num, 3).value or '',  # C열: 쿼리
                            'change_date': test_sheet.cell(row_num, 4).value or ''  # D열: 실행일자
                        }
                    }
                    samples.append(sample)

                    # 테스트 결과 (E,F,G 컬럼 삭제 후 템플릿 구조에 맞춤)
                    test_result = {
                        'sample_index': row_num - start_row,
                        'population_data': sample['data'],
                        'request_number': test_sheet.cell(row_num, 5).value or '',  # E열: 변경 요청서 번호
                        'requester_id': '',  # 템플릿에 없음
                        'requester_name': test_sheet.cell(row_num, 6).value or '',  # F열: 요청자명
                        'requester_department': test_sheet.cell(row_num, 7).value or '',  # G열: 요청부서
                        'approver_id': '',  # 템플릿에 없음
                        'approver_name': test_sheet.cell(row_num, 8).value or '',  # H열: 승인자명
                        'approver_department': test_sheet.cell(row_num, 9).value or '',  # I열: 승인자부서
                        'approval_date': test_sheet.cell(row_num, 10).value or '',  # J열: 승인일자
                        'notes': test_sheet.cell(row_num, 14).value or '',  # N열: 비고
                        'has_exception': False
                    }
                    test_results.append(test_result)

                elif control_code in ['APD09', 'APD12']:
                    # 표본 데이터
                    sample = {
                        'index': row_num - start_row,
                        'data': {
                            'account': test_sheet.cell(row_num, 3).value or '',  # C열: OS/DB 접근권한명
                            'grant_date': test_sheet.cell(row_num, 4).value or ''  # D열: 접근권한부여일
                        }
                    }
                    samples.append(sample)

                    # 테스트 결과 (E,F,G 컬럼 삭제 후 APD07과 동일한 템플릿 구조)
                    test_result = {
                        'sample_index': row_num - start_row,
                        'population_data': sample['data'],
                        'request_number': test_sheet.cell(row_num, 5).value or '',  # E열: 권한 요청서 번호
                        'requester_id': '',  # 템플릿에 없음
                        'requester_name': test_sheet.cell(row_num, 6).value or '',  # F열: 요청자명
                        'requester_department': test_sheet.cell(row_num, 7).value or '',  # G열: 요청부서
                        'approver_id': '',  # 템플릿에 없음
                        'approver_name': test_sheet.cell(row_num, 8).value or '',  # H열: 승인자명
                        'approver_department': test_sheet.cell(row_num, 9).value or '',  # I열: 승인자부서
                        'approval_date': test_sheet.cell(row_num, 10).value or '',  # J열: 승인일자
                        'notes': test_sheet.cell(row_num, 14).value or '',  # N열: 비고
                        'has_exception': False
                    }
                    test_results.append(test_result)

                elif control_code == 'PC01':
                    # 표본 데이터
                    sample = {
                        'index': row_num - start_row,
                        'data': {
                            'program_name': test_sheet.cell(row_num, 3).value or '',  # C열: 프로그램명
                            'deploy_date': test_sheet.cell(row_num, 4).value or ''  # D열: 반영일자
                        }
                    }
                    samples.append(sample)

                    # 테스트 결과
                    test_result = {
                        'sample_index': row_num - start_row,
                        'population_data': sample['data'],
                        'request_number': test_sheet.cell(row_num, 5).value or '',  # E열: 요청번호
                        'requester_id': '',
                        'requester_name': test_sheet.cell(row_num, 6).value or '',  # F열: 요청자
                        'requester_department': test_sheet.cell(row_num, 7).value or '',  # G열: 요청자부서
                        'approver_id': '',
                        'approver_name': test_sheet.cell(row_num, 8).value or '',  # H열: 승인자
                        'approver_department': test_sheet.cell(row_num, 9).value or '',  # I열: 승인자부서
                        'approval_date': test_sheet.cell(row_num, 10).value or '',  # J열: 승인일자
                        'developer': test_sheet.cell(row_num, 11).value or '',  # K열: 개발담당자
                        'deployer': test_sheet.cell(row_num, 12).value or '',  # L열: 배포담당자
                        'notes': test_sheet.cell(row_num, 14).value or '',  # N열: 비고
                        'has_exception': False
                    }
                    test_results.append(test_result)

                elif control_code == 'PC02':
                    # 표본 데이터
                    sample = {
                        'index': row_num - start_row,
                        'data': {
                            'program_name': test_sheet.cell(row_num, 3).value or '',  # C열: 프로그램명
                            'deploy_date': test_sheet.cell(row_num, 4).value or ''  # D열: 반영일자
                        }
                    }
                    samples.append(sample)

                    # 테스트 결과
                    test_result = {
                        'sample_index': row_num - start_row,
                        'population_data': sample['data'],
                        'request_number': test_sheet.cell(row_num, 5).value or '',  # E열: 변경 요청서 번호
                        'user_test_yn': test_sheet.cell(row_num, 6).value or '',  # F열: 사용자테스트 유무
                        'user_test_person': test_sheet.cell(row_num, 7).value or '',  # G열: 사용자테스트담당자
                        'user_test_date': test_sheet.cell(row_num, 8).value or '',  # H열: 사용자테스트일자
                        'notes': test_sheet.cell(row_num, 10).value or '',  # J열: 비고
                        'has_exception': False
                    }
                    test_results.append(test_result)

                elif control_code == 'PC03':
                    # 표본 데이터
                    sample = {
                        'index': row_num - start_row,
                        'data': {
                            'program_name': test_sheet.cell(row_num, 3).value or '',  # C열: 프로그램명
                            'deploy_date': test_sheet.cell(row_num, 4).value or ''  # D열: 반영일자
                        }
                    }
                    samples.append(sample)

                    # 테스트 결과
                    test_result = {
                        'sample_index': row_num - start_row,
                        'population_data': sample['data'],
                        'request_number': test_sheet.cell(row_num, 5).value or '',  # E열: 변경 요청서 번호
                        'deploy_requester': test_sheet.cell(row_num, 6).value or '',  # F열: 배포요청자
                        'deploy_requester_dept': test_sheet.cell(row_num, 7).value or '',  # G열: 배포요청자부서
                        'deploy_approver': test_sheet.cell(row_num, 8).value or '',  # H열: 배포승인자
                        'deploy_approver_dept': test_sheet.cell(row_num, 9).value or '',  # I열: 배포승인자부서
                        'deploy_approval_date': test_sheet.cell(row_num, 10).value or '',  # J열: 배포승인일자
                        'notes': test_sheet.cell(row_num, 12).value or '',  # L열: 비고
                        'has_exception': False
                    }
                    test_results.append(test_result)

                elif control_code == 'CO01':
                    # 표본 데이터
                    sample = {
                        'index': row_num - start_row,
                        'data': {
                            'batch_schedule_name': test_sheet.cell(row_num, 3).value or '',  # C열: 배치스케줄이름
                            'register_date': test_sheet.cell(row_num, 4).value or ''  # D열: 등록일자
                        }
                    }
                    samples.append(sample)

                    # 테스트 결과
                    test_result = {
                        'sample_index': row_num - start_row,
                        'population_data': sample['data'],
                        'request_number': test_sheet.cell(row_num, 5).value or '',  # E열: 요청번호
                        'requester_id': '',
                        'requester_name': test_sheet.cell(row_num, 6).value or '',  # F열: 요청자
                        'requester_department': test_sheet.cell(row_num, 7).value or '',  # G열: 요청자부서
                        'approver_id': '',
                        'approver_name': test_sheet.cell(row_num, 8).value or '',  # H열: 승인자
                        'approver_department': test_sheet.cell(row_num, 9).value or '',  # I열: 승인자부서
                        'approval_date': test_sheet.cell(row_num, 10).value or '',  # J열: 승인일자
                        'notes': test_sheet.cell(row_num, 12).value or '',  # L열: 비고
                        'has_exception': False
                    }
                    test_results.append(test_result)

            result['samples_data'] = {
                'samples': samples,
                'population_count': len(population_data),
                'sample_size': len(samples),
                'test_results': {
                    'test_results': test_results,
                    'test_type': control_code
                }
            }

    except Exception as e:
        print(f"엑셀 파일 읽기 오류: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if wb:
            wb.close()

    return result


def get_operation_test_file_paths(rcm_id, design_session, control_code):
    """
    운영평가 데이터 파일 경로 정보만 반환 (DB 저장용)

    Args:
        rcm_id: RCM ID
        design_session: 설계평가 차수 (헤더 ID)
        control_code: 통제 코드

    Returns:
        dict: {
            'population_path': '...',
            'samples_path': '...',
            'test_results_path': '...'
        }
        파일이 없으면 해당 키의 값은 None
    """
    base_dir = os.path.join('static', 'uploads', 'operation_evaluations',
                            str(rcm_id), str(design_session), control_code)

    result = {
        'population_path': None,
        'samples_path': None,
        'test_results_path': None
    }

    if not os.path.exists(base_dir):
        return result

    # 모집단 파일 경로
    for filename in os.listdir(base_dir):
        if filename.startswith('population') and filename.endswith(('.xlsx', '.xls', '.csv')):
            result['population_path'] = f"uploads/operation_evaluations/{rcm_id}/{design_session}/{control_code}/{filename}"
            break

    # 표본 데이터 경로
    if os.path.exists(os.path.join(base_dir, 'samples.json')):
        result['samples_path'] = f"uploads/operation_evaluations/{rcm_id}/{design_session}/{control_code}/samples.json"

    # 테스트 결과 경로
    if os.path.exists(os.path.join(base_dir, 'test_results.json')):
        result['test_results_path'] = f"uploads/operation_evaluations/{rcm_id}/{design_session}/{control_code}/test_results.json"

    return result


def get_test_results_path(rcm_id, operation_header_id, control_code):
    """
    테스트 결과 엑셀 파일 경로 반환 (저장용)

    Args:
        rcm_id: RCM ID
        operation_header_id: 운영평가 헤더 ID
        control_code: 통제 코드

    Returns:
        str: 테스트 결과 엑셀 파일의 절대 경로
    """
    base_dir = os.path.join('static', 'uploads', 'operation_evaluations',
                            str(rcm_id), str(operation_header_id), control_code)

    # 디렉토리 생성 (여기서는 생성하지 않고 경로만 반환)
    filename = f'test_results_{control_code}.xlsx'
    return os.path.join(base_dir, filename)
