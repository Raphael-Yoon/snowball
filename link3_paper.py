from flask import Flask, request
import openpyxl.utils
from werkzeug.utils import secure_filename
import os
import openpyxl
import random
import datetime

def paper_template_download(form_data):
    print("Paper Template called")

    param1 = form_data.get('param1')
    param2 = form_data.get('param2')
    param3 = form_data.get('param3')
    param4 = form_data.get('param4')
    param5 = form_data.get('param5')

    print("Param1 = ", param1)
    print("Param2 = ", param2)
    print("Param3 = ", param3)
    print("Param4 = ", param4)
    print("Param5 = ", param5)

    output_path = './paper_templates/' + param2 + '_template.xlsx'

    print("output = ", output_path)

    return output_path

def paper_generate(form_data):
    print("Paper Generate called")

    param3 = form_data.get('param3')
    param4 = form_data.get('param4')
    param5 = form_data.get('param5')

    print("Param3 = ", param3)
    print("Param4 = ", param4)
    print("Param5 = ", param5)

    uploaded_file = request.files['param4']
    filename = secure_filename(uploaded_file.filename)
    file_path = os.path.join('uploads', uploaded_file.filename)
    uploaded_file.save(file_path)
    print('upload complete: ', param3)

    if(param3=="APD01"):
        output_path = paper_generate_population(param3, file_path)
        output_path = paper_generate_apd01(output_path)
    elif(param3=="APD02"):
        output_path = paper_generate_population(param3, file_path)
        output_path = paper_generate_apd02(output_path)
    elif(param3=="APD03"):
        output_path = paper_generate_population(param3, file_path)
        output_path = paper_generate_apd03(output_path)
    elif(param3=="APD07"):
        output_path = paper_generate_population(param3, file_path)
        output_path = paper_generate_apd07(output_path)
    elif(param3=="APD09"):
        output_path = paper_generate_population(param3, file_path)
        output_path = paper_generate_apd09(output_path)
    elif(param3=="APD12"):
        output_path = paper_generate_population(param3, file_path)
        output_path = paper_generate_apd12(output_path)
    else:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook['모집단']
        a1_value = sheet['A1'].value
        print("A1 = ", a1_value)
            
        #output_path = './paper_templates/' + param2 + '.xlsx'
        output_path = ''
        workbook.save(output_path)
        workbook.close()

    print("output = ", output_path)

    return output_path

def get_sample_size(max_row, risk_level):
    print("into_sample_size")

    sample_size = 0
    if(risk_level == "LOW"):
        if(max_row > 250): # multiple
            sample_size = 25
        elif(max_row > 52): # daily
            sample_size = 20
        elif(max_row > 12): # weekly
            sample_size = 5
        elif(max_row > 4): # monthly
            sample_size = 2
        elif(max_row > 1): # quarterly
            sample_size = 2
        elif(max_row == 1):
            sample_size = 1 # annually
        else:
            sample_size = 0
    elif(risk_level == "MIDDLE"):
        if(max_row > 250): # multiple
            sample_size = 25
        elif(max_row > 52): # daily
            sample_size = 20
        elif(max_row > 12): # weekly
            sample_size = 5
        elif(max_row > 4): # monthly
            sample_size = 2
        elif(max_row > 1): # quarterly
            sample_size = 2
        elif(max_row == 1): # annually
            sample_size = 1
        else:
            sample_size = 0
    elif(risk_level == "HIGH"):
        if(max_row > 250): # multiple
            sample_size = 25
        elif(max_row > 52): # daily
            sample_size = 20
        elif(max_row > 12): # weekly
            sample_size = 5
        elif(max_row > 4): # monthly
            sample_size = 2
        elif(max_row > 1): # quarterly
            sample_size = 2
        elif(max_row == 1): # annually
            sample_size = 1
        else:
            sample_size = 0

    return sample_size

def paper_generate_population(control_code, upload_file):
    print("into paper generate population")
    
    open_path = './paper_templates/' + control_code + '_paper.xlsx'
    print("open_path = ", open_path)
    
    paper_workbook = openpyxl.load_workbook(open_path)
    paper_sheet = paper_workbook['Population']

    upload_workbook = openpyxl.load_workbook(upload_file)
    upload_sheet = upload_workbook['모집단']

    for row in upload_sheet.iter_rows():
        for cell in row:
            paper_sheet[cell.coordinate].value = cell.value
    
    output_path = './uploads/' + control_code + '_paper.xlsx'
    print("output_path = ", output_path)
    paper_workbook.save(output_path)
    paper_workbook.close()
    
    print("end papaer generate population")
    
    return output_path

# Application 권한 부여 승인
def paper_generate_apd01(output_path):
    print("into paper generate apd01")

    paper_workbook = openpyxl.load_workbook(output_path)
    sheet_pop = paper_workbook['Population']
    sheet_test = paper_workbook['Testing Table']

    max_row = sheet_pop.max_row-1
    sample_size = get_sample_size(max_row, 'LOW')
    print("sample Size = ", sample_size)

    random_row_indices = random.sample(range(1, max_row), sample_size)
    selected_row_data = []

    if sample_size == 0:
        sheet_test.delete_rows(6, 29)
        sheet_test.delete_cols(10, 20)
        sheet_test.merge_cells("C5:G5")
        sheet_test["C5"] = '당기 발생건이 존재하지 않음'
    else:
        sheet_pop.insert_cols(1)
        sheet_pop['A1'] = "Sample"
        i=1
        for row_index in random_row_indices:
            row_data = [cell.value for cell in sheet_pop[row_index]]
            selected_row_data.append(row_data)
            sheet_pop['A'+str(row_index)] = "#" + str(i).zfill(2)
            i=i+1
        print("Random index = ", random_row_indices)
        print("Selected:")
        i=5
        for row_data in selected_row_data:
            print(row_data)
            sheet_test["C" + str(i)] = str(row_data[1]) # 사용자ID
            sheet_test["D" + str(i)] = str(row_data[2]) # 사용자명
            sheet_test["E" + str(i)] = str(row_data[3]) # 부서명
            sheet_test["F" + str(i)] = str(row_data[4]) # 권한명
            if isinstance(row_data[5], datetime.date):
                sheet_test["G" + str(i)] = str(row_data[5].date()) # 권한부여일
            else:
                sheet_test["G" + str(i)] = str(row_data[5]) # 권한부여일
            i=i+1
   
    paper_workbook.save(output_path)
    paper_workbook.close()

    print("end papaer generate apd01")
    
    return output_path

# 부서이동자 권한 회수
def paper_generate_apd02(output_path):
    print("into paper generate apd02")

    paper_workbook = openpyxl.load_workbook(output_path)
    sheet_pop = paper_workbook['Population']
    sheet_test = paper_workbook['Testing Table']

    max_row = sheet_pop.max_row-1
    print("max row = ", max_row)
    sample_size = get_sample_size(max_row, 'LOW')
    print("sample Size = ", sample_size)

    random_row_indices = random.sample(range(1, max_row), sample_size)
    selected_row_data = []

    if sample_size == 0:
        sheet_test.delete_rows(6, 29)
        sheet_test.delete_cols(11, 13)
        sheet_test.merge_cells("C5:H5")
        sheet_test["C5"] = '당기 발생건이 존재하지 않음'
    else:
        sheet_pop.insert_cols(1)
        sheet_pop['A1'] = "Sample"
        i=1
        for row_index in random_row_indices:
            row_data = [cell.value for cell in sheet_pop[row_index]]
            selected_row_data.append(row_data)
            sheet_pop['A'+str(row_index)] = "#" + str(i).zfill(2)
            i=i+1
        print("Random index = ", random_row_indices)
        print("Selected:")
        i=5
        for row_data in selected_row_data:
            print(row_data)
            sheet_test["C" + str(i)] = str(row_data[1]) # 사번
            sheet_test["D" + str(i)] = str(row_data[2]) # 이름
            sheet_test["E" + str(i)] = str(row_data[3]) # 부서명
            sheet_test["F" + str(i)] = str(row_data[4]) # 이전 부서명
            if isinstance(row_data[5], datetime.date):
                sheet_test["G" + str(i)] = str(row_data[5].date()) # 부서 이동일
            else:
                sheet_test["G" + str(i)] = str(row_data[5]) # 부서 이동일
            sheet_test["H" + str(i)] = str(row_data[6]) # 시스템ID
            i=i+1
    
    paper_workbook.save(output_path)
    paper_workbook.close()

    print("end papaer generate apd02")
    
    return output_path

# 퇴사자 계정 삭제
def paper_generate_apd03(output_path):
    print("into paper generate apd03")

    paper_workbook = openpyxl.load_workbook(output_path)
    sheet_pop = paper_workbook['Population']
    sheet_test = paper_workbook['Testing Table']

    max_row = sheet_pop.max_row-1
    print("max row = ", max_row)
    sample_size = get_sample_size(max_row, 'LOW')
    print("sample Size = ", sample_size)

    random_row_indices = random.sample(range(1, max_row), sample_size)
    selected_row_data = []

    if sample_size == 0:
        sheet_test.delete_rows(6, 29)
        sheet_test.delete_cols(9, 11)
        sheet_test.merge_cells("C5:F5")
        sheet_test["C5"] = '당기 발생건이 존재하지 않음'
    else:
        sheet_pop.insert_cols(1)
        sheet_pop['A1'] = "Sample"
        i=1
        for row_index in random_row_indices:
            row_data = [cell.value for cell in sheet_pop[row_index]]
            selected_row_data.append(row_data)
            sheet_pop['A'+str(row_index)] = "#" + str(i).zfill(2)
            i=i+1
        print("Random index = ", random_row_indices)
        print("Selected:")
        i=5
        for row_data in selected_row_data:
            print(row_data)
            sheet_test["C" + str(i)] = str(row_data[1]) # 사번
            sheet_test["D" + str(i)] = str(row_data[2]) # 이름
            if isinstance(row_data[3], datetime.date):
                sheet_test["E" + str(i)] = str(row_data[3].date()) # 퇴직일
            else:
                sheet_test["E" + str(i)] = str(row_data[3]) # 퇴직일
            sheet_test["F" + str(i)] = str(row_data[4]) # 시스템ID
            i=i+1
    
    paper_workbook.save(output_path)
    paper_workbook.close()

    print("end papaer generate apd03")
    
    return output_path

# Data 직접변경 승인
def paper_generate_apd07(output_path):
    print("into paper generate apd07")

    paper_workbook = openpyxl.load_workbook(output_path)
    sheet_pop = paper_workbook['Population']
    sheet_test = paper_workbook['Testing Table']

    max_row = sheet_pop.max_row-1
    print("max row = ", max_row)
    sample_size = get_sample_size(max_row, 'LOW')
    print("sample Size = ", sample_size)

    random_row_indices = random.sample(range(1, max_row), sample_size)
    selected_row_data = []

    if sample_size == 0:
        sheet_test.delete_rows(6, 29)
        sheet_test.delete_cols(8, 17)
        sheet_test.merge_cells("C5:E5")
        sheet_test["C5"] = '당기 발생건이 존재하지 않음'
    else:
        sheet_pop.insert_cols(1)
        sheet_pop['A1'] = "Sample"
        i=1
        for row_index in random_row_indices:
            row_data = [cell.value for cell in sheet_pop[row_index]]
            selected_row_data.append(row_data)
            sheet_pop['A'+str(row_index)] = "#" + str(i).zfill(2)
            i=i+1 
        print("Random index = ", random_row_indices)
        print("Selected:")
        i=5
        for row_data in selected_row_data:
            print(row_data)
            sheet_test["C" + str(i)] = str(row_data[1]) # Data 변경
            sheet_test["D" + str(i)] = str(row_data[2]) # 실행자
            if isinstance(row_data[3], datetime.date):
                sheet_test["E" + str(i)] = str(row_data[3].date()) # 실행일자
            else:
                sheet_test["E" + str(i)] = str(row_data[3]) # 실행일자
            i=i+1
    
    paper_workbook.save(output_path)
    paper_workbook.close()

    print("end papaer generate apd07")
    
    return output_path

# DB 접근권한 승인
def paper_generate_apd09(output_path):
    print("into paper generate apd09")

    paper_workbook = openpyxl.load_workbook(output_path)
    sheet_pop = paper_workbook['Population']
    sheet_test = paper_workbook['Testing Table']

    max_row = sheet_pop.max_row-1
    print("max row = ", max_row)
    sample_size = get_sample_size(max_row, 'LOW')
    print("sample Size = ", sample_size)

    random_row_indices = random.sample(range(1, max_row), sample_size)
    selected_row_data = []

    if sample_size == 0:
        sheet_test.delete_rows(6, 29)
        sheet_test.delete_cols(7, 16)
        sheet_test.merge_cells("C5:D5")
        sheet_test["C5"] = '당기 발생건이 존재하지 않음'
    else:
        sheet_pop.insert_cols(1)
        sheet_pop['A1'] = "Sample"
        i=1
        for row_index in random_row_indices:
            row_data = [cell.value for cell in sheet_pop[row_index]]
            selected_row_data.append(row_data)
            sheet_pop['A'+str(row_index)] = "#" + str(i).zfill(2)
            i=i+1 
        print("Random index = ", random_row_indices)
        print("Selected:")
        i=5
        for row_data in selected_row_data:
            print(row_data)
            sheet_test["C" + str(i)] = str(row_data[1]) # 접근권한ID
            if isinstance(row_data[3], datetime.date):
                sheet_test["D" + str(i)] = str(row_data[3].date()) # 생성일자
            else:
                sheet_test["D" + str(i)] = str(row_data[3]) # 생성일자
            i=i+1
    
    paper_workbook.save(output_path)
    paper_workbook.close()

    print("end papaer generate apd09")
    
    return output_path

# OS 접근권한 승인
def paper_generate_apd12(output_path):
    print("into paper generate apd12")

    paper_workbook = openpyxl.load_workbook(output_path)
    sheet_pop = paper_workbook['Population']
    sheet_test = paper_workbook['Testing Table']

    max_row = sheet_pop.max_row-1
    print("max row = ", max_row)
    sample_size = get_sample_size(max_row, 'LOW')
    print("sample Size = ", sample_size)

    random_row_indices = random.sample(range(1, max_row), sample_size)
    selected_row_data = []

    if sample_size == 0:
        sheet_test.delete_rows(6, 29)
        sheet_test.delete_cols(7, 16)
        sheet_test.merge_cells("C5:D5")
        sheet_test["C5"] = '당기 발생건이 존재하지 않음'
    else:
        sheet_pop.insert_cols(1)
        sheet_pop['A1'] = "Sample"
        i=1
        for row_index in random_row_indices:
            row_data = [cell.value for cell in sheet_pop[row_index]]
            selected_row_data.append(row_data)
            sheet_pop['A'+str(row_index)] = "#" + str(i).zfill(2)
            i=i+1 
        print("Random index = ", random_row_indices)
        print("Selected:")
        i=5
        for row_data in selected_row_data:
            print(row_data)
            sheet_test["C" + str(i)] = str(row_data[1]) # 접근권한ID
            if isinstance(row_data[3], datetime.date):
                sheet_test["D" + str(i)] = str(row_data[3].date()) # 생성일자
            else:
                sheet_test["D" + str(i)] = str(row_data[3]) # 생성일자
            i=i+1
    
    paper_workbook.save(output_path)
    paper_workbook.close()

    print("end papaer generate apd12")
    
    return output_path