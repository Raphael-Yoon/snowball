"""
통합 운영평가 API (Generic Implementation)
모든 수동통제(APD01, APD07, APD09, APD12 등)에 대해 공통으로 사용
"""

from flask import Blueprint, request, jsonify, render_template, session, redirect, url_for, flash
from auth import login_required, get_db
from snowball_link5 import get_user_info
import file_manager
import json
import tempfile
import os
from control_config import get_control_config, is_manual_control
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

bp_generic = Blueprint('operation_evaluation_generic', __name__)


@bp_generic.route('/operation-evaluation/manual/<control_code>')
@login_required
def manual_control_evaluation(control_code):
    """
    통합 운영평가 페이지
    모든 수동통제에 대해 공통으로 사용
    """
    user_info = get_user_info()

    # 통제 설정 확인
    config = get_control_config(control_code)
    if not config:
        flash(f'지원하지 않는 통제 코드입니다: {control_code}', 'error')
        return redirect(url_for('link7.user_operation_evaluation_index'))

    # URL 파라미터
    rcm_id = request.args.get('rcm_id')
    control_code_param = request.args.get('control_code', control_code)
    control_name = request.args.get('control_name', config['name'])
    design_evaluation_session = request.args.get('design_evaluation_session')

    if not all([rcm_id, design_evaluation_session]):
        flash('필수 정보가 누락되었습니다.', 'error')
        return redirect(url_for('link7.user_operation_evaluation_index'))

    # 기존 운영평가 데이터 조회
    operation_evaluation_session = f"OP_{design_evaluation_session}"
    existing_data = None
    pc01_data = None  # PC02용 PC01 데이터

    try:
        with get_db() as conn:
            # 운영평가 헤더 조회
            header = conn.execute('''
                SELECT header_id FROM sb_operation_evaluation_header
                WHERE rcm_id = ? AND user_id = ? AND evaluation_session = ? AND design_evaluation_session = ?
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if header:
                operation_header_id = header['header_id']

                # 기존 데이터 로드
                loaded_data = file_manager.load_operation_test_data(
                    rcm_id=rcm_id,
                    operation_header_id=operation_header_id,
                    control_code=control_code_param
                )

                if loaded_data and loaded_data.get('samples_data'):
                    existing_data = loaded_data

                # 데이터베이스에서 no_occurrence 정보 추가 로드
                if existing_data:
                    eval_line = conn.execute('''
                        SELECT operating_effectiveness, conclusion, no_occurrence, no_occurrence_reason
                        FROM sb_operation_evaluation_line
                        WHERE header_id = ? AND control_code = ?
                        ORDER BY line_id DESC
                        LIMIT 1
                    ''', (operation_header_id, control_code_param)).fetchone()

                    if eval_line:
                        existing_data['no_occurrence'] = eval_line['no_occurrence'] if eval_line['no_occurrence'] else False
                        existing_data['no_occurrence_reason'] = eval_line['no_occurrence_reason'] or ''
                        existing_data['operating_effectiveness'] = eval_line['operating_effectiveness']
                        existing_data['conclusion'] = eval_line['conclusion']
                elif not loaded_data:
                    # 엑셀 파일은 없지만 DB에 no_occurrence 데이터가 있을 수 있음
                    eval_line = conn.execute('''
                        SELECT operating_effectiveness, conclusion, no_occurrence, no_occurrence_reason
                        FROM sb_operation_evaluation_line
                        WHERE header_id = ? AND control_code = ?
                        ORDER BY line_id DESC
                        LIMIT 1
                    ''', (operation_header_id, control_code_param)).fetchone()

                    if eval_line and eval_line['no_occurrence']:
                        existing_data = {
                            'no_occurrence': True,
                            'no_occurrence_reason': eval_line['no_occurrence_reason'] or '',
                            'operating_effectiveness': eval_line['operating_effectiveness'],
                            'conclusion': eval_line['conclusion'],
                            'samples_data': None,
                            'population_data': []
                        }

                # PC02, PC03인 경우 PC01 데이터도 로드
                if control_code in ['PC02', 'PC03']:
                    pc01_loaded = file_manager.load_operation_test_data(
                        rcm_id=rcm_id,
                        operation_header_id=operation_header_id,
                        control_code='PC01'
                    )
                    if pc01_loaded and pc01_loaded.get('samples_data'):
                        pc01_data = pc01_loaded

                        # PC02/PC03 파일이 없으면 PC01 데이터로 파일 생성
                        if not existing_data:
                            # PC01의 표본 데이터로 엑셀 파일 생성
                            pc01_samples = pc01_data['samples_data']['samples']
                            pc01_population = pc01_data.get('population_data', [])

                            # PC02/PC03 파일 생성
                            file_manager.save_operation_test_data(
                                rcm_id=rcm_id,
                                operation_header_id=operation_header_id,
                                control_code=control_code,
                                population_data=pc01_population,
                                field_mapping={'program_name': '0', 'deploy_date': '1'},
                                samples=pc01_samples,
                                test_results_data=None  # 테스트 결과는 아직 없음
                            )

                            # 생성된 파일 로드
                            existing_data = file_manager.load_operation_test_data(
                                rcm_id=rcm_id,
                                operation_header_id=operation_header_id,
                                control_code=control_code
                            )
                    else:
                        flash('PC01 모집단 데이터가 없습니다. PC01을 먼저 수행해주세요.', 'error')
                        return redirect(url_for('link7.user_operation_evaluation_rcm',
                                              rcm_id=rcm_id,
                                              evaluation_session=design_evaluation_session))
    except Exception as e:
        print(f"기존 데이터 로드 오류: {e}")

    return render_template('link7_manual_generic.jsp',
                         rcm_id=rcm_id,
                         control_code=control_code_param,
                         control_name=control_name,
                         design_evaluation_session=design_evaluation_session,
                         config=config,
                         existing_data=existing_data,
                         pc01_data=pc01_data)


@bp_generic.route('/api/operation-evaluation/manual/<control_code>/upload-population', methods=['POST'])
@login_required
def upload_manual_population(control_code):
    """
    통합 모집단 업로드 API
    모든 수동통제에 대해 공통으로 사용
    """
    user_info = get_user_info()

    # 통제 설정 확인
    config = get_control_config(control_code)
    if not config:
        return jsonify({'success': False, 'message': f'지원하지 않는 통제 코드입니다: {control_code}'})

    # 파일 받기
    if 'population_file' not in request.files:
        return jsonify({'success': False, 'message': '파일이 없습니다.'})

    file = request.files['population_file']
    if not file.filename:
        return jsonify({'success': False, 'message': '파일을 선택해주세요.'})

    # 필드 매핑 정보 받기
    field_mapping_str = request.form.get('field_mapping')
    if not field_mapping_str:
        return jsonify({'success': False, 'message': '필드 매핑 정보가 없습니다.'})

    try:
        field_mapping = json.loads(field_mapping_str)
    except:
        return jsonify({'success': False, 'message': '필드 매핑 형식이 올바르지 않습니다.'})

    # RCM 정보
    rcm_id = request.form.get('rcm_id')
    design_evaluation_session = request.form.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        return jsonify({'success': False, 'message': '필수 정보가 누락되었습니다.'})

    try:
        # 운영평가 헤더 조회
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        with get_db() as conn:
            header = conn.execute('''
                SELECT header_id FROM sb_operation_evaluation_header
                WHERE rcm_id = ? AND user_id = ? AND evaluation_session = ? AND design_evaluation_session = ?
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if not header:
                return jsonify({'success': False, 'message': '운영평가 세션을 찾을 수 없습니다. RCM 페이지에서 다시 시작해주세요.'})

            operation_header_id = header['header_id']

        # 임시 파일로 저장
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()

        # 모집단 파싱 (통합 함수 사용)
        population_data = file_manager.parse_population_excel(temp_file.name, field_mapping)
        count = len(population_data)
        sample_size = file_manager.calculate_sample_size(count)

        # 표본 선택
        samples = file_manager.select_random_samples(population_data, sample_size)

        # 임시 파일 삭제 (Windows에서 파일 핸들 문제로 실패할 수 있으므로 무시)
        try:
            os.unlink(temp_file.name)
        except Exception as e:
            print(f"임시 파일 삭제 실패 (무시됨): {e}")

        # 템플릿 기반 엑셀 파일 생성 및 저장
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=population_data,
            field_mapping=field_mapping,
            samples=samples,
            test_results_data=None
        )

        # 세션에 정보 저장
        session_key = f'manual_test_{rcm_id}_{control_code}'
        session[session_key] = {
            'file_paths': file_paths,
            'rcm_id': rcm_id,
            'control_code': control_code,
            'design_evaluation_session': design_evaluation_session,
            'operation_header_id': operation_header_id,
            'field_mapping': field_mapping,
            'population_count': count,
            'sample_size': sample_size
        }

        return jsonify({
            'success': True,
            'population_count': count,
            'sample_size': sample_size,
            'samples': samples
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


@bp_generic.route('/api/operation-evaluation/manual/<control_code>/save-test-results', methods=['POST'])
@login_required
def save_manual_test_results(control_code):
    """
    통합 테스트 결과 저장 API
    모든 수동통제에 대해 공통으로 사용
    """
    user_info = get_user_info()

    # 통제 설정 확인
    config = get_control_config(control_code)
    if not config:
        return jsonify({'success': False, 'message': f'지원하지 않는 통제 코드입니다: {control_code}'})

    try:
        data = request.json
        rcm_id = data.get('rcm_id')
        design_evaluation_session = data.get('design_evaluation_session')
        test_results = data.get('test_results', [])

        # 세션에서 정보 가져오기
        session_key = f'manual_test_{rcm_id}_{control_code}'
        session_data = session.get(session_key)
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        # 세션이 없으면 DB에서 operation_header_id 조회
        if not session_data:
            with get_db() as conn:
                header = conn.execute('''
                    SELECT header_id FROM sb_operation_evaluation_header
                    WHERE rcm_id = ? AND user_id = ? AND evaluation_session = ? AND design_evaluation_session = ?
                ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

                if not header:
                    return jsonify({'success': False, 'message': '운영평가 세션을 찾을 수 없습니다.'})

                operation_header_id = header['header_id']
        else:
            operation_header_id = session_data['operation_header_id']

        # 저장된 파일에서 표본 데이터 로드
        loaded_data = file_manager.load_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code
        )

        if not loaded_data or not loaded_data['samples_data']:
            return jsonify({'success': False, 'message': '저장된 표본 데이터를 찾을 수 없습니다.'})

        samples_data = loaded_data['samples_data']

        # field_mapping 가져오기 (세션 또는 기본 매핑)
        if session_data and 'field_mapping' in session_data:
            field_mapping = session_data['field_mapping']
        else:
            # 기본 field_mapping 생성 (인덱스 기반)
            config = get_control_config(control_code)
            field_mapping = {field: str(idx) for idx, field in enumerate(config['population_fields'])}

        # 엑셀 파일에 테스트 결과 저장
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=loaded_data.get('population_data', []),
            field_mapping=field_mapping,
            samples=samples_data['samples'],
            test_results_data={
                'test_results': test_results,
                'exceptions': [r for r in test_results if r.get('has_exception')],
                'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception',
                'test_type': control_code
            }
        )

        # 운영평가 데이터 저장
        from auth import save_operation_evaluation
        evaluation_data = {
            'test_type': control_code,
            'population_count': samples_data['population_count'],
            'sample_size': samples_data['sample_size'],
            'population_path': None,
            'samples_path': file_paths.get('samples_path'),
            'test_results_path': file_paths.get('excel_path'),
            'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception'
        }

        save_operation_evaluation(rcm_id, control_code, user_info['user_id'],
                                 operation_evaluation_session, design_evaluation_session, evaluation_data)

        from auth import log_user_activity
        log_user_activity(user_info, 'OPERATION_EVALUATION', f'{control_code} 테스트 저장',
                         f'/api/operation-evaluation/manual/{control_code}/save-test-results',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({'success': True, 'message': f'{control_code} 테스트 결과가 저장되었습니다.'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'저장 오류: {str(e)}'})


@bp_generic.route('/api/operation-evaluation/save-no-occurrence', methods=['POST'])
@login_required
def save_no_occurrence():
    """
    당기 발생사실 없음 저장 API
    모든 수동통제에 대해 공통으로 사용
    """
    user_info = get_user_info()

    try:
        data = request.get_json()
        rcm_id = data.get('rcm_id')
        control_code = data.get('control_code')
        design_evaluation_session = data.get('design_evaluation_session')
        no_occurrence_reason = data.get('no_occurrence_reason', '')

        if not all([rcm_id, control_code, design_evaluation_session]):
            return jsonify({'success': False, 'message': '필수 정보가 누락되었습니다.'})

        # 운영평가 헤더 조회 또는 생성
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        with get_db() as conn:
            header = conn.execute('''
                SELECT header_id FROM sb_operation_evaluation_header
                WHERE rcm_id = ? AND user_id = ? AND evaluation_session = ? AND design_evaluation_session = ?
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if not header:
                # 헤더 생성
                cursor = conn.execute('''
                    INSERT INTO sb_operation_evaluation_header
                    (rcm_id, user_id, evaluation_session, design_evaluation_session, created_date)
                    VALUES (?, ?, ?, ?, datetime('now'))
                ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session))
                operation_header_id = cursor.lastrowid
                conn.commit()
            else:
                operation_header_id = header['header_id']

        # 통제 설정 조회
        config = get_control_config(control_code)
        if not config:
            config = get_control_config('GENERIC')  # GENERIC 설정 사용

        # "당기 발생사실 없음" 엑셀 파일 생성
        wb = Workbook()

        # Population 시트
        ws_pop = wb.active
        ws_pop.title = 'Population'

        # 헤더 스타일
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        # Population 시트에 "당기 발생사실 없음" 메시지
        ws_pop['A1'] = '당기 발생사실 없음'
        ws_pop['A1'].font = Font(bold=True, size=14, color='FF0000')
        ws_pop['A1'].alignment = Alignment(horizontal='center', vertical='center')

        if no_occurrence_reason:
            ws_pop['A2'] = f'사유: {no_occurrence_reason}'
            ws_pop['A2'].font = Font(size=11)

        ws_pop.column_dimensions['A'].width = 50

        # Test Table 시트
        ws_test = wb.create_sheet('Test Table')

        # 헤더 행 작성
        excel_headers = config['excel_headers']['testing']
        for col_idx, header in enumerate(excel_headers, 1):
            cell = ws_test.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 한 줄의 빈 데이터 행 추가 (No = 1)
        ws_test.cell(row=2, column=1, value=1)  # No
        for col_idx in range(2, len(excel_headers) + 1):
            ws_test.cell(row=2, column=col_idx, value='')

        # 비고 컬럼에 "당기 발생사실 없음" 표시
        remark_col_idx = len(excel_headers)  # 마지막 컬럼이 비고
        ws_test.cell(row=2, column=remark_col_idx, value='당기 발생사실 없음')

        # 컬럼 너비 자동 조정
        for col_idx, header in enumerate(excel_headers, 1):
            ws_test.column_dimensions[ws_test.cell(row=1, column=col_idx).column_letter].width = 15

        # 엑셀 파일 저장
        excel_path = file_manager.get_test_results_path(rcm_id, operation_header_id, control_code)
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        wb.save(excel_path)

        # 운영평가 데이터 저장 (당기 발생사실 없음)
        evaluation_data = {
            'operating_effectiveness': 'not_applicable',
            'sample_size': 0,
            'exception_count': 0,
            'exception_details': '',
            'conclusion': 'not_applicable',
            'improvement_plan': '',
            'no_occurrence': True,
            'no_occurrence_reason': no_occurrence_reason,
            'population_count': 0,
            'population_path': None,
            'samples_path': None,
            'test_results_path': excel_path
        }

        from snowball_link7 import save_operation_evaluation
        save_operation_evaluation(rcm_id, control_code, user_info['user_id'],
                                 operation_evaluation_session, design_evaluation_session, evaluation_data)

        from auth import log_user_activity
        log_user_activity(user_info, 'OPERATION_EVALUATION', f'{control_code} 당기 발생사실 없음 저장',
                         '/api/operation-evaluation/save-no-occurrence',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({'success': True, 'message': '당기 발생사실 없음이 저장되었습니다.'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'저장 오류: {str(e)}'})
