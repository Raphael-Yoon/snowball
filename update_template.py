#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""운영평가 템플릿 파일의 행 레이블 업데이트"""

from openpyxl import load_workbook

template_path = r'c:\Pythons\snowball\paper_templates\Template_Manual.xlsx'
temp_output_path = r'c:\Pythons\snowball\paper_templates\Template_Manual_new.xlsx'

# 워크북 로드
wb = load_workbook(template_path)
ws = wb['Template']

# 행 레이블 업데이트
print('수정 전:')
print(f'  Row 12: {ws["B12"].value}')
print(f'  Row 13: {ws["B13"].value}')
print(f'  Row 14: {ws["B14"].value}')

ws['B13'] = '검토 결과(운영평가 의견)'
ws['B14'] = '결론(운영평가)'

print('\n수정 후:')
print(f'  Row 12: {ws["B12"].value}')
print(f'  Row 13: {ws["B13"].value}')
print(f'  Row 14: {ws["B14"].value}')

# 새 파일로 저장
wb.save(temp_output_path)
print(f'\n[OK] 템플릿 파일 생성 완료: {temp_output_path}')

# 기존 파일 삭제 및 이름 변경
import os
import shutil

try:
    if os.path.exists(template_path):
        os.remove(template_path)
        print(f'[OK] 기존 파일 삭제: {template_path}')

    shutil.move(temp_output_path, template_path)
    print(f'[OK] 새 파일로 교체 완료: {template_path}')
except Exception as e:
    print(f'[ERROR] 파일 교체 실패: {e}')
    print(f'  수정된 파일은 {temp_output_path}에 저장되었습니다.')
    print(f'  수동으로 파일을 교체해주세요.')
