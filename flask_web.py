from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
import openpyxl

app = Flask(__name__)

@app.route('/')
def index():
    return render_template("index.html")

def main():
    app.run(host='0.0.0.0', debug=False, port=8001)
    #app.run(host='127.0.0.1', debug=False, port=8001)

@app.route('/generate', methods=['POST'])
def calculate():
    print("Python function called")
    form_data = request.form.to_dict()

    param1 = form_data.get('param1')
    param2 = form_data.get('param2')
    param3 = form_data.get('param3')
    param4 = form_data.get('param4')

    print("Param1 = ", param1)
    print("Param2 = ", param2)
    print("Param3 = ", param3)
    print("Param4 = ", param4)

    output_path = 'pbc.xlsx' 
    workbook = openpyxl.load_workbook('PBC_template.xlsx')


    if param1 == 'SAP':
        sheets_to_delete = ['APD01_Oracle', 'APD06_Oracle']
    elif param1 == 'Oracle':
        sheets_to_delete = ['APD01_SAP', 'APD06_SAP']

    # 시트 삭제
    for sheet_name in sheets_to_delete:
        if sheet_name in workbook.sheetnames:
            sheet_to_delete = workbook[sheet_name]
            workbook.remove(sheet_to_delete)
    
    # 시트명 변경
    for sheet in workbook:
        if '_' in sheet.title:
            index = sheet.title.index('_')
            sheet.title = sheet.title[:index]

    workbook.save(output_path)
    workbook.close()

    return send_file(output_path, as_attachment=True)

if __name__ == '__main__':
    main()