from flask import Blueprint, request, jsonify, render_template, redirect, url_for, flash, session
from auth import login_required, get_current_user, get_user_rcms, get_rcm_details, save_design_evaluation, get_design_evaluations, get_design_evaluations_by_header_id, get_user_evaluation_sessions, delete_evaluation_session, create_evaluation_structure, log_user_activity, get_db, get_or_create_evaluation_header, get_rcm_detail_mappings, archive_design_evaluation_session, unarchive_design_evaluation_session
from snowball_link5 import get_user_info, is_logged_in
import sys
import os
import json
from werkzeug.utils import secure_filename
from flask import send_file
import tempfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles.colors import Color

bp_link6 = Blueprint('link6', __name__)

# 설계평가 관련 기능들

@bp_link6.route('/design-evaluation')
@login_required
def user_design_evaluation():
    """설계평가 페이지"""
    user_info = get_user_info()
    
    log_user_activity(user_info, 'PAGE_ACCESS', '설계평가', '/user/design-evaluation', 
                     request.remote_addr, request.headers.get('User-Agent'))
    
    return render_template('link6_design_evaluation.jsp',
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr)

@bp_link6.route('/design-evaluation/rcm', methods=['GET', 'POST'])
@login_required
def user_design_evaluation_rcm():
    """사용자 디자인 평가 페이지 (RCM별)"""
    user_info = get_user_info()

    # POST로 전달된 RCM ID 받기 또는 세션에서 가져오기
    if request.method == 'POST':
        rcm_id = request.form.get('rcm_id')
        evaluation_type = request.form.get('evaluation_type', 'ITGC')  # 기본값 ITGC
        evaluation_session = request.form.get('session')  # 평가 세션
        if not rcm_id:
            flash('RCM 정보가 없습니다.', 'error')
            return redirect(url_for('link6.user_design_evaluation'))

        # 세션에 저장
        session['current_design_rcm_id'] = int(rcm_id)
        session['current_evaluation_type'] = evaluation_type
        if evaluation_session:
            session['current_evaluation_session'] = evaluation_session
    else:
        # GET 요청인 경우 URL 파라미터 우선, 없으면 세션에서 가져오기
        rcm_id = request.args.get('rcm_id')
        evaluation_session = request.args.get('session')

        if rcm_id:
            # URL 파라미터가 있으면 세션에 저장
            rcm_id = int(rcm_id)
            session['current_design_rcm_id'] = rcm_id
            if evaluation_session:
                session['current_evaluation_session'] = evaluation_session

            # RCM의 실제 control_category를 조회하여 evaluation_type 설정
            with get_db() as conn:
                rcm_category = conn.execute('''
                    SELECT control_category FROM sb_rcm WHERE rcm_id = %s
                ''', (rcm_id,)).fetchone()

                if rcm_category and rcm_category['control_category']:
                    evaluation_type = rcm_category['control_category']
                    session['current_evaluation_type'] = evaluation_type
                else:
                    evaluation_type = 'ITGC'  # 기본값
                    session['current_evaluation_type'] = evaluation_type
        else:
            # URL 파라미터가 없으면 세션에서 가져오기
            rcm_id = session.get('current_design_rcm_id')
            evaluation_session = session.get('current_evaluation_session')
            evaluation_type = session.get('current_evaluation_type', 'ITGC')

        if not rcm_id:
            flash('RCM 정보가 없습니다. 다시 선택해주세요.', 'error')
            return redirect(url_for('link6.user_design_evaluation'))

    # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
    if user_info.get('admin_flag') != 'Y':
        user_rcms = get_user_rcms(user_info['user_id'])
        rcm_ids = [rcm['rcm_id'] for rcm in user_rcms]

        if rcm_id not in rcm_ids:
            flash('해당 RCM에 대한 접근 권한이 없습니다.', 'error')
            return redirect(url_for('link6.user_design_evaluation'))
    else:
        # 관리자는 모든 RCM에 접근 가능
        user_rcms = get_user_rcms(user_info['user_id'])
    
    # RCM 정보 조회
    rcm_info = None
    if user_info.get('admin_flag') == 'Y':
        # 관리자인 경우 모든 RCM 정보를 조회
        with get_db() as conn:
            rcm_data = conn.execute('''
                SELECT r.rcm_id, r.rcm_name, r.description, r.upload_date, r.user_id,
                       u.user_name as owner_name, u.company_name, 'admin' as permission_type
                FROM sb_rcm r
                LEFT JOIN sb_user u ON r.user_id = u.user_id
                WHERE r.rcm_id = %s
            ''', (rcm_id,)).fetchone()
            if rcm_data:
                rcm_info = dict(rcm_data)
    else:
        # 일반 사용자인 경우 권한이 있는 RCM 중에서 조회
        for rcm in user_rcms:
            if rcm['rcm_id'] == rcm_id:
                rcm_info = rcm
                break
    
    # evaluation_type에 따라 자동으로 control_category 설정
    # ITGC, ELC, TLC는 해당 카테고리만 보여줌
    control_category = evaluation_type if evaluation_type in ['ITGC', 'ELC', 'TLC'] else None

    # RCM 세부 데이터 조회
    rcm_details = get_rcm_details(rcm_id, control_category=control_category)

    # 매핑 정보 조회
    rcm_mappings = get_rcm_detail_mappings(rcm_id)

    # 통제 카테고리별 통계
    with get_db() as conn:
        category_stats = conn.execute('''
            SELECT control_category, COUNT(*) as count
            FROM sb_rcm_detail
            WHERE rcm_id = %s
            GROUP BY control_category
            ORDER BY control_category
        ''', (rcm_id,)).fetchall()
        category_stats = {row['control_category']: row['count'] for row in category_stats}

    log_user_activity(user_info, 'PAGE_ACCESS', 'RCM 디자인 평가', '/design-evaluation/rcm',
                     request.remote_addr, request.headers.get('User-Agent'))

    return render_template('link6_design_rcm_detail.jsp',
                         rcm_id=rcm_id,
                         rcm_info=rcm_info,
                         rcm_details=rcm_details,
                         rcm_mappings=rcm_mappings,
                         control_category=control_category,
                         category_stats=category_stats,
                         evaluation_type=evaluation_type,
                         evaluation_session=evaluation_session,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr)

@bp_link6.route('/api/design-evaluation/get', methods=['GET'])
@login_required
def get_design_evaluation_api():
    """설계평가 결과 조회 API"""
    user_info = get_user_info()

    rcm_id = request.args.get('rcm_id')
    design_evaluation_session = request.args.get('design_evaluation_session')
    control_code = request.args.get('control_code')

    # control_code가 없으면 전체 설계평가 목록 조회
    if not control_code:
        if not all([rcm_id, design_evaluation_session]):
            return jsonify({'success': False, 'message': '필수 파라미터가 누락되었습니다.'}), 400

        try:
            # 전체 설계평가 목록 조회
            with get_db() as conn:
                evaluations = conn.execute('''
                    SELECT
                        l.control_code,
                        rd.control_name,
                        rd.control_description,
                        rd.control_frequency,
                        rd.control_type,
                        rd.control_nature,
                        rd.key_control,
                        l.description_adequacy as design_adequacy,
                        l.improvement_suggestion,
                        l.overall_effectiveness as conclusion,
                        l.evaluation_evidence,
                        l.evaluation_rationale,
                        l.recommended_actions,
                        l.evaluation_date
                    FROM sb_design_evaluation_line l
                    JOIN sb_design_evaluation_header h ON l.header_id = h.header_id
                    JOIN sb_rcm_detail_v rd ON h.rcm_id = rd.rcm_id AND l.control_code = rd.control_code
                    WHERE h.rcm_id = %s AND h.evaluation_session = %s
                    ORDER BY l.control_code
                ''', (rcm_id, design_evaluation_session)).fetchall()

                return jsonify({
                    'success': True,
                    'evaluations': [dict(row) for row in evaluations]
                })

        except Exception as e:
            print(f"설계평가 목록 조회 오류: {str(e)}")
            return jsonify({'success': False, 'message': str(e)}), 500

    # control_code가 있으면 단일 설계평가 조회
    if not all([rcm_id, design_evaluation_session, control_code]):
        return jsonify({'success': False, 'message': '필수 파라미터가 누락되었습니다.'}), 400

    try:
        # 단일 통제의 설계평가 데이터 조회
        with get_db() as conn:
            design_eval = conn.execute('''
                SELECT
                    l.line_id,
                    l.description_adequacy,
                    l.improvement_suggestion,
                    l.overall_effectiveness,
                    l.evaluation_evidence,
                    l.evaluation_rationale,
                    l.recommended_actions,
                    l.design_comment,
                    l.evaluation_date,
                    l.no_occurrence,
                    l.no_occurrence_reason,
                    h.header_id
                FROM sb_design_evaluation_line l
                JOIN sb_design_evaluation_header h ON l.header_id = h.header_id
                WHERE h.rcm_id = %s AND h.evaluation_session = %s AND l.control_code = %s
            ''', (rcm_id, design_evaluation_session, control_code)).fetchone()

            if design_eval:
                # 이미지 DB에서 조회
                images = conn.execute('''
                    SELECT image_id, file_path, file_name, file_size, uploaded_at
                    FROM sb_evaluation_image
                    WHERE evaluation_type = %s AND line_id = %s
                    ORDER BY uploaded_at
                ''', ('design', design_eval['line_id'])).fetchall()

                return jsonify({
                    'success': True,
                    'evaluation': dict(design_eval),
                    'images': [dict(img) for img in images],
                    'design_evaluation': {
                        'conclusion': design_eval['overall_effectiveness'],
                        'deficiency_details': design_eval['improvement_suggestion'],
                        'improvement_plan': design_eval['recommended_actions'],
                        'test_procedure': design_eval['evaluation_rationale'],
                        'evaluated_at': design_eval['evaluation_date']
                    }
                })
            else:
                return jsonify({'success': False, 'message': '설계평가 데이터를 찾을 수 없습니다.'}), 404

    except Exception as e:
        print(f"설계평가 조회 오류: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp_link6.route('/api/design-evaluation/save', methods=['POST'])
@login_required
def save_design_evaluation_api():
    """설계평가 결과 저장 API"""
    user_info = get_user_info()
    
    
    # FormData와 JSON 모두 처리
    if request.content_type and 'multipart/form-data' in request.content_type:
        # FormData 처리 (이미지 포함)
        rcm_id = request.form.get('rcm_id')
        control_code = request.form.get('control_code')
        evaluation_data_str = request.form.get('evaluation_data')
        evaluation_session = request.form.get('evaluation_session')

        # JSON 문자열을 파싱
        evaluation_data = json.loads(evaluation_data_str) if evaluation_data_str else None
        
        # 이미지 파일 처리
        uploaded_images = []
        for key in request.files:
            if key.startswith('evaluation_image_'):
                file = request.files[key]
                if file and file.filename:
                    uploaded_images.append(file)
    else:
        # 기존 JSON 처리
        data = request.get_json()
        rcm_id = data.get('rcm_id') if data else None
        control_code = data.get('control_code') if data else None
        evaluation_data = data.get('evaluation_data') if data else None
        evaluation_session = data.get('evaluation_session') if data else None
        uploaded_images = []

    # 필수 데이터 검증
    
    if not all([rcm_id, control_code, evaluation_data, evaluation_session]):
        missing_fields = []
        if not rcm_id: missing_fields.append('RCM ID')
        if not control_code: missing_fields.append('통제 코드')
        if not evaluation_data: missing_fields.append('평가 데이터')
        if not evaluation_session: missing_fields.append('세션명')
        
        error_msg = f'필수 데이터가 누락되었습니다: {", ".join(missing_fields)}'
        
        return jsonify({
            'success': False,
            'message': error_msg
        })
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()
                
                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })
        
        # 이미지 파일 저장
        saved_images = []
        if uploaded_images:
            # 헤더 ID 가져오기 (이미지 저장 폴더명으로 사용)
            with get_db() as conn:
                header_id = get_or_create_evaluation_header(conn, rcm_id, user_info['user_id'], evaluation_session)
            
            # 이미지 저장 디렉토리 생성 (header_id로 분리)
            upload_dir = os.path.join('static', 'uploads', 'design_evaluations', str(rcm_id), str(header_id), control_code)
            os.makedirs(upload_dir, exist_ok=True)
            
            for i, image_file in enumerate(uploaded_images):
                if image_file and image_file.filename:
                    # 디버깅 로그
                    
                    # 확장자를 먼저 분리
                    original_name, original_ext = os.path.splitext(image_file.filename)
                    
                    # 안전한 파일명 생성 (확장자 제외)
                    safe_name = secure_filename(original_name)
                    
                    # secure_filename이 빈 문자열을 반환하면 기본 이름 사용
                    if not safe_name or safe_name.strip() == '' or safe_name == '_':
                        safe_name = 'evaluation_image'
                    
                    # 중복 방지를 위해 타임스탬프 추가
                    import time
                    timestamp = str(int(time.time()))
                    
                    # 최종 파일명: 안전한이름_타임스탬프_인덱스.확장자
                    safe_filename = f"{safe_name}_{timestamp}_{i}{original_ext.lower()}"
                    
                    # 파일 저장
                    file_path = os.path.join(upload_dir, safe_filename)
                    image_file.save(file_path)

                    # 상대 경로로 저장 (DB 저장용)
                    relative_path = f"static/uploads/design_evaluations/{rcm_id}/{header_id}/{control_code}/{safe_filename}"

                    # 파일 크기
                    file_size = os.path.getsize(file_path)

                    saved_images.append({
                        'path': relative_path,
                        'name': safe_filename,
                        'size': file_size
                    })

        # 설계평가 결과 저장
        save_design_evaluation(rcm_id, control_code, user_info['user_id'], evaluation_data, evaluation_session)

        # 이미지를 DB에 저장
        if saved_images:
            with get_db() as conn:
                # line_id 조회
                line_record = conn.execute('''
                    SELECT l.line_id FROM sb_design_evaluation_line l
                    JOIN sb_design_evaluation_header h ON l.header_id = h.header_id
                    WHERE h.rcm_id = %s AND h.evaluation_session = %s AND l.control_code = %s
                ''', (rcm_id, evaluation_session, control_code)).fetchone()

                if line_record:
                    line_id = line_record['line_id']

                    for img_info in saved_images:
                        # 중복 체크
                        existing = conn.execute('''
                            SELECT image_id FROM sb_evaluation_image
                            WHERE evaluation_type = %s AND line_id = %s AND file_path = %s
                        ''', ('design', line_id, img_info['path'])).fetchone()

                        if not existing:
                            conn.execute('''
                                INSERT INTO sb_evaluation_image
                                (evaluation_type, line_id, file_path, file_name, file_size, uploaded_at)
                                VALUES (%s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
                            ''', ('design', line_id, img_info['path'], img_info['name'], img_info['size']))

                    conn.commit()
        
        
        # 활동 로그 기록
        log_user_activity(user_info, 'DESIGN_EVALUATION', f'설계평가 저장 - {control_code}', 
                         f'/api/design-evaluation/save', 
                         request.remote_addr, request.headers.get('User-Agent'))
        
        return jsonify({
            'success': True,
            'message': '새로운 설계평가 결과가 저장되었습니다.'
        })
        
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        return jsonify({
            'success': False,
            'message': f'저장 중 오류가 발생했습니다: {str(e)}'
        })

@bp_link6.route('/api/design-evaluation/reset', methods=['POST'])
@login_required
def reset_design_evaluations_api():
    """설계평가 결과 초기화 API"""
    user_info = get_user_info()
    data = request.get_json()
    
    rcm_id = data.get('rcm_id')
    
    if not rcm_id:
        return jsonify({
            'success': False,
            'message': 'RCM ID가 누락되었습니다.'
        })
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()
                
                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })
            
            # 해당 사용자의 모든 설계평가 결과 삭제 (Header-Line 구조)
            # 1. 먼저 line 레코드들 삭제
            conn.execute('''
                DELETE FROM sb_design_evaluation_line 
                WHERE header_id IN (
                    SELECT header_id FROM sb_design_evaluation_header 
                    WHERE rcm_id = %s AND user_id = %s
                )
            ''', (rcm_id, user_info['user_id']))
            
            # 2. header 레코드 삭제
            cursor = conn.execute('''
                DELETE FROM sb_design_evaluation_header 
                WHERE rcm_id = %s AND user_id = %s
            ''', (rcm_id, user_info['user_id']))
            deleted_count = cursor.rowcount
            
            conn.commit()
        
        # 활동 로그 기록
        log_user_activity(user_info, 'DESIGN_EVALUATION_RESET', f'설계평가 초기화 - RCM ID: {rcm_id}', 
                         f'/api/design-evaluation/reset', 
                         request.remote_addr, request.headers.get('User-Agent'))
        
        return jsonify({
            'success': True,
            'message': f'{deleted_count}개의 설계평가 결과가 초기화되었습니다.',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': '초기화 중 오류가 발생했습니다.'
        })

# 임시 비활성화 - 테이블 구조 문제로 인해
# @bp_link6.route('/api/design-evaluation/versions/<int:rcm_id>/<control_code>')
# @login_required
# def get_evaluation_versions_api(rcm_id, control_code):
#     return jsonify({'success': False, 'message': '기능 준비 중입니다.'}), 503

@bp_link6.route('/api/design-evaluation/sessions/<int:rcm_id>')
@login_required
def get_evaluation_sessions_api(rcm_id):
    """사용자의 설계평가 세션 목록 조회 API"""
    user_info = get_user_info()
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()
                
                if not access_check:
                    return jsonify({'success': False, 'message': '접근 권한이 없습니다.'}), 403
        
        # 평가 세션 목록 조회
        sessions = get_user_evaluation_sessions(rcm_id, user_info['user_id'])
        
        return jsonify({
            'success': True,
            'sessions': sessions,
            'total_count': len(sessions)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': '세션 조회 중 오류가 발생했습니다.'}), 500

@bp_link6.route('/api/design-evaluation/load/<int:rcm_id>')
@login_required
def load_evaluation_data_api(rcm_id):
    """특정 평가 세션의 데이터 로드 API"""
    user_info = get_user_info()
    evaluation_session = request.args.get('session')
    header_id = request.args.get('header_id')
    
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()
                
                if not access_check:
                    return jsonify({'success': False, 'message': '접근 권한이 없습니다.'}), 403
        
        # 평가 데이터 로드
        if header_id:
            evaluations = get_design_evaluations_by_header_id(rcm_id, user_info['user_id'], int(header_id))
        else:
            evaluations = get_design_evaluations(rcm_id, user_info['user_id'], evaluation_session)
            
            # 세션명으로 로드할 때 실제 header_id를 찾아서 반환
            if evaluations and evaluation_session:
                # 첫 번째 평가 데이터에서 header_id 추출
                actual_header_id = evaluations[0].get('header_id')
        
        # 통제별로 정리
        evaluation_dict = {}
        for eval_data in evaluations:
            control_code = eval_data['control_code']
            evaluation_date = eval_data.get('evaluation_date')
            
            # 디버깅용 로그 추가
            
            # 해당 통제의 이미지 파일 찾기
            saved_images = []
            
            # header_id를 찾기 위해 현재 평가 데이터에서 확인
            current_header_id = header_id
            if not current_header_id and evaluation_session:
                # header_id가 없으면 세션명으로 찾기
                try:
                    with get_db() as conn:
                        result = conn.execute('''
                            SELECT header_id FROM sb_design_evaluation_header
                            WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s
                        ''', (rcm_id, user_info['user_id'], evaluation_session)).fetchone()
                        if result:
                            current_header_id = result['header_id']
                except Exception as e:
                    pass

            if current_header_id:
                # DB에서 이미지 정보 조회
                images_found = False
                try:
                    # line_id 조회
                    line_result = conn.execute('''
                        SELECT line_id FROM sb_design_evaluation_line
                        WHERE header_id = %s AND control_code = %s
                    ''', (current_header_id, control_code)).fetchone()

                    if line_result:
                        images = conn.execute('''
                            SELECT file_path, file_name
                            FROM sb_evaluation_image
                            WHERE evaluation_type = %s AND line_id = %s
                            ORDER BY uploaded_at
                        ''', ('design', line_result['line_id'])).fetchall()

                        for img in images:
                            # file_path는 이미 "static/uploads/..."로 시작
                            # static/ 제거하여 상대 경로로 변환
                            relative_path = img['file_path'].replace('static/', '', 1) if img['file_path'].startswith('static/') else img['file_path']
                            saved_images.append({
                                'filename': img['file_name'],
                                'path': relative_path,
                                'url': f"/static/{relative_path}"
                            })
                            images_found = True
                except Exception as e:
                    pass

                # DB에 이미지가 없으면 파일 시스템에서 폴백 검색
                if not images_found:
                    # 다른 header_id로 시도해보기
                    base_dir = os.path.join('static', 'uploads', 'design_evaluations', str(rcm_id))
                    if os.path.exists(base_dir):
                        for folder_name in os.listdir(base_dir):
                            test_dir = os.path.join(base_dir, folder_name, control_code)
                            if os.path.exists(test_dir):
                                for filename in os.listdir(test_dir):
                                    if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp')):
                                        relative_path = f"uploads/design_evaluations/{rcm_id}/{folder_name}/{control_code}/{filename}"
                                        saved_images.append({
                                            'filename': filename,
                                            'path': relative_path,
                                            'url': f"/static/{relative_path}"
                                        })
            else:
                pass

            evaluation_dict[control_code] = {
                'description_adequacy': eval_data['description_adequacy'],
                'improvement_suggestion': eval_data['improvement_suggestion'],
                'evaluation_evidence': eval_data.get('evaluation_evidence'),
                'overall_effectiveness': eval_data['overall_effectiveness'],
                'evaluation_rationale': eval_data['evaluation_rationale'],
                'design_comment': eval_data.get('design_comment'),
                'recommended_actions': eval_data['recommended_actions'],
                'evaluation_date': evaluation_date,
                'images': saved_images
            }
        
        response_data = {
            'success': True,
            'evaluations': evaluation_dict,
            'session_name': evaluation_session
        }

        # 세션명으로 로드할 때 실제 header_id를 응답에 포함
        if evaluations and evaluation_session and not header_id:
            actual_header_id = evaluations[0].get('header_id')
            response_data['header_id'] = actual_header_id
            print(f"[DEBUG] API Response - header_id: {actual_header_id}")

        # header의 completed_date 정보도 포함
        try:
            with get_db() as conn:
                # header_id가 있으면 해당 header의 completed_date 조회
                if header_id:
                    result = conn.execute('''
                        SELECT completed_date FROM sb_design_evaluation_header
                        WHERE header_id = %s
                    ''', (int(header_id),)).fetchone()
                elif evaluation_session:
                    result = conn.execute('''
                        SELECT completed_date FROM sb_design_evaluation_header
                        WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s
                    ''', (rcm_id, user_info['user_id'], evaluation_session)).fetchone()
                else:
                    result = None


                if result:
                    response_data['header_completed_date'] = result['completed_date']
                    print(f"[DEBUG] API Response - header_completed_date: {result['completed_date']}")
                else:
                    response_data['header_completed_date'] = None
                    print(f"[DEBUG] API Response - header_completed_date: None (no result found)")
        except Exception as e:
            response_data['header_completed_date'] = None
            print(f"[DEBUG] API Response - Error getting completed_date: {e}")

        # 평가 데이터 개수 로그
        print(f"[DEBUG] API Response - Total evaluations: {len(evaluation_dict)}")
        print(f"[DEBUG] API Response - Control codes: {list(evaluation_dict.keys())}")

        # 첫 3개 평가 데이터 샘플 로그
        for i, (code, data) in enumerate(list(evaluation_dict.items())[:3]):
            print(f"[DEBUG] Sample {i+1} - {code}: evaluation_date={data.get('evaluation_date')}, adequacy={data.get('description_adequacy')}, effectiveness={data.get('overall_effectiveness')}")

        return jsonify(response_data)
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'데이터 로드 중 오류가 발생했습니다: {str(e)}'}), 500

@bp_link6.route('/api/design-evaluation/delete-session', methods=['POST'])
@login_required
def delete_evaluation_session_api():
    """평가 세션 삭제 API"""
    user_info = get_user_info()
    data = request.get_json()
    
    rcm_id = data.get('rcm_id')
    evaluation_session = data.get('evaluation_session')
    
    if not all([rcm_id, evaluation_session]):
        return jsonify({
            'success': False,
            'message': '필수 데이터가 누락되었습니다.'
        })
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()

                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })

            # 해당 설계평가 세션을 기반으로 한 운영평가가 있는지 확인
            operation_check = conn.execute('''
                SELECT COUNT(*) as count
                FROM sb_operation_evaluation_header
                WHERE rcm_id = %s AND design_evaluation_session = %s
            ''', (rcm_id, evaluation_session)).fetchone()

            if operation_check and operation_check['count'] > 0:
                return jsonify({
                    'success': False,
                    'message': '⛔ 이 설계평가를 기반으로 진행 중인 운영평가가 있습니다. 운영평가를 먼저 삭제해주세요.'
                })

        # 세션 삭제
        deleted_count = delete_evaluation_session(rcm_id, user_info['user_id'], evaluation_session)
        
        # 활동 로그 기록
        log_user_activity(user_info, 'DESIGN_EVALUATION_DELETE', f'설계평가 세션 삭제 - {evaluation_session}', 
                         f'/api/design-evaluation/delete-session', 
                         request.remote_addr, request.headers.get('User-Agent'))
        
        return jsonify({
            'success': True,
            'message': f'평가 세션 "{evaluation_session}"이 삭제되었습니다.',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': '삭제 중 오류가 발생했습니다.'
        })

@bp_link6.route('/api/design-evaluation/create-evaluation', methods=['POST'])
@login_required
def create_design_evaluation_api():
    """새로운 설계평가 생성 API (Header + 모든 Line 생성)"""
    user_info = get_user_info()
    data = request.get_json()
    
    rcm_id = data.get('rcm_id')
    evaluation_session = data.get('evaluation_session')
    
    if not all([rcm_id, evaluation_session]):
        return jsonify({
            'success': False,
            'message': '필수 데이터가 누락되었습니다.'
        })
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()
                
                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })
        
        # 중복 세션명 체크
        existing_sessions = get_user_evaluation_sessions(rcm_id, user_info['user_id'])
        session_names = [session['evaluation_session'] for session in existing_sessions]
        
        if evaluation_session in session_names:
            return jsonify({
                'success': False,
                'message': f'"{evaluation_session}" 세션이 이미 존재합니다. 다른 이름을 사용해주세요.',
                'exists': True
            })
        
        # 평가 구조 생성 (Header + 모든 빈 Line 생성)
        header_id = create_evaluation_structure(rcm_id, user_info['user_id'], evaluation_session)
        
        # 활동 로그 기록
        log_user_activity(user_info, 'DESIGN_EVALUATION_CREATE', 
                         f'설계평가 생성 - {evaluation_session}', 
                         f'/api/design-evaluation/create-evaluation', 
                         request.remote_addr, request.headers.get('User-Agent'))
        
        return jsonify({
            'success': True,
            'message': f'설계평가 "{evaluation_session}"이 생성되었습니다.',
            'header_id': header_id,
            'evaluation_session': evaluation_session
        })
        
    except ValueError as ve:
        return jsonify({
            'success': False,
            'message': str(ve)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': '[EVAL-001] 평가 생성 중 오류가 발생했습니다.'
        })

@bp_link6.route('/api/design-evaluation/complete', methods=['POST'])
@login_required
def complete_design_evaluation_api():
    """설계평가 완료 처리 API - header 테이블에 완료일시 업데이트"""
    user_info = get_user_info()
    data = request.get_json()
    
    rcm_id = data.get('rcm_id')
    evaluation_session = data.get('evaluation_session')
    
    if not rcm_id or not evaluation_session:
        return jsonify({
            'success': False,
            'message': 'RCM ID와 평가 세션이 필요합니다.'
        })
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()
                
                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })
        
        # header 테이블에서 해당 평가 세션의 완료일시 업데이트
        with get_db() as conn:
            # 현재 평가 세션이 존재하는지 확인
            header = conn.execute('''
                SELECT header_id FROM sb_design_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], evaluation_session)).fetchone()
            
            if not header:
                return jsonify({
                    'success': False,
                    'message': '해당 평가 세션을 찾을 수 없습니다.'
                })
            
            # completed_date를 현재 시간으로 업데이트
            from datetime import datetime
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            conn.execute('''
                UPDATE sb_design_evaluation_header
                SET completed_date = %s, evaluation_status = 'COMPLETED'
                WHERE header_id = %s
            ''', (current_time, header['header_id']))
            
            conn.commit()
        
        # 활동 로그 기록
        log_user_activity(user_info, 'DESIGN_EVALUATION_COMPLETE', 
                         f'설계평가 완료 - {evaluation_session}', 
                         f'/api/design-evaluation/complete', 
                         request.remote_addr, request.headers.get('User-Agent'))
        
        return jsonify({
            'success': True,
            'message': '설계평가가 완료 처리되었습니다.',
            'completed_date': current_time
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': '완료 처리 중 오류가 발생했습니다.'
        })

@bp_link6.route('/api/design-evaluation/cancel', methods=['POST'])
@login_required
def cancel_design_evaluation_api():
    """설계평가 완료 취소 API - header 테이블의 completed_date를 NULL로 설정"""
    user_info = get_user_info()
    data = request.get_json()
    
    rcm_id = data.get('rcm_id')
    evaluation_session = data.get('evaluation_session')
    
    if not rcm_id or not evaluation_session:
        return jsonify({
            'success': False,
            'message': 'RCM ID와 평가 세션이 필요합니다.'
        })
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()
                
                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })
        
        # header 테이블에서 해당 평가 세션의 완료일시를 NULL로 설정
        with get_db() as conn:
            # 현재 평가 세션이 존재하는지 확인
            header = conn.execute('''
                SELECT header_id, completed_date FROM sb_design_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], evaluation_session)).fetchone()
            
            if not header:
                return jsonify({
                    'success': False,
                    'message': '해당 평가 세션을 찾을 수 없습니다.'
                })
            
            if not header['completed_date']:
                return jsonify({
                    'success': False,
                    'message': '완료되지 않은 평가입니다.'
                })

            # 운영평가 진행 여부 확인
            operation_check = conn.execute('''
                SELECT COUNT(*) as cnt
                FROM sb_operation_evaluation_header
                WHERE design_evaluation_session = %s AND rcm_id = %s AND user_id = %s
            ''', (evaluation_session, rcm_id, user_info['user_id'])).fetchone()

            has_operation_evaluation = operation_check and operation_check['cnt'] > 0

            # completed_date를 NULL로 설정하고 status를 IN_PROGRESS로 변경
            conn.execute('''
                UPDATE sb_design_evaluation_header
                SET completed_date = NULL, evaluation_status = 'IN_PROGRESS'
                WHERE header_id = %s
            ''', (header['header_id'],))

            conn.commit()
        
        # 활동 로그 기록
        log_user_activity(user_info, 'DESIGN_EVALUATION_CANCEL', 
                         f'설계평가 완료 취소 - {evaluation_session}', 
                         f'/api/design-evaluation/cancel', 
                         request.remote_addr, request.headers.get('User-Agent'))
        
        message = '설계평가 완료가 취소되었습니다.'
        if has_operation_evaluation:
            message += '\n\n⚠️ 참고: 이 설계평가를 기반으로 한 운영평가가 진행중입니다.'

        return jsonify({
            'success': True,
            'message': message,
            'has_operation_evaluation': has_operation_evaluation
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': '완료 취소 중 오류가 발생했습니다.'
        })


@bp_link6.route('/api/design-evaluation/download-excel/<int:rcm_id>')
@login_required
def download_evaluation_excel(rcm_id):
    """설계평가 엑셀 다운로드 - 원본 파일의 Template 시트를 통제별로 복사해서 생성"""
    # mimetypes 초기화 및 .mpo 오류 방지
    import mimetypes
    try:
        # 강제로 mimetypes 재초기화
        mimetypes.init()
        # 여러 방법으로 .mpo 확장자 등록 시도
        mimetypes.add_type('application/octet-stream', '.mpo')
        mimetypes.add_type('image/jpeg', '.mpo')  # MPO는 Multi Picture Object 포맷
        
        # types_map에 직접 추가
        if hasattr(mimetypes, 'types_map'):
            if True in mimetypes.types_map:
                mimetypes.types_map[True]['.mpo'] = 'image/jpeg'
            if False in mimetypes.types_map:
                mimetypes.types_map[False]['.mpo'] = 'image/jpeg'
        
    except Exception as mt_error:
        # 최후의 수단: monkey patch
        try:
            import openpyxl.packaging.manifest
            original_register = openpyxl.packaging.manifest.Manifest._register_mimetypes
            def patched_register(self, filenames):
                try:
                    return original_register(self, filenames)
                except KeyError as e:
                    if '.mpo' in str(e):
                        # .mpo 파일을 제외하고 다시 시도
                        filtered_filenames = [f for f in filenames if not f.lower().endswith('.mpo')]
                        return original_register(self, filtered_filenames)
                    raise
            openpyxl.packaging.manifest.Manifest._register_mimetypes = patched_register
        except Exception as patch_error:
            pass

    user_info = get_user_info()
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()
                
                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    }), 403

            # RCM 정보 및 파일 경로 가져오기 (업로드한 사용자의 회사명, 최신 설계평가 세션명 포함)
            rcm_info = conn.execute('''
                SELECT r.rcm_name, r.original_filename, u.company_name
                FROM sb_rcm r
                LEFT JOIN sb_user u ON r.user_id = u.user_id
                WHERE r.rcm_id = %s
            ''', (rcm_id,)).fetchone()
            
            # 현재 사용자의 가장 최신 설계평가 세션명 조회
            evaluation_session = None
            try:
                session_info = conn.execute('''
                    SELECT evaluation_session 
                    FROM sb_design_evaluation_header 
                    WHERE rcm_id = %s AND user_id = %s 
                    ORDER BY start_date DESC 
                    LIMIT 1
                ''', (rcm_id, user_info['user_id'])).fetchone()
                
                if session_info:
                    evaluation_session = session_info['evaluation_session']
            except:
                pass  # 설계평가 세션이 없는 경우 기본값 사용
            
            if not rcm_info or not rcm_info['original_filename']:
                return jsonify({
                    'success': False,
                    'message': '원본 엑셀 파일을 찾을 수 없습니다.'
                }), 404
            
            # RCM 상세 정보 가져오기
            rcm_details = conn.execute('''
                SELECT detail_id, control_code, control_name, control_frequency, control_type, 
                       test_procedure, control_description
                FROM sb_rcm_detail 
                WHERE rcm_id = %s 
                ORDER BY detail_id
            ''', (rcm_id,)).fetchall()
            
            if not rcm_details:
                return jsonify({
                    'success': False,
                    'message': 'RCM 상세 정보를 찾을 수 없습니다.'
                }), 404
            
            # 설계평가 데이터 조회 (평가 근거 및 첫부 이미지)
            evaluation_data = {}
            if evaluation_session:
                try:
                    # 해당 세션의 설계평가 결과 조회 (overall_effectiveness 포함)
                    eval_results = conn.execute('''
                        SELECT l.control_code, l.evaluation_evidence, l.evaluation_rationale, l.overall_effectiveness, h.header_id
                        FROM sb_design_evaluation_line l
                        JOIN sb_design_evaluation_header h ON l.header_id = h.header_id
                        WHERE h.rcm_id = %s AND h.user_id = %s AND h.evaluation_session = %s
                        ORDER BY l.control_sequence, l.control_code
                    ''', (rcm_id, user_info['user_id'], evaluation_session)).fetchall()

                    for eval_result in eval_results:
                        control_code = eval_result['control_code']
                        evidence = eval_result['evaluation_evidence'] or ''
                        rationale = eval_result['evaluation_rationale'] or ''
                        effectiveness = eval_result['overall_effectiveness'] or ''
                        header_id = eval_result['header_id']
                        
                        # DB에서 이미지 경로 찾기
                        images_info = ''
                        try:
                            # line_id 조회
                            line_result = conn.execute('''
                                SELECT line_id FROM sb_design_evaluation_line
                                WHERE header_id = %s AND control_code = %s
                            ''', (header_id, control_code)).fetchone()

                            if line_result:
                                images = conn.execute('''
                                    SELECT file_path, file_name
                                    FROM sb_evaluation_image
                                    WHERE evaluation_type = %s AND line_id = %s
                                    ORDER BY uploaded_at
                                ''', ('design', line_result['line_id'])).fetchall()

                                if images:
                                    images_info = json.dumps([{'file': img['file_name'], 'path': img['file_path']} for img in images])
                        except Exception as e:
                            pass
                        
                        evaluation_data[control_code] = {
                            'evidence': evidence,
                            'rationale': rationale,
                            'effectiveness': effectiveness,
                            'images': images_info
                        }
                except Exception as e:
                    pass

        # 원본 엑셀 파일 로드 (회사별 폴더 구조 지원)
        # rcm_info['original_filename']이 이미 상대 경로(company_name/filename.xlsx)인 경우와
        # 기존 방식(filename.xlsx)인 경우 모두 지원
        if os.path.sep in rcm_info['original_filename'] or '/' in rcm_info['original_filename']:
            # 이미 회사별 경로가 포함된 경우
            original_file_path = os.path.join('uploads', rcm_info['original_filename'])
        else:
            # 기존 방식 - uploads 루트에서 찾기
            original_file_path = os.path.join('uploads', rcm_info['original_filename'])
        
        if not os.path.exists(original_file_path):
            return jsonify({
                'success': False,
                'message': f'원본 파일이 존재하지 않습니다: {original_file_path}'
            }), 404
            
        workbook = load_workbook(original_file_path)
        
        # Template 시트가 있는지 확인
        if 'Template' not in workbook.sheetnames:
            return jsonify({
                'success': False,
                'message': '원본 파일에 Template 시트가 없습니다.'
            }), 404
            
        template_sheet = workbook['Template']
        
        # 각 통제별로 Template 시트를 복사해서 새 시트 생성
        for detail in rcm_details:
            control_code = detail['control_code']
            
            # Template 시트 복사
            new_sheet = workbook.copy_worksheet(template_sheet)
            new_sheet.title = control_code
            
            # B3, B5, C7~C11 셀에 정보 입력
            new_sheet['B3'] = user_info.get('company_name', '')  # 회사명
            new_sheet['B5'] = user_info.get('user_name', '')  # 사용자명
            new_sheet['C7'] = detail['control_code']  # 통제코드
            new_sheet['C8'] = detail['control_name']  # 통제명
            new_sheet['C9'] = detail['control_frequency']  # 통제주기
            new_sheet['C10'] = detail['control_type']  # 통제구분
            new_sheet['C11'] = detail['test_procedure'] or ''  # 테스트 절차
            
            # C12에 증빙, C13에 평가 근거, C14에 첨부 이미지, C15에 효과성 추가
            eval_info = evaluation_data.get(control_code, {})
            evidence_value = eval_info.get('evidence', '')
            rationale_value = eval_info.get('rationale', '')
            effectiveness_value = eval_info.get('effectiveness', '')

            new_sheet['C12'] = evidence_value  # 증빙
            new_sheet['C13'] = rationale_value  # 평가 근거

            # 여러 라인 텍스트가 있는 셀들의 높이 자동 조정
            # C11 (테스트 절차) 셀 설정
            c11_cell = new_sheet['C11']
            c11_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # C12 (증빙) 셀 설정
            c12_cell = new_sheet['C12']
            c12_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # C13 (평가 근거) 셀 설정
            c13_cell = new_sheet['C13']
            c13_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # 텍스트 길이에 따라 행 높이 자동 조정
            test_procedure_text = detail['test_procedure'] or ''
            evidence_text = evidence_value or ''
            rationale_text = rationale_value or ''

            # 줄바꿈 개수와 텍스트 길이를 기준으로 높이 계산
            c11_lines = max(1, len(test_procedure_text.split('\n'))) if test_procedure_text else 1
            c12_lines = max(1, len(evidence_text.split('\n'))) if evidence_text else 1
            c13_lines = max(1, len(rationale_text.split('\n'))) if rationale_text else 1
            
            # 긴 텍스트의 경우 줄바꿈이 없어도 자동 줄바꿈으로 인해 여러 줄이 될 수 있음
            # 약 50자당 1줄로 가정 (엑셀 C열 기준)
            if test_procedure_text:
                estimated_c11_lines = max(c11_lines, (len(test_procedure_text) // 50) + 1)
            else:
                estimated_c11_lines = c11_lines

            if evidence_text:
                estimated_c12_lines = max(c12_lines, (len(evidence_text) // 50) + 1)
            else:
                estimated_c12_lines = c12_lines

            if rationale_text:
                estimated_c13_lines = max(c13_lines, (len(rationale_text) // 50) + 1)
            else:
                estimated_c13_lines = c13_lines

            # 최소 높이는 25pt, 최대 높이는 150pt로 제한 (기본 20pt + 줄 수 * 18pt)
            c11_height = max(25, min(300, 20 + (estimated_c11_lines - 1) * 18))
            c12_height = max(25, min(150, 20 + (estimated_c12_lines - 1) * 18))
            c13_height = max(25, min(150, 20 + (estimated_c13_lines - 1) * 18))

            new_sheet.row_dimensions[11].height = c11_height
            new_sheet.row_dimensions[12].height = c12_height
            new_sheet.row_dimensions[13].height = c13_height

            # C15에 효과성 및 시트 탭 색상 설정
            if effectiveness_value == '효과적' or effectiveness_value.lower() == 'effective':
                new_sheet['C15'] = 'Effective' # 효과적인 경우
                new_sheet.sheet_properties.tabColor = Color(rgb="90EE90")
            else:
                new_sheet['C15'] = 'Ineffective'
                # 비효과적인 경우 시트 탭을 연한 빨간색으로 설정
                new_sheet.sheet_properties.tabColor = Color(rgb="FFA0A0")  # 연한 빨간색
            
            # 첨부 이미지 처리 (C14 셀에 이미지 삽입)
            images_info = eval_info.get('images', '')
            if images_info:
                try:
                    from openpyxl.drawing.image import Image as ExcelImage

                    # JSON 형태로 저장된 이미지 정보 파싱
                    if images_info.startswith('['):
                        image_list = json.loads(images_info)
                        if image_list and len(image_list) > 0:
                            # 최대 3개 이미지를 C14, D14, E14에 가로로 배치
                            target_cells = ['C14', 'D14', 'E14']
                            max_images = min(len(image_list), 3)

                            # C14 셀 크기 기준으로 통일된 이미지 크기 계산
                            col_width = new_sheet.column_dimensions['C'].width or 8.43
                            cell_width_px = int(col_width * 7)  # 문자 단위를 픽셀로 변환

                            row_height = new_sheet.row_dimensions[14].height or 15
                            cell_height_px = int(row_height * 1.33)  # 포인트를 픽셀로 변환

                            # 셀 크기의 90%로 최대 크기 설정 (모든 이미지에 동일 적용)
                            uniform_max_width = max(int(cell_width_px * 0.9), 150)
                            uniform_max_height = max(int(cell_height_px * 0.9), 150)

                            for idx in range(max_images):
                                image_path = image_list[idx].get('path', '')
                                if image_path and os.path.exists(image_path):
                                    try:
                                        # PIL을 사용해 이미지를 정리한 후 임시 파일로 저장
                                        from PIL import Image as PILImage
                                        import tempfile

                                        # 원본 이미지 로드 (메타데이터 제거)
                                        pil_image = PILImage.open(image_path)
                                        # RGB로 변환 (EXIF 데이터 제거)
                                        if pil_image.mode != 'RGB':
                                            pil_image = pil_image.convert('RGB')

                                        # 임시 파일로 저장 (메타데이터 없이)
                                        temp_image_file = tempfile.NamedTemporaryFile(delete=False, suffix='.jpg')
                                        temp_image_file.close()
                                        pil_image.save(temp_image_file.name, 'JPEG', quality=85)
                                        pil_image.close()

                                        # 정리된 이미지를 엑셀에 삽입
                                        img = ExcelImage(temp_image_file.name)

                                        # 원본 이미지 크기 비율 계산
                                        original_width = img.width
                                        original_height = img.height

                                        # 가로세로 비율 유지하며 통일된 크기로 조정
                                        width_ratio = uniform_max_width / original_width
                                        height_ratio = uniform_max_height / original_height
                                        scale_ratio = min(width_ratio, height_ratio)

                                        img.width = int(original_width * scale_ratio)
                                        img.height = int(original_height * scale_ratio)

                                        # 해당 셀에 이미지 추가
                                        new_sheet.add_image(img, target_cells[idx])

                                        # 임시 파일은 나중에 정리
                                        if not hasattr(workbook, 'temp_files'):
                                            workbook.temp_files = []
                                        workbook.temp_files.append(temp_image_file.name)

                                    except Exception as img_error:
                                        # 오류 발생시 해당 셀에 메시지 표시
                                        new_sheet[target_cells[idx]] = f'이미지 {idx+1} 오류'
                    else:
                        # 단순 문자열 경우
                        new_sheet['C13'] = images_info
                except Exception as e:
                    new_sheet['C13'] = f'이미지 처리 오류: {str(e)}'
            else:
                new_sheet['C13'] = ''
        
        # 원본 Template 시트 삭제
        workbook.remove(template_sheet)
        
        # downloads 폴더에 파일 저장
        company_name = rcm_info['company_name'] or '회사명없음'
        evaluation_name = evaluation_session or '설계평가'
        safe_filename = f"{company_name}_{evaluation_name}.xlsx"
        downloads_path = os.path.join('downloads', safe_filename)
        
        # 기존 파일이 있으면 삭제
        if os.path.exists(downloads_path):
            os.remove(downloads_path)
        
        # .mpo 파일 문제 해결을 위한 대체 저장 방법
        try:
            workbook.save(downloads_path)
        except Exception as save_error:
            
            # 임시 디렉토리에서 작업
            import tempfile
            import shutil
            
            with tempfile.TemporaryDirectory() as temp_dir:
                temp_path = os.path.join(temp_dir, "evaluation.xlsx")
                
                try:
                    # 임시 경로에 저장 시도
                    workbook.save(temp_path)
                    # 성공하면 최종 경로로 복사
                    shutil.copy2(temp_path, downloads_path)
                    
                except Exception as temp_save_error:
                    
                    # 최후의 수단: BytesIO를 사용한 메모리 저장
                    try:
                        from io import BytesIO
                        buffer = BytesIO()
                        workbook.save(buffer)
                        buffer.seek(0)
                        
                        with open(downloads_path, 'wb') as f:
                            f.write(buffer.read())
                        
                    except Exception as bytesio_error:
                        raise save_error  # 원래 오류를 다시 발생시킴
        
        # 임시 이미지 파일들 정리
        if hasattr(workbook, 'temp_files'):
            for temp_file in workbook.temp_files:
                try:
                    os.unlink(temp_file)
                except:
                    pass  # 임시 파일 삭제 실패 시 무시
        
        workbook.close()
        
        def remove_download_file():
            try:
                if os.path.exists(downloads_path):
                    os.unlink(downloads_path)
            except:
                pass
        
        # 활동 로그 기록
        log_user_activity(user_info, 'DESIGN_EVALUATION_DOWNLOAD', 
                         f'설계평가 엑셀 다운로드 - {rcm_info["rcm_name"]}', 
                         f'/api/design-evaluation/download-excel/{rcm_id}', 
                         request.remote_addr, request.headers.get('User-Agent'))
        
        # 다운로드 파일 전송
        response = send_file(
            downloads_path,
            as_attachment=True,
            download_name=safe_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # 요청 완료 후 다운로드 파일 삭제
        response.call_on_close(remove_download_file)
        
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        
        # 구체적인 오류 메시지 반환
        error_msg = str(e)
        if "No such file or directory" in error_msg:
            error_msg = "원본 RCM 파일을 찾을 수 없습니다."
        elif "Template" in error_msg:
            error_msg = "원본 파일에 Template 시트가 없습니다."
        elif "Permission denied" in error_msg:
            error_msg = "파일 접근 권한이 없습니다."
        else:
            error_msg = f"엑셀 생성 중 오류 발생: {error_msg}"
            
        return jsonify({
            'success': False,
            'message': error_msg
        }), 500

@bp_link6.route('/api/design-evaluation/archive', methods=['POST'])
@login_required
def archive_design_evaluation_api():
    """설계평가 세션 Archive 처리 API"""
    user_info = get_user_info()
    data = request.get_json()

    rcm_id = data.get('rcm_id')
    evaluation_session = data.get('evaluation_session')

    if not rcm_id or not evaluation_session:
        return jsonify({
            'success': False,
            'message': 'RCM ID와 평가 세션이 필요합니다.'
        })

    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()

                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })

        # 세션 Archive 처리
        success = archive_design_evaluation_session(rcm_id, user_info['user_id'], evaluation_session)

        if success:
            # 활동 로그 기록
            log_user_activity(user_info, 'DESIGN_EVALUATION_ARCHIVE',
                             f'설계평가 세션 Archive - {evaluation_session}',
                             f'/api/design-evaluation/archive',
                             request.remote_addr, request.headers.get('User-Agent'))

            return jsonify({
                'success': True,
                'message': f'설계평가 세션 "{evaluation_session}"이 Archive 처리되었습니다.'
            })
        else:
            return jsonify({
                'success': False,
                'message': '해당 평가 세션을 찾을 수 없습니다.'
            })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': 'Archive 처리 중 오류가 발생했습니다.'
        })

@bp_link6.route('/api/design-evaluation/unarchive', methods=['POST'])
@login_required
def unarchive_design_evaluation_api():
    """설계평가 세션 Unarchive 처리 API"""
    user_info = get_user_info()
    data = request.get_json()

    rcm_id = data.get('rcm_id')
    evaluation_session = data.get('evaluation_session')

    if not rcm_id or not evaluation_session:
        return jsonify({
            'success': False,
            'message': 'RCM ID와 평가 세션이 필요합니다.'
        })

    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()

                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })

        # 세션 Unarchive 처리
        success = unarchive_design_evaluation_session(rcm_id, user_info['user_id'], evaluation_session)

        if success:
            # 활동 로그 기록
            log_user_activity(user_info, 'DESIGN_EVALUATION_UNARCHIVE',
                             f'설계평가 세션 Unarchive - {evaluation_session}',
                             f'/api/design-evaluation/unarchive',
                             request.remote_addr, request.headers.get('User-Agent'))

            return jsonify({
                'success': True,
                'message': f'설계평가 세션 "{evaluation_session}"이 복원되었습니다.'
            })
        else:
            return jsonify({
                'success': False,
                'message': '해당 평가 세션을 찾을 수 없습니다.'
            })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': 'Unarchive 처리 중 오류가 발생했습니다.'
        })
# ============================================================================
# ELC 설계평가 (수동통제만, 기준통제 매핑 없음)
# ============================================================================

@bp_link6.route('/elc/design-evaluation')
@login_required
def elc_design_evaluation():
    """ELC 설계평가 페이지"""
    user_info = get_user_info()

    # ELC RCM 목록만 필터링
    all_rcms = get_user_rcms(user_info['user_id'])
    elc_rcms = [rcm for rcm in all_rcms if rcm.get('control_category') == 'ELC']

    # 최근 완료된 설계평가 결과 조회 (최근 5개)
    with get_db() as conn:
        recent_evaluations = conn.execute('''
            SELECT
                h.evaluation_session,
                h.rcm_id,
                r.rcm_name,
                h.total_controls,
                h.evaluated_controls,
                h.completed_date,
                h.evaluation_status
            FROM sb_design_evaluation_header h
            JOIN sb_rcm r ON h.rcm_id = r.rcm_id
            WHERE h.user_id = %s
            AND r.control_category = 'ELC'
            AND h.evaluation_status = 'COMPLETED'
            ORDER BY h.completed_date DESC
            LIMIT 5
        ''', (user_info['user_id'],)).fetchall()

        recent_evaluations = [dict(row) for row in recent_evaluations]

    log_user_activity(user_info, 'PAGE_ACCESS', 'ELC 설계평가', '/elc/design-evaluation',
                     request.remote_addr, request.headers.get('User-Agent'))

    return render_template('link6_elc_design_evaluation.jsp',
                         elc_rcms=elc_rcms,
                         recent_evaluations=recent_evaluations,
                         is_logged_in=is_logged_in(),
                         user_info=user_info)

# ============================================================================
# TLC 설계평가 (자동통제 포함, 기준통제 매핑 없음)
# ============================================================================

@bp_link6.route('/tlc/design-evaluation')
@login_required
def tlc_design_evaluation():
    """TLC 설계평가 페이지"""
    user_info = get_user_info()

    # TLC RCM 목록만 필터링
    all_rcms = get_user_rcms(user_info['user_id'])
    tlc_rcms = [rcm for rcm in all_rcms if rcm.get('control_category') == 'TLC']

    log_user_activity(user_info, 'PAGE_ACCESS', 'TLC 설계평가', '/tlc/design-evaluation',
                     request.remote_addr, request.headers.get('User-Agent'))

    return render_template('link6_tlc_design_evaluation.jsp',
                         tlc_rcms=tlc_rcms,
                         is_logged_in=is_logged_in(),
                         user_info=user_info)

# ============================================================================
# 설계평가 다운로드 기능
# ============================================================================

def format_evidence_for_excel(evaluation_evidence_json, attributes_dict, population_count):
    """
    증빙 데이터를 "컬럼명: 증빙값" 형식으로 포맷팅

    Args:
        evaluation_evidence_json (str): JSON 형식의 증빙 데이터
        attributes_dict (dict): {'attribute0': '거래일자', 'attribute1': '거래금액', ...}
        population_count (int): 모집단 항목 개수 (이 값 이후부터 증빙으로 간주)

    Returns:
        tuple: (formatted_evidence, line_count)
            - formatted_evidence (str): 줄바꿈으로 구분된 증빙 내용
            - line_count (int): 줄 수 (셀 높이 계산용)
    """
    # 빈 값 처리
    if not evaluation_evidence_json or not evaluation_evidence_json.strip():
        return '', 0

    try:
        # JSON 파싱 시도
        evidence_data = json.loads(evaluation_evidence_json)

        # JSON 객체가 아니면 원본 반환
        if not isinstance(evidence_data, dict):
            # 줄 수 계산
            line_count = evaluation_evidence_json.count('\n') + 1
            return evaluation_evidence_json, line_count

        # 증빙 항목만 추출 (모집단 제외)
        evidence_lines = []
        for i in range(population_count, 10):  # attribute{population_count}부터 attribute9까지
            attr_key = f'attribute{i}'
            attr_name = attributes_dict.get(attr_key)
            attr_value = evidence_data.get(attr_key, '')

            # 속성명과 값이 모두 있을 때만 추가
            if attr_name and attr_value:
                evidence_lines.append(f'{attr_name}: {attr_value}')

        # 줄바꿈으로 연결
        formatted_evidence = '\n'.join(evidence_lines)
        line_count = len(evidence_lines)

        return formatted_evidence, line_count

    except (json.JSONDecodeError, ValueError, TypeError):
        # JSON 파싱 실패 시 (레거시 데이터) 원본 텍스트 반환
        line_count = evaluation_evidence_json.count('\n') + 1
        return evaluation_evidence_json, line_count


@bp_link6.route('/design-evaluation/download')
@login_required
def download_design_evaluation():
    """설계평가 결과를 Design_Template.xlsx 양식으로 다운로드"""
    user_info = get_user_info()

    rcm_id = request.args.get('rcm_id')
    evaluation_session = request.args.get('evaluation_session')

    if not rcm_id or not evaluation_session:
        flash('RCM ID와 평가 세션이 필요합니다.', 'error')
        return redirect(url_for('link6.user_design_evaluation'))

    try:
        # 템플릿 파일 경로
        template_path = os.path.join(os.path.dirname(__file__), 'paper_templates', 'Design_Template.xlsx')

        # 템플릿 로드
        wb = load_workbook(template_path)

        # RCM 정보 조회
        with get_db() as conn:
            rcm_info = conn.execute('''
                SELECT rcm_name, description
                FROM sb_rcm
                WHERE rcm_id = %s
            ''', (rcm_id,)).fetchone()

            if not rcm_info:
                flash('RCM 정보를 찾을 수 없습니다.', 'error')
                return redirect(url_for('link6.user_design_evaluation'))

            # 설계평가 결과 조회 (RCM attribute 메타 정보 포함)
            evaluations = conn.execute('''
                SELECT
                    l.line_id,
                    l.control_code,
                    rd.control_name,
                    rd.control_description,
                    rd.control_frequency,
                    rd.control_type,
                    rd.control_nature,
                    l.description_adequacy as design_adequacy,
                    l.improvement_suggestion,
                    l.overall_effectiveness as conclusion,
                    l.evaluation_rationale,
                    l.evaluation_evidence,
                    l.design_comment,
                    l.evaluation_date,
                    d.attribute0, d.attribute1, d.attribute2, d.attribute3, d.attribute4,
                    d.attribute5, d.attribute6, d.attribute7, d.attribute8, d.attribute9,
                    d.population_attribute_count
                FROM sb_design_evaluation_line l
                JOIN sb_design_evaluation_header h ON l.header_id = h.header_id
                JOIN sb_rcm_detail_v rd ON h.rcm_id = rd.rcm_id AND l.control_code = rd.control_code
                JOIN sb_rcm_detail d ON d.rcm_id = h.rcm_id AND d.control_code = l.control_code
                WHERE h.rcm_id = %s AND h.evaluation_session = %s
                ORDER BY l.control_code
            ''', (rcm_id, evaluation_session)).fetchall()

        # Template 시트를 사용하여 각 통제별 시트 생성
        template_sheet = wb['Template']

        # 각 통제별 시트 생성
        for eval_data in evaluations:
            eval_dict = dict(eval_data)
            control_code = eval_dict['control_code']

            # 템플릿 시트 복사
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = control_code

            # Client 정보 (C2) - 회사명
            new_sheet['C2'] = user_info.get('company_name', '')

            # Prepared by (C4) - 사용자명
            new_sheet['C4'] = user_info.get('user_name', '')

            # 통제번호 (C7)
            new_sheet['C7'] = control_code

            # 통제명 (C8)
            new_sheet['C8'] = eval_dict.get('control_name', '')

            # 주기 (C9)
            new_sheet['C9'] = eval_dict.get('control_frequency', '')

            # 구분 (C10)
            new_sheet['C10'] = eval_dict.get('control_type', '')

            # 테스트 절차 (C11)
            control_description = eval_dict.get('control_description', '')
            new_sheet['C11'] = control_description

            # C11 셀의 행 높이 자동 조정 (텍스트 길이에 따라)
            if control_description:
                # 줄바꿈 개수 계산
                line_count = control_description.count('\n') + 1
                # 기본 행 높이(15) + 각 줄당 추가 높이(15)
                row_height = 15 + (line_count * 15)
                # 최대 높이 제한 (300)
                row_height = min(row_height, 300)
                new_sheet.row_dimensions[11].height = row_height

            # RCM attribute 메타 정보 수집 (C12, C13에서 공통 사용)
            attributes_dict = {}
            for i in range(10):
                attr_key = f'attribute{i}'
                attr_value = eval_dict.get(attr_key)
                if attr_value:  # None이 아닌 경우만 저장
                    attributes_dict[attr_key] = attr_value

            population_count = eval_dict.get('population_attribute_count', 2)
            evaluation_evidence_json = eval_dict.get('evaluation_evidence', '')

            # 설계 평가 코멘트 (C12)
            design_comment = eval_dict.get('design_comment', '')

            # 설계평가 샘플의 증빙 항목을 검토 결과에 추가
            line_id = eval_dict.get('line_id')
            if line_id:
                # 설계평가 샘플 조회
                with get_db() as conn2:
                    samples = conn2.execute('''
                        SELECT sample_id, sample_number,
                               attribute0, attribute1, attribute2, attribute3, attribute4,
                               attribute5, attribute6, attribute7, attribute8, attribute9
                        FROM sb_evaluation_sample
                        WHERE line_id = %s AND evaluation_type = 'design'
                        ORDER BY sample_number
                        LIMIT 1
                    ''', (line_id,)).fetchall()

                    if samples and len(samples) > 0:
                        sample = dict(samples[0])
                        # 증빙 항목 (population_count부터) 수집
                        attribute_lines = []
                        for i in range(population_count, 10):
                            attr_key = f'attribute{i}'
                            attr_name = attributes_dict.get(attr_key)
                            attr_value = sample.get(attr_key, '')

                            if attr_name and attr_value:
                                attribute_lines.append(f'{attr_name}: {attr_value}')

                        # 검토 결과에 attribute 값 추가
                        if attribute_lines:
                            if design_comment:
                                design_comment += '\n\n'
                            design_comment += '\n'.join(attribute_lines)

            new_sheet['C12'] = design_comment

            # C12 셀의 행 높이 자동 조정
            if design_comment:
                line_count = design_comment.count('\n') + 1
                row_height = 15 + (line_count * 15)
                row_height = min(row_height, 300)
                new_sheet.row_dimensions[12].height = row_height

            # 증빙 내용 (C13) - 포맷팅된 텍스트

            # 증빙 데이터 포맷팅 (컬럼명:증빙명 형식)
            formatted_evidence, evidence_line_count = format_evidence_for_excel(
                evaluation_evidence_json,
                attributes_dict,
                population_count
            )

            new_sheet['C13'] = formatted_evidence

            # C13 셀 높이 자동 조정
            if evidence_line_count > 0:
                # 기본 높이(15) + 각 줄당 높이(15)
                row_height = 15 + (evidence_line_count * 15)
                # 최대 높이 제한 (300)
                row_height = min(row_height, 300)
                new_sheet.row_dimensions[13].height = row_height

            # 증빙 이미지 (C14부터 시작)
            # 이미지 파일이 저장된 경로: static/uploads/design_evaluations/{rcm_id}/{header_id}/{control_code}/
            from openpyxl.drawing.image import Image as OpenpyxlImage

            image_count = 0  # 삽입된 이미지 개수
            with get_db() as conn:
                header = conn.execute('''
                    SELECT header_id
                    FROM sb_design_evaluation_header
                    WHERE rcm_id = %s AND evaluation_session = %s
                ''', (rcm_id, evaluation_session)).fetchone()

                if header:
                    header_id = header['header_id']

                    # DB에서 이미지 정보 조회
                    try:
                        # line_id 조회
                        line_result = conn.execute('''
                            SELECT line_id FROM sb_design_evaluation_line
                            WHERE header_id = %s AND control_code = %s
                        ''', (header_id, control_code)).fetchone()

                        if line_result:
                            images = conn.execute('''
                                SELECT file_path, file_name
                                FROM sb_evaluation_image
                                WHERE evaluation_type = %s AND line_id = %s
                                ORDER BY uploaded_at
                            ''', ('design', line_result['line_id'])).fetchall()

                            if images:
                                # C14 셀부터 이미지 삽입
                                row_offset = 14  # C14부터 시작
                                for idx, img_data in enumerate(images):
                                    image_path = img_data['file_path']
                                    cell_position = f'C{row_offset + idx}'
                                    try:
                                        # 이미지 객체 생성
                                        img = OpenpyxlImage(image_path)
                                        # 이미지 크기 조정 (너비 280px로 제한하여 셀보다 작게)
                                        max_width = 280
                                        if img.width > max_width:
                                            ratio = max_width / img.width
                                            img.width = max_width
                                            img.height = int(img.height * ratio)
                                        else:
                                            # 원본이 280px보다 작으면 90%로 축소
                                            img.width = int(img.width * 0.9)
                                            img.height = int(img.height * 0.9)
                                        # 이미지를 C14, C15, C16... 위치에 삽입
                                        new_sheet.add_image(img, cell_position)
                                        # 행 높이 조정 (이미지 높이보다 약간 크게 여백 추가)
                                        new_sheet.row_dimensions[row_offset + idx].height = (img.height * 0.75) + 5
                                        image_count += 1
                                        print(f"이미지 삽입 성공: {img_data['file_name']} -> {cell_position}")
                                    except Exception as e:
                                        print(f"이미지 삽입 오류 ({img_data['file_name']}): {e}")
                                        import traceback
                                        traceback.print_exc()
                                        # 오류 발생 시 파일명만 표시
                                        new_sheet[cell_position] = img_data['file_name']
                                        image_count += 1
                    except Exception as e:
                        print(f"이미지 조회 오류: {e}")

            # 결론 (이미지 이후 행)
            conclusion_row = 15 if image_count == 0 else 14 + image_count
            new_sheet[f'C{conclusion_row}'] = eval_dict.get('conclusion', '')

        # Template 시트 삭제
        if 'Template' in wb.sheetnames:
            wb.remove(wb['Template'])

        # 임시 파일로 저장
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()

        # 다운로드 파일명 생성 (평가 세션명 사용)
        # 한글을 유지하면서 안전하지 않은 문자만 제거
        filename = f"{evaluation_session}.xlsx"
        # 파일명에서 위험한 문자만 제거 (/, \, :, *, %s, ", <, >, |)
        unsafe_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
        for char in unsafe_chars:
            filename = filename.replace(char, '_')

        # 파일 전송 (한글 파일명 지원)
        from flask import make_response
        import urllib.parse

        response = make_response(send_file(
            temp_file.name,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ))

        # 한글 파일명을 위한 Content-Disposition 헤더 설정
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"

        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'다운로드 중 오류가 발생했습니다: {str(e)}', 'error')
        return redirect(url_for('link6.user_design_evaluation'))
@bp_link6.route('/api/client-log', methods=['POST'])
def client_log():
    """클라이언트 JavaScript 로그를 Flask 콘솔에 출력"""
    try:
        data = request.get_json()
        message = data.get('message', 'No message')
        log_data = data.get('data', {})
        
        print(f'[CLIENT LOG] {message}')
        if log_data:
            import json
            print(f'[CLIENT DATA] {json.dumps(log_data, ensure_ascii=False, indent=2)}')
        
        return jsonify({'success': True})
    except Exception as e:
        print(f'[CLIENT LOG ERROR] {e}')
        return jsonify({'success': False})
