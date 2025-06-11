from flask import render_template
import openpyxl
import datetime
import snowball_db

def rcm_generate(form_data, file_name=None):
    print("RCM Generate called")

    param1 = form_data.get('param1')
    param2 = form_data.get('param2')
    param3 = form_data.get('param3')
    param4 = form_data.get('param4')
    param5 = form_data.get('param5')

    print("Param1 = ", param1)
    print("Param2 = ", param2)
    print("Param3 = ", param3)
    print("Param4 = ", param4)
    print("Param5 = ", param5)

    if file_name:
        output_path = file_name
    else:
        if param1=="":
            output_path = 'rcm.xlsx'
        else:
            if param2=="":
                output_path = param1 + '_rcm.xlsx'
            else:
                output_path = param1 + '_' + param2 + '_rcm.xlsx'

    try:
        workbook = openpyxl.load_workbook('./paper_templates/RCM_generate.xlsx')
    except FileNotFoundError:
        print("./paper_templates/PBC_Template.xlsx: 파일열기 오류")
        return ''

    sheet = workbook["RCM"]

    # Application 조정
    for row in sheet.iter_rows():
        if row[0].value == 'Application':
            if row[1].value != param3:
                sheet.delete_rows(row[0].row)

    # DB 조정
    for row in sheet.iter_rows():
        if row[0].value == 'DB':
            if row[1].value != param4:
                sheet.delete_rows(row[0].row)

    # OS 조정
    for row in sheet.iter_rows():
        if row[0].value == 'OS':
            if row[1].value != param5:
                sheet.delete_rows(row[0].row)

    workbook.save('./downloads/' + output_path)
    workbook.close()

    return './downloads/' + output_path

def rcm_request(form_data):
    print("RCM Request called")

    client_name = form_data.get('param1')
    email_address = form_data.get('param2')
    file_name = form_data.get('param3')
    
    print("client = ", client_name)
    print("email = ", email_address)
    print("file = ", file_name)

    #snowball_db.set_rcm_request('RCM', file_name, client_name, email_address)

    return

'''    
    # Application 시트 조정
    if param4 == 'SAP':
        sheets_to_delete = ['APD01_Oracle', 'APD01_Douzone', 'APD01_KSystem', 'APD01_ETC', 'APD04_Oracle', 'APD04_Douzone', 'APD04_KSystem', 'APD04_ETC', 'APD06_Oracle', 'APD06_Douzone', 'APD06_KSystem', 'APD06_ETC', 'PC01_ETC', 'PC04_ETC', 'PC05_ETC', 'CO01_Oracle', 'CO01_ETC', 'CO02_Oracle', 'CO02_ETC']
    elif param4 == 'Oracle':
        sheets_to_delete = ['APD01_SAP', 'APD01_Douzone', 'APD01_KSystem', 'APD01_ETC', 'APD04_SAP', 'APD04_Douzone', 'APD04_KSystem', 'APD04_ETC', 'APD06_SAP', 'APD06_Douzone', 'APD06_KSystem', 'APD06_ETC', 'APD07_SAP', 'APD08_SAP', 'PC01_SAP', 'PC04_SAP', 'PC05_SAP', 'CO01_SAP', 'CO01_ETC', 'CO02_SAP', 'CO02_ETC']
    elif param4 == 'Douzone':
        sheets_to_delete = ['APD01_SAP', 'APD01_Oracle', 'APD01_KSystem', 'APD01_ETC', 'APD04_SAP', 'APD04_Oracle', 'APD04_KSystem', 'APD04_ETC', 'APD06_SAP', 'APD06_Oracle', 'APD06_KSystem', 'APD06_ETC', 'APD07_SAP', 'APD08_SAP', 'PC01_SAP', 'PC04_SAP', 'PC05_SAP', 'CO01_SAP', 'CO01_Oracle', 'CO02_SAP', 'CO02_Oracle']
    elif param4 == 'KSystem':
        sheets_to_delete = ['APD01_SAP', 'APD01_Oracle', 'APD01_Douzone', 'APD01_ETC', 'APD04_SAP', 'APD04_Oracle', 'APD04_Douzone', 'APD04_ETC', 'APD06_SAP', 'APD06_Oracle', 'APD06_Douzone', 'APD06_ETC', 'APD07_SAP', 'APD08_SAP', 'PC01_SAP', 'PC04_SAP', 'PC05_SAP', 'CO01_SAP', 'CO01_Oracle', 'CO02_SAP', 'CO02_Oracle']
    else:
        sheets_to_delete = ['APD01_SAP', 'APD01_Oracle', 'APD01_Douzone', 'APD01_KSystem', 'APD04_SAP', 'APD04_Oracle', 'APD04_Douzone', 'APD04_KSystem', 'APD06_SAP', 'APD06_Oracle', 'APD06_Douzone', 'APD06_KSystem', 'APD07_SAP', 'APD08_SAP', 'PC01_SAP','PC04_SAP', 'PC05_SAP', 'CO01_SAP', 'CO01_Oracle', 'CO02_SAP', 'CO02_Oracle']

    # Application 시트 삭제
    for sheet_name in sheets_to_delete:
        if sheet_name in workbook.sheetnames:
            sheet_to_delete = workbook[sheet_name]
            workbook.remove(sheet_to_delete)

    # OS 시트 조정
    if param5 == 'Unix':
        sheets_to_delete = ['APD12_Windows', 'APD12_Linux', 'APD12_Tool', 'APD12_ETC', 'APD13_Windows', 'APD13_Linux', 'APD13_Tool', 'APD13_ETC', 'APD14_Windows', 'APD14_Linux', 'APD14_Tool', 'APD14_ETC', 'PC07_Windows', 'PC07_Linux', 'PC07_ETC']
    elif param5 == 'Windows':
        sheets_to_delete = ['APD12_Unix', 'APD12_Linux', 'APD12_Tool', 'APD12_ETC', 'APD13_Unix', 'APD13_Linux', 'APD13_Tool', 'APD13_ETC', 'APD14_Unix', 'APD14_Linux', 'APD14_Tool', 'APD14_ETC', 'PC07_Unix', 'PC07_Linux', 'PC07_ETC']
    elif param5 == 'Linux':
        sheets_to_delete = ['APD12_Unix', 'APD12_Windows', 'APD12_Tool', 'APD12_ETC', 'APD13_Unix', 'APD13_Windows', 'APD13_Tool', 'APD13_ETC', 'APD14_Unix', 'APD14_Windows', 'APD14_Tool', 'APD14_ETC', 'PC07_Unix', 'PC07_Windows', 'PC07_ETC']
    else:
        sheets_to_delete = ['APD12_Unix', 'APD12_Windows', 'APD12_Linux', 'APD12_Tool', 'APD13_Unix', 'APD13_Windows', 'APD13_Linux', 'APD13_Tool', 'APD14_Unix', 'APD14_Windows', 'APD14_Linux', 'APD14_Tool', 'PC07_Unix', 'PC07_Windows', 'PC07_Linux']

    # OS 시트 삭제
    for sheet_name in sheets_to_delete:
        if sheet_name in workbook.sheetnames:
            sheet_to_delete = workbook[sheet_name]
            workbook.remove(sheet_to_delete)

    # DB 시트 조정
    if param6 == 'Oracle':
        sheets_to_delete = ['APD09_MSSQL', 'APD09_ETC', 'APD10_MSSQL', 'APD10_ETC', 'APD11_MSSQL', 'APD11_ETC', 'PC06_MSSQL', 'PC06_ETC']
    elif param6 == 'MSSQL':
        sheets_to_delete = ['APD09_Oracle', 'APD09_ETC', 'APD10_Oracle', 'APD10_ETC', 'APD11_Oracle', 'APD11_ETC', 'PC06_Oracle', 'PC06_ETC']
    else:
        sheets_to_delete = ['APD09_Oracle', 'APD09_MSSQL', 'APD10_Oracle', 'APD10_MSSQL', 'APD11_Oracle', 'APD11_MSSQL', 'PC06_Oracle', 'PC06_MSSQL']

    # DB 시트 삭제
    for sheet_name in sheets_to_delete:
        if sheet_name in workbook.sheetnames:
            sheet_to_delete = workbook[sheet_name]
            workbook.remove(sheet_to_delete)
    
    # 시트명 변경
    for sheet in workbook:
        if '_' in sheet.title:
            index = sheet.title.index('_')
            sheet.title = sheet.title[:index]

    workbook.save('./downloads/' + output_path)
    workbook.close()

    return './downloads/' + output_path
'''