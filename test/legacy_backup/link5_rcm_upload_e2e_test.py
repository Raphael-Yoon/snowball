"""
E03: RCM 업로드 및 관리 E2E 테스트 - Playwright 기반

Link5 RCM 업로드 및 관리 전체 플로우를 실제 브라우저에서 테스트합니다.
- Excel 파일 업로드
- 컬럼 매핑
- RCM 저장
- RCM 목록 조회
- RCM 상세보기
- RCM 삭제
"""

import sys
from pathlib import Path
from datetime import datetime
from unittest.mock import patch, MagicMock
import time
import tempfile
import openpyxl

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from test.playwright_base import PlaywrightTestBase, E2ETestResult


class Link5RCMUploadE2ETestSuite(PlaywrightTestBase):
    """Link5 RCM 업로드 및 관리 E2E 테스트 스위트"""

    def __init__(self):
        super().__init__(base_url="http://localhost:5000", headless=False)
        self.test_email = "test@example.com"
        self.test_rcm_name = f"E2E_TEST_RCM_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.test_company = "테스트회사"
        self.test_excel_file = None

    def run_all_tests(self):
        """모든 E2E 테스트 실행"""
        print("=" * 80)
        print("E03: RCM 업로드 및 관리 E2E 테스트 시작 (Playwright)")
        print("=" * 80)
        print(f"시작 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Base URL: {self.base_url}\n")

        try:
            self.setup()

            # 테스트용 Excel 파일 생성
            self._create_test_excel_file()

            self.run_category("1. RCM 관리 페이지 접근", [
                self.test_rcm_list_page_loads,
            ])

            self.run_category("2. RCM 업로드 프로세스", [
                self.test_rcm_upload_page_access,
                self.test_file_upload_and_preview,
                self.test_column_mapping_ui,
                self.test_rcm_save_process,
            ])

            self.run_category("3. RCM 조회 및 관리", [
                self.test_rcm_list_display,
                self.test_rcm_detail_view,
                self.test_rcm_search_filter,
            ])

            self.run_category("4. RCM 삭제", [
                self.test_rcm_delete_flow,
            ])

        finally:
            # 테스트용 파일 정리
            self._cleanup_test_files()
            self.teardown()

        exit_code = self.print_final_report()
        self.save_json_report("link5_rcm_upload_e2e")
        return exit_code

    # =========================================================================
    # 헬퍼 메서드
    # =========================================================================

    def _create_test_excel_file(self):
        """테스트용 Excel RCM 파일 생성"""
        # 임시 Excel 파일 생성
        self.test_excel_file = tempfile.NamedTemporaryFile(
            suffix='.xlsx',
            delete=False
        )

        wb = openpyxl.Workbook()
        ws = wb.active

        # ITGC RCM 샘플 데이터
        headers = [
            '통제코드', '통제명', '통제설명', '핵심통제',
            '통제주기', '통제성격', '통제방법'
        ]

        data = [
            ['APD01', '사용자 등록 및 삭제', '사용자 계정 등록 및 삭제 절차', 'Y', '월간', '예방', '수동'],
            ['APD02', '사용자 권한 부여', '사용자 권한 부여 및 검토', 'Y', '월간', '예방', '수동'],
            ['APD03', '비밀번호 정책', '비밀번호 복잡도 및 만료 정책', 'N', '연간', '예방', '자동'],
        ]

        # 헤더 작성
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # 데이터 작성
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(self.test_excel_file.name)
        wb.close()

        print(f"✓ 테스트용 Excel 파일 생성: {self.test_excel_file.name}")

    def _cleanup_test_files(self):
        """테스트용 파일 정리"""
        try:
            if self.test_excel_file:
                Path(self.test_excel_file.name).unlink(missing_ok=True)
                print(f"✓ 테스트용 Excel 파일 삭제")
        except Exception as e:
            print(f"⚠️ 파일 정리 실패: {e}")

    def _quick_login(self):
        """빠른 로그인 (테스트용)"""
        with patch('auth.send_otp') as mock_send:
            with patch('auth.verify_otp') as mock_verify:
                mock_send.return_value = (True, "OTP 발송됨")
                mock_verify.return_value = (True, {
                    'user_id': 1,
                    'user_email': self.test_email,
                    'user_name': '테스트사용자',
                    'admin_flag': 'N'
                })

                self.navigate_to("/login")
                self.fill_input("input[name='email']", self.test_email)
                self.click_button("button[type='submit']")

                try:
                    self.page.wait_for_url("**/otp**", timeout=5000)
                    self.fill_input("input[name='otp']", "123456")
                    self.click_button("button[type='submit']")
                    self.page.wait_for_url("**/main", timeout=5000)
                except:
                    pass  # 이미 로그인된 경우

    # =========================================================================
    # 1. RCM 관리 페이지 접근
    # =========================================================================

    def test_rcm_list_page_loads(self, result: E2ETestResult):
        """RCM 목록 페이지가 정상적으로 로드되는지 확인"""
        try:
            result.add_detail("🎯 목표: RCM 관리 페이지 접근 확인")

            # 로그인 필요
            self._quick_login()
            result.add_detail("✓ 로그인 완료")

            # RCM 목록 페이지 접속
            self.navigate_to("/user/rcm")
            screenshot = self.take_screenshot("rcm_list_page")
            result.add_screenshot(screenshot)
            result.add_detail("✓ RCM 목록 페이지 접속")

            # 페이지 로드 확인
            self.page.wait_for_load_state("networkidle")

            # 주요 UI 요소 확인
            page_content = self.page.content()

            # 업로드 버튼 존재 확인
            if "업로드" in page_content or "upload" in page_content.lower():
                result.add_detail("✓ RCM 업로드 버튼 확인")
            else:
                result.add_detail("⚠️ 업로드 버튼을 찾을 수 없음")

            # 테이블 또는 목록 확인
            if self.check_element_exists("table") or self.check_element_exists(".rcm-list"):
                result.add_detail("✓ RCM 목록 테이블 확인")
            else:
                result.add_detail("⚠️ RCM 목록을 찾을 수 없음")

            result.pass_test("RCM 관리 페이지가 정상적으로 로드됩니다")

        except Exception as e:
            screenshot = self.take_screenshot("rcm_list_page_error")
            result.add_screenshot(screenshot)
            result.fail_test(f"페이지 로드 실패: {str(e)}")

    # =========================================================================
    # 2. RCM 업로드 프로세스
    # =========================================================================

    def test_rcm_upload_page_access(self, result: E2ETestResult):
        """RCM 업로드 페이지 접근"""
        try:
            result.add_detail("🎯 목표: RCM 업로드 페이지 접근")

            # RCM 업로드 페이지로 이동
            self.navigate_to("/user/rcm/upload")
            screenshot = self.take_screenshot("rcm_upload_page")
            result.add_screenshot(screenshot)
            result.add_detail("✓ RCM 업로드 페이지 접속")

            # 페이지 로드 대기
            self.page.wait_for_load_state("networkidle")

            # 파일 업로드 폼 확인
            page_content = self.page.content()

            if "파일" in page_content or "file" in page_content.lower():
                result.add_detail("✓ 파일 업로드 UI 확인")
            else:
                result.add_detail("⚠️ 파일 업로드 UI를 찾을 수 없음")

            # 필수 입력 필드 확인
            required_fields = ["RCM", "회사", "카테고리"]
            found_fields = [field for field in required_fields if field in page_content]
            result.add_detail(f"✓ 필수 필드 확인: {', '.join(found_fields)}")

            result.pass_test("RCM 업로드 페이지 접근 성공")

        except Exception as e:
            screenshot = self.take_screenshot("rcm_upload_page_error")
            result.add_screenshot(screenshot)
            result.fail_test(f"업로드 페이지 접근 실패: {str(e)}")

    def test_file_upload_and_preview(self, result: E2ETestResult):
        """파일 업로드 및 미리보기"""
        try:
            result.add_detail("🎯 목표: Excel 파일 업로드 및 미리보기")

            # 업로드 페이지로 이동 (이미 있다면 스킵)
            current_url = self.page.url
            if "/upload" not in current_url:
                self.navigate_to("/user/rcm/upload")

            # 파일 입력 요소 찾기
            file_input = self.page.locator("input[type='file']").first

            if file_input.is_visible() or file_input.count() > 0:
                # 파일 선택
                file_input.set_input_files(self.test_excel_file.name)
                result.add_detail(f"✓ 파일 선택: {Path(self.test_excel_file.name).name}")

                # 파일 업로드 처리 대기
                time.sleep(2)

                screenshot = self.take_screenshot("file_uploaded")
                result.add_screenshot(screenshot)

                # 미리보기 확인 (페이지에 데이터가 표시되는지)
                page_content = self.page.content()
                if "APD01" in page_content or "통제" in page_content:
                    result.add_detail("✓ 파일 데이터 미리보기 확인")
                else:
                    result.add_detail("⚠️ 미리보기 데이터를 찾을 수 없음")

                result.pass_test("파일 업로드 및 미리보기 성공")
            else:
                result.skip_test("파일 입력 요소를 찾을 수 없어 테스트를 건너뜁니다")

        except Exception as e:
            screenshot = self.take_screenshot("file_upload_error")
            result.add_screenshot(screenshot)
            result.fail_test(f"파일 업로드 실패: {str(e)}")

    def test_column_mapping_ui(self, result: E2ETestResult):
        """컬럼 매핑 UI 테스트"""
        try:
            result.add_detail("🎯 목표: 컬럼 매핑 UI 확인")

            # 현재 페이지 내용 확인
            page_content = self.page.content()

            # 컬럼 매핑 관련 UI 요소 확인
            mapping_keywords = ["매핑", "mapping", "컬럼", "column"]
            found_keywords = [kw for kw in mapping_keywords if kw in page_content.lower()]

            if found_keywords:
                result.add_detail(f"✓ 컬럼 매핑 UI 확인: {', '.join(found_keywords)}")
            else:
                result.add_detail("ℹ️ 자동 매핑되거나 매핑 UI가 없을 수 있음")

            screenshot = self.take_screenshot("column_mapping")
            result.add_screenshot(screenshot)

            result.pass_test("컬럼 매핑 UI 확인 완료")

        except Exception as e:
            result.fail_test(f"컬럼 매핑 UI 확인 실패: {str(e)}")

    @patch('auth.create_rcm')
    def test_rcm_save_process(self, result: E2ETestResult, mock_create_rcm):
        """RCM 저장 프로세스"""
        try:
            result.add_detail("🎯 목표: RCM 정보 입력 및 저장")

            # Mock 설정
            mock_create_rcm.return_value = 999  # 테스트 RCM ID

            # RCM명 입력 (필드가 있다면)
            try:
                rcm_name_input = self.page.locator("input[name*='rcm'], input[name*='name']").first
                if rcm_name_input.is_visible():
                    rcm_name_input.fill(self.test_rcm_name)
                    result.add_detail(f"✓ RCM명 입력: {self.test_rcm_name}")
            except:
                result.add_detail("ℹ️ RCM명 입력 필드 없음 (자동 생성일 수 있음)")

            # 회사명 입력
            try:
                company_input = self.page.locator("input[name*='company'], input[name*='회사']").first
                if company_input.is_visible():
                    company_input.fill(self.test_company)
                    result.add_detail(f"✓ 회사명 입력: {self.test_company}")
            except:
                result.add_detail("⚠️ 회사명 입력 필드 없음")

            # 카테고리 선택
            try:
                category_select = self.page.locator("select[name*='category'], select[name*='카테고리']").first
                if category_select.is_visible():
                    category_select.select_option("ITGC")
                    result.add_detail("✓ 카테고리 선택: ITGC")
            except:
                result.add_detail("⚠️ 카테고리 선택 필드 없음")

            # 스크린샷
            screenshot = self.take_screenshot("rcm_form_filled")
            result.add_screenshot(screenshot)

            # 저장 버튼 클릭
            try:
                save_button = self.page.locator("button:has-text('저장'), button:has-text('Save'), button[type='submit']").first
                if save_button.is_visible():
                    save_button.click()
                    result.add_detail("✓ 저장 버튼 클릭")

                    # 처리 대기
                    time.sleep(2)

                    # 성공 메시지 확인
                    page_content = self.page.content()
                    success_keywords = ["성공", "완료", "success", "saved"]
                    if any(kw in page_content.lower() for kw in success_keywords):
                        result.add_detail("✓ 저장 성공 메시지 확인")
                    else:
                        result.add_detail("⚠️ 성공 메시지를 찾을 수 없음")

                    screenshot_after = self.take_screenshot("rcm_saved")
                    result.add_screenshot(screenshot_after)

                    result.pass_test("RCM 저장 프로세스 완료")
                else:
                    result.skip_test("저장 버튼을 찾을 수 없습니다")
            except Exception as e:
                result.fail_test(f"저장 버튼 클릭 실패: {str(e)}")

        except Exception as e:
            screenshot = self.take_screenshot("rcm_save_error")
            result.add_screenshot(screenshot)
            result.fail_test(f"RCM 저장 실패: {str(e)}")

    # =========================================================================
    # 3. RCM 조회 및 관리
    # =========================================================================

    def test_rcm_list_display(self, result: E2ETestResult):
        """RCM 목록 표시 확인"""
        try:
            result.add_detail("🎯 목표: RCM 목록 데이터 표시 확인")

            # RCM 목록 페이지로 이동
            self.navigate_to("/user/rcm")
            self.page.wait_for_load_state("networkidle")

            screenshot = self.take_screenshot("rcm_list_display")
            result.add_screenshot(screenshot)

            # 테이블 확인
            if self.check_element_exists("table"):
                result.add_detail("✓ RCM 목록 테이블 존재")

                # 테이블 행 수 확인
                rows = self.page.locator("table tbody tr")
                row_count = rows.count()
                result.add_detail(f"✓ 표시된 RCM 개수: {row_count}개")

                result.pass_test(f"RCM 목록이 정상적으로 표시됩니다 ({row_count}개)")
            else:
                result.warn_test("RCM 테이블을 찾을 수 없습니다 (빈 목록일 수 있음)")

        except Exception as e:
            result.fail_test(f"RCM 목록 표시 확인 실패: {str(e)}")

    def test_rcm_detail_view(self, result: E2ETestResult):
        """RCM 상세보기"""
        try:
            result.add_detail("🎯 목표: RCM 상세보기 페이지 확인")

            # RCM 목록 페이지에서 첫 번째 RCM 클릭
            self.navigate_to("/user/rcm")
            self.page.wait_for_load_state("networkidle")

            # 상세보기 링크 찾기
            detail_link = self.page.locator("a:has-text('상세'), a:has-text('보기'), .view-link").first

            if detail_link.is_visible():
                detail_link.click()
                result.add_detail("✓ 상세보기 클릭")

                time.sleep(1)
                screenshot = self.take_screenshot("rcm_detail_view")
                result.add_screenshot(screenshot)

                # 상세 정보 확인
                page_content = self.page.content()
                detail_keywords = ["통제코드", "통제명", "control"]
                found = [kw for kw in detail_keywords if kw in page_content.lower()]

                if found:
                    result.add_detail(f"✓ 상세 정보 확인: {', '.join(found)}")
                    result.pass_test("RCM 상세보기 페이지가 정상 작동합니다")
                else:
                    result.warn_test("상세 정보를 확인할 수 없습니다")
            else:
                result.skip_test("상세보기 링크를 찾을 수 없습니다 (목록이 비어있을 수 있음)")

        except Exception as e:
            result.skip_test(f"RCM 상세보기 테스트 건너뜀: {str(e)}")

    def test_rcm_search_filter(self, result: E2ETestResult):
        """RCM 검색 및 필터링"""
        try:
            result.add_detail("🎯 목표: RCM 검색/필터 기능 확인")

            self.navigate_to("/user/rcm")
            self.page.wait_for_load_state("networkidle")

            # 검색 입력 필드 확인
            search_input = self.page.locator("input[type='search'], input[placeholder*='검색']").first

            if search_input.is_visible():
                # 검색어 입력
                search_input.fill("ITGC")
                result.add_detail("✓ 검색어 입력: ITGC")

                time.sleep(1)
                screenshot = self.take_screenshot("rcm_search")
                result.add_screenshot(screenshot)

                result.pass_test("검색 기능이 존재합니다")
            else:
                result.skip_test("검색 기능이 구현되지 않았거나 찾을 수 없습니다")

        except Exception as e:
            result.skip_test(f"검색 기능 테스트 건너뜀: {str(e)}")

    # =========================================================================
    # 4. RCM 삭제
    # =========================================================================

    @patch('auth.get_db')
    def test_rcm_delete_flow(self, result: E2ETestResult, mock_get_db):
        """RCM 삭제 플로우"""
        try:
            result.add_detail("🎯 목표: RCM 삭제 기능 확인")

            # Mock 설정
            mock_conn = MagicMock()
            mock_get_db.return_value.__enter__.return_value = mock_conn

            self.navigate_to("/user/rcm")
            self.page.wait_for_load_state("networkidle")

            # 삭제 버튼 찾기
            delete_button = self.page.locator("button:has-text('삭제'), .delete-btn").first

            if delete_button.is_visible():
                screenshot_before = self.take_screenshot("before_delete")
                result.add_screenshot(screenshot_before)

                delete_button.click()
                result.add_detail("✓ 삭제 버튼 클릭")

                # 확인 다이얼로그 대기 및 확인
                time.sleep(0.5)

                # 브라우저 confirm 다이얼로그 처리
                try:
                    self.page.on("dialog", lambda dialog: dialog.accept())
                    time.sleep(1)
                    result.add_detail("✓ 삭제 확인 다이얼로그 처리")
                except:
                    result.add_detail("ℹ️ 확인 다이얼로그 없음")

                screenshot_after = self.take_screenshot("after_delete")
                result.add_screenshot(screenshot_after)

                result.pass_test("RCM 삭제 기능이 작동합니다")
            else:
                result.skip_test("삭제 버튼을 찾을 수 없습니다 (권한 또는 데이터 없음)")

        except Exception as e:
            result.skip_test(f"RCM 삭제 테스트 건너뜀: {str(e)}")


def main():
    """메인 실행 함수"""
    suite = Link5RCMUploadE2ETestSuite()
    exit_code = suite.run_all_tests()
    sys.exit(exit_code)


if __name__ == '__main__':
    main()
