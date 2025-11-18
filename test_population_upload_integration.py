"""
통합 테스트: 모집단 업로드 기능 전체 워크플로우 검증
"""
import sys
import os

print("=" * 70)
print("모집단 업로드 기능 통합 테스트")
print("=" * 70)
print()

# 테스트 1: 필수 모듈 임포트
print("[테스트 1] 필수 모듈 임포트 검증")
try:
    import flask
    import openpyxl
    from werkzeug.utils import secure_filename
    import file_manager
    print("  [OK] All required modules imported successfully")
    print()
except Exception as e:
    print(f"  [FAIL] Module import failed: {e}")
    sys.exit(1)

# 테스트 2: 표본 크기 계산 로직
print("[테스트 2] 표본 크기 계산 로직 검증")
test_cases = [
    (0, 0), (1, 1), (4, 2), (5, 2), (12, 2),
    (13, 5), (30, 5), (52, 5),
    (53, 20), (100, 20), (250, 20),
    (251, 25), (1000, 25)
]

all_passed = True
for pop_count, expected in test_cases:
    result = file_manager.calculate_sample_size(pop_count)
    if result == expected:
        print(f"  [OK] {pop_count:4d} population -> {result:2d} samples")
    else:
        print(f"  [FAIL] {pop_count:4d} population -> {result:2d} samples (expected: {expected:2d})")
        all_passed = False

if all_passed:
    print("  All test cases passed!")
else:
    print("  Some tests failed!")
    sys.exit(1)
print()

# 테스트 3: Excel 파일 생성 및 읽기
print("[테스트 3] Excel 파일 생성 및 읽기")
try:
    from openpyxl import Workbook, load_workbook

    # 테스트 파일 생성
    test_file = 'test_integration_population.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = '모집단'
    ws.append(['번호', '설명', '비고'])
    for i in range(1, 31):
        ws.append([f'TEST-{i:03d}', f'테스트 항목 {i}', f'비고 {i}'])
    wb.save(test_file)
    print(f"  [OK] Test file created: {test_file}")

    # 파일 읽기
    wb = load_workbook(test_file, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    print(f"  [OK] Headers: {headers}")

    # 데이터 읽기
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"  [OK] Data rows: {len(rows)}")

    wb.close()
    print()

except Exception as e:
    print(f"  [FAIL] Excel processing failed: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# 테스트 4: 무작위 표본 추출
print("[테스트 4] 무작위 표본 추출")
try:
    import random

    # 모집단 생성
    population = [{'number': f'ITEM-{i:03d}', 'description': f'설명 {i}'}
                  for i in range(1, 31)]

    # 표본 크기 계산 (30개 -> 5개)
    sample_size = file_manager.calculate_sample_size(len(population))
    print(f"  [OK] Population: {len(population)}, Sample: {sample_size}")

    # 무작위 추출
    sample_indices = random.sample(range(len(population)), sample_size)
    samples = [population[i] for i in sorted(sample_indices)]

    print(f"  [OK] Extracted samples:")
    for i, sample in enumerate(samples, 1):
        print(f"     {i}. {sample['number']}: {sample['description']}")
    print()

except Exception as e:
    print(f"  [FAIL] Sampling failed: {e}")
    sys.exit(1)

# 테스트 5: 한글 파일명 처리
print("[테스트 5] 한글 파일명 처리")
try:
    test_filenames = [
        ('모집단.xlsx', '.xlsx'),
        ('테스트_파일.xlsm', '.xlsm'),
        ('population.xlsx', '.xlsx'),
        ('test file.xlsx', '.xlsx'),
    ]

    for original, expected_ext in test_filenames:
        file_ext = os.path.splitext(original)[1]
        filename = secure_filename(original)

        if not filename or filename == file_ext.replace('.', ''):
            filename = f"population{file_ext}"

        if not os.path.splitext(filename)[1]:
            filename = filename + file_ext

        print(f"  [OK] '{original}' -> '{filename}'")

    print()

except Exception as e:
    print(f"  [FAIL] Filename handling failed: {e}")
    sys.exit(1)

# 테스트 6: 데이터베이스 연결 및 테이블 확인
print("[테스트 6] 데이터베이스 테이블 확인")
try:
    import sqlite3

    db_path = 'snowball.db'
    if not os.path.exists(db_path):
        print(f"  [WARN] Database file not found: {db_path}")
        print()
    else:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # 필수 테이블 확인
        required_tables = [
            'sb_operation_evaluation_header',
            'sb_operation_evaluation_line',
            'sb_operation_evaluation_sample'
        ]

        for table in required_tables:
            cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table}'")
            result = cursor.fetchone()
            if result:
                print(f"  [OK] Table exists: {table}")
            else:
                print(f"  [FAIL] Table missing: {table}")

        conn.close()
        print()

except Exception as e:
    print(f"  [WARN] Database check failed: {e}")
    print()

# 테스트 7: Flask 애플리케이션 임포트
print("[테스트 7] Flask 애플리케이션 모듈 확인")
try:
    import snowball_link7
    print("  [OK] snowball_link7 module imported successfully")

    from snowball_link7 import bp_link7
    print("  [OK] bp_link7 blueprint imported successfully")
    print()

except Exception as e:
    print(f"  [FAIL] Module import failed: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# 테스트 8: 업로드 디렉토리 확인
print("[테스트 8] 업로드 디렉토리 확인")
try:
    upload_folder = os.path.join('uploads', 'populations')

    if not os.path.exists(upload_folder):
        os.makedirs(upload_folder, exist_ok=True)
        print(f"  [OK] Directory created: {upload_folder}")
    else:
        print(f"  [OK] Directory exists: {upload_folder}")

    # 쓰기 권한 테스트
    test_file_path = os.path.join(upload_folder, '_test_write.tmp')
    with open(test_file_path, 'w') as f:
        f.write('test')
    os.remove(test_file_path)
    print(f"  [OK] Write permission confirmed")
    print()

except Exception as e:
    print(f"  [FAIL] Directory check failed: {e}")
    sys.exit(1)

# 테스트 9: 파일 형식 검증
print("[테스트 9] 파일 형식 검증 로직")
try:
    valid_extensions = ['.xlsx', '.xlsm']
    test_files = [
        ('test.xlsx', True),
        ('test.xlsm', True),
        ('test.xls', False),
        ('test.csv', False),
        ('test.txt', False),
    ]

    for filename, should_be_valid in test_files:
        ext = os.path.splitext(filename)[1].lower()
        is_valid = ext in valid_extensions

        if is_valid == should_be_valid:
            status = "ALLOWED" if is_valid else "REJECTED"
            print(f"  [OK] '{filename}' -> {status}")
        else:
            print(f"  [FAIL] '{filename}' -> Validation failed")

    print()

except Exception as e:
    print(f"  [FAIL] File format validation failed: {e}")
    sys.exit(1)

# 최종 결과
print("=" * 70)
print("통합 테스트 완료!")
print("=" * 70)
print()
print("다음 사항을 수동으로 확인하세요:")
print("  1. Flask 서버 실행: python snowball.py")
print("  2. 브라우저에서 http://127.0.0.1:5001 접속")
print("  3. RCM 상세 페이지 → 운영평가 탭")
print("  4. 표본 크기가 0인 통제(C-EL-RA-08-01) 선택")
print("  5. 모집단 업로드 섹션 표시 확인")
print("  6. test_integration_population.xlsx 파일 업로드")
print("  7. 필드 매핑 설정 (번호, 설명)")
print("  8. 업로드 및 표본 추출")
print("  9. Attribute 설정")
print("  10. 저장 및 완료")
print()
print("테스트 파일:")
print(f"  - {os.path.abspath(test_file)}")
print()

# 정리
try:
    # 테스트 파일은 유지 (수동 테스트에 사용)
    pass
except:
    pass
