# -*- coding: utf-8 -*-
"""
Snowball Link11 - 정보보호공시 모듈

정보보호공시를 위한 질문-답변 시스템과 증빙 자료 관리 기능을 제공합니다.
ITGC 인터뷰 모듈과 유사한 구조로 설계되었습니다.
"""
import sys
import os
import json
import uuid
from datetime import datetime
from flask import Blueprint, render_template, jsonify, request, session, send_file
from werkzeug.utils import secure_filename

# Windows 콘솔 UTF-8 설정
if os.name == 'nt':
    os.system('chcp 65001 > nul')
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')

bp_link11 = Blueprint('link11', __name__)

# 파일 업로드 설정
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads', 'disclosure')
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx', 'jpg', 'jpeg', 'png', 'gif', 'txt', 'hwp'}
MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB

# 업로드 폴더 생성
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ============================================================================
# Helper Functions
# ============================================================================

def is_logged_in():
    """로그인 상태 확인"""
    return 'user_id' in session


def get_user_info():
    """현재 로그인한 사용자 정보 반환 (세션 우선)"""
    if is_logged_in():
        if 'user_info' in session:
            return session['user_info']
        try:
            from auth import get_current_user
            return get_current_user()
        except ImportError:
            pass
    return None


def get_db():
    """데이터베이스 연결 가져오기"""
    try:
        from auth import get_db as auth_get_db
        return auth_get_db()
    except ImportError:
        import sqlite3
        conn = sqlite3.connect('snowball.db')
        conn.row_factory = sqlite3.Row
        return conn


def allowed_file(filename):
    """허용된 파일 확장자 확인"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def generate_uuid():
    """UUID 생성"""
    return str(uuid.uuid4())


def get_company_name_by_user_id(user_id):
    """user_id로 company_name 조회"""
    try:
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT company_name FROM sb_user WHERE user_id = ?
            ''', (user_id,))
            result = cursor.fetchone()
            return result['company_name'] if result else 'default'
    except Exception as e:
        print(f"회사명 조회 오류: {e}")
        return 'default'


# 카테고리 ID ↔ 이름 매핑
CATEGORY_MAP = {
    1: '정보보호 투자 현황',
    2: '정보보호 인력 현황',
    3: '정보보호 관련 인증',
    4: '정보보호 관련 활동'
}

CATEGORY_REVERSE_MAP = {v: k for k, v in CATEGORY_MAP.items()}


def get_category_name(category_id):
    """카테고리 ID로 이름 조회"""
    return CATEGORY_MAP.get(category_id)


def get_category_id(category_name):
    """카테고리 이름으로 ID 조회"""
    return CATEGORY_REVERSE_MAP.get(category_name)


# ============================================================================
# Routes - 메인 페이지
# ============================================================================

@bp_link11.route('/link11')
def index():
    """정보보호공시 메인 페이지 (대시보드)"""
    user_logged_in = is_logged_in()
    user_info = get_user_info()
    return render_template('link11.jsp',
                           is_logged_in=user_logged_in,
                           user_info=user_info)


@bp_link11.route('/link11/category/<category_id>')
def category_view(category_id):
    """카테고리별 질문 응답 페이지"""
    user_logged_in = is_logged_in()
    user_info = get_user_info()
    return render_template('link11.jsp',
                           is_logged_in=user_logged_in,
                           user_info=user_info,
                           view='category',
                           category_id=category_id)


@bp_link11.route('/link11/progress')
def progress_view():
    """진행 상황 추적 페이지"""
    user_logged_in = is_logged_in()
    user_info = get_user_info()
    return render_template('link11.jsp',
                           is_logged_in=user_logged_in,
                           user_info=user_info,
                           view='progress')


@bp_link11.route('/link11/evidence')
def evidence_view():
    """증빙 자료 관리 페이지"""
    user_logged_in = is_logged_in()
    user_info = get_user_info()
    return render_template('link11_evidence.jsp',
                           is_logged_in=user_logged_in,
                           user_info=user_info)


@bp_link11.route('/link11/report')
def report_view():
    """공시 자료 생성/검토 페이지"""
    user_logged_in = is_logged_in()
    user_info = get_user_info()
    return render_template('link11_report.jsp',
                           is_logged_in=user_logged_in,
                           user_info=user_info)


# ============================================================================
# API Endpoints - 질문 관리
# ============================================================================

@bp_link11.route('/link11/api/questions', methods=['GET'])
def get_questions():
    """모든 질문 조회 (카테고리별 필터링 가능)"""
    try:
        category_id = request.args.get('category', type=int)

        with get_db() as conn:
            if category_id:
                category_name = get_category_name(category_id)
                if not category_name:
                    return jsonify({'success': False, 'message': '유효하지 않은 카테고리 ID입니다.'}), 400
                cursor = conn.execute('''
                    SELECT * FROM disclosure_questions
                    WHERE category = ?
                    ORDER BY sort_order
                ''', (category_name,))
            else:
                cursor = conn.execute('''
                    SELECT * FROM disclosure_questions
                    ORDER BY sort_order
                ''')

            questions = []
            for row in cursor.fetchall():
                q = dict(row)
                # JSON 필드 파싱
                if q.get('options'):
                    q['options'] = json.loads(q['options'])
                if q.get('dependent_question_ids'):
                    q['dependent_question_ids'] = json.loads(q['dependent_question_ids'])
                if q.get('evidence_list'):
                    q['evidence_list'] = json.loads(q['evidence_list'])
                questions.append(q)

            return jsonify({
                'success': True,
                'questions': questions,
                'count': len(questions)
            })

    except Exception as e:
        print(f"질문 조회 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/questions/<question_id>', methods=['GET'])
def get_question(question_id):
    """특정 질문 상세 조회"""
    try:
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT * FROM disclosure_questions WHERE id = ?
            ''', (question_id,))
            row = cursor.fetchone()

            if not row:
                return jsonify({'success': False, 'message': '질문을 찾을 수 없습니다.'}), 404

            q = dict(row)
            # JSON 필드 파싱
            if q.get('options'):
                q['options'] = json.loads(q['options'])
            if q.get('dependent_question_ids'):
                q['dependent_question_ids'] = json.loads(q['dependent_question_ids'])
            if q.get('evidence_list'):
                q['evidence_list'] = json.loads(q['evidence_list'])

            return jsonify({'success': True, 'question': q})

    except Exception as e:
        print(f"질문 상세 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/categories', methods=['GET'])
def get_categories():
    """카테고리 목록 조회"""
    try:
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT category, MIN(sort_order) as min_order
                FROM disclosure_questions
                GROUP BY category
                ORDER BY min_order
            ''')
            categories = [row['category'] for row in cursor.fetchall()]

            # 카테고리별 질문 수 계산
            category_stats = []
            for cat in categories:
                cursor = conn.execute('''
                    SELECT COUNT(*) as total,
                           SUM(CASE WHEN level = 1 THEN 1 ELSE 0 END) as level1_count
                    FROM disclosure_questions WHERE category = ?
                ''', (cat,))
                stats = cursor.fetchone()
                category_stats.append({
                    'id': get_category_id(cat),  # 카테고리 ID 추가
                    'name': cat,
                    'total': stats['total'],
                    'level1_count': stats['level1_count']
                })

            return jsonify({
                'success': True,
                'categories': category_stats
            })

    except Exception as e:
        print(f"카테고리 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# API Endpoints - 답변 관리
# ============================================================================

@bp_link11.route('/link11/api/answers', methods=['POST'])
def save_answer():
    """답변 저장"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401

    try:
        data = request.get_json()
        user_info = get_user_info()

        question_id = data.get('question_id')
        value = data.get('value')
        company_id = data.get('company_id', user_info.get('company_name', 'default'))
        year = data.get('year', datetime.now().year)

        if not question_id:
            return jsonify({'success': False, 'message': '질문 ID가 필요합니다.'}), 400

        # 값이 리스트인 경우 JSON으로 직렬화
        if isinstance(value, list):
            value = json.dumps(value, ensure_ascii=False)

        with get_db() as conn:
            # 기존 답변 확인
            cursor = conn.execute('''
                SELECT id FROM disclosure_answers
                WHERE question_id = ? AND company_id = ? AND year = ?
            ''', (question_id, company_id, year))
            existing = cursor.fetchone()

            if existing:
                # 기존 답변 업데이트
                conn.execute('''
                    UPDATE disclosure_answers
                    SET value = ?, status = 'completed', updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (value, existing['id']))
                answer_id = existing['id']
            else:
                # 새 답변 생성
                answer_id = generate_uuid()
                conn.execute('''
                    INSERT INTO disclosure_answers
                    (id, question_id, company_id, user_id, year, value, status)
                    VALUES (?, ?, ?, ?, ?, ?, 'completed')
                ''', (answer_id, question_id, company_id, str(user_info.get('user_id', '')), year, value))

            # "아니요" 또는 "NO" 선택 시 하위 질문들을 N/A로 자동 처리
            if value in ('아니요', 'NO', 'no', 'No'):
                _mark_dependent_questions_na(conn, question_id, company_id, year, str(user_info.get('user_id', '')))

            conn.commit()

            # 진행률 업데이트
            _update_session_progress(conn, company_id, year)

            return jsonify({
                'success': True,
                'message': '답변이 저장되었습니다.',
                'answer_id': answer_id
            })

    except Exception as e:
        print(f"답변 저장 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/answers/<int:user_id>/<int:year>', methods=['GET'])
def get_answers(user_id, year):
    """특정 회사의 특정 연도 모든 답변 조회"""
    try:
        company_id = get_company_name_by_user_id(user_id)
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT a.*, q.text as question_text, q.type as question_type
                FROM disclosure_answers a
                JOIN disclosure_questions q ON a.question_id = q.id
                WHERE a.company_id = ? AND a.year = ?
                ORDER BY q.sort_order
            ''', (company_id, year))

            answers = []
            for row in cursor.fetchall():
                a = dict(row)
                # JSON 값 파싱 시도
                if a.get('value'):
                    try:
                        a['value'] = json.loads(a['value'])
                    except (json.JSONDecodeError, TypeError):
                        pass  # 일반 문자열인 경우 그대로 유지
                answers.append(a)

            return jsonify({
                'success': True,
                'answers': answers,
                'count': len(answers)
            })

    except Exception as e:
        print(f"답변 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# API Endpoints - 증빙 자료 관리
# ============================================================================

@bp_link11.route('/link11/api/evidence', methods=['POST'])
def upload_evidence():
    """증빙 자료 업로드"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401

    try:
        user_info = get_user_info()

        # 파일 확인
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '파일이 없습니다.'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '파일이 선택되지 않았습니다.'}), 400

        if not allowed_file(file.filename):
            return jsonify({'success': False, 'message': '허용되지 않는 파일 형식입니다.'}), 400

        # 파일 크기 확인
        file.seek(0, 2)
        file_size = file.tell()
        file.seek(0)

        if file_size > MAX_FILE_SIZE:
            return jsonify({'success': False, 'message': '파일 크기가 100MB를 초과합니다.'}), 400

        # 폼 데이터 가져오기
        question_id = request.form.get('question_id')
        answer_id = request.form.get('answer_id')
        evidence_type = request.form.get('evidence_type', '')
        company_id = user_info.get('company_name', 'default')
        user_id = user_info.get('user_id', 0)
        year = int(request.form.get('year', datetime.now().year))

        # 안전한 파일명 생성
        original_filename = secure_filename(file.filename)
        file_ext = original_filename.rsplit('.', 1)[1].lower() if '.' in original_filename else ''
        stored_filename = f"{generate_uuid()}.{file_ext}"

        # user_id/연도별 폴더 생성 (숫자 기반으로 안전한 경로)
        upload_dir = os.path.join(UPLOAD_FOLDER, str(user_id), str(year))
        os.makedirs(upload_dir, exist_ok=True)

        # 파일 저장
        file_path = os.path.join(upload_dir, stored_filename)
        file.save(file_path)

        # 데이터베이스에 기록
        evidence_id = generate_uuid()
        with get_db() as conn:
            conn.execute('''
                INSERT INTO disclosure_evidence
                (id, answer_id, question_id, company_id, year, file_name, file_url,
                 file_size, file_type, evidence_type, uploaded_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                evidence_id, answer_id, question_id, company_id, year,
                original_filename, file_path, file_size, file_ext,
                evidence_type, str(user_info.get('user_id', ''))
            ))
            conn.commit()

        return jsonify({
            'success': True,
            'message': '파일이 업로드되었습니다.',
            'evidence': {
                'id': evidence_id,
                'file_name': original_filename,
                'file_size': file_size,
                'evidence_type': evidence_type
            }
        })

    except Exception as e:
        print(f"파일 업로드 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/evidence/<evidence_id>', methods=['DELETE'])
def delete_evidence(evidence_id):
    """증빙 자료 삭제"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401

    try:
        with get_db() as conn:
            # 증빙 자료 조회
            cursor = conn.execute('''
                SELECT * FROM disclosure_evidence WHERE id = ?
            ''', (evidence_id,))
            evidence = cursor.fetchone()

            if not evidence:
                return jsonify({'success': False, 'message': '증빙 자료를 찾을 수 없습니다.'}), 404

            # 파일 삭제
            file_path = evidence['file_url']
            if os.path.exists(file_path):
                os.remove(file_path)

            # 데이터베이스에서 삭제
            conn.execute('DELETE FROM disclosure_evidence WHERE id = ?', (evidence_id,))
            conn.commit()

        return jsonify({'success': True, 'message': '증빙 자료가 삭제되었습니다.'})

    except Exception as e:
        print(f"증빙 자료 삭제 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/evidence/<int:user_id>/<int:year>', methods=['GET'])
def get_evidence_list(user_id, year):
    """특정 회사/연도의 모든 증빙 자료 조회"""
    try:
        company_id = get_company_name_by_user_id(user_id)
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT e.*, q.text as question_text
                FROM disclosure_evidence e
                LEFT JOIN disclosure_questions q ON e.question_id = q.id
                WHERE e.company_id = ? AND e.year = ?
                ORDER BY e.uploaded_at DESC
            ''', (company_id, year))

            evidence_list = [dict(row) for row in cursor.fetchall()]

            return jsonify({
                'success': True,
                'evidence': evidence_list,
                'count': len(evidence_list)
            })

    except Exception as e:
        print(f"증빙 자료 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/evidence/missing/<int:user_id>/<int:year>', methods=['GET'])
def get_missing_evidence(user_id, year):
    """누락된 증빙 자료 조회"""
    try:
        company_id = get_company_name_by_user_id(user_id)
        with get_db() as conn:
            # 답변이 완료된 질문 중 증빙 자료가 필요하지만 없는 것 찾기
            cursor = conn.execute('''
                SELECT q.id, q.text, q.evidence_list
                FROM disclosure_questions q
                JOIN disclosure_answers a ON q.id = a.question_id
                    AND a.company_id = ? AND a.year = ? AND a.status = 'completed'
                WHERE q.evidence_list IS NOT NULL AND q.evidence_list != '[]'
                    AND NOT EXISTS (
                        SELECT 1 FROM disclosure_evidence e
                        WHERE e.question_id = q.id
                            AND e.company_id = ? AND e.year = ?
                    )
                ORDER BY q.sort_order
            ''', (company_id, year, company_id, year))

            missing = []
            for row in cursor.fetchall():
                q = dict(row)
                if q.get('evidence_list'):
                    q['evidence_list'] = json.loads(q['evidence_list'])
                missing.append(q)

            return jsonify({
                'success': True,
                'missing': missing,
                'count': len(missing)
            })

    except Exception as e:
        print(f"누락 증빙 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/evidence/download/<evidence_id>', methods=['GET'])
def download_evidence(evidence_id):
    """증빙 자료 다운로드"""
    try:
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT * FROM disclosure_evidence WHERE id = ?
            ''', (evidence_id,))
            evidence = cursor.fetchone()

            if not evidence:
                return jsonify({'success': False, 'message': '증빙 자료를 찾을 수 없습니다.'}), 404

            file_path = evidence['file_url']
            if not os.path.exists(file_path):
                return jsonify({'success': False, 'message': '파일을 찾을 수 없습니다.'}), 404

            return send_file(
                file_path,
                as_attachment=True,
                download_name=evidence['file_name']
            )

    except Exception as e:
        print(f"파일 다운로드 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/evidence/stats/<int:user_id>/<int:year>', methods=['GET'])
def get_evidence_stats(user_id, year):
    """증빙 자료 통계 조회"""
    try:
        company_id = get_company_name_by_user_id(user_id)
        with get_db() as conn:
            # 전체 파일 수
            cursor = conn.execute('''
                SELECT COUNT(*) as total FROM disclosure_evidence
                WHERE company_id = ? AND year = ?
            ''', (company_id, year))
            total = cursor.fetchone()['total']

            # 연결된 질문 수
            cursor = conn.execute('''
                SELECT COUNT(DISTINCT question_id) as questions FROM disclosure_evidence
                WHERE company_id = ? AND year = ? AND question_id IS NOT NULL
            ''', (company_id, year))
            questions = cursor.fetchone()['questions']

            # 총 용량
            cursor = conn.execute('''
                SELECT COALESCE(SUM(file_size), 0) as total_size FROM disclosure_evidence
                WHERE company_id = ? AND year = ?
            ''', (company_id, year))
            total_size = cursor.fetchone()['total_size']

            # 최근 7일 업로드
            cursor = conn.execute('''
                SELECT COUNT(*) as recent FROM disclosure_evidence
                WHERE company_id = ? AND year = ?
                AND uploaded_at >= datetime('now', '-7 days')
            ''', (company_id, year))
            recent = cursor.fetchone()['recent']

            return jsonify({
                'success': True,
                'stats': {
                    'total': total,
                    'questions': questions,
                    'total_size': total_size,
                    'recent': recent
                }
            })

    except Exception as e:
        print(f"증빙 통계 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# API Endpoints - 진행 상황
# ============================================================================

@bp_link11.route('/link11/api/progress/<int:user_id>/<int:year>', methods=['GET'])
def get_progress(user_id, year):
    """공시 진행 상황 조회"""
    try:
        company_id = get_company_name_by_user_id(user_id)
        with get_db() as conn:
            # 카테고리별 진행 상황 계산 (세션 유무와 관계없이 항상 계산)
            cursor = conn.execute('''
                SELECT q.category,
                       COUNT(DISTINCT q.id) as total,
                       COUNT(DISTINCT CASE WHEN a.status = 'completed' THEN a.question_id END) as completed
                FROM disclosure_questions q
                LEFT JOIN disclosure_answers a ON q.id = a.question_id
                    AND a.company_id = ? AND a.year = ?
                GROUP BY q.category
            ''', (company_id, year))

            categories = {}
            total_questions = 0
            answered_questions = 0

            for row in cursor.fetchall():
                categories[row['category']] = {
                    'total': row['total'],
                    'completed': row['completed'],
                    'rate': round((row['completed'] / row['total']) * 100) if row['total'] > 0 else 0
                }
                total_questions += row['total']
                answered_questions += row['completed']

            # 전체 진행률 계산
            completion_rate = round((answered_questions / total_questions) * 100) if total_questions > 0 else 0

            # 세션 조회
            cursor = conn.execute('''
                SELECT * FROM disclosure_sessions
                WHERE company_id = ? AND year = ?
            ''', (company_id, year))
            session_data = cursor.fetchone()

            session_dict = dict(session_data) if session_data else None

            return jsonify({
                'success': True,
                'session': session_dict,
                'progress': {
                    'total_questions': total_questions,
                    'answered_questions': answered_questions,
                    'completion_rate': completion_rate
                },
                'categories': categories
            })

    except Exception as e:
        print(f"진행 상황 조회 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/session', methods=['POST'])
def create_or_update_session():
    """공시 세션 생성/업데이트"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401

    try:
        data = request.get_json()
        user_info = get_user_info()

        company_id = data.get('company_id', user_info.get('company_name', 'default'))
        year = data.get('year', datetime.now().year)
        status = data.get('status', 'draft')

        with get_db() as conn:
            # 기존 세션 확인
            cursor = conn.execute('''
                SELECT id FROM disclosure_sessions
                WHERE company_id = ? AND year = ?
            ''', (company_id, year))
            existing = cursor.fetchone()

            if existing:
                # 세션 업데이트
                conn.execute('''
                    UPDATE disclosure_sessions
                    SET status = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (status, existing['id']))
                session_id = existing['id']
            else:
                # 새 세션 생성
                session_id = generate_uuid()
                conn.execute('''
                    INSERT INTO disclosure_sessions
                    (id, company_id, user_id, year, status)
                    VALUES (?, ?, ?, ?, ?)
                ''', (session_id, company_id, str(user_info.get('user_id', '')), year, status))

            conn.commit()

            # 진행률 업데이트
            _update_session_progress(conn, company_id, year)

            return jsonify({
                'success': True,
                'message': '세션이 저장되었습니다.',
                'session_id': session_id
            })

    except Exception as e:
        print(f"세션 저장 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


def _mark_dependent_questions_na(conn, parent_question_id, company_id, year, user_id):
    """부모 질문이 '아니요'일 때 하위 질문들을 N/A로 자동 처리 (내부 함수)"""
    try:
        print(f"[DEBUG] _mark_dependent_questions_na 호출: parent={parent_question_id}, company={company_id}, year={year}")

        # 해당 질문의 dependent_question_ids 조회
        cursor = conn.execute('''
            SELECT dependent_question_ids FROM disclosure_questions
            WHERE id = ?
        ''', (parent_question_id,))
        result = cursor.fetchone()

        if not result or not result['dependent_question_ids']:
            print(f"[DEBUG] {parent_question_id}에 하위 질문 없음")
            return

        # dependent_question_ids 파싱
        dependent_ids_str = result['dependent_question_ids']
        print(f"[DEBUG] dependent_ids_str: {dependent_ids_str}")
        try:
            dependent_ids = json.loads(dependent_ids_str)
        except (json.JSONDecodeError, TypeError):
            # JSON이 아닌 경우 콤마로 분리 시도
            dependent_ids = [id.strip() for id in dependent_ids_str.split(',') if id.strip()]

        print(f"[DEBUG] 파싱된 dependent_ids: {dependent_ids}")

        if not dependent_ids:
            return

        # 각 하위 질문에 대해 N/A 답변 저장
        for dep_id in dependent_ids:
            print(f"[DEBUG] N/A 처리 중: {dep_id}")
            # 기존 답변 확인
            cursor = conn.execute('''
                SELECT id FROM disclosure_answers
                WHERE question_id = ? AND company_id = ? AND year = ?
            ''', (dep_id, company_id, year))
            existing = cursor.fetchone()

            if existing:
                # 기존 답변을 N/A로 업데이트
                conn.execute('''
                    UPDATE disclosure_answers
                    SET value = 'N/A', status = 'completed', updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (existing['id'],))
            else:
                # 새 N/A 답변 생성
                answer_id = generate_uuid()
                conn.execute('''
                    INSERT INTO disclosure_answers
                    (id, question_id, company_id, user_id, year, value, status)
                    VALUES (?, ?, ?, ?, ?, 'N/A', 'completed')
                ''', (answer_id, dep_id, company_id, user_id, year))

            # 재귀적으로 해당 하위 질문의 하위 질문도 처리
            _mark_dependent_questions_na(conn, dep_id, company_id, year, user_id)

    except Exception as e:
        print(f"하위 질문 N/A 처리 오류: {e}")
        import traceback
        traceback.print_exc()


def _update_session_progress(conn, company_id, year, user_id=None):
    """세션 진행률 자동 업데이트 (내부 함수)"""
    try:
        # 답변 완료된 질문 수 계산 (N/A 포함)
        cursor = conn.execute('''
            SELECT COUNT(DISTINCT question_id) as answered
            FROM disclosure_answers
            WHERE company_id = ? AND year = ? AND status = 'completed'
        ''', (company_id, year))
        result = cursor.fetchone()
        answered = result['answered'] if result else 0

        # 전체 질문 수
        cursor = conn.execute('SELECT COUNT(*) as total FROM disclosure_questions')
        total_result = cursor.fetchone()
        total = total_result['total'] if total_result else 65

        # 진행률 계산
        completion_rate = round((answered / total) * 100) if total > 0 else 0

        print(f"[DEBUG] 진행률 계산: answered={answered}, total={total}, rate={completion_rate}%")

        # 세션 존재 여부 확인
        cursor = conn.execute('''
            SELECT id FROM disclosure_sessions
            WHERE company_id = ? AND year = ?
        ''', (company_id, year))
        existing = cursor.fetchone()

        if existing:
            # 세션 업데이트
            conn.execute('''
                UPDATE disclosure_sessions
                SET answered_questions = ?, completion_rate = ?, updated_at = CURRENT_TIMESTAMP,
                    status = CASE WHEN ? = 100 THEN 'completed' ELSE 'in_progress' END
                WHERE company_id = ? AND year = ?
            ''', (answered, completion_rate, completion_rate, company_id, year))
        else:
            # 세션 자동 생성
            session_id = generate_uuid()
            status = 'completed' if completion_rate == 100 else ('in_progress' if answered > 0 else 'draft')
            conn.execute('''
                INSERT INTO disclosure_sessions
                (id, company_id, user_id, year, status, total_questions, answered_questions, completion_rate)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (session_id, company_id, user_id or '', year, status, total, answered, completion_rate))

        conn.commit()

    except Exception as e:
        print(f"진행률 업데이트 오류: {e}")
        import traceback
        traceback.print_exc()


# ============================================================================
# API Endpoints - 보고서 생성
# ============================================================================

@bp_link11.route('/link11/api/report/generate/<int:user_id>/<int:year>', methods=['GET'])
def generate_report(user_id, year):
    """공시 자료 보고서 생성"""
    try:
        company_id = get_company_name_by_user_id(user_id)
        format_type = request.args.get('format', 'json')  # json, pdf, excel

        with get_db() as conn:
            # 모든 답변 조회
            cursor = conn.execute('''
                SELECT q.*, a.value, a.status
                FROM disclosure_questions q
                LEFT JOIN disclosure_answers a ON q.id = a.question_id
                    AND a.company_id = ? AND a.year = ?
                ORDER BY q.sort_order
            ''', (company_id, year))

            report_data = {
                'company_id': company_id,
                'year': year,
                'generated_at': datetime.now().isoformat(),
                'categories': {}
            }

            for row in cursor.fetchall():
                category = row['category']
                if category not in report_data['categories']:
                    report_data['categories'][category] = {
                        'questions': [],
                        'completed': 0,
                        'total': 0
                    }

                q_data = {
                    'id': row['id'],
                    'text': row['text'],
                    'type': row['type'],
                    'value': row['value'],
                    'status': row['status'] or 'pending'
                }

                # JSON 값 파싱
                if q_data['value']:
                    try:
                        q_data['value'] = json.loads(q_data['value'])
                    except (json.JSONDecodeError, TypeError):
                        pass

                report_data['categories'][category]['questions'].append(q_data)
                report_data['categories'][category]['total'] += 1
                if row['status'] == 'completed':
                    report_data['categories'][category]['completed'] += 1

            if format_type == 'json':
                return jsonify({
                    'success': True,
                    'report': report_data
                })
            else:
                # PDF/Excel 생성은 추후 구현
                return jsonify({
                    'success': False,
                    'message': f'{format_type} 형식은 아직 지원되지 않습니다.'
                }), 501

    except Exception as e:
        print(f"보고서 생성 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/report/download/<int:user_id>/<int:year>', methods=['GET'])
def download_report(user_id, year):
    """공시 자료 Excel 다운로드"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from io import BytesIO

        company_id = get_company_name_by_user_id(user_id)

        with get_db() as conn:
            # 모든 답변 조회
            cursor = conn.execute('''
                SELECT q.id, q.category, q.subcategory, q.text, q.type, q.level,
                       a.value, a.status
                FROM disclosure_questions q
                LEFT JOIN disclosure_answers a ON q.id = a.question_id
                    AND a.company_id = ? AND a.year = ?
                ORDER BY q.sort_order
            ''', (company_id, year))

            questions = [dict(row) for row in cursor.fetchall()]

        # Excel 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "정보보호공시"

        # 스타일 정의
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        category_font = Font(bold=True, size=11)
        category_fill = PatternFill(start_color="e0f2fe", end_color="e0f2fe", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 제목
        ws.merge_cells('A1:D1')
        ws['A1'] = f"정보보호공시 - {company_id} ({year}년)"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:D2')
        ws['A2'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal='center')

        # 헤더
        headers = ['질문ID', '질문', '답변', '상태']
        ws.append([])  # 빈 줄
        ws.append(headers)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # 데이터
        current_category = None
        row_num = 5

        for q in questions:
            # 카테고리 헤더
            if q['category'] != current_category:
                current_category = q['category']
                ws.merge_cells(f'A{row_num}:D{row_num}')
                ws[f'A{row_num}'] = current_category
                ws[f'A{row_num}'].font = category_font
                ws[f'A{row_num}'].fill = category_fill
                ws[f'A{row_num}'].border = thin_border
                row_num += 1

            # 질문 데이터
            value = q['value'] or ''
            if value.startswith('[') and value.endswith(']'):
                try:
                    value = ', '.join(json.loads(value))
                except:
                    pass

            status = '완료' if q['status'] == 'completed' else ('N/A' if q['value'] == 'N/A' else '미완료')

            ws.append([q['id'], q['text'], value, status])

            for col in range(1, 5):
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

            row_num += 1

        # 열 너비 조정
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 10

        # ============================================
        # 증빙자료 시트 추가 (내부 관리용)
        # ============================================
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT e.id, e.question_id, e.file_name, e.file_size, e.file_type,
                       e.evidence_type, e.uploaded_at, e.uploaded_by,
                       q.category, q.text as question_text
                FROM disclosure_evidence e
                LEFT JOIN disclosure_questions q ON e.question_id = q.id
                WHERE e.company_id = ? AND e.year = ?
                ORDER BY q.category, e.uploaded_at
            ''', (company_id, year))
            evidence_list = [dict(row) for row in cursor.fetchall()]

        if evidence_list:
            ws_evidence = wb.create_sheet(title="증빙자료")

            # 제목
            ws_evidence.merge_cells('A1:F1')
            ws_evidence['A1'] = f"증빙자료 목록 - {company_id} ({year}년)"
            ws_evidence['A1'].font = Font(bold=True, size=16)
            ws_evidence['A1'].alignment = Alignment(horizontal='center')

            ws_evidence.merge_cells('A2:F2')
            ws_evidence['A2'] = f"총 {len(evidence_list)}개 파일"
            ws_evidence['A2'].alignment = Alignment(horizontal='center')

            # 헤더
            evidence_headers = ['카테고리', '관련 질문', '파일명', '파일크기', '유형', '업로드일']
            ws_evidence.append([])  # 빈 줄
            ws_evidence.append(evidence_headers)

            for col, header in enumerate(evidence_headers, 1):
                cell = ws_evidence.cell(row=4, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # 데이터
            row_num = 5
            for ev in evidence_list:
                # 파일 크기 포맷
                size = ev['file_size'] or 0
                if size < 1024:
                    size_str = f"{size} B"
                elif size < 1024 * 1024:
                    size_str = f"{size / 1024:.1f} KB"
                else:
                    size_str = f"{size / (1024 * 1024):.1f} MB"

                # 업로드일 포맷
                uploaded_at = ev['uploaded_at'] or ''
                if uploaded_at and len(uploaded_at) >= 10:
                    uploaded_at = uploaded_at[:10]  # YYYY-MM-DD만

                ws_evidence.append([
                    ev['category'] or '',
                    ev['question_text'] or '',
                    ev['file_name'] or '',
                    size_str,
                    ev['evidence_type'] or '',
                    uploaded_at
                ])

                for col in range(1, 7):
                    cell = ws_evidence.cell(row=row_num, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)

                row_num += 1

            # 열 너비 조정
            ws_evidence.column_dimensions['A'].width = 20
            ws_evidence.column_dimensions['B'].width = 50
            ws_evidence.column_dimensions['C'].width = 35
            ws_evidence.column_dimensions['D'].width = 12
            ws_evidence.column_dimensions['E'].width = 15
            ws_evidence.column_dimensions['F'].width = 12

        # 파일 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'정보보호공시_{company_id}_{year}.xlsx'
        )

    except ImportError:
        return jsonify({'success': False, 'message': 'openpyxl 라이브러리가 필요합니다.'}), 500
    except Exception as e:
        print(f"보고서 다운로드 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# API Endpoints - 공시 제출
# ============================================================================

@bp_link11.route('/link11/api/submit/<int:user_id>/<int:year>', methods=['POST'])
def submit_disclosure(user_id, year):
    """공시 제출"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401

    try:
        data = request.get_json() or {}
        user_info = get_user_info()

        company_id = get_company_name_by_user_id(user_id)
        submission_details = data.get('details', '')

        with get_db() as conn:
            # 세션 조회
            cursor = conn.execute('''
                SELECT * FROM disclosure_sessions
                WHERE company_id = ? AND year = ?
            ''', (company_id, year))
            session_data = cursor.fetchone()

            if not session_data:
                return jsonify({'success': False, 'message': '공시 세션을 찾을 수 없습니다.'}), 404

            # 완료율 확인 (필수 질문 모두 완료 여부)
            if session_data['completion_rate'] < 100:
                return jsonify({
                    'success': False,
                    'message': f'모든 질문에 답변해야 제출할 수 있습니다. (현재 {session_data["completion_rate"]}%)'
                }), 400

            # 제출 기록 생성
            submission_id = generate_uuid()
            confirmation_number = f"DISC-{year}-{datetime.now().strftime('%Y%m%d%H%M%S')}"

            conn.execute('''
                INSERT INTO disclosure_submissions
                (id, session_id, company_id, year, submitted_by, submission_details, confirmation_number, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'submitted')
            ''', (
                submission_id, session_data['id'], company_id, year,
                str(user_info.get('user_id', '')), submission_details, confirmation_number
            ))

            # 세션 상태 업데이트
            conn.execute('''
                UPDATE disclosure_sessions
                SET status = 'submitted', submitted_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (session_data['id'],))

            conn.commit()

            return jsonify({
                'success': True,
                'message': '공시가 제출되었습니다.',
                'submission_id': submission_id,
                'confirmation_number': confirmation_number
            })

    except Exception as e:
        print(f"공시 제출 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/submissions/<int:user_id>', methods=['GET'])
def get_submissions(user_id):
    """제출 이력 조회"""
    try:
        company_id = get_company_name_by_user_id(user_id)
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT * FROM disclosure_submissions
                WHERE company_id = ?
                ORDER BY submitted_at DESC
            ''', (company_id,))

            submissions = [dict(row) for row in cursor.fetchall()]

            return jsonify({
                'success': True,
                'submissions': submissions,
                'count': len(submissions)
            })

    except Exception as e:
        print(f"제출 이력 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# API Endpoints - 데이터 관리 (새로하기, 이전 자료 불러오기)
# ============================================================================

@bp_link11.route('/link11/api/reset/<int:user_id>/<int:year>', methods=['POST'])
def reset_disclosure(user_id, year):
    """새로하기 - 현재 연도 데이터 초기화"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401

    try:
        company_id = get_company_name_by_user_id(user_id)

        with get_db() as conn:
            # 답변 삭제
            conn.execute('''
                DELETE FROM disclosure_answers
                WHERE company_id = ? AND year = ?
            ''', (company_id, year))

            # 세션 삭제
            conn.execute('''
                DELETE FROM disclosure_sessions
                WHERE company_id = ? AND year = ?
            ''', (company_id, year))

            # 증빙자료는 유지 (필요시 별도 삭제)

            conn.commit()

            return jsonify({
                'success': True,
                'message': f'{year}년 데이터가 초기화되었습니다.'
            })

    except Exception as e:
        print(f"데이터 초기화 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/available-years/<int:user_id>', methods=['GET'])
def get_available_years(user_id):
    """이전 자료가 있는 연도 목록 조회"""
    try:
        company_id = get_company_name_by_user_id(user_id)
        current_year = datetime.now().year

        with get_db() as conn:
            cursor = conn.execute('''
                SELECT DISTINCT year, COUNT(*) as answer_count
                FROM disclosure_answers
                WHERE company_id = ? AND year < ?
                GROUP BY year
                ORDER BY year DESC
            ''', (company_id, current_year))

            years = []
            for row in cursor.fetchall():
                years.append({
                    'year': row['year'],
                    'answer_count': row['answer_count']
                })

            return jsonify({
                'success': True,
                'years': years
            })

    except Exception as e:
        print(f"연도 목록 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/copy-from-year/<int:user_id>/<int:source_year>/<int:target_year>', methods=['POST'])
def copy_from_year(user_id, source_year, target_year):
    """이전 자료 불러오기 - 이전 연도 데이터를 현재 연도로 복사"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401

    try:
        user_info = get_user_info()
        company_id = get_company_name_by_user_id(user_id)

        if source_year == target_year:
            return jsonify({'success': False, 'message': '같은 연도로는 복사할 수 없습니다.'}), 400

        with get_db() as conn:
            # 기존 타겟 연도 데이터 삭제
            conn.execute('''
                DELETE FROM disclosure_answers
                WHERE company_id = ? AND year = ?
            ''', (company_id, target_year))

            conn.execute('''
                DELETE FROM disclosure_sessions
                WHERE company_id = ? AND year = ?
            ''', (company_id, target_year))

            # 이전 연도 답변 복사
            cursor = conn.execute('''
                SELECT question_id, value, status
                FROM disclosure_answers
                WHERE company_id = ? AND year = ?
            ''', (company_id, source_year))

            copied_count = 0
            for row in cursor.fetchall():
                answer_id = generate_uuid()
                conn.execute('''
                    INSERT INTO disclosure_answers
                    (id, question_id, company_id, user_id, year, value, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (
                    answer_id, row['question_id'], company_id,
                    str(user_info.get('user_id', '')), target_year,
                    row['value'], row['status']
                ))
                copied_count += 1

            conn.commit()

            # 진행률 업데이트
            _update_session_progress(conn, company_id, target_year, str(user_info.get('user_id', '')))

            return jsonify({
                'success': True,
                'message': f'{source_year}년 자료를 {target_year}년으로 복사했습니다. ({copied_count}개 답변)',
                'copied_count': copied_count
            })

    except Exception as e:
        print(f"자료 복사 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
