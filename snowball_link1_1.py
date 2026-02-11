from flask import Blueprint, request, render_template, session, url_for, jsonify
from datetime import datetime
import os
import json
import time
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from auth import log_user_activity, login_required
from openai import OpenAI
from google import genai

# Blueprint 생성
bp_link1_1 = Blueprint('link1_1', __name__)

# ================================
# 마스터 통제 항목 정의 (표준 ITGC 36개)
# ================================
MASTER_ITGC_CONTROLS = [
    {
        "id": "PWC01", "name": "IT 정책서 관리", "category": "IT정책", "objective": "IT 정책의 명문화 및 준수 보장",
        "risk_code": "R-01", "risk_description": "IT 프로세스에 대한 정책이 수립되지 않을 위험 - ITGCs",
        "control_description": "1. 하기 영역에 대한 IT운영프로세스에 관한 정책이 IT정책서에 명시되어 있음\n - Access to programs and data\n - Changes to application Programs\n - Program Development\n - Computer Operations\n2. IT정책서는 매년 경영정보팀장의 검토 및 승인을 득함\n3. 검토된 IT정책서는 그룹웨어 게시판에 게시됨",
        "type": "Manual", "frequency": "연", "method": "예방",
        "test_procedure_manual": "1. IT운영프로세스가 명시된 '운영보안 규정' 정책서를 입수함\n2. 입수한 운영보안 규정상 제·개정 이력을 검토하고 매년 경영정보팀장의 검토 및 승인을 득하였는지 여부를 검토함\n3. 검토 및 승인을 득한 운영보안 규정이 그룹웨어 내 문서관리-규정문서 항목상 전체 공지 되고 있는지 여부를 확인함",
        "test_procedure_auto": "1. IT정책서 관리 시스템의 자동 알림/승인 워크플로우 설정을 확인함\n2. 모집단에서 임의의 샘플 1건을 추출하여 시스템에 의해 자동으로 검토 요청 및 게시가 수행되었는지 확인함"
    },
    {
        "id": "APD01", "name": "Application 권한 승인", "category": "계정관리", "objective": "인가된 사용자에게 적절한 권한 부여",
        "risk_code": "R-02", "risk_description": "시스템 사용자가 시스템에서 규정하는 승인과 업무분장을 우회 할 위험",
        "control_description": "권한 요청에 대해 요청 부서장 또는 IT팀장 등 책임자의 승인을 득한다",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "아래의 순서에 따라 통제 테스트를 수행한다.\n\n1. 모집단에서 샘플을 추출한다.\n2. 추출된 샘플의 권한 요청서 작성 여부와 부서장 등 책임권자의 승인 여부를 확인한다.\n3. 권한 요청 및 승인 완료 후 권한 부여가 수행되었는지 확인한다\n  - 권한 요청서 작성/승인 일자 < 권한부여 일자",
        "test_procedure_auto": "1. 권한 부여 시 시스템에 의한 자동 승인 워크플로우 로직을 확인한다\n2. 모집단에서 임의의 샘플 1건을 추출하여 시스템에 의해 승인 후 권한이 부여되었는지 확인한다"
    },
    {
        "id": "APD02", "name": "Application 부서이동자 권한 회수", "category": "계정관리", "objective": "부서 이동시 불필요 권한 즉시 회수",
        "risk_code": "R-03", "risk_description": "시스템 사용자가 시스템에서 규정하는 승인과 업무분장을 우회 할 위험",
        "control_description": "전배 등 기존 권한에 대한 회수 사유 발생시 즉시 회수한다",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "부서이동(전배)자의 기존 권한 회수 여부를 검토한다\n\n1. 모집단에서 샘플을 추출한다.\n2. 기존 권한의 회수 여부를 확인한다\n  - 권한 회수 일자 < 부서이동(전배) 일자",
        "test_procedure_auto": "1. 인사시스템과 연동된 권한 자동회수 프로그램 로직을 확인한다\n2. 모집단(부서이동자)에서 임의의 샘플 1건을 추출하여 기존 권한이 자동 회수되었는지 확인한다"
    },
    {
        "id": "APD03", "name": "Application 퇴사자 접근권한 회수", "category": "계정관리", "objective": "퇴사자 시스템 접근 차단",
        "risk_code": "R-04", "risk_description": "시스템 사용자가 시스템에서 규정하는 승인과 업무분장을 우회 할 위험",
        "control_description": "퇴사 등 계정삭제(비활성화) 사유 발생시 시스템에 의해 계정이 자동 삭제(비활성화)된다",
        "type": "Auto", "frequency": "기타", "method": "예방",
        "test_procedure_auto": "Application 계정 삭제(비활성화)에 대해 아래와 같이 검토한다\n\n1. 권한회수 프로그램 Source에 Logic이 정확히 구현되어 있는지 확인한다\n2. 모집단(퇴사자)에서 임의의 샘플 1건을 추출하여 퇴사자 계정이 모두 삭제(비활성화)되었는지 확인한다",
        "test_procedure_manual": "1. 모집단(퇴사자)에서 샘플을 추출한다\n2. 퇴사자 계정 삭제 요청서 작성 및 승인 여부를 확인한다\n3. 승인 후 계정이 삭제(비활성화)되었는지 확인한다"
    },
    {
        "id": "APD04", "name": "Application 권한 Monitoring", "category": "계정관리", "objective": "권한 부여의 적정성 주기적 검토",
        "risk_code": "R-05", "risk_description": "시스템 사용자가 시스템에서 규정하는 승인과 업무분장을 우회 할 위험",
        "control_description": "전체 사용자가 보유한 권한에 대해 Monitoring이 수행되며 결과는 책임자에 의해 승인된다",
        "type": "Manual", "frequency": "분기", "method": "적발",
        "test_procedure_manual": "권한 Monitoring 통제에 대해 아래와 같이 검토한다\n\n1. 모집단에서 샘플을 추출한다\n2. Application/서버(OS/DB)/VPN에 대해 보유권한 적정성이 검토되었는지 확인한다\n3. 승인권자의 승인 여부를 확인한다",
        "test_procedure_auto": "1. 시스템에 의한 권한 모니터링 자동 수행 로직을 확인한다\n2. 자동 생성된 모니터링 보고서 샘플 1건을 추출하여 정상 생성 여부를 확인한다"
    },
    {
        "id": "APD05", "name": "Application 관리자 권한 제한", "category": "계정관리", "objective": "관리자 권한 오남용 방지",
        "risk_code": "R-06", "risk_description": "위험하고 강력한 계정사용자가 시스템에서 규정하는 승인과 업무분장을 우회 할 위험",
        "control_description": "관리자 권한은 지정된 담당자로 제한된다",
        "type": "Auto", "frequency": "기타", "method": "예방",
        "test_procedure_auto": "1. 관리자 권한을 보유한 인원을 추출한다.\n2. 해당 인원의 부서, 직책, 직무 등을 검토하여 적절성 여부를 확인한다",
        "test_procedure_manual": "1. 관리자 권한 보유 현황을 추출한다\n2. 모집단에서 샘플을 추출하여 관리자 권한 부여에 대한 승인 여부를 확인한다"
    },
    {
        "id": "APD06", "name": "Application 패스워드", "category": "계정관리", "objective": "강력한 인증 체계 유지",
        "risk_code": "R-07", "risk_description": "취약한 패스워드 정책이나 보안 설정으로 인해 시스템 접근 통제를 우회할 위험",
        "control_description": "1. Application 패스워드는 최소 아래 기준 이상으로 설정함\n 1) 복잡성: 문자, 숫자, 특수문자 포함\n 2) 최소 길이: 8자리 이상",
        "type": "Auto", "frequency": "기타", "method": "예방",
        "test_procedure_auto": "1. 시스템의 패스워드 정책 설정(최소 길이, 복잡성, 만료 기간 등) 환경설정 값을 확인한다.\n2. 설정값이 회사의 정보보안 정책(최소 8자리 이상, 문자/숫자/특수문자 조합 등)과 일치하는지 검토한다.",
        "test_procedure_manual": "1. 패스워드 정책 변경 이력을 조회한다\n2. 모집단에서 샘플을 추출하여 변경 요청 및 승인 여부를 확인한다"
    },
    {
        "id": "APD07", "name": "Data 변경 승인", "category": "데이터관리", "objective": "DB 데이터 직접 변경의 승인 체계 확보",
        "risk_code": "R-08", "risk_description": "기본적인 거래정보나 마스터 데이터를 부적절하게 변경 할 위험",
        "control_description": "데이터 직접 변경은 적절한 승인권자의 승인을 득한다",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "1. 데이터 변경 이력을 조회하고, 조회화면(조회 조건, 조회된 전체 데이터 수 확인 가능한 화면) 및 조회 결과를 문서화함\n2. 위의 1.에서 조회된 데이터 변경의 발생빈도에 따라 샘플 Guide의 샘플 수를 적용하여 샘플을 선정함\n3. 샘플에 대하여 승인권자의 승인을 득하였는지 여부를 확인함",
        "test_procedure_auto": "1. 데이터 변경 시 시스템에 의한 자동 승인 워크플로우 로직을 확인한다\n2. 모집단에서 임의의 샘플 1건을 추출하여 자동 승인 후 변경이 수행되었는지 확인한다"
    },
    {
        "id": "APD08", "name": "Data 변경 권한 제한", "category": "데이터관리", "objective": "DB 직접 접근 권한 제한",
        "risk_code": "R-09", "risk_description": "인가받지 않은 DB 접근으로 인하여 거래정보나 마스터 데이터가 부적절하게 변경 될 위험",
        "control_description": "데이터 변경 권한은 지정된 담당자로 제한된다",
        "type": "Auto", "frequency": "기타", "method": "예방",
        "test_procedure_auto": "1. 데이터 변경 권한을 보유한 인원을 추출한다.\n2. 해당 인원의 부서, 직책, 직무 등을 검토하여 적절성 여부를 확인한다.",
        "test_procedure_manual": "1. 데이터 변경 권한 부여 이력을 조회한다\n2. 모집단에서 샘플을 추출하여 권한 부여에 대한 승인 여부를 확인한다"
    },
    {
        "id": "APD09", "name": "OS 접근권한 승인", "category": "OS관리", "objective": "서버 인프라 접근 승인 절차 준수",
        "risk_code": "R-10", "risk_description": "인가받지 않은 OS 접근으로 인하여 운영서버 상의 프로그램이 부적절하게 변경 될 위험",
        "control_description": "OS에 접근권한은 적절한 승인권자의 승인을 득한 후 부여한다",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "1. 테스트 기간 동안 신규 부여된 OS 접근권한 목록을 조회하고, 조회화면 및 조회 결과를 문서화한다.\n2. 위의 1.에서 조회된 모집단의 발생빈도에 따라 샘플 Guide의 샘플 수를 적용하여 샘플을 선정한다.\n3. 샘플에 대하여 승인권자의 승인을 득하였는지 여부를 확인한다.",
        "test_procedure_auto": "1. OS 접근권한 자동 부여 시스템 로직을 확인한다\n2. 모집단에서 임의의 샘플 1건을 추출하여 자동 승인 후 권한이 부여되었는지 확인한다"
    },
    {
        "id": "APD10", "name": "OS 패스워드", "category": "OS관리", "objective": "인프라 계정 보안 강화",
        "risk_code": "R-11", "risk_description": "취약한 패스워드 정책이나 보안 설정으로 인해 시스템 접근 통제를 우회할 위험",
        "control_description": "OS 패스워드는 회사의 규정에 따라 최소자리, 복잡성 등을 설정한다",
        "type": "Auto", "frequency": "기타", "method": "예방",
        "test_procedure_auto": "1. OS서버 또는 OS 접근제어 Tool에 설정된 패스워드를 확인하여 아래 기준으로 설정되어 있는지 여부를 검토한다.\n 1) 복잡성: 문자, 숫자, 특수문자 포함\n 2) 최소 길이: 8자리 이상",
        "test_procedure_manual": "1. OS 패스워드 정책 변경 이력을 조회한다\n2. 모집단에서 샘플을 추출하여 변경 요청 및 승인 여부를 확인한다"
    },
    {
        "id": "APD11", "name": "OS 관리자 권한 제한", "category": "OS관리", "objective": "서버 관리자 권한 최소화",
        "risk_code": "R-12", "risk_description": "인가받지 않은 OS 접근으로 인하여 운영서버 상의 프로그램이 부적절하게 변경 될 위험",
        "control_description": "OS Super User 권한은 지정된 인원으로 제한한다",
        "type": "Auto", "frequency": "기타", "method": "예방",
        "test_procedure_auto": "1. OS 또는 OS 접근제어 Tool에서 슈퍼권한자 목록을 확인함\n2. 해당 권한을 보유한 인원의 부서, 직책, 직무 등을 검토하여 적절성 여부를 확인한다.",
        "test_procedure_manual": "1. OS 관리자 권한 부여 이력을 조회한다\n2. 모집단에서 샘플을 추출하여 권한 부여에 대한 승인 여부를 확인한다"
    },
    {
        "id": "APD12", "name": "DB 접근권한 승인", "category": "DB관리", "objective": "DB 접근 승인 절차 준수",
        "risk_code": "R-13", "risk_description": "인가받지 않은 DB 접근으로 인하여 거래정보나 마스터 데이터가 부적절하게 변경 될 위험",
        "control_description": "DB에 접근권한은 적절한 승인권자의 승인을 득한 후 부여한다",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "1. DB 또는 DB 접근제어 Tool에서 테스트 기간 동안 신규 부여된 접근권한 목록을 조회하고, 조회화면 및 조회 결과를 문서화함\n2. 위의 1.에서 조회된 모집단의 발생빈도에 따라 샘플 Guide의 샘플 수를 적용하여 샘플을 선정함\n3. 샘플에 대하여 적절한 승인권자의 승인을 득하였는지 여부를 확인함",
        "test_procedure_auto": "1. DB 접근권한 자동 부여 시스템 로직을 확인한다\n2. 모집단에서 임의의 샘플 1건을 추출하여 자동 승인 후 권한이 부여되었는지 확인한다"
    },
    {
        "id": "APD13", "name": "DB 패스워드", "category": "DB관리", "objective": "DB 계정 보안 강화",
        "risk_code": "R-14", "risk_description": "취약한 패스워드 정책이나 보안 설정으로 인해 시스템 접근 통제를 우회할 위험",
        "control_description": "DB 패스워드는 회사의 규정에 따라 최소자리, 복잡성 등을 설정한다",
        "type": "Auto", "frequency": "기타", "method": "예방",
        "test_procedure_auto": "1. DB의 패스워드 정책을 확인하여 아래 기준으로 설정되어 있는지 여부를 검토함.\n 1) 복잡성: 문자, 숫자, 특수문자 포함\n 2) 최소 길이: 8자리 이상",
        "test_procedure_manual": "1. DB 패스워드 정책 변경 이력을 조회한다\n2. 모집단에서 샘플을 추출하여 변경 요청 및 승인 여부를 확인한다"
    },
    {
        "id": "APD14", "name": "DB 관리자 권한 제한", "category": "DB관리", "objective": "DB 관리자 권한 최소화",
        "risk_code": "R-15", "risk_description": "인가받지 않은 DB 접근으로 인하여 거래정보나 마스터 데이터가 부적절하게 변경 될 위험",
        "control_description": "DB Super User 권한은 지정된 인원으로 제한한다",
        "type": "Auto", "frequency": "기타", "method": "예방",
        "test_procedure_auto": "1. DB 시스템 또는 접근제어 서비스의 권한 정책을 확인하여 DB Super User(관리자) 권한자 목록을 추출함\n2. 추출된 관리자 권한이 사전에 지정된 최소한의 담당자에게만 부여되어 있는지 검토함",
        "test_procedure_manual": "1. DB 관리자 권한 부여 이력을 조회한다\n2. 모집단에서 샘플을 추출하여 권한 부여에 대한 승인 여부를 확인한다"
    },
    {
        "id": "PC01", "name": "프로그램 변경 승인", "category": "변경관리", "objective": "프로그램 변경 전 승인 득함",
        "risk_code": "R-16", "risk_description": "승인 또는 테스트 절차 없이 수행된 프로그램 변경이 주요 거래정보를 왜곡 할 위험",
        "control_description": "프로그램 변경 필요시 변경요청서를 작성하고 요청부서장 등 책임자의 승인을 득한다",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "프로그램 변경 요청에 대해 SR(Service Request)의 작성/승인 완료 여부를 검토한다\n\n1. 모집단에서 샘플을 추출한다\n2. 샘플건이 SR의 작성 및 승인권자(요청 부서장)의 승인 여부를 확인한다\n3. SR 작성/승인 이후 프로그램 이관이 수행되었는지 확인한다\n  - SR 작성/승인 완료 일자 < 프로그램 이관 일자",
        "test_procedure_auto": "1. 프로그램 변경 시 시스템에 의한 자동 승인 워크플로우 로직을 확인한다\n2. 모집단에서 임의의 샘플 1건을 추출하여 자동 승인 후 이관이 수행되었는지 확인한다"
    },
    {
        "id": "PC02", "name": "프로그램 변경 사용자 테스트", "category": "변경관리", "objective": "사용자 검증을 통한 품질 확보",
        "risk_code": "R-17", "risk_description": "승인 또는 테스트 절차 없이 수행된 프로그램 변경이 주요 거래정보를 왜곡 할 위험",
        "control_description": "프로그램 변경 후 요청자에 의한 사용자인수테스트가 수행되며 결과는 문서화된다",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "프로그램 변경(개발) 완료 후 테스트 통제에 대해 사용자 테스트 결과서 작성 여부를 검토한다\n\n1. 모집단에서 샘플을 추출한다\n2. 샘플건이 프로그램 변경(개발) 완료 후 사용자 테스트가 수행되었고 결과의 문서화 여부를 확인한다\n3. 사용자 테스트 완료 이후 프로그램 이관이 수행되었는지 확인한다\n  - 사용자 테스트 완료 일자 < 프로그램 이관 일자",
        "test_procedure_auto": "1. 자동화된 테스트 시스템의 로직 및 테스트 케이스 설정을 확인한다\n2. 모집단에서 임의의 샘플 1건을 추출하여 자동 테스트가 수행되었는지 확인한다"
    },
    {
        "id": "PC03", "name": "프로그램 변경 이관 승인", "category": "변경관리", "objective": "운영 반영 전 최종 승인 확인",
        "risk_code": "R-18", "risk_description": "승인 또는 테스트 절차 없이 수행된 프로그램 변경이 주요 거래정보를 왜곡 할 위험",
        "control_description": "개발완료 후 개발자는 이관을 요청하고 IT팀장 등 책임자의 승인을 득한다",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "프로그램 이관 요청/승인 통제에 대해 이관 요청서의 작성/승인 완료 여부를 검토한다\n\n1. 모집단에서 샘플을 추출한다\n2. 샘플건이 프로그램 변경(개발) 완료 후 이관 요청서가 작성되었고 승인권자(IT운영팀장)의 승인 여부를 확인한다\n3. 이관 요청 및 승인 완료 후 프로그램 이관이 수행되었는지 확인한다\n  - 이관 요청/승인 완료 일자 < 프로그램 이관 일자",
        "test_procedure_auto": "1. 자동 이관 시스템의 승인 워크플로우 로직을 확인한다\n2. 모집단에서 임의의 샘플 1건을 추출하여 자동 승인 후 이관이 수행되었는지 확인한다"
    },
    {
        "id": "PC04", "name": "개발/운영 환경의 분리", "category": "변경관리", "objective": "개발과 운영 독립성 확보",
        "risk_code": "R-19", "risk_description": "승인 또는 테스트 절차 없이 수행된 프로그램 변경이 주요 거래정보를 왜곡 할 위험",
        "control_description": "개발과 운영 환경은 논리적으로 분리되어 있다",
        "type": "Auto", "frequency": "기타", "method": "예방",
        "test_procedure_auto": "1. 시스템 아키텍처 및 서버 구성을 확인하여 개발(DEV)/스테이징(STG)/운영(PRD) 서버가 논리적 또는 물리적으로 분리되어 있는지 확인함\n2. 각 환경별 IP 정보 및 접근 통제 룰을 검토하여 개발환경에서 운영환경으로의 직접적인 접근이 차단되어 있는지 확인함",
        "test_procedure_manual": "1. 개발/운영 환경 분리에 대한 정책서를 입수한다\n2. 환경 분리 변경 이력에 대한 승인 여부를 확인한다"
    },
    {
        "id": "PC05", "name": "이관담당자 권한 제한", "category": "변경관리", "objective": "배포 권한의 분리 및 최소화",
        "risk_code": "R-20", "risk_description": "승인 또는 테스트 절차 없이 수행된 프로그램 변경이 주요 거래정보를 왜곡 할 위험",
        "control_description": "이관담당자과 개발담당자는 직무상 분리되어 있으며 시스템 권한도 분리되어 있다",
        "type": "Auto", "frequency": "기타", "method": "예방",
        "test_procedure_auto": "1. 개발자와 이관담당자의 권한을 각각 확인함\n2. 개발자가 운영환경에 직접 이관할 수 없도록 권한이 분리되어 있는지 확인함",
        "test_procedure_manual": "1. 이관 권한 부여 이력을 조회한다\n2. 모집단에서 샘플을 추출하여 권한 부여에 대한 승인 여부를 확인한다"
    },
    {
        "id": "PC06", "name": "OS 설정변경", "category": "변경관리", "objective": "인프라 설정 변경의 통제",
        "risk_code": "R-21", "risk_description": "인가받지 않은 OS/DB 설정 변경으로 인해 프로그램이 부적절하게 변경 될 위험",
        "control_description": "OS 설정변경시 적절한 승인권자의 승인을 득한다",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "1. In-scope 시스템 OS/DB의 변경이력(업데이트/패치 이력)을 아래 조건으로 추출함\n  - 작업일자 : 테스트 대상 기간\n\n2. 조회 된 건수의 빈도에 따라 샘플 Guide의 샘플 수를 적용하여 무작위로 샘플링 함\n\n3. 샘플로 선정된 건에 대하여 적절한 근거가 존재하는지 확인함\n  - OS/DB 변경(업데이트/패치 이력)에 대한 요청 및 승인\n  - 변경(업데이트/패치 이력) 건에 대한 테스트 결과",
        "test_procedure_auto": "1. OS 설정변경 자동 승인 시스템 로직을 확인한다\n2. 모집단에서 임의의 샘플 1건을 추출하여 자동 승인 후 변경이 수행되었는지 확인한다"
    },
    {
        "id": "PC07", "name": "DB 설정변경", "category": "변경관리", "objective": "DB 설정 변경의 통제",
        "risk_code": "R-22", "risk_description": "인가받지 않은 OS/DB 설정 변경으로 인해 프로그램이 부적절하게 변경 될 위험",
        "control_description": "DB 설정변경시 적절한 승인권자의 승인을 득한다",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "1. In-scope 시스템 OS/DB의 변경이력(업데이트/패치 이력)을 아래 조건으로 추출함\n  - 작업일자 : 테스트 대상 기간\n\n2. 조회 된 건수의 빈도에 따라 샘플 Guide의 샘플 수를 적용하여 무작위로 샘플링 함\n\n3. 샘플로 선정된 건에 대하여 적절한 근거가 존재하는지 확인함\n  - OS/DB 변경(업데이트/패치 이력)에 대한 요청 및 승인\n  - 변경(업데이트/패치 이력) 건에 대한 테스트 결과",
        "test_procedure_auto": "1. DB 설정변경 자동 승인 시스템 로직을 확인한다\n2. 모집단에서 임의의 샘플 1건을 추출하여 자동 승인 후 변경이 수행되었는지 확인한다"
    },
    {
        "id": "CO01", "name": "Batch Schedule 등록/변경 요청 및 승인", "category": "운영관리", "objective": "배치 작업 변경의 무결성 확보",
        "risk_code": "R-23", "risk_description": "배치 프로세스가 부적절하게 변경되거나 실패할 위험",
        "control_description": "새로운 배치 작업 등록 시 요청서를 작성하고 IT팀장 등 책임자의 승인을 득한다",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "1. Batch 작업 목록 중 아래 조건의 배치 등록/변경 이력을 추출함\n2. 조회 된 건수의 빈도에 따라 샘플 Guide의 샘플 수를 적용하여 무작위로 샘플링 함\n3. 샘플로 선정된 Batch 작업 건에 대하여 하기 사항을 확인함\n - 샘플로 선정된 Batch에 대한 IT 시스템 설정 변경 승인서\n - Batch 등록, 변경에 대한 승인권자의 승인 여부\n - Batch 등록 결과가 요청상의 내용과 동일한지 여부",
        "test_procedure_auto": "1. Batch 등록/변경 자동 승인 시스템 로직을 확인한다\n2. 모집단에서 임의의 샘플 1건을 추출하여 자동 승인 후 등록이 수행되었는지 확인한다"
    },
    {
        "id": "CO02", "name": "Batch Schedule 등록/변경 권한 제한", "category": "운영관리", "objective": "스케줄링 권한 오남용 방지",
        "risk_code": "R-24", "risk_description": "배치 프로세스가 부적절하게 변경되거나 실패할 위험",
        "control_description": "Batch Schedule 등록/변경 권한은 지정된 담당자로 제한된다",
        "type": "Auto", "frequency": "기타", "method": "예방",
        "test_procedure_auto": "1. 배치 작업 스케줄링 관리 시스템에서 등록/변경 권한을 보유한 사용자 목록을 추출함\n2. 해당 권한 보유자가 업무분장(SoD)에 따라 적절한 인원(운영 담당자 등)으로 제한되어 있는지 검토함",
        "test_procedure_manual": "1. Batch 권한 부여 이력을 조회한다\n2. 모집단에서 샘플을 추출하여 권한 부여에 대한 승인 여부를 확인한다"
    },
    {
        "id": "CO03", "name": "Batch Schedule 모니터링", "category": "운영관리", "objective": "배치 작업 실패 시 적시 대응",
        "risk_code": "R-23", "risk_description": "배치 프로세스가 부적절하게 변경되거나 실패할 위험",
        "control_description": "Batch 오류는 담당자가 인식할 수 있어야 하며 오류 내역 및 조치결과가 문서화된다",
        "type": "Manual", "frequency": "일", "method": "적발",
        "test_procedure_manual": "1. IT 주간점검보고 5건에 대하여 하기 사항을 확인함\n - 주별 Batch 모니터링 수행 여부\n - 점검 결과에 대한 문서화 여부\n - 이상 건 식별 시, 해당 건에 대한 조치 완료 여부 및 결과 문서화 여부\n - 모니터링 결과에 대한 보고 여부",
        "test_procedure_auto": "1. 자동 Batch 모니터링 시스템의 알림 설정을 확인한다\n2. 자동 생성된 모니터링 보고서 샘플 1건을 추출하여 정상 생성 여부를 확인한다"
    },
    {
        "id": "CO04", "name": "백업 모니터링", "category": "운영관리", "objective": "백업 성공 여부 모니터링",
        "risk_code": "R-24", "risk_description": "시스템 오류 등으로 거래 데이터가 손상되거나 복구되지 않을 위험",
        "control_description": "주기적으로 백업 및 복구가 수행되고, 백업미디어는 소산보관된다",
        "type": "Manual", "frequency": "일", "method": "적발",
        "test_procedure_manual": "1. IT 주간점검보고 5건에 대하여 하기 사항을 확인함\n - 주별 백업 모니터링 수행 여부\n - 점검 결과에 대한 문서화 여부\n - 이상 건 식별 시, 해당 건에 대한 조치 완료 여부 및 결과 문서화 여부\n - 모니터링 결과에 대한 보고 여부",
        "test_procedure_auto": "1. 자동 백업 시스템의 스케줄 및 설정을 확인한다\n2. 자동 생성된 백업 로그 샘플 1건을 추출하여 정상 수행 여부를 확인한다"
    },
    {
        "id": "CO05", "name": "장애 관리", "category": "운영관리", "objective": "시스템 장애 기록 및 원인 분석",
        "risk_code": "R-25", "risk_description": "시스템 운영 중 발생한 오류가 적시에 해결되지 않을 위험",
        "control_description": "복구절차의 효과성을 정기적으로 테스트하기 위한 절차가 존재한다",
        "type": "Manual", "frequency": "수시", "method": "적발",
        "test_procedure_manual": "1. 서버유지보수업체로부터 매월 보고 받는 월간 보고서를 입수함\n2. 시스템 오류와 관련하여 원인, 해결방안, 재발방지 대책이 문서화 되었는지 여부를 검토함\n3. 작성된 장애보고서가 IT부서장에게 보고 되었는지 여부를 확인함",
        "test_procedure_auto": "1. 자동 장애 감지 및 알림 시스템의 설정을 확인한다\n2. 자동 생성된 장애 로그 샘플 1건을 추출하여 정상 기록 여부를 확인한다"
    },
    {
        "id": "CO06", "name": "데이터센터 접근 제한", "category": "운영관리", "objective": "서버실 물리적 출입 통제",
        "risk_code": "R-26", "risk_description": "IT 시설, 장비, 자원에 승인되지 않은 접근이 발생 할 위험",
        "control_description": "외부 인원의 데이터 센터 출입시 방문일지를 작성하고 서버담당자의 동행하에 출입한다",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "1. 분기 통제이므로 데이터센터 출입이력 검토 내역 중 임의의 샘플을 무작위로 샘플링함\n\n2. 샘플 건에 대해서 하기 사항을 확인함\n - 데이터센터 출입이력 검토내역에 주요 서버실 포함 여부\n - 출입인원에 대한 적절한 검토 여부 (특정인원 출입빈도 / 사유파악 등)\n - 출입이력 검토내역에 대한 IT부서장의 승인 여부",
        "test_procedure_auto": "1. 자동 출입통제 시스템의 설정 및 권한을 확인한다\n2. 출입 로그 샘플 1건을 추출하여 권한에 따른 접근 통제가 정상 작동하는지 확인한다"
    },
    {
        "id": "PD01", "name": "사용자 인수 테스트", "category": "개발관리", "objective": "최종 사용자의 시스템 수용 검증",
        "risk_code": "R-27", "risk_description": "주요 정보가 불완전하고 부정확하게 이관 될 위험",
        "control_description": "1. 프로젝트 진행 확정 시 품의 진행 및 승인 득함\n2. 개발 과정에서 단위/통합/사용자 테스트 수행 및 결과 문서화\n3. 프로젝트 완료 시 결과보고 진행",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "1. 경영정보팀에서 관리하고 있는 SI완료 건 중 아래 내역을 대상으로 하여 발생빈도에 따라 샘플 Guide의 샘플 수를 적용하여 무작위로 샘플링 함\n - 내부회계운영 지원 시스템 (In-scope)\n - 재무보고와 관련있는 개발건\n2. 샘플로 선정한 시스템 추가/기능 개선 등의 프로젝트 건에 대하여 개발 과정에서 인수테스트(단위테스트/통합테스트/사용자 테스트)가 수행되고 테스트 결과가 문서화 되었는 지 여부를 검토함\n사용자 테스트 결과 특이사항이 없는 경우 검토승인->검토완료 변경처리됨.",
        "test_procedure_auto": "1. 자동화된 테스트 시스템의 로직 및 테스트 케이스 설정을 확인한다\n2. 자동 테스트 결과 로그 샘플 1건을 추출하여 정상 수행 여부를 확인한다"
    },
    {
        "id": "PD02", "name": "데이터 이관", "category": "개발관리", "objective": "이관 데이터의 완전성 및 정확성 보장",
        "risk_code": "R-28", "risk_description": "새로운 시스템에 불완전하거나 부정확하게 데이터가 생성 될 위험",
        "control_description": "1. 중요한 데이터 이관 시 별도의 데이터 테스트를 수행하며, 테스트 결과 및 승인 이력이 문서화 관리됨",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "1. IT부서에서 관리하고 있는 SI완료 건 중 아래 내역을 대상으로 하여 발생빈도에 따라 샘플 Guide의 샘플 수를 적용하여 무작위로 샘플링 함\n - 내부회계운영 지원 시스템 (In-scope)\n - 재무보고와 관련있는 개발건\n - 데이터 마이그레이션이 동반된 건\n\n2. 개발 중 유효한 마이그레이션 테스트가 수행되었는지 아래 사항을 기준으로 검토함.\n - 기존 시스템의 데이터 필드와 목표 시스템의 데이터 필드의 적절한 mapping에 대한 확인 여부\n - 이관 결과에 대한 승인 존재 여부\n - 이관된 데이터의 완전성, 정확성 및 유효성 검증 여부",
        "test_procedure_auto": "1. 자동 데이터 이관 시스템의 매핑 설정 및 검증 로직을 확인한다\n2. 자동 이관 로그 샘플 1건을 추출하여 정상 수행 여부를 확인한다"
    },
    {
        "id": "PD03", "name": "이슈 관리", "category": "개발관리", "objective": "개발 이슈의 누적 관리 및 해결",
        "risk_code": "R-29", "risk_description": "새로운 시스템에 불완전하거나 부정확하게 데이터가 생성 될 위험",
        "control_description": "1. 프로젝트 중 발생하는 이슈 및 오류는 주기적으로 집계 및 관리됨\n2. 관리되는 이슈는 적시에 해결되어 프로젝트 관리자에게 보고 됨",
        "type": "Manual", "frequency": "수시", "method": "적발",
        "test_procedure_manual": "1. IT부서에서 관리하고 있는 SI완료 건 중 아래 내역을 대상으로 하여 발생빈도에 따라 샘플 Guide의 샘플 수를 적용하여 무작위로 샘플링 함\n - 내부회계운영 지원 시스템 (In-scope)\n - 재무보고와 관련있는 개발건\n2. 샘플로 선정한 SI건에 대해 이슈 및 오류내역이 문서화 되어 있는지 검토함\n3. 이슈 및 오류내역이 주기적으로 프로젝트 관리자에게 보고 및 조치되었는지 검토함",
        "test_procedure_auto": "1. 자동 이슈 추적 시스템의 설정 및 알림 로직을 확인한다\n2. 자동 생성된 이슈 보고서 샘플 1건을 추출하여 정상 생성 여부를 확인한다"
    },
    {
        "id": "PD04", "name": "사용자 교육", "category": "개발관리", "objective": "운영 준비를 위한 사용자 교육 실시",
        "risk_code": "R-30", "risk_description": "새로운 시스템에 불완전하거나 부정확하게 데이터가 생성 될 위험",
        "control_description": "1. 시스템 추가/기능 개선 프로젝트 종료 시 사용자 교육 또는 매뉴얼 배포가 수행 됨",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "1. 경영정보팀에서 관리하고 있는 SI완료 건 중 아래 내역을 대상으로 하여 발생빈도에 따라 샘플 Guide의 샘플 수를 적용하여 무작위로 샘플링 함\n - 내부회계운영 지원 시스템 (In-scope)\n - 재무보고와 관련있는 개발건\n2. 샘플로 선정한 SI건에 대한 교육 및 매뉴얼 배포가 설계 및 수행되었는지 검토함",
        "test_procedure_auto": "1. 자동 교육 시스템(LMS)의 설정 및 교육 이수 추적 로직을 확인한다\n2. 교육 이수 로그 샘플 1건을 추출하여 정상 기록 여부를 확인한다"
    },
    {
        "id": "ST01", "name": "Supporting Tool Super User 권한 제한", "category": "지원툴관리", "objective": "보조 툴 관리자 권한 통제",
        "risk_code": "R-31", "risk_description": "위험하고 강력한 계정사용자가 시스템 통제를 우회 할 위험",
        "control_description": "1. Supporting tool의 Super User는 지정된 담당자로 제한되어 있음",
        "type": "Auto", "frequency": "기타", "method": "예방",
        "test_procedure_auto": "1. Supporting Tool의 관리자 권한 목록을 확인함\n2. 관리자 권한이 지정된 담당자로 제한되어 있음을 확인함",
        "test_procedure_manual": "1. 관리자 권한 부여 이력을 조회한다\n2. 모집단에서 샘플을 추출하여 권한 부여에 대한 승인 여부를 확인한다"
    },
    {
        "id": "ST02", "name": "Supporting Tool 패스워드 설정", "category": "지원툴관리", "objective": "보조 툴 보안 설정 유지",
        "risk_code": "R-32", "risk_description": "취약한 패스워드 정책으로 인해 시스템 접근 통제를 우회할 위험",
        "control_description": "1. Supporting tool의 패스워드는 최소 8자리 이상, 문자/숫자/특수문자 조합, 분기별 변경 적용",
        "type": "Auto", "frequency": "기타", "method": "예방",
        "test_procedure_auto": "1. Supporting Tool 패스워드 환경설정 정보를 확인하여 아래 기준으로 설정되어 있는지 여부를 검토함.\n 1) 복잡성: 문자, 숫자, 특수문자 포함\n 2) 최소 길이: 8자리 이상",
        "test_procedure_manual": "1. 패스워드 정책 변경 이력을 조회한다\n2. 모집단에서 샘플을 추출하여 변경 요청 및 승인 여부를 확인한다"
    },
    {
        "id": "ST03", "name": "Supporting Tool 기능 변경 요청 및 승인", "category": "지원툴관리", "objective": "지원 도구 변경의 이력 관리",
        "risk_code": "R-33", "risk_description": "승인 없이 수행된 시스템 설정 변경이 데이터를 왜곡 할 위험",
        "control_description": "1. Supporting tool(DBSafer 등) 변경 시 요청서 작성 및 IT부서장 승인 득함\n2. 테스트 수행 및 최종 적용 검토 후 운영환경 적용",
        "type": "Manual", "frequency": "수시", "method": "예방",
        "test_procedure_manual": "1. 테스트기간 동안 발생한 Supporting Tool 프로그램 변경 이력을 조회하고, 조회화면(조회 조건, 조회된 전체 데이터 수 확인 가능한 화면) 및 조회 결과를 문서화함\n2. 위의 1.에서 조회된Supporting Tool 변경의 발생빈도에 따라 샘플 Guide의 샘플 수를 적용하여 샘플을 선정함\n3. 샘플에 대하여 전자결재상 \"IT 시스템 설정 변경 승인서\"를 확인하고 승인권자의 승인을 득하였는지 여부를 확인함",
        "test_procedure_auto": "1. Supporting Tool 변경 자동 승인 시스템 로직을 확인한다\n2. 모집단에서 임의의 샘플 1건을 추출하여 자동 승인 후 변경이 수행되었는지 확인한다"
    },
    {
        "id": "PWC02", "name": "정보보안 교육", "category": "IT정책", "objective": "전사 보안 의식 함양 및 법적 준수",
        "risk_code": "R-34", "risk_description": "정보보안 교육이 적절하게 이루어지지 않을 위험",
        "control_description": "1. 모든 임직원 대상 매분기 1회 개인정보보호 및 정보보안 교육 실시 및 관련 자료 게시",
        "type": "Manual", "frequency": "분기", "method": "예방",
        "test_procedure_manual": "1. 경영정보팀에서 모든 임직원을 대상으로 분기별 개인정보보호 및 정보보안 교육을 위해 정보보안 게시판에 게시물을 게시하였는지 확인함.",
        "test_procedure_auto": "1. 자동 교육 시스템(LMS)의 설정 및 교육 공지 로직을 확인한다\n2. 자동 발송된 교육 알림 로그 샘플 1건을 추출하여 정상 발송 여부를 확인한다"
    },
]

# ================================
# 수시 통제용 모집단/완전성 템플릿 (SW/OS/DB별)
# ================================

# SW(Application)별 템플릿 - APD01~03, APD07, PC01~03, CO01, ST03, PD01~04, CO05
SW_POPULATION_TEMPLATES = {
    "SAP": {
        "APD01": {"population": "SU01 권한 변경 이력", "completeness": "T-Code SUIM에서 테스트 기간 내 권한 변경 건 조회\n[Query] SUIM > 사용자 > 권한 변경 이력\n\n[버전별 참고]\n• ECC 6.0: T-Code SUIM, USH04 테이블\n• S/4HANA: Fiori App 'Display Authorization Changes' 또는 SUIM\n• S/4HANA Cloud: IAM 앱 > Security Audit Log"},
        "APD02": {"population": "SU01 부서이동자 권한 변경 이력", "completeness": "인사시스템 전배자 목록과 SU01 권한 변경 이력 대사\n\n[버전별 참고]\n• ECC 6.0: PA20/PA30 인사정보와 SU01 대사\n• S/4HANA: Employee Central 또는 PA 데이터와 SU01 대사\n• S/4HANA Cloud: Workforce Person Integration 데이터 활용"},
        "APD03": {"population": "SU01 퇴사자 계정 삭제 이력", "completeness": "인사시스템 퇴사자 목록과 SU01 계정 삭제 이력 대사\n\n[버전별 참고]\n• ECC 6.0: USR02 테이블의 USTYP, UFLAG 필드 확인\n• S/4HANA: 동일 + Fiori 'Manage Users' 앱\n• S/4HANA Cloud: IAS(Identity Authentication Service) 연동 확인"},
        "APD07": {"population": "Table 로그(DBTABLOG) 및 SE16N 직접 변경 로그", "completeness": "T-Code: SCU3 및 SE16N 편집 로그 조회\n\n[버전별 참고]\n• ECC 6.0: DBTABLOG(rec/client 설정 필수), SE16N_CD_HDR/DATA\n• S/4HANA: 동일 + SAL (Security Audit Log) 내 Table Access 감사\n• S/4HANA Cloud: CSP 영역으로 직접 변경 원천 차단"},
        "PC01": {"population": "E070 테이블 조회 결과 (당기 전체 TR 목록)", "completeness": "SE16에서 E070 테이블의 AS4DATE(생성일) 기준 당기분 추출\n\n[버전별 참고]\n• ECC 6.0: STMS, SE01, E070/E071 테이블\n• S/4HANA: CTS+ 또는 gCTS(Git-enabled CTS) 사용 가능\n• S/4HANA Cloud: SAP Cloud Transport Management 전용"},
        "PC02": {"population": "E070 테이블 조회 결과 (테스트 대상 TR)", "completeness": "E070 목록 중 커스터마이징(K/W) TR에 대한 테스트 문서 대사\n\n[버전별 참고]\n• ECC 6.0: SPRO 커스터마이징 TR\n• S/4HANA: Fiori 'Custom Fields and Logic' 앱 변경 포함\n• S/4HANA Cloud: Key User Extensibility 변경 이력"},
        "PC03": {"population": "E070 테이블 조회 결과 (운영 이관 TR)", "completeness": "STMS 이관 로그와 E070 테이블의 이관 승인 상태 대사\n\n[버전별 참고]\n• ECC 6.0: STMS QA 승인 로그\n• S/4HANA: STMS 또는 CTS+ 이관 로그\n• S/4HANA Cloud: Cloud Transport Management 승인 이력"},
        "CO01": {"population": "T-Code: SM36 (Batch Job 등록 현황)", "completeness": "SM37에서 테스트 기간 내 신규/변경된 Job 실시간 실행 로그 조회\n\n[버전별 참고]\n• ECC 6.0: SM36/SM37\n• S/4HANA: SM36/SM37 + Fiori 'Application Jobs' 앱\n• S/4HANA Cloud: Job Scheduling Service 전용"},
        "CO02": {"population": "T-Code: SM36 (Batch 스케줄링 권한 보유자)", "completeness": "SUIM에서 S_BTCH_ADM 권한 및 SM36 실행 권한자 추출"},
        "ST03": {"population": "지원도구 변경 요청 이력", "completeness": "IT 시스템 설정 변경 승인서 목록 조회"},
        "PD01": {"population": "SI 프로젝트 완료 건", "completeness": "경영정보팀 SI 완료 목록 입수\n\n[버전별 참고]\n• ECC 6.0: SAP Solution Manager 프로젝트 관리\n• S/4HANA: SAP Activate 방법론 기반\n• S/4HANA Cloud: SAP Cloud ALM 프로젝트 관리"},
        "PD02": {"population": "데이터 마이그레이션 수행 건", "completeness": "SI 프로젝트 중 마이그레이션 동반 건 식별\n\n[버전별 참고]\n• ECC 6.0: LSMW, BDC, BAPI\n• S/4HANA: SAP S/4HANA Migration Cockpit\n• S/4HANA Cloud: Migration Cockpit 전용 (파일 업로드 방식)"},
        "PD03": {"population": "SI 프로젝트 이슈 목록", "completeness": "프로젝트 관리 시스템의 이슈 등록 건 조회"},
        "PD04": {"population": "SI 프로젝트 교육 수행 건", "completeness": "교육 수행 결과 또는 매뉴얼 배포 이력 조회\n\n[버전별 참고]\n• S/4HANA Cloud: SAP Enable Now 활용 권장"},
        "CO05": {"population": "시스템 장애 발생 건", "completeness": "월간 유지보수 보고서 또는 장애 접수 시스템 조회\n\n[버전별 참고]\n• ECC 6.0: ST22 Dump 분석, SM21 시스템 로그\n• S/4HANA: SAP Solution Manager 또는 Cloud ALM\n• S/4HANA Cloud: SAP for Me 인시던트 관리"},
        "APD04": {"population": "T-Code: SUIM (권한 보유자 분석)", "completeness": "[자동통제 확인방법]\nT-Code: SUIM > 사용자별 권한 분석 > 주요 권한(SAP_ALL, SAP_NEW) 보유자 추출\n[Query] Transaction SUIM > Users by Complex Selection Criteria\n\n[버전별 참고]\n• ECC 6.0: SUIM 동일\n• S/4HANA: Fiori App 'IAM Information System' 추가 활용 가능"},
        "APD05": {"population": "T-Code: SU01, SUIM (관리자 권한 보유자)", "completeness": "[자동통제 확인방법]\nT-Code: SUIM > 권한/프로파일별 사용자 조회 > SAP_ALL, SAP_NEW 프로파일 보유자\nS_USER_ADMIN, S_ADMI_FCD 권한 보유자 추출\n관리자 계정: SAP*, DDIC 계정 활성화 여부 확인\nT-Code: SU01 > User Lock Status 조회\n\n[버전별 참고]\n• ECC 6.0: SAP*, DDIC 계정 잠금 상태 확인\n• S/4HANA: 동일 + Fiori 'Manage Users' 앱에서 관리자 확인\n• S/4HANA Cloud: 표준 관리자 계정 없음, BTP Admin 권한 확인"},
        "APD06": {"population": "T-Code: RSPARAM, RZ11 (패스워드 정책)", "completeness": "[자동통제 확인방법]\nT-Code: RSPARAM > login/* 파라미터 조회\n- login/min_password_lng: 최소 길이\n- login/password_expiration_time: 만료 기간\n- login/fails_to_user_lock: 잠금 임계값\nT-Code: RZ11 > 파라미터 상세 설정 확인\n\n[버전별 참고]\n• ECC 6.0: 위 T-Code 동일\n• S/4HANA: Fiori App 'Maintain Password Rules' 추가 확인"},
        "PC04": {"population": "T-Code: SCC4, SE06 (변경 옵션 설정)", "completeness": "[자동통제 확인방법]\nT-Code: SCC4 > 클라이언트 설정 > 변경 및 이관 옵션\n- Production: 변경 불가 (Not Modifiable)\nT-Code: SE06 > 시스템 변경 옵션 > 운영계 설정 확인\n\n[버전별 참고]\n• ECC 6.0: SCC4, SE06 동일\n• S/4HANA: SCC4, SE06 동일 + Software Component 설정\n• S/4HANA Cloud: 운영계 변경 원천 차단 (CSP 관리)"},
        "PC05": {"population": "T-Code: STMS (이관 권한 보유자)", "completeness": "[자동통제 확인방법]\nT-Code: STMS > Overview > Systems > Import Queue 권한자\nT-Code: SUIM > S_CTS_ADMI, S_TRANSPRT 권한 보유자 조회\n\n[버전별 참고]\n• ECC 6.0: STMS 권한자\n• S/4HANA: STMS 또는 gCTS 권한자\n• S/4HANA Cloud: Cloud Transport Management 권한자 (BTP)"},
        "CO02": {"population": "T-Code: SM36 (배치 스케줄링 권한)", "completeness": "[자동통제 확인방법]\nT-Code: SUIM > S_BTCH_ADM, S_BTCH_JOB 권한 보유자 추출\nSM36 실행 권한자 = 배치 스케줄 등록/변경 권한자\n\n[버전별 참고]\n• ECC 6.0: SM36 권한 (S_BTCH_ADM)\n• S/4HANA: SM36 + Fiori 'Schedule Application Jobs' 권한\n• S/4HANA Cloud: Job Scheduling Service 권한 (BTP IAM)"},
        "CO03": {"population": "T-Code: SM37 (배치 모니터링)", "completeness": "[자동통제 확인방법]\nT-Code: SM37 > Job Selection > 기간 내 실패(Canceled) Job 조회\nCCMS 알림 설정: RZ20 > 배치 모니터링 Alert 설정 확인\n\n[버전별 참고]\n• ECC 6.0: SM37, RZ20 CCMS Alert\n• S/4HANA: SM37 + Fiori 'Monitor Application Jobs' + SAP Solution Manager\n• S/4HANA Cloud: Cloud ALM Job Monitoring"},
        "CO04": {"population": "T-Code: DB13 (백업 스케줄)", "completeness": "[자동통제 확인방법]\nT-Code: DB13 > Backup Logs > 백업 성공/실패 이력 조회\nT-Code: DB12 > Backup Logs Overview 확인\n\n[버전별 참고]\n• ECC 6.0: DB13 (Oracle/DB2/MSSQL별 상이)\n• S/4HANA: HANA Studio > Backup 또는 HANA Cockpit\n• S/4HANA Cloud: CSP 책임 영역 (SAP 자동 관리)"},
        "ST01": {"population": "DB 접근통제 솔루션 관리자 설정", "completeness": "[자동통제 확인방법]\n접근통제 솔루션(DBSafer/Chakra 등) 관리 콘솔 > 관리자 계정 목록\n관리자 권한 부여 이력 조회"},
        "ST02": {"population": "DB 접근통제 솔루션 정책 설정", "completeness": "[자동통제 확인방법]\n접근통제 솔루션 관리 콘솔 > 보안 정책 설정 화면\n- 접근 가능 IP 대역\n- 허용 시간대\n- 민감정보 마스킹 정책"},
    },
    "ORACLE": {
        "APD01": {"population": "FND_USER 권한 변경 이력", "completeness": "FND_USER_RESP_GROUPS에서 테스트 기간 내 권한 부여 건 조회\n\n[버전별 참고]\n• R12: FND_USER_RESP_GROUPS_DIRECT 테이블\n• Fusion Cloud: Security Console > Users > Role Assignments\n• JDE: F0092 (User Profile), F00950 (User Role)"},
        "APD02": {"population": "FND_USER 부서이동자 권한 변경", "completeness": "인사시스템 전배자 목록과 FND_USER 권한 변경 이력 대사\n\n[버전별 참고]\n• R12: HRMS 연동 또는 별도 인사시스템\n• Fusion Cloud: HCM Cloud 자동 연동\n• JDE: Address Book과 User Profile 연동"},
        "APD03": {"population": "FND_USER 퇴사자 계정 비활성화", "completeness": "인사시스템 퇴사자 목록과 FND_USER.END_DATE 대사\n\n[버전별 참고]\n• R12: FND_USER.END_DATE 필드\n• Fusion Cloud: User 상태 'Inactive' 확인\n• JDE: F0092.USERSTS 필드 (D=Disabled)"},
        "APD07": {"population": "데이터 직접 변경 이력 (DML 이력)", "completeness": "DBA_AUDIT_TRAIL 또는 시스템 감사 로그에서 직접 DB 접근 및 DML 이력 조회\n\n[버전별 참고]\n• R12: Database Vault, Fine-Grained Auditing\n• Fusion Cloud: 직접 DB 접근 불가 (CSP 관리)\n• JDE: Data Browser 접근 로그"},
        "PC01": {"population": "당기 프로그램 변경 이력", "completeness": "FND_EXECUTABLES 및 FND_CONCURRENT_PROGRAMS의 LAST_UPDATE_DATE 기준 조회 결과\n\n[버전별 참고]\n• R12: FND_EXECUTABLES 테이블\n• Fusion Cloud: Setup and Maintenance > Scheduled Processes\n• JDE: Object Management Workbench (OMW)"},
        "PC02": {"population": "당기 프로그램 변경 이력 (테스트)", "completeness": "테스트 결과 보고서 또는 사용자 인수 서명(UAT) 확인\n\n[버전별 참고]\n• R12: Vision 또는 별도 테스트 인스턴스\n• Fusion Cloud: Test 환경 (Sandbox)\n• JDE: Prototype/Development 환경"},
        "PC03": {"population": "당기 프로그램 변경 이력 (이관)", "completeness": "CR(Change Request) 승인 목록과 FND_EXECUTABLES의 업데이트 일자 대사\n\n[버전별 참고]\n• R12: AD Utility, Patch History\n• Fusion Cloud: Configuration Migration 이력\n• JDE: Package Deployment, OMW Transfer Activity"},
        "CO01": {"population": "Concurrent Program 등록/변경", "completeness": "FND_CONCURRENT_PROGRAMS에서 신규/변경 건 조회\n\n[버전별 참고]\n• R12: Concurrent > Program > Define\n• Fusion Cloud: Enterprise Scheduler Service (ESS)\n• JDE: Batch Versions (P98305)"},
        "ST03": {"population": "지원도구 변경 요청 이력", "completeness": "IT 시스템 설정 변경 승인서 목록 조회"},
        "PD01": {"population": "SI 프로젝트 완료 건", "completeness": "경영정보팀 SI 완료 목록 입수\n\n[버전별 참고]\n• R12: AIM/OUM 방법론 산출물\n• Fusion Cloud: Oracle Soar 방법론\n• JDE: OneMethodology 기반"},
        "PD02": {"population": "데이터 마이그레이션 수행 건", "completeness": "SI 프로젝트 중 마이그레이션 동반 건 식별\n\n[버전별 참고]\n• R12: Data Loader, SQL*Loader\n• Fusion Cloud: FBDI (File-Based Data Import)\n• JDE: Z-Tables, Batch Input"},
        "PD03": {"population": "SI 프로젝트 이슈 목록", "completeness": "프로젝트 관리 시스템의 이슈 등록 건 조회"},
        "PD04": {"population": "SI 프로젝트 교육 수행 건", "completeness": "교육 수행 결과 또는 매뉴얼 배포 이력 조회"},
        "CO05": {"population": "시스템 장애 발생 건", "completeness": "월간 유지보수 보고서 또는 장애 접수 시스템 조회\n\n[버전별 참고]\n• R12: Alert Log, OEM (Enterprise Manager)\n• Fusion Cloud: Oracle Cloud Support (SR)\n• JDE: Server Manager Log"},
        "APD04": {"population": "EBS User Monitor (FND_USER_RESP_GROUPS)", "completeness": "[자동통제 확인방법]\n[메뉴] System Administrator > Security > User > Define\n[Query] SELECT fu.user_name, fr.responsibility_name, furg.start_date, furg.end_date\nFROM fnd_user fu, fnd_user_resp_groups_direct furg, fnd_responsibility_tl fr\nWHERE fu.user_id = furg.user_id AND furg.responsibility_id = fr.responsibility_id\n\n[버전별 참고]\n• R12: 위 쿼리 동일\n• Fusion Cloud: Security Console > Users 메뉴\n• JDE: User Profile Revisions (P0092)"},
        "APD05": {"population": "EBS 관리자 권한 보유자 (SYSADMIN)", "completeness": "[자동통제 확인방법]\n[메뉴] System Administrator > Security > User > Define\nSYSADMIN, Application DBA Responsibility 보유자 조회\n[Query] SELECT * FROM fnd_user_resp_groups_direct WHERE responsibility_id IN (SELECT responsibility_id FROM fnd_responsibility WHERE responsibility_key LIKE '%SYSADMIN%')\n\n[버전별 참고]\n• R12: SYSADMIN, Application Developer 권한\n• Fusion Cloud: IT Security Manager, Application Implementation Consultant 역할\n• JDE: *SECURITY, *ALL 권한 보유자"},
        "APD06": {"population": "EBS Profile Option (패스워드 정책)", "completeness": "[자동통제 확인방법]\n[메뉴] System Administrator > Profile > System\n- SIGNON_PASSWORD_LENGTH: 최소 길이\n- SIGNON_PASSWORD_HARD_TO_GUESS: 복잡도\n- SIGNON_PASSWORD_NO_REUSE: 재사용 금지\n[Query] SELECT * FROM fnd_profile_options WHERE profile_option_name LIKE 'SIGNON%'\n\n[버전별 참고]\n• R12: Profile Options 동일\n• Fusion Cloud: Setup and Maintenance > Security Console > Password Policy\n• JDE: System Control (P99410) > Security Workbench"},
        "PC04": {"population": "EBS 환경 설정 (Context File)", "completeness": "[자동통제 확인방법]\n$CONTEXT_FILE 위치의 XML 파일 설정 확인\n운영계 변경 제한: s_status=ENABLED, s_isForms=ENABLED\n[Query] cat $CONTEXT_FILE | grep -i 'modifiable'\n\n[버전별 참고]\n• R12: AutoConfig Context File\n• Fusion Cloud: 직접 변경 불가 (CSP 관리)\n• JDE: INI 파일 설정, PathCode 관리"},
        "PC05": {"population": "EBS 이관 권한 보유자", "completeness": "[자동통제 확인방법]\n[메뉴] Application Developer > Application > Register\nPATCH, AD Administration 권한 보유자 조회\n[Query] SELECT * FROM fnd_user_resp_groups_direct WHERE responsibility_id IN (SELECT responsibility_id FROM fnd_responsibility WHERE responsibility_key LIKE '%PATCH%')\n\n[버전별 참고]\n• R12: AD Utility 실행 권한자\n• Fusion Cloud: FSM Migration, Sandbox Publish 권한자\n• JDE: OMW Transfer Activity 권한자"},
        "CO02": {"population": "Concurrent Manager 권한 보유자", "completeness": "[자동통제 확인방법]\n[메뉴] System Administrator > Concurrent > Manager > Define\nConcurrent Manager 관리 권한 보유자 조회\n\n[버전별 참고]\n• R12: Concurrent Manager 관리 권한\n• Fusion Cloud: ESS Administrator 역할\n• JDE: Batch Server Administration 권한"},
        "CO03": {"population": "Concurrent Request 모니터링", "completeness": "[자동통제 확인방법]\n[메뉴] System Administrator > Concurrent > Requests\n[Query] SELECT * FROM fnd_concurrent_requests WHERE status_code IN ('E','G') AND request_date BETWEEN '시작일' AND '종료일'\n실패(Error/Warning) 건 조회\n\n[버전별 참고]\n• R12: FND_CONCURRENT_REQUESTS 테이블\n• Fusion Cloud: Scheduled Processes > Monitor\n• JDE: Work With Servers (P986116)"},
        "CO04": {"population": "RMAN 백업 스케줄", "completeness": "[자동통제 확인방법]\n[Query] SELECT * FROM v$rman_backup_job_details WHERE start_time BETWEEN '시작일' AND '종료일'\nRMAN 백업 로그 또는 백업 솔루션 콘솔에서 확인\n\n[버전별 참고]\n• R12: RMAN, OEM Backup\n• Fusion Cloud: CSP 자동 관리 (Oracle Cloud Backup)\n• JDE: DB 레벨 RMAN 또는 백업 솔루션"},
    },
    "DOUZONE": {
        "APD01": {"population": "더존 ERP 사용자권한 변경 이력", "completeness": "시스템관리 > 사용자권한관리에서 테스트 기간 내 권한 변경 건 조회\n[Query] SELECT * FROM TB_SYS_USER_AUTH WHERE UPD_DT BETWEEN '시작일' AND '종료일'\n\n[버전별 참고]\n• iCUBE: TB_SYS_USER_AUTH 테이블\n• iU ERP: 사용자관리 > 권한이력 메뉴\n• WEHAGO: 클라우드 관리자 > 사용자/권한 > 변경이력\n• Amaranth: 사용자관리 > 권한변경로그"},
        "APD02": {"population": "더존 ERP 부서이동자 권한회수 이력", "completeness": "인사시스템 전배자 목록과 권한변경이력 대사\n\n[버전별 참고]\n• iCUBE/iU: 인사모듈(HR) 연동 또는 별도 인사시스템\n• WEHAGO: 더존 HR Cloud 자동 연동\n• Amaranth: 인사관리 연동"},
        "APD03": {"population": "더존 ERP 퇴사자 계정 비활성화 이력", "completeness": "인사시스템 퇴사자 목록과 계정 상태 대사\n[Query] SELECT USER_ID, USER_NM, USE_YN FROM TB_SYS_USER WHERE USE_YN='N'\n\n[버전별 참고]\n• iCUBE: TB_SYS_USER.USE_YN='N'\n• iU ERP: 사용자상태='비활성'\n• WEHAGO: 사용자 상태 '사용안함' (클라우드 콘솔)\n• Amaranth: 사용자 상태 필드"},
        "APD07": {"population": "더존 ERP DB 직접 변경 로그", "completeness": "더존 유지보수팀 데이터 변경 요청서 목록 입수\n(Package 특성상 고객사 직접 변경 미발생 확인)\n\n[버전별 참고]\n• iCUBE/iU: 더존 엔지니어 통한 변경만 가능\n• WEHAGO: CSP 관리 영역 (직접 접근 불가)\n• Amaranth: 데이터 변경요청서 확인"},
        "PC01": {"population": "더존 유지보수 변경요청 건", "completeness": "더존 고객지원센터 SR(Service Request) 목록 입수\n\n[버전별 참고]\n• iCUBE: 더존 고객지원센터 > 패치이력\n• iU ERP: 더존 고객포털 > 서비스요청\n• WEHAGO: WEHAGO 관리자 > 업데이트 이력 (자동)\n• Amaranth: 고객지원 > 버전업이력"},
        "PC02": {"population": "더존 패치 테스트 결과", "completeness": "더존 패치 적용 전 테스트서버(DEV/STG) 검증 결과 확인\n[메뉴] 더존 고객지원센터 > 패치이력 > 테스트결과\n\n[버전별 참고]\n• iCUBE/iU: 테스트서버 구성 시 별도 검증\n• WEHAGO: 더존 자체 테스트 후 일괄 배포 (고객 테스트 불필요)\n• Amaranth: Sandbox 환경 테스트"},
        "PC03": {"population": "더존 패치 적용 승인 건", "completeness": "시스템관리 > 패치관리에서 패치 적용 이력 조회\n[Query] SELECT PATCH_NO, PATCH_NM, APPLY_DT, APPLY_ID FROM TB_SYS_PATCH_HIST\n\n[버전별 참고]\n• iCUBE/iU: 패치관리 메뉴 또는 고객포털 이력\n• WEHAGO: 자동 업데이트 (관리자 승인 프로세스)\n• Amaranth: 버전관리 이력"},
        "CO01": {"population": "더존 ERP 자동분개 스케줄 변경", "completeness": "시스템관리 > 배치관리에서 등록/변경 이력 조회\n[Query] SELECT BATCH_ID, BATCH_NM, EXEC_CYCLE, UPD_DT, UPD_ID FROM TB_SYS_BATCH\n\n[버전별 참고]\n• iCUBE: TB_SYS_BATCH 테이블\n• iU ERP: 배치관리 > 스케줄이력\n• WEHAGO: 배치 스케줄 관리 (클라우드 기반)\n• Amaranth: 자동화관리 > 스케줄"},
        "ST03": {"population": "더존 ERP 시스템 설정 변경 이력", "completeness": "시스템관리 > 환경설정에서 설정 변경 이력 조회\n[Query] SELECT CONFIG_ID, CONFIG_VAL, UPD_DT, UPD_ID FROM TB_SYS_CONFIG_LOG\n\n[버전별 참고]\n• iCUBE: TB_SYS_CONFIG_LOG 테이블\n• iU/WEHAGO/Amaranth: 환경설정 변경이력 메뉴"},
        "PD01": {"population": "더존 SI 프로젝트 완료 건", "completeness": "더존 SI 프로젝트 완료보고서 또는 모듈 추가 계약서 입수\n[메뉴] 더존 고객지원센터 > 프로젝트이력"},
        "PD02": {"population": "더존 데이터 마이그레이션 수행 건", "completeness": "마이그레이션 수행 로그 조회\n[Query] SELECT MIG_ID, MIG_NM, EXEC_DT FROM TB_SYS_MIGRATION_LOG\n\n[버전별 참고]\n• iCUBE/iU: 마이그레이션 도구 이력\n• WEHAGO: 클라우드 마이그레이션 서비스 이력\n• Amaranth: 데이터 이관 로그"},
        "PD03": {"population": "더존 SI 프로젝트 이슈 목록", "completeness": "더존 고객지원센터 > 프로젝트관리 > 이슈등록 목록 조회\n또는 내부 프로젝트 관리 시스템 이슈 목록"},
        "PD04": {"population": "더존 사용자 교육 수행 건", "completeness": "더존 고객지원센터 > 교육이력 또는 내부 교육 완료 보고서\n\n[버전별 참고]\n• WEHAGO: 온라인 교육 콘텐츠 제공 (LMS 연동)\n• Amaranth: 더존 아카데미 교육 이력"},
        "CO05": {"population": "더존 시스템 장애 발생 건", "completeness": "시스템관리 > 오류로그에서 장애 발생 건 조회\n[Query] SELECT ERR_ID, ERR_MSG, ERR_DT, USER_ID FROM TB_SYS_ERROR_LOG\n\n[버전별 참고]\n• iCUBE/iU: 오류로그 테이블 + 고객지원센터 장애접수\n• WEHAGO: 클라우드 모니터링 (SLA 기반)\n• Amaranth: 시스템로그 + 고객지원"},
        "APD04": {"population": "더존 ERP 권한 모니터링 화면", "completeness": "[자동통제 확인방법]\n[메뉴] 시스템관리 > 사용자권한관리 > 권한현황\n[Query] SELECT USER_ID, AUTH_CD, AUTH_NM, GRANT_DT FROM TB_SYS_USER_AUTH\nWHERE AUTH_CD IN ('ADMIN', 'SUPER')\n\n[버전별 참고]\n• iCUBE: TB_SYS_USER_AUTH 쿼리\n• iU ERP: 권한현황 리포트\n• WEHAGO: 클라우드 관리자 > 권한현황\n• Amaranth: 사용자권한 조회"},
        "APD05": {"population": "더존 ERP 관리자 권한 보유자", "completeness": "[자동통제 확인방법]\n[메뉴] 시스템관리 > 관리자설정\n[Query] SELECT USER_ID, USER_NM, ADMIN_YN FROM TB_SYS_USER WHERE ADMIN_YN = 'Y'\n\n[버전별 참고]\n• iCUBE/iU: ADMIN_YN='Y' 필터\n• WEHAGO: 클라우드 관리자 역할 보유자\n• Amaranth: 시스템관리자 그룹"},
        "APD06": {"population": "더존 ERP 패스워드 정책 설정", "completeness": "[자동통제 확인방법]\n[메뉴] 시스템관리 > 환경설정 > 보안설정\n[Query] SELECT CONFIG_ID, CONFIG_VAL FROM TB_SYS_CONFIG WHERE CONFIG_ID LIKE 'PWD_%'\n- PWD_MIN_LEN: 최소 길이 (8자리 이상)\n- PWD_EXPIRE_DAY: 만료 기간 (90일)\n- PWD_LOCK_CNT: 잠금 임계값 (5회)\n\n[버전별 참고]\n• iCUBE: TB_SYS_CONFIG 테이블 조회\n• iU ERP: 환경설정 > 보안정책 메뉴\n• WEHAGO: 클라우드 관리자 콘솔 > 보안설정 (CSP 정책 포함)\n• Amaranth: 시스템설정 > 비밀번호 정책"},
        "PC04": {"population": "iCUBE 운영계 변경 제한 설정", "completeness": "[자동통제 확인방법]\n[메뉴] 시스템관리 > 환경설정 > 시스템설정\n운영환경 Direct 수정 불가 설정 확인\n더존 ERP는 Package 솔루션으로 소스 수정 불가"},
        "PC05": {"population": "iCUBE 패치 적용 권한 보유자", "completeness": "[자동통제 확인방법]\n[메뉴] 시스템관리 > 관리자설정 > 패치권한\n[Query] SELECT USER_ID, USER_NM FROM TB_SYS_USER WHERE PATCH_AUTH = 'Y'\n패치 적용 권한 보유자 = 시스템 관리자로 제한"},
        "CO02": {"population": "iCUBE 배치 스케줄링 권한 (TB_SYS_BATCH_AUTH)", "completeness": "[자동통제 확인방법]\n[메뉴] 시스템관리 > 배치관리 > 권한설정\n[Query] SELECT USER_ID, BATCH_AUTH FROM TB_SYS_USER WHERE BATCH_AUTH = 'Y'\n배치 등록/변경 권한자 목록 조회"},
        "CO03": {"population": "iCUBE 배치 실행 모니터링 (TB_SYS_BATCH_LOG)", "completeness": "[자동통제 확인방법]\n[메뉴] 시스템관리 > 배치관리 > 실행이력\n[Query] SELECT BATCH_ID, BATCH_NM, EXEC_DT, RESULT_CD FROM TB_SYS_BATCH_LOG\nWHERE RESULT_CD = 'F' AND EXEC_DT BETWEEN '시작일' AND '종료일'\n실패 건 알림 설정 확인"},
        "CO04": {"population": "iCUBE 백업 모니터링", "completeness": "[자동통제 확인방법]\nDB 백업은 DBMS 레벨에서 수행 (DB 템플릿 참조)\n[메뉴] 시스템관리 > 백업관리 > 백업이력\n백업 성공/실패 로그 조회"},
        "ST01": {"population": "DB 접근통제 솔루션 관리자 설정", "completeness": "[자동통제 확인방법]\n접근통제 솔루션(DBSafer/Chakra/Petra 등) 관리 콘솔\n관리자 계정 목록 및 권한 부여 이력 조회"},
        "ST02": {"population": "DB 접근통제 솔루션 정책 설정", "completeness": "[자동통제 확인방법]\n접근통제 솔루션 관리 콘솔 > 보안 정책 설정\n- iCUBE 접근 허용 IP 대역\n- 업무시간 외 접근 차단 정책\n- 민감정보 컬럼 마스킹 정책"},
    },
    "YOUNG": {
        "APD01": {"population": "K-System 사용자권한 변경 이력 (TBSM_USER_AUTH)", "completeness": "시스템관리 > 사용자관리 > 권한부여이력에서 테스트 기간 내 변경 건 조회\n[Query] SELECT * FROM TBSM_USER_AUTH WHERE MOD_DATE BETWEEN '시작일' AND '종료일'\n\n[버전별 참고]\n• K-System: TBSM_USER_AUTH 테이블 직접 조회\n• K-System Plus: 권한관리 강화 - 상세 이력 조회 메뉴 제공\n• SystemEver: 클라우드 관리포털 > 사용자/권한 > 변경이력"},
        "APD02": {"population": "K-System 부서이동자 권한회수 이력", "completeness": "인사시스템 전배자 목록과 TBSM_USER_AUTH 권한변경이력 대사\n\n[버전별 참고]\n• K-System: HR 모듈 연동 또는 별도 인사시스템 대사\n• K-System Plus: 인사 연동 자동화 지원 (HR Interface)\n• SystemEver: 클라우드 HR 연동 - 자동 권한 조정 가능"},
        "APD03": {"population": "K-System 퇴사자 계정 비활성화 이력 (TBSM_USER)", "completeness": "인사시스템 퇴사자 목록과 TBSM_USER.USE_FLAG='N' 대사\n[Query] SELECT USER_ID, USER_NAME, USE_FLAG FROM TBSM_USER WHERE USE_FLAG='N'\n\n[버전별 참고]\n• K-System: USE_FLAG='N' 확인\n• K-System Plus: 계정상태 '퇴직' 필드 추가\n• SystemEver: 클라우드 사용자 상태 '비활성' (SSO 연동 시 자동 처리)"},
        "APD07": {"population": "K-System DB 직접 변경 로그", "completeness": "영림원 유지보수팀 데이터 변경 요청서 목록 입수 (Package 특성상 직접 변경 미발생 확인)\n\n[버전별 참고]\n• K-System/Plus: 영림원 엔지니어 통한 변경만 가능\n• SystemEver: CSP 관리 영역 (고객사 직접 DB 접근 불가)"},
        "PC01": {"population": "영림원 유지보수 변경요청 건", "completeness": "영림원 고객지원 SR(Service Request) 목록 또는 K-System 패치 적용 이력 조회\n\n[버전별 참고]\n• K-System: 영림원 고객지원포털 > SR 목록\n• K-System Plus: 고객포털 > 패치/업데이트 요청\n• SystemEver: 클라우드 관리포털 > 업데이트 이력 (자동 적용)"},
        "PC02": {"population": "K-System 패치 테스트 결과 (테스트서버)", "completeness": "영림원 패치 적용 전 테스트서버(DEV) 검증 결과 확인\n[메뉴] 영림원 고객지원포털 > 패치이력 > 테스트결과\n테스트 완료 후 검수확인서 입수\n\n[버전별 참고]\n• K-System/Plus: DEV서버 구성 시 별도 테스트\n• SystemEver: 영림원 자체 테스트 후 일괄 배포 (고객 테스트 불필요)"},
        "PC03": {"population": "K-System 패치 적용 승인 건 (TBSM_PATCH_HIST)", "completeness": "시스템관리 > 패치관리에서 패치 적용 이력 조회\n[Query] SELECT PATCH_NO, PATCH_NM, APPLY_DATE, APPLY_USER FROM TBSM_PATCH_HIST WHERE APPLY_DATE BETWEEN '시작일' AND '종료일'\n패치 적용 품의서 또는 변경 승인서와 대사\n\n[버전별 참고]\n• K-System: TBSM_PATCH_HIST 테이블\n• K-System Plus: 패치관리 메뉴 강화\n• SystemEver: 클라우드 업데이트 이력 (자동 적용 승인 프로세스)"},
        "CO01": {"population": "K-System 자동분개 스케줄 변경 (TBSM_BATCH_JOB)", "completeness": "시스템관리 > 배치관리에서 등록/변경 이력 조회\n[Query] SELECT JOB_ID, JOB_NAME, EXEC_CYCLE, MOD_DATE, MOD_USER FROM TBSM_BATCH_JOB WHERE MOD_DATE BETWEEN '시작일' AND '종료일'\n\n[버전별 참고]\n• K-System: TBSM_BATCH_JOB 테이블 조회\n• K-System Plus: 배치관리 UI 개선\n• SystemEver: 클라우드 스케줄러 관리 (AWS EventBridge 등)"},
        "ST03": {"population": "K-System 시스템 설정 변경 이력 (TBSM_SYS_CONFIG)", "completeness": "시스템관리 > 환경설정에서 설정 변경 이력 조회\n[Query] SELECT CONFIG_ID, CONFIG_VAL, MOD_DATE, MOD_USER FROM TBSM_SYS_CONFIG_LOG WHERE MOD_DATE BETWEEN '시작일' AND '종료일'\n\n[버전별 참고]\n• K-System/Plus: TBSM_SYS_CONFIG_LOG 테이블\n• SystemEver: 클라우드 설정 변경 감사로그"},
        "PD01": {"population": "K-System 신규 모듈 도입 건", "completeness": "영림원 SI 프로젝트 완료보고서 또는 모듈 추가 계약서 입수\n[메뉴] 영림원 고객지원포털 > 프로젝트이력\n\n[버전별 참고]\n• K-System: 모듈 추가 계약 및 설치\n• K-System Plus: 확장 모듈 라이선스\n• SystemEver: 클라우드 모듈 구독 (SaaS)"},
        "PD02": {"population": "K-System 데이터 마이그레이션 수행 건", "completeness": "마이그레이션 수행 로그 조회\n[Query] SELECT MIG_ID, MIG_NAME, EXEC_DATE, EXEC_USER FROM TBSM_MIGRATION_LOG WHERE EXEC_DATE BETWEEN '시작일' AND '종료일'\n데이터 검증 결과서와 대사\n\n[버전별 참고]\n• K-System/Plus: 마이그레이션 도구 제공\n• SystemEver: 클라우드 마이그레이션 서비스"},
        "PD03": {"population": "K-System SI 프로젝트 이슈 목록", "completeness": "영림원 고객지원포털 > 프로젝트관리 > 이슈등록 목록 조회\n또는 내부 프로젝트 관리 시스템(Redmine/Jira) 이슈 목록\n\n[버전별 참고]\n• K-System/Plus: 고객포털 이슈 관리\n• SystemEver: 클라우드 티켓 시스템"},
        "PD04": {"population": "K-System 사용자 교육 수행 건", "completeness": "영림원 고객지원포털 > 교육이력 또는 내부 교육 완료 보고서\n매뉴얼 배포 이력 확인\n\n[버전별 참고]\n• K-System/Plus: 집합/방문 교육\n• SystemEver: 온라인 LMS 교육 제공"},
        "CO05": {"population": "K-System 시스템 장애 발생 건 (TBSM_ERROR_LOG)", "completeness": "시스템관리 > 오류로그에서 장애 발생 건 조회\n[Query] SELECT ERR_ID, ERR_MSG, ERR_DATE, USER_ID FROM TBSM_ERROR_LOG WHERE ERR_DATE BETWEEN '시작일' AND '종료일'\n영림원 고객지원포털 장애접수 이력과 대사\n\n[버전별 참고]\n• K-System/Plus: 오류로그 + 장애접수\n• SystemEver: 클라우드 SLA 모니터링 (99.5%)"},
        "APD04": {"population": "K-System 권한 모니터링 화면 (TBSM_USER_AUTH)", "completeness": "[자동통제 확인방법]\n[메뉴] 시스템관리 > 사용자관리 > 권한현황조회\n[Query] SELECT USER_ID, AUTH_CODE, AUTH_NAME, GRANT_DATE FROM TBSM_USER_AUTH\nWHERE AUTH_CODE IN ('SYS_ADMIN', 'SUPER_USER') -- 주요 권한 보유자 조회\n\n[버전별 참고]\n• K-System: TBSM_USER_AUTH 쿼리\n• K-System Plus: 권한현황 리포트 메뉴\n• SystemEver: 클라우드 관리포털 > 권한현황"},
        "APD05": {"population": "K-System 관리자 권한 보유자 (TBSM_ADMIN_USER)", "completeness": "[자동통제 확인방법]\n[메뉴] 시스템관리 > 관리자설정 > 관리자목록\n[Query] SELECT USER_ID, USER_NAME, ADMIN_FLAG, REG_DATE FROM TBSM_USER WHERE ADMIN_FLAG = 'Y'\n시스템 관리자 계정 목록 및 권한 범위 확인\n\n[버전별 참고]\n• K-System/Plus: ADMIN_FLAG='Y' 필터\n• SystemEver: 클라우드 관리자 역할 보유자"},
        "APD06": {"population": "영림원 ERP 패스워드 정책 설정", "completeness": "[자동통제 확인방법]\n[메뉴] 시스템관리 > 환경설정 > 보안정책\n[Query] SELECT CONFIG_ID, CONFIG_VALUE FROM TBSM_SYS_CONFIG WHERE CONFIG_ID LIKE 'SEC_PWD_%'\n- SEC_PWD_MIN_LENGTH: 최소 길이 (8자리 이상)\n- SEC_PWD_EXPIRE_DAYS: 만료 기간 (90일)\n- SEC_PWD_LOCK_COUNT: 잠금 임계값 (5회)\n- SEC_PWD_COMPLEXITY: 복잡도 설정\n\n[버전별 참고]\n• K-System: TBSM_SYS_CONFIG 테이블 조회\n• K-System Plus: 확장된 보안정책 메뉴 (2FA 지원)\n• SystemEver: 클라우드 관리자 포털 > 보안설정 (Azure AD 연동 가능)"},
        "PC04": {"population": "K-System 운영계 변경 제한 설정", "completeness": "[자동통제 확인방법]\n[메뉴] 시스템관리 > 환경설정 > 시스템모드\n운영환경 소스 수정 불가 설정 확인\n영림원 K-System은 Package 솔루션으로 고객사 직접 소스 수정 불가\n\n[버전별 참고]\n• K-System/Plus: Package - 소스 수정 불가\n• SystemEver: SaaS - 커스터마이징 제한"},
        "PC05": {"population": "K-System 패치 적용 권한 보유자", "completeness": "[자동통제 확인방법]\n[메뉴] 시스템관리 > 관리자설정 > 패치관리권한\n[Query] SELECT USER_ID, USER_NAME FROM TBSM_USER WHERE PATCH_AUTH_FLAG = 'Y'\n패치 적용 권한 보유자 = 시스템 관리자로 제한 확인\n\n[버전별 참고]\n• K-System/Plus: 패치권한자 쿼리 확인\n• SystemEver: 자동 업데이트 (영림원 관리)"},
        "CO02": {"population": "K-System 배치 스케줄링 권한 (TBSM_BATCH_AUTH)", "completeness": "[자동통제 확인방법]\n[메뉴] 시스템관리 > 배치관리 > 권한설정\n[Query] SELECT USER_ID, BATCH_AUTH_FLAG FROM TBSM_USER WHERE BATCH_AUTH_FLAG = 'Y'\n배치 Job 등록/변경 권한 보유자 목록 조회\n\n[버전별 참고]\n• K-System/Plus: BATCH_AUTH_FLAG 조회\n• SystemEver: 클라우드 배치 권한 설정"},
        "CO03": {"population": "K-System 배치 실행 모니터링 (TBSM_BATCH_LOG)", "completeness": "[자동통제 확인방법]\n[메뉴] 시스템관리 > 배치관리 > 실행결과조회\n[Query] SELECT JOB_ID, JOB_NAME, EXEC_DATE, RESULT_CODE, ERR_MSG FROM TBSM_BATCH_LOG\nWHERE RESULT_CODE = 'FAIL' AND EXEC_DATE BETWEEN '시작일' AND '종료일'\n실패 건 이메일/SMS 알림 설정 확인\n\n[버전별 참고]\n• K-System/Plus: TBSM_BATCH_LOG 조회\n• SystemEver: 클라우드 모니터링 대시보드"},
        "CO04": {"population": "K-System 백업 모니터링", "completeness": "[자동통제 확인방법]\nDB 백업은 DBMS 레벨에서 수행 (DB 템플릿 참조)\n[메뉴] 시스템관리 > 운영관리 > 백업현황\n백업 스케줄 및 성공/실패 로그 조회\n\n[버전별 참고]\n• K-System/Plus: DB 레벨 백업 확인\n• SystemEver: CSP 자동 백업 (AWS/Azure)"},
        "ST01": {"population": "DB 접근통제 솔루션 관리자 설정", "completeness": "[자동통제 확인방법]\n접근통제 솔루션(DBSafer/Chakra/Petra 등) 관리 콘솔\n관리자 계정 목록 및 권한 부여 이력 조회\n\n[버전별 참고]\n• K-System/Plus: 접근통제 솔루션 도입 시\n• SystemEver: 클라우드 보안 그룹 설정"},
        "ST02": {"population": "DB 접근통제 솔루션 정책 설정", "completeness": "[자동통제 확인방법]\n접근통제 솔루션 관리 콘솔 > 보안 정책 설정\n- K-System 접근 허용 IP 대역\n- 업무시간 외 접근 차단 정책\n- 개인정보 컬럼 마스킹 정책\n\n[버전별 참고]\n• K-System/Plus: On-Premise 접근통제\n• SystemEver: VPC/보안그룹 + WAF 정책"},
    },
    "ETC": {
        "APD01": {"population": "사용자 권한 부여 이력", "completeness": "권한 관리 시스템에서 테스트 기간 내 권한 부여 건 조회"},
        "APD02": {"population": "부서이동자 권한 회수 이력", "completeness": "인사시스템 전배자 목록과 권한 변경 이력 대사"},
        "APD03": {"population": "퇴사자 계정 비활성화 이력", "completeness": "인사시스템 퇴사자 목록과 사용자 상태 대사"},
        "APD07": {"population": "데이터 직접 변경 이력", "completeness": "시스템 감사 로그에서 직접 변경 건 조회"},
        "PC01": {"population": "프로그램 변경 요청 건", "completeness": "형상관리 시스템(Git/SVN)에서 커밋 이력 조회"},
        "PC02": {"population": "변경 테스트 결과", "completeness": "테스트 관리 도구 또는 테스트 결과서 입수"},
        "PC03": {"population": "프로그램 이관 승인 건", "completeness": "배포 시스템 또는 이관 승인서 목록 입수"},
        "CO01": {"population": "Batch Job 등록/변경 이력", "completeness": "스케줄러에서 테스트 기간 내 변경 건 조회"},
        "ST03": {"population": "지원도구 변경 요청 이력", "completeness": "IT 시스템 설정 변경 승인서 목록 조회"},
        "PD01": {"population": "SI 프로젝트 완료 건", "completeness": "경영정보팀 SI 완료 목록 입수"},
        "PD02": {"population": "데이터 마이그레이션 수행 건", "completeness": "SI 프로젝트 중 마이그레이션 동반 건 식별"},
        "PD03": {"population": "SI 프로젝트 이슈 목록", "completeness": "프로젝트 관리 시스템의 이슈 등록 건 조회"},
        "PD04": {"population": "SI 프로젝트 교육 수행 건", "completeness": "교육 수행 결과 또는 매뉴얼 배포 이력 조회"},
        "CO05": {"population": "시스템 장애 발생 건", "completeness": "장애 관리 시스템 또는 월간 보고서 조회"},
    },
}

# OS별 템플릿 - APD09~11, PC06
OS_POPULATION_TEMPLATES = {
    "RHEL": {
        "APD09": {"population": "OS 계정 생성/변경 이력", "completeness": "/var/log/secure에서 useradd/usermod 명령 이력 조회\n[Query] grep -E 'useradd|usermod|userdel' /var/log/secure\n\n[버전별 참고]\n• RHEL 7: /var/log/secure, systemd-journald\n• RHEL 8/9: journalctl -u systemd-logind + /var/log/secure\n• RHEL 9: 강화된 감사 로그 (auditd 기본 활성화)"},
        "APD10": {"population": "OS 패스워드 정책 설정", "completeness": "/etc/login.defs 및 /etc/pam.d/system-auth 설정 확인\n[Query] cat /etc/login.defs | grep -E 'PASS_MAX|PASS_MIN|PASS_WARN'\ncat /etc/pam.d/system-auth | grep pam_pwquality\n\n[버전별 참고]\n• RHEL 7: pam_cracklib 또는 pam_pwquality\n• RHEL 8/9: pam_pwquality 기본 (pwquality.conf)\n• RHEL 9: crypto-policies 통합 관리"},
        "APD11": {"population": "root 및 sudo 권한 보유자", "completeness": "/etc/passwd, /etc/sudoers에서 관리자 권한자 목록 추출\n[Query] awk -F: '$3==0 {print $1}' /etc/passwd\ncat /etc/sudoers /etc/sudoers.d/*\n\n[버전별 참고]\n• RHEL 7/8/9: wheel 그룹 기반 sudo 권한\n• RHEL 8/9: sudoers.d 디렉토리 분리 관리 권장"},
        "PC06": {"population": "OS 패치/업데이트 이력", "completeness": "yum history 또는 /var/log/yum.log에서 패치 이력 조회\n[Query] yum history list\n\n[버전별 참고]\n• RHEL 7: yum history, /var/log/yum.log\n• RHEL 8/9: dnf history, /var/log/dnf.log\n• RHEL 9: dnf5 도입 예정"},
    },
    "UBUNTU": {
        "APD09": {"population": "OS 계정 생성/변경 이력", "completeness": "/var/log/auth.log에서 useradd/usermod 명령 이력 조회\n[Query] grep -E 'useradd|usermod|userdel' /var/log/auth.log\n\n[버전별 참고]\n• Ubuntu 18.04 LTS: /var/log/auth.log\n• Ubuntu 20.04/22.04 LTS: journalctl + auth.log\n• Ubuntu 22.04+: 강화된 AppArmor 프로파일"},
        "APD10": {"population": "OS 패스워드 정책 설정", "completeness": "/etc/login.defs 및 /etc/pam.d/common-password 설정 확인\n[Query] cat /etc/login.defs | grep PASS\ncat /etc/pam.d/common-password\n\n[버전별 참고]\n• Ubuntu 18.04: pam_cracklib\n• Ubuntu 20.04/22.04: pam_pwquality 기본\n• Ubuntu 22.04: libpwquality 강화"},
        "APD11": {"population": "root 및 sudo 권한 보유자", "completeness": "/etc/passwd, /etc/sudoers에서 관리자 권한자 목록 추출\n[Query] getent group sudo\ncat /etc/sudoers /etc/sudoers.d/*\n\n[버전별 참고]\n• Ubuntu 전버전: sudo 그룹 기반 권한\n• Ubuntu 20.04+: sudoers.d 분리 관리"},
        "PC06": {"population": "OS 패치/업데이트 이력", "completeness": "/var/log/apt/history.log에서 패치 이력 조회\n[Query] cat /var/log/apt/history.log\n\n[버전별 참고]\n• Ubuntu 18.04: apt history.log\n• Ubuntu 20.04/22.04: unattended-upgrades 자동 패치\n• Ubuntu 22.04: Ubuntu Pro 보안 업데이트"},
    },
    "WINDOWS": {
        "APD09": {"population": "AD 계정 생성/변경 이력", "completeness": "Event Log(Security) 4720/4722/4738 이벤트 조회\n[Query] Get-WinEvent -FilterHashtable @{LogName='Security';ID=4720,4722,4738}\n\n[버전별 참고]\n• Server 2016: 기본 감사 정책\n• Server 2019: Advanced Audit Policy 강화\n• Server 2022: Azure AD 연동 강화, Defender 통합"},
        "APD10": {"population": "AD 패스워드 정책 설정", "completeness": "GPO(그룹 정책) 패스워드 정책 설정 확인\n[Query] Get-ADDefaultDomainPasswordPolicy\nnet accounts (로컬)\n\n[버전별 참고]\n• Server 2016: 기본 GPO 정책\n• Server 2019: Fine-Grained Password Policy\n• Server 2022: Azure AD Password Protection 연동"},
        "APD11": {"population": "Administrators 그룹 멤버", "completeness": "AD에서 Domain Admins, Local Admins 그룹 멤버 조회\n[Query] Get-ADGroupMember -Identity 'Domain Admins'\nnet localgroup Administrators\n\n[버전별 참고]\n• Server 2016/2019: 기본 AD 그룹 관리\n• Server 2022: Privileged Access Management (PAM)"},
        "PC06": {"population": "Windows Update 이력", "completeness": "WSUS 또는 Windows Update 히스토리 조회\n[Query] Get-WmiObject -Class Win32_QuickFixEngineering\nGet-WindowsUpdateLog\n\n[버전별 참고]\n• Server 2016: WSUS/SCCM\n• Server 2019: Windows Update for Business\n• Server 2022: Azure Update Management 통합"},
    },
    "UNIX": {
        "APD09": {"population": "OS 계정 생성/변경 이력", "completeness": "/var/adm/syslog에서 계정 변경 이력 조회\n\n[버전별 참고]\n• AIX: /var/adm/syslog, /etc/security/failedlogin\n• HP-UX: /var/adm/syslog, /var/adm/wtmps\n• Solaris: /var/adm/messages, /var/log/authlog"},
        "APD10": {"population": "OS 패스워드 정책 설정", "completeness": "/etc/security/user 또는 /etc/default/passwd 설정 확인\n\n[버전별 참고]\n• AIX: /etc/security/user (minlen, maxage 등)\n• HP-UX: /etc/default/security\n• Solaris: /etc/default/passwd, pam.conf"},
        "APD11": {"population": "root 권한 보유자", "completeness": "/etc/passwd에서 UID 0 계정 및 sudo 권한자 목록 추출\n\n[버전별 참고]\n• AIX: /etc/security/user roles 확인\n• HP-UX: /etc/privgroup\n• Solaris: RBAC (Role-Based Access Control)"},
        "PC06": {"population": "OS 패치/업데이트 이력", "completeness": "instfix -i 또는 oslevel -s 명령으로 패치 이력 조회\n\n[버전별 참고]\n• AIX: instfix -i, oslevel -s\n• HP-UX: swlist -l patch\n• Solaris: pkg info, patchadd -p"},
    },
    "N/A": {
        "APD09": {"population": "N/A (CSP Managed)", "completeness": "SOC 1/2 Type II 리포트로 갈음"},
        "APD10": {"population": "N/A (CSP Managed)", "completeness": "SOC 1/2 Type II 리포트로 갈음"},
        "APD11": {"population": "N/A (CSP Managed)", "completeness": "SOC 1/2 Type II 리포트로 갈음"},
        "PC06": {"population": "N/A (CSP Managed)", "completeness": "SOC 1/2 Type II 리포트로 갈음"},
    },
}

# DB별 템플릿 - APD08, APD12~14, PC07
DB_POPULATION_TEMPLATES = {
    "ORACLE": {
        "APD08": {"population": "DB 직접 변경 권한 보유자", "completeness": "DBA_TAB_PRIVS에서 INSERT/UPDATE/DELETE 권한자 조회\n[Query] SELECT * FROM DBA_TAB_PRIVS WHERE PRIVILEGE IN ('INSERT','UPDATE','DELETE')\n\n[버전별 참고]\n• 11g: DBA_TAB_PRIVS 기본 조회\n• 12c/19c: CDB/PDB 환경 시 CDB_TAB_PRIVS 활용\n• 19c/21c: Data Redaction, Real Application Security 추가 확인"},
        "APD12": {"population": "DB 계정 생성/변경 이력", "completeness": "DBA_USERS에서 테스트 기간 내 생성/변경된 계정 조회\n[Query] SELECT USERNAME, CREATED, ACCOUNT_STATUS FROM DBA_USERS WHERE CREATED BETWEEN '시작일' AND '종료일'\n\n[버전별 참고]\n• 11g: DBA_USERS 기본 조회\n• 12c+: Common User(C##) vs Local User 구분\n• 19c/21c: Last Login 추적 가능 (DBA_USERS.LAST_LOGIN)"},
        "APD13": {"population": "DB 패스워드 정책 설정", "completeness": "DBA_PROFILES에서 PASSWORD 관련 파라미터 확인\n[Query] SELECT * FROM DBA_PROFILES WHERE RESOURCE_NAME LIKE 'PASSWORD%'\n\n[버전별 참고]\n• 11g: PASSWORD_LIFE_TIME, PASSWORD_REUSE_TIME 등\n• 12c+: PASSWORD_VERIFY_FUNCTION 강화 (ora12c_verify_function)\n• 19c/21c: 대소문자 구분 기본값 변경, Gradual Password Rollover"},
        "APD14": {"population": "DBA 권한 보유자", "completeness": "DBA_ROLE_PRIVS에서 DBA Role 부여자 목록 조회\n[Query] SELECT * FROM DBA_ROLE_PRIVS WHERE GRANTED_ROLE='DBA'\n\n[버전별 참고]\n• 11g/12c: DBA_ROLE_PRIVS, DBA_SYS_PRIVS 조회\n• 19c+: Privilege Analysis 기능 활용 가능\n• 21c: Schema-Only 계정 도입 (보안 강화)"},
        "PC07": {"population": "DB 패치/업데이트 이력", "completeness": "DBA_REGISTRY_HISTORY 또는 opatch lsinventory로 패치 이력 조회\n[Query] SELECT * FROM DBA_REGISTRY_HISTORY ORDER BY ACTION_TIME DESC\n\n[버전별 참고]\n• 11g: opatch lsinventory\n• 12c+: datapatch, DBA_REGISTRY_SQLPATCH\n• 19c/21c: Release Update (RU) 분기별 패치"},
    },
    "TIBERO": {
        "APD08": {"population": "DB 직접 변경 권한 보유자", "completeness": "DBA_TAB_PRIVS에서 INSERT/UPDATE/DELETE 권한자 조회\n[Query] SELECT * FROM DBA_TAB_PRIVS WHERE PRIVILEGE IN ('INSERT','UPDATE','DELETE')\n\n[버전별 참고]\n• Tibero 5: Oracle 호환 딕셔너리 뷰\n• Tibero 6: 향상된 보안 기능, Audit 강화\n• Tibero 7: 클라우드 지원, 컨테이너 환경"},
        "APD12": {"population": "DB 계정 생성/변경 이력", "completeness": "DBA_USERS에서 테스트 기간 내 생성/변경된 계정 조회\n[Query] SELECT USERNAME, CREATED, ACCOUNT_STATUS FROM DBA_USERS\n\n[버전별 참고]\n• Tibero 5/6/7: Oracle 호환 DBA_USERS 뷰 동일"},
        "APD13": {"population": "DB 패스워드 정책 설정", "completeness": "DBA_PROFILES에서 PASSWORD 관련 파라미터 확인\n[Query] SELECT * FROM DBA_PROFILES WHERE RESOURCE_NAME LIKE 'PASSWORD%'\n\n[버전별 참고]\n• Tibero 5/6: Oracle 호환 프로파일\n• Tibero 7: 강화된 패스워드 검증 함수"},
        "APD14": {"population": "DBA 권한 보유자", "completeness": "DBA_ROLE_PRIVS에서 DBA Role 부여자 목록 조회\n[Query] SELECT * FROM DBA_ROLE_PRIVS WHERE GRANTED_ROLE='DBA'\n\n[버전별 참고]\n• Tibero 전버전: Oracle 호환 권한 관리"},
        "PC07": {"population": "DB 패치/업데이트 이력", "completeness": "tbpatch 명령 또는 패치 적용 이력 문서 확인\n\n[버전별 참고]\n• Tibero 5/6: tbpatch 명령 또는 티맥스 고객포털\n• Tibero 7: 자동 패치 관리 도구 제공"},
    },
    "MSSQL": {
        "APD08": {"population": "DB 직접 변경 권한 보유자", "completeness": "sys.database_permissions에서 DML 권한자 조회\n[Query] SELECT pr.name, pe.permission_name FROM sys.database_permissions pe JOIN sys.database_principals pr ON pe.grantee_principal_id = pr.principal_id\n\n[버전별 참고]\n• SQL Server 2016: 기본 권한 뷰\n• SQL Server 2019: 행 수준 보안(RLS) 확인\n• SQL Server 2022: Ledger 테이블, Always Encrypted 강화"},
        "APD12": {"population": "DB 로그인/사용자 생성 이력", "completeness": "sys.server_principals에서 create_date 기준 조회\n[Query] SELECT name, create_date, modify_date FROM sys.server_principals WHERE create_date BETWEEN '시작일' AND '종료일'\n\n[버전별 참고]\n• SQL Server 2016: 기본 조회\n• SQL Server 2019: Azure AD 통합 인증\n• SQL Server 2022: Microsoft Entra ID 지원"},
        "APD13": {"population": "DB 패스워드 정책 설정", "completeness": "sys.sql_logins의 is_policy_checked 설정 확인\n[Query] SELECT name, is_policy_checked, is_expiration_checked FROM sys.sql_logins\n\n[버전별 참고]\n• SQL Server 2016+: Windows 정책 연동 (CHECK_POLICY)\n• SQL Server 2019/2022: Azure AD 정책 연동 가능"},
        "APD14": {"population": "sysadmin 권한 보유자", "completeness": "sys.server_role_members에서 sysadmin 멤버 조회\n[Query] SELECT m.name FROM sys.server_role_members r JOIN sys.server_principals m ON r.member_principal_id = m.principal_id WHERE r.role_principal_id = SUSER_ID('sysadmin')\n\n[버전별 참고]\n• SQL Server 전버전: sysadmin 고정 서버 역할\n• SQL Server 2022: 세분화된 서버 역할 추가"},
        "PC07": {"population": "DB 패치/업데이트 이력", "completeness": "@@VERSION 또는 Windows Update 히스토리 조회\n[Query] SELECT @@VERSION\nSELECT * FROM sys.dm_os_sys_info\n\n[버전별 참고]\n• SQL Server 2016: SP/CU 패치\n• SQL Server 2019/2022: CU 누적 업데이트 방식"},
    },
    "MYSQL": {
        "APD08": {"population": "DB 직접 변경 권한 보유자", "completeness": "mysql.db에서 Insert/Update/Delete_priv 권한자 조회\n[Query] SELECT User, Host, Db, Insert_priv, Update_priv, Delete_priv FROM mysql.db WHERE Insert_priv='Y' OR Update_priv='Y' OR Delete_priv='Y'\n\n[버전별 참고]\n• MySQL 5.7: mysql.db, mysql.user 테이블\n• MySQL 8.0: 역할 기반 권한 관리 도입\n• MySQL 8.0.30+: 동적 권한 확장"},
        "APD12": {"population": "DB 계정 생성/변경 이력", "completeness": "감사로그(Audit Log) 또는 mysql.user의 password_last_changed 기준 조회\n[Query] SELECT User, Host, password_last_changed FROM mysql.user\n(※ MySQL은 mysql.user 테이블에 Create_time을 저장하지 않으므로 감사로그 확인 필수)\n\n[버전별 참고]\n• MySQL 5.7: mysql.user 기본 조회\n• MySQL 8.0: audit_log 플러그인 또는 General Log 활용\n• MySQL 8.0+: CREATE USER ... FAILED_LOGIN_ATTEMPTS 옵션"},
        "APD13": {"population": "DB 패스워드 정책 설정", "completeness": "validate_password 플러그인 설정 확인\n[Query] SHOW VARIABLES LIKE 'validate_password%'\n\n[버전별 참고]\n• MySQL 5.7: validate_password 플러그인\n• MySQL 8.0: validate_password 컴포넌트 (기본 활성화)\n• MySQL 8.0+: password_history, password_reuse_interval 추가"},
        "APD14": {"population": "root 및 SUPER 권한 보유자", "completeness": "mysql.user에서 Super_priv='Y' 계정 조회\n[Query] SELECT User, Host FROM mysql.user WHERE Super_priv='Y'\n\n[버전별 참고]\n• MySQL 5.7: SUPER 권한\n• MySQL 8.0: SUPER 권한 세분화 (CONNECTION_ADMIN 등)\n• MySQL 8.0+: 동적 권한으로 SUPER 대체 권장"},
        "PC07": {"population": "DB 패치/업데이트 이력", "completeness": "SELECT VERSION() 또는 패키지 업데이트 이력 조회\n[Query] SELECT VERSION()\n\n[버전별 참고]\n• MySQL 5.7: Innovation/LTS 릴리스 구분 없음\n• MySQL 8.0: 분기별 릴리스\n• MySQL 8.0 LTS: 장기 지원 버전 (8.0.34+)"},
    },
    "POSTGRES": {
        "APD08": {"population": "DB 직접 변경 권한 보유자", "completeness": "information_schema.table_privileges에서 DML 권한자 조회\n[Query] SELECT grantee, table_name, privilege_type FROM information_schema.table_privileges WHERE privilege_type IN ('INSERT','UPDATE','DELETE')\n\n[버전별 참고]\n• PostgreSQL 12/13: 기본 권한 뷰\n• PostgreSQL 14: 역할 기반 권한 강화\n• PostgreSQL 15/16: Row-Level Security 개선"},
        "APD12": {"population": "DB Role 생성/변경 이력", "completeness": "pg_roles에서 rolvaliduntil 등 기준 조회\n[Query] SELECT rolname, rolcreatedb, rolcreaterole, rolvaliduntil FROM pg_roles\n\n[버전별 참고]\n• PostgreSQL 12+: pg_roles, pg_user 뷰 동일\n• PostgreSQL 15+: GRANT ... WITH ADMIN OPTION 추적"},
        "APD13": {"population": "DB 패스워드 정책 설정", "completeness": "pg_hba.conf 및 passwordcheck 모듈 설정 확인\n[Query] SHOW password_encryption\n\n[버전별 참고]\n• PostgreSQL 12/13: md5 또는 scram-sha-256\n• PostgreSQL 14+: scram-sha-256 기본 권장\n• PostgreSQL 15+: passwordcheck 모듈 강화"},
        "APD14": {"population": "superuser 권한 보유자", "completeness": "pg_roles에서 rolsuper=true 계정 조회\n[Query] SELECT rolname FROM pg_roles WHERE rolsuper=true\n\n[버전별 참고]\n• PostgreSQL 전버전: superuser 개념 동일\n• PostgreSQL 15+: pg_read_all_data 등 사전정의 역할 확장"},
        "PC07": {"population": "DB 패치/업데이트 이력", "completeness": "SELECT version() 또는 패키지 업데이트 이력 조회\n[Query] SELECT version()\n\n[버전별 참고]\n• PostgreSQL 12-16: 연간 메이저 릴리스\n• 마이너 릴리스: 분기별 보안 패치"},
    },
    "HANA": {
        "APD08": {"population": "DB 직접 변경 권한 보유자", "completeness": "SYS.GRANTED_PRIVILEGES에서 DML 권한자 조회\n[Query] SELECT * FROM SYS.GRANTED_PRIVILEGES WHERE PRIVILEGE IN ('INSERT','UPDATE','DELETE')\n\n[버전별 참고]\n• HANA 1.0: Studio 기반 조회\n• HANA 2.0: Cockpit 및 System Views 활용\n• HANA Cloud: Managed DB Policy 적용"},
        "APD12": {"population": "DB 계정 생성/변경 이력", "completeness": "SYS.USERS에서 생성/변경된 계정 조회\n[Query] SELECT USER_NAME, CREATE_TIME, USER_DEACTIVATED FROM SYS.USERS WHERE CREATE_TIME BETWEEN '시작일' AND '종료일'\n\n[버전별 참고]\n• HANA 전버전: SYS.USERS 시스템 뷰 사용"},
        "APD13": {"population": "DB 패스워드 정책 설정", "completeness": "SYS.USER_PARAMETERS에서 PASSWORD 관련 설정 확인\n[Query] SELECT * FROM SYS.USER_PARAMETERS WHERE PARAMETER LIKE 'PASSWORD%'\n\n[버전별 참고]\n• HANA 2.0: 'Password Policy' 메뉴에서 시각적 확인 가능"},
        "APD14": {"population": "시스템 관리자 권한 보유자", "completeness": "SYS.GRANTED_ROLES에서 CONTENT_ADMIN, USER_ADMIN 등 역할 조회\n[Query] SELECT * FROM SYS.GRANTED_ROLES WHERE ROLE_NAME IN ('CONTENT_ADMIN', 'USER_ADMIN')\n\n[버전별 참고]\n• HANA 전버전: SYSTEM 계정 및 관리자 역할 보유자 확인"},
        "PC07": {"population": "DB 패치/업데이트 이력", "completeness": "HANA 버전에 따른 패치 이력 확인\n[Query] SELECT * FROM SYS.M_DATABASE\n\n[버전별 참고]\n• HANA 2.0: SPS (Support Package Stack) 업데이트 이력"},
    },
    "NOSQL": {
        "APD08": {"population": "DB 직접 변경 권한 보유자", "completeness": "MongoDB: db.getUsers() 또는 역할 기반 접근 제어 설정 확인\n[Query] db.getUsers() 또는 db.getRoles()\n\n[버전별 참고]\n• MongoDB 4.x: RBAC 기본\n• MongoDB 5.x: 필드 레벨 암호화 강화\n• MongoDB 6.x/7.x: Queryable Encryption"},
        "APD12": {"population": "DB 사용자 생성/변경 이력", "completeness": "시스템 로그에서 사용자 생성 이력 조회\n\n[버전별 참고]\n• MongoDB 4.x+: Audit Log 활성화 필요\n• MongoDB 5.x+: 강화된 감사 로그\n• MongoDB Atlas: 클라우드 감사 로그"},
        "APD13": {"population": "DB 인증 설정", "completeness": "mongod.conf의 authorization 설정 확인\n[설정 파일] security.authorization: enabled\n\n[버전별 참고]\n• MongoDB 4.x: SCRAM-SHA-256 기본\n• MongoDB 5.x+: LDAP/Kerberos 연동 강화\n• MongoDB 6.x+: OIDC 인증 지원"},
        "APD14": {"population": "관리자 권한 보유자", "completeness": "admin DB의 userAdminAnyDatabase 역할 보유자 조회\n[Query] use admin; db.getUsers()\n\n[버전별 참고]\n• MongoDB 전버전: Built-in 역할 구조 동일\n• MongoDB 5.x+: 커스텀 역할 강화"},
        "PC07": {"population": "DB 버전 업데이트 이력", "completeness": "db.version() 또는 패키지 업데이트 이력 조회\n[Query] db.version()\n\n[버전별 참고]\n• MongoDB 4.x/5.x/6.x/7.x: 연간 메이저 릴리스\n• MongoDB Atlas: 자동 업그레이드 옵션"},
    },
    "N/A": {
        "APD08": {"population": "N/A (CSP Managed)", "completeness": "SOC 1/2 Type II 리포트로 갈음"},
        "APD12": {"population": "N/A (CSP Managed)", "completeness": "SOC 1/2 Type II 리포트로 갈음"},
        "APD13": {"population": "N/A (CSP Managed)", "completeness": "SOC 1/2 Type II 리포트로 갈음"},
        "APD14": {"population": "N/A (CSP Managed)", "completeness": "SOC 1/2 Type II 리포트로 갈음"},
        "PC07": {"population": "N/A (CSP Managed)", "completeness": "SOC 1/2 Type II 리포트로 갈음"},
    },
}

def get_user_info():
    """현재 로그인한 사용자 정보 반환"""
    if 'user_id' in session:
        return session.get('user_info')
    return None

# ================================
# AI 엔진: 통제 활동 자동 생성
# ================================
def generate_ai_rcm_content(system_info):
    """
    AI를 사용하여 시스템 환경에 맞는 통제 활동 및 기술적 증적 생성
    """
    api_key = os.getenv('OPENAI_API_KEY')
    if not api_key:
        return {"error": "API Key not found"}

    client = OpenAI(api_key=api_key)
    
    # 컨텍스트 요약
    context = (
        f"시스템명: {system_info.get('system_name')}, "
        f"유형: {system_info.get('system_type')}, "
        f"SW: {system_info.get('software')}, "
        f"OS: {system_info.get('os')}, "
        f"DB: {system_info.get('db')}"
    )

    prompt = f"""당신은 IT 감사 및 ITGC 전문가입니다.
다음 시스템 환경에 대해 주어진 ITGC 통제 항목별 '통제 활동'과 '테스트 절차'를 생성하십시오.

[시스템 환경]
{context}

[작성 지침]
1. 각 통제 항목에 대해 해당 기술 스택에 적합한 '기술적 증적(Technical Objects)'을 명시하십시오.
   - 예: SAP -> SU01, PFCG, STAD, SE16
   - 예: Oracle ERP -> FND_USER, FND_RESPONSIBILITY, AD_APPL_TOP
   - 예: DB가 Oracle일 경우 → DBA_USERS, DBA_TAB_PRIVS, Audit Trail
   - 예: Linux일 경우 → /etc/passwd, sudoers, syslog
2. 시스템 유형(System Type)에 따라 다음 통제 범위를 적용하십시오:
   - In-house (자체개발): 개발 및 변경관리(SDLC) 전체 통제 항목을 상세히 포함.
   - Package-Modifiable: 패키지 표준 기능 + 커스터마이징 영역에 대한 변경 통제 포함.
   - Package-Non-modifiable: 운영 및 권한 통제 위주로 작성하고, 개발/변경 통제는 제외하거나 최소화.
3. 응답은 반드시 아래 JSON 형식으로만 작성하십시오.

[JSON 형식]
{{
  "controls": [
    {{
      "id": "APD01",
      "activity": "통제 활동 내용...",
      "procedure": "테스트 절차 내용..."
    }},
    ...
  ]
}}
"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "당신은 IT 전문 감사인이며 정확한 기술 용어를 사용합니다."},
                {"role": "user", "content": prompt}
            ],
            response_format={ "type": "json_object" },
            temperature=0.0
        )
        return json.loads(response.choices[0].message.content)
    except Exception as e:
        return {"error": str(e)}

# ================================
# AI 엔진: Gemini를 통한 통제 활동 자동 생성
# ================================
def generate_gemini_rcm_content(system_info):
    """
    Google Gemini AI를 사용하여 시스템 환경에 맞는 통제 활동 및 기술적 증적 생성
    """
    api_key = os.getenv('GEMINI_API_KEY')
    if not api_key:
        return {"error": "Gemini API Key not found"}

    # 컨텍스트 요약
    context = (
        f"시스템명: {system_info.get('system_name')}, "
        f"유형: {system_info.get('system_type')}, "
        f"SW: {system_info.get('software')}, "
        f"OS: {system_info.get('os')}, "
        f"DB: {system_info.get('db')}"
    )

    prompt = f"""당신은 IT 감사 및 ITGC 전문가입니다.
다음 시스템 환경에 대해 주어진 ITGC 통제 항목별 '통제 활동'과 '테스트 절차'를 생성하십시오.

[시스템 환경]
{context}

[작성 지침]
1. 각 통제 항목에 대해 해당 기술 스택에 적합한 '기술적 증적(Technical Objects)'을 명시하십시오.
   - 예: SAP -> SU01, PFCG, STAD, SE16
   - 예: Oracle ERP -> FND_USER, FND_RESPONSIBILITY, AD_APPL_TOP
   - 예: DB가 Oracle일 경우 → DBA_USERS, DBA_TAB_PRIVS, Audit Trail
   - 예: Linux일 경우 → /etc/passwd, sudoers, syslog
2. 시스템 유형(System Type)에 따라 다음 통제 범위를 적용하십시오:
   - In-house (자체개발): 개발 및 변경관리(SDLC) 전체 통제 항목을 상세히 포함.
   - Package-Modifiable: 패키지 표준 기능 + 커스터마이징 영역에 대한 변경 통제 포함.
   - Package-Non-modifiable: 운영 및 권한 통제 위주로 작성하고, 개발/변경 통제는 제외하거나 최소화.
3. 응답은 반드시 아래 JSON 형식으로만 작성하십시오. 다른 텍스트 없이 순수 JSON만 반환하세요.

[JSON 형식]
{{
  "controls": [
    {{
      "id": "APD01",
      "activity": "통제 활동 내용...",
      "procedure": "테스트 절차 내용..."
    }}
  ]
}}
"""

    try:
        client = genai.Client(api_key=api_key)
        model_id = 'gemini-2.0-flash'

        max_retries = 3
        for attempt in range(max_retries):
            try:
                response = client.models.generate_content(
                    model=model_id,
                    contents=prompt
                )

                # 응답 텍스트에서 JSON 추출
                response_text = response.text.strip()

                # ```json ... ``` 형식으로 감싸져 있을 경우 처리
                if response_text.startswith('```'):
                    lines = response_text.split('\n')
                    # 첫 줄(```json)과 마지막 줄(```) 제거
                    json_lines = []
                    in_json = False
                    for line in lines:
                        if line.startswith('```') and not in_json:
                            in_json = True
                            continue
                        elif line.startswith('```') and in_json:
                            break
                        elif in_json:
                            json_lines.append(line)
                    response_text = '\n'.join(json_lines)

                return json.loads(response_text)

            except Exception as e:
                err_msg = str(e)
                if ("429" in err_msg or "RESOURCE_EXHAUSTED" in err_msg) and attempt < max_retries - 1:
                    wait_time = (attempt + 1) * 5
                    print(f"Gemini AI 분석 제한 발생 (시도 {attempt+1}/{max_retries}). {wait_time}초 후 재시도합니다...")
                    time.sleep(wait_time)
                    continue
                raise e

    except Exception as e:
        return {"error": str(e)}

# ================================
# Blueprint 라우트
# ================================

@bp_link1_1.route('/link1_1')
def link1_1():
    """AI RCM Builder 메인 페이지"""
    user_info = get_user_info()
    users = user_info['user_name'] if user_info else "Guest"
    user_email = user_info.get('user_email', '') if user_info else ''

    return render_template('link1_1.jsp',
                         users=users,
                         is_logged_in='user_id' in session,
                         user_info=user_info,
                         user_email=user_email,
                         master_controls=MASTER_ITGC_CONTROLS)

@bp_link1_1.route('/api/rcm/ai_generate', methods=['POST'])
@login_required
def api_rcm_ai_generate():
    """AI를 통한 RCM 내용 생성 API (Gemini 기본 사용)"""
    data = request.json

    # Gemini를 기본으로 사용, 실패 시 OpenAI로 fallback
    result = generate_gemini_rcm_content(data)

    if "error" in result:
        # Gemini 실패 시 OpenAI로 재시도
        print(f"Gemini 실패: {result['error']}, OpenAI로 재시도...")
        result = generate_ai_rcm_content(data)

        if "error" in result:
            return jsonify({"success": False, "message": result["error"]}), 500

    return jsonify({"success": True, "data": result["controls"]})

@bp_link1_1.route('/api/rcm/population_templates', methods=['GET'])
def api_get_population_templates():
    """모집단/완전성 템플릿 데이터 반환"""
    return jsonify({
        "success": True,
        "sw_templates": SW_POPULATION_TEMPLATES,
        "os_templates": OS_POPULATION_TEMPLATES,
        "db_templates": DB_POPULATION_TEMPLATES
    })

# 주기별 모집단/표본수 기준 (백엔드 계산용)
FREQUENCY_POPULATION = {'연': 1, '분기': 4, '월': 12, '주': 52, '일': 250, '수시': 0, '기타': 0}
FREQUENCY_SAMPLE = {'연': 1, '분기': 2, '월': 2, '주': 5, '일': 20, '수시': 25, '기타': 0}

@bp_link1_1.route('/api/rcm/export_excel', methods=['POST'])
@login_required
def api_rcm_export_excel():
    """최종 RCM 엑셀 내보내기 - 마스터 데이터 기반"""
    import base64
    from openpyxl import Workbook

    data = request.json
    user_overrides = {item['id']: item for item in data.get('rcm_data', [])}
    system_info = data.get('system_info', {})

    # 마스터 데이터를 ID로 조회할 수 있게 딕셔너리로 변환
    master_dict = {ctrl['id']: ctrl for ctrl in MASTER_ITGC_CONTROLS}

    wb = Workbook()
    ws = wb.active
    ws.title = "RCM"

    # 헤더 작성 (UI의 모든 컬럼 반영)
    headers = [
        "Risk Code", "Risk Name", "Category", "Control Code", "Control Name",
        "Control Description", "구분", "주기", "성격", "모집단", "모집단 수", "모집단 완전성", "표본수", "테스트 절차"
    ]


    # 헤더 스타일 적용
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 데이터 작성 - 마스터 데이터 순서대로
    for row_idx, control in enumerate(MASTER_ITGC_CONTROLS, 2):
        ctrl_id = control['id']
        override = user_overrides.get(ctrl_id, {})

        # 1. Risk Code / 2. Risk Name / 3. Category
        ws.cell(row=row_idx, column=1, value=control.get('risk_code', ''))
        ws.cell(row=row_idx, column=2, value=control.get('risk_description', ''))
        ws.cell(row=row_idx, column=3, value=control.get('category', ''))
        
        # 4. Control Code / 5. Control Name
        ws.cell(row=row_idx, column=4, value=ctrl_id)
        ws.cell(row=row_idx, column=5, value=control.get('name', ''))
        
        # 6. Control Description (AI 분석 결과 우선)
        activity = override.get('activity', control.get('control_description', ''))
        ws.cell(row=row_idx, column=6, value=activity)

        # 테스트 절차 초기화 및 설정 (UI 선택값 우선)
        procedure = override.get('procedure', '')

        # 7. 구분 / 8. 주기 / 9. 성격 (UI 선택값 우선)
        selected_type = override.get('type', control.get('type', ''))
        selected_freq = override.get('frequency', control.get('frequency', ''))
        selected_method = override.get('method', control.get('method', ''))

        ws.cell(row=row_idx, column=7, value=selected_type)
        ws.cell(row=row_idx, column=8, value=selected_freq)
        ws.cell(row=row_idx, column=9, value=selected_method)

        # 10. 모집단 / 11. 모집단 수 / 12. 모집단 완전성 / 13. 표본수
        pop_names = {'연': '연간 모니터링 문서', '분기': '분기별 모니터링 문서', '월': '월별 모니터링 문서', '주': '주별 모니터링 문서', '일': '일별 모니터링 문서'}
        
        # 시스템 정보를 기반으로 수시 통제 템플릿 적용
        sw = system_info.get('software', 'ETC')
        os_type = system_info.get('os', 'RHEL')
        db_type = system_info.get('db', 'ORACLE')
        
        pop_name = ""
        completeness = ""
        pop_count = 0
        
        # 통제 ID에 따라 기본 템플릿 조회 (모집단/완전성 템플릿)
        template = None
        if ctrl_id in ['APD01', 'APD02', 'APD03', 'APD04', 'APD05', 'APD06', 'APD07', 'PC01', 'PC02', 'PC03', 'PC04', 'PC05', 'CO01', 'CO02', 'ST01', 'ST02', 'ST03', 'PD01', 'PD02', 'PD03', 'PD04', 'CO05']:
            template = SW_POPULATION_TEMPLATES.get(sw, {}).get(ctrl_id)
        elif ctrl_id in ['APD09', 'APD10', 'APD11', 'PC06']:
            template = OS_POPULATION_TEMPLATES.get(os_type, {}).get(ctrl_id)
        elif ctrl_id in ['APD08', 'APD12', 'APD13', 'APD14', 'PC07']:
            template = DB_POPULATION_TEMPLATES.get(db_type, {}).get(ctrl_id)

        # 주기에 따른 기본 설정
        if selected_freq in pop_names:
            pop_name = pop_names[selected_freq]
            pop_count = FREQUENCY_POPULATION.get(selected_freq, 0)
            
            # 템플릿 정보가 있으면 기술적 정보를 완전성에 결합
            if template:
                pop_name = template.get('population', pop_name)
                completeness = f"{template.get('completeness', '')}\n\n[결과] {pop_name}이므로 {pop_count}건을 완전성 있는 것으로 확인함"
            elif pop_count > 0:
                completeness = f"{pop_name}이므로 {pop_count}건을 완전성 있는 것으로 확인함"
        elif selected_freq == '수시':
            # APD02, APD03 수동 설정 시 예외 처리 (사용자 요청 사항)
            if selected_type in ['Manual', '수동'] and ctrl_id == 'APD02':
                pop_name = "부서이동자리스트"
                completeness = "인사시스템 상의 부서이동자 명단과 권한 회수 내역 전수 대사"
            elif selected_type in ['Manual', '수동'] and ctrl_id == 'APD03':
                pop_name = "퇴사자리스트"
                completeness = "인사시스템 상의 퇴사자 명단과 계정 비활성화 내역 전수 대사"
            else:
                # 일반적인 수시 통제인 경우 템플릿 사용
                if template:
                    pop_name = template.get('population', '수시 발생 건')
                    completeness = template.get('completeness', '전수 대사 또는 샘플 테스트를 통해 완전성 확인')
                else:
                    pop_name = "수시 발생 건"
                    completeness = "수시 발생 건이므로 모집단 완전성을 별도로 확인해야 함"
            pop_count = 0 # 수시는 보통 가변적임
                
        # 자동통제이거나 주기가 '기타'인 경우
        if selected_type in ['Auto', '자동'] or selected_freq == '기타':
            # 자동통제의 경우 템플릿에서 조회 화면 정보 가져오기
            if template:
                pop_name = template.get('population', 'N/A')
            else:
                pop_name = "N/A"
            
            # 자동통제일 경우 완전성 확인은 '제외'로 표시
            completeness = "제외 (자동통제)"
            pop_count = 0
            sample_count = 0
        else:
            sample_count = FREQUENCY_SAMPLE.get(selected_freq, 0)

        # 시스템 제외 통제 처리 (Cloud Managed)
        cloud_env = system_info.get('cloud_env', 'None')
        is_cloud_excluded = False
        if cloud_env != 'None' and ctrl_id == 'CO06': is_cloud_excluded = True
        if cloud_env in ['SaaS', 'PaaS'] and ctrl_id in ['APD09', 'APD10', 'APD11', 'PC06']: is_cloud_excluded = True
        if cloud_env == 'SaaS' and ctrl_id in ['APD12', 'APD13', 'APD14', 'PC01', 'PC02', 'PC03', 'PC04', 'PC05', 'PC07']: is_cloud_excluded = True

        if is_cloud_excluded:
            pop_name = "N/A (CSP Managed)"
            pop_count = 0
            completeness = "SOC 1/2 Type II 리포트로 갈음"
            sample_count = 0
            if not procedure:
                procedure = "[CSP Managed] 본 통제는 클라우드 서비스 제공자의 책임 영역에 해당하므로, 당해년도 CSP의 SOC 1/2 Type II 리포트 상의 물리적/환경적 보안 적정성 검토 결과로 갈음함."

        ws.cell(row=row_idx, column=10, value=pop_name)
        ws.cell(row=row_idx, column=11, value=pop_count if pop_count > 0 else ('' if selected_freq == '수시' else 0))
        ws.cell(row=row_idx, column=12, value=completeness)
        ws.cell(row=row_idx, column=13, value=sample_count)

        # 14. 테스트 절차
        if not procedure:
            if selected_type in ['Auto', '자동']:
                procedure = control.get('test_procedure_auto', '')
            else:
                procedure = control.get('test_procedure_manual', '')
        
        ws.cell(row=row_idx, column=14, value=procedure)

        # 셀 스타일 적용
        for col in range(1, 15):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border

    # 열 너비 조정
    column_widths = {
        'A': 10, 'B': 40, 'C': 12, 'D': 12, 'E': 25,
        'F': 50, 'G': 10, 'H': 10, 'I': 10, 'J': 20, 'K': 10, 'L': 45, 'M': 10, 'N': 60
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 첫 행 고정
    ws.freeze_panes = 'A2'

    # 결과 전송
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.now().strftime("%Y%m%d")
    system_name = system_info.get('system_name', 'System').replace(' ', '_')
    filename = f"ITGC_RCM_{system_name}_{today}.xlsx"

    xlsx_base64 = base64.b64encode(output.getvalue()).decode('utf-8')

    return jsonify({
        "success": True,
        "filename": filename,
        "file_data": xlsx_base64
    })
