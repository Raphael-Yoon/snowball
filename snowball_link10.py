# -*- coding: utf-8 -*-
import sys
import os

# Windows 콘솔 UTF-8 설정
if os.name == 'nt':
    os.system('chcp 65001 > nul')
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')

# [System Config] 불필요한 프록시 설정 제거 (서버 환경 호환성)
for key in ['HTTP_PROXY', 'HTTPS_PROXY', 'http_proxy', 'https_proxy']:
    os.environ.pop(key, None)

from flask import Blueprint, render_template, jsonify, send_file, request, session
import threading
import uuid
from datetime import datetime, timedelta
import json
import warnings
warnings.filterwarnings('ignore')

# Data collection dependencies (optional imports)
try:
    import OpenDartReader
    OPENDART_AVAILABLE = True
except ImportError:
    OPENDART_AVAILABLE = False
    print("Warning: OpenDartReader not installed. DART financial data will not be available.")

try:
    import pandas as pd
    from pykrx import stock
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    PYKRX_AVAILABLE = True
except ImportError:
    PYKRX_AVAILABLE = False
    print("Warning: pykrx or related packages not installed. Stock data collection will not be available.")

import time

bp_link10 = Blueprint('link10', __name__)

# 작업 상태 저장 (실제 운영 환경에서는 Redis 등 사용 권장)
tasks = {}

# 결과 파일 저장 디렉토리 - snowball 프로젝트 기준으로 설정
RESULTS_DIR = os.path.join(os.path.dirname(__file__), 'trade_results')
try:
    if not os.path.exists(RESULTS_DIR):
        os.makedirs(RESULTS_DIR, exist_ok=True)
except Exception as e:
    print(f"trade_results 디렉토리 생성 실패: {e}")
    # 임시 디렉토리 사용
    import tempfile
    RESULTS_DIR = tempfile.gettempdir()

# DART API KEY (환경변수로 관리)
API_KEY = os.environ.get('DART_API_KEY', '')

# ============================================================================
# Data Collection Functions (from data_collect.py)
# ============================================================================

def get_latest_business_day():
    """가장 최근의 영업일을 반환합니다."""
    try:
        end_date = datetime.now().strftime("%Y%m%d")
        start_date = (datetime.now() - timedelta(days=10)).strftime("%Y%m%d")
        ohlcv = stock.get_market_ohlcv(start_date, end_date, "005930")
        if ohlcv.empty:
            return (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
        return ohlcv.index[-1].strftime("%Y%m%d")
    except Exception as e:
        print(f"영업일 조회 오류: {e}. 어제 날짜 사용")
        return (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")

def get_sector_from_pykrx(ticker, sector_df):
    """pykrx 업종 분류 데이터에서 업종 정보를 추출합니다."""
    try:
        if ticker in sector_df.index:
            sector = sector_df.loc[ticker].iloc[1]
            return sector if pd.notna(sector) else 'N/A'
        return 'N/A'
    except:
        return 'N/A'

def get_dart_financials(dart, ticker, year):
    """OpenDARTReader를 사용하여 재무 데이터를 추출합니다."""
    try:
        df = dart.finstate_all(ticker, year)
        if df is None or df.empty:
            return 0, 0, 0, 0, 0, 0, 0, 0, 0

        revenue = op = re = cash = liabilities = equity = ocf = capex = da = 0

        for _, row in df.iterrows():
            acc_id = str(row.get('account_id', ''))
            acc_name = str(row['account_nm']).replace(" ", "")
            val = pd.to_numeric(row['thstrm_amount'], errors='coerce')
            if pd.isna(val):
                val = 0

            # 1. 매출액
            if acc_id == 'ifrs-full_Revenue' or acc_name == '매출액' or acc_name == '수익(매출액)':
                if revenue == 0 or acc_id == 'ifrs-full_Revenue':
                    revenue = val

            # 2. 영업이익
            elif acc_id == 'dart_OperatingIncomeLoss' or acc_name == '영업이익' or acc_name == '영업이익(손실)':
                if op == 0 or acc_id == 'dart_OperatingIncomeLoss':
                    op = val

            # 3. 이익잉여금
            elif acc_id == 'ifrs-full_RetainedEarnings' or (re == 0 and '이익잉여금' in acc_name and '기타' not in acc_name):
                if re == 0 or acc_id == 'ifrs-full_RetainedEarnings':
                    re = val

            # 4. 현금및현금성자산
            elif acc_id == 'ifrs-full_CashAndCashEquivalents' or (cash == 0 and '현금및현금성자산' in acc_name):
                if cash == 0 or acc_id == 'ifrs-full_CashAndCashEquivalents':
                    cash = val

            # 5. 부채총계
            elif acc_id == 'ifrs-full_Liabilities' or acc_name == '부채총계':
                liabilities = val

            # 6. 자본총계
            elif acc_id == 'ifrs-full_Equity' or acc_name == '자본총계':
                if equity == 0 or acc_id == 'ifrs-full_Equity':
                    equity = val

            # 7. 영업활동현금흐름
            elif acc_id == 'ifrs-full_CashFlowsFromUsedInOperatingActivities' or acc_name == '영업활동현금흐름':
                ocf = val

            # 8. 유형/무형자산 취득
            elif 'PurchaseOfPropertyPlantAndEquipment' in acc_id or 'PurchaseOfIntangibleAssets' in acc_id:
                capex += val
            elif acc_name in ['유형자산의취득', '무형자산의취득']:
                capex += val

            # 9. 감가상각비
            if 'Depreciation' in acc_id or 'Amortisation' in acc_id or '감가상각' in acc_name:
                da += val

        return revenue, op, re, cash, liabilities, equity, ocf, capex, da
    except Exception as e:
        print(f"[DART] {ticker} 재무제표 조회 실패: {e}")
        return 0, 0, 0, 0, 0, 0, 0, 0, 0

def collect_stock_data(task_id, stock_count=100, selected_fields=None):
    """종목 데이터 수집 메인 함수"""
    try:
        tasks[task_id]['status'] = 'running'
        tasks[task_id]['progress'] = 0
        tasks[task_id]['message'] = '데이터 수집 시작...'

        # 필수 라이브러리 확인
        if not PYKRX_AVAILABLE:
            tasks[task_id]['status'] = 'error'
            tasks[task_id]['message'] = '필수 라이브러리(pykrx, pandas, openpyxl)가 설치되지 않았습니다. 서버 관리자에게 문의하세요.'
            return

        # DART API 초기화 (실패해도 계속 진행)
        dart = None
        if OPENDART_AVAILABLE and API_KEY:
            try:
                # file_cache 사용 안함으로 초기화
                dart = OpenDartReader(API_KEY)
                # 기업 코드 캐시 비활성화
                dart.corp_codes = None
                tasks[task_id]['message'] = 'DART API 연결 성공'
            except Exception as e:
                tasks[task_id]['message'] = f'DART API 연결 실패 (재무 데이터 제외): {str(e)[:100]}'
                print(f"DART API 초기화 실패: {e}")
        else:
            if not OPENDART_AVAILABLE:
                tasks[task_id]['message'] = 'OpenDartReader 미설치 (재무 데이터 제외)'
            else:
                tasks[task_id]['message'] = 'DART API 키 없음 (재무 데이터 제외)'

        # 1. 최근 영업일 조회
        try:
            latest_date = get_latest_business_day()
            tasks[task_id]['message'] = f'최근 영업일: {latest_date}'
            tasks[task_id]['progress'] = 5
        except Exception as e:
            tasks[task_id]['status'] = 'error'
            tasks[task_id]['message'] = f'영업일 조회 실패: {str(e)[:100]}'
            print(f"영업일 조회 오류: {e}")
            return

        # 2. pykrx로 시가총액 및 기본 데이터 수집
        try:
            tasks[task_id]['message'] = 'pykrx 데이터 수집 중...'
            tasks[task_id]['progress'] = 10
            df_cap = stock.get_market_cap_by_ticker(latest_date, market="KOSPI")
            tasks[task_id]['progress'] = 15
            df_fundamental = stock.get_market_fundamental(latest_date, market="KOSPI")
            tasks[task_id]['progress'] = 20
            df_sector = stock.get_market_sector_classifications(latest_date, market="KOSPI")
            tasks[task_id]['progress'] = 25
            tasks[task_id]['message'] = f'pykrx 데이터 수집 완료 ({len(df_cap)}개 종목)'
        except Exception as e:
            tasks[task_id]['status'] = 'error'
            tasks[task_id]['message'] = f'pykrx 데이터 수집 실패: {str(e)[:150]}'
            print(f"pykrx 오류: {e}")
            import traceback
            traceback.print_exc()
            return

        # 상위 N개 종목
        if stock_count == 0:
            df_top = df_cap.sort_values(by='시가총액', ascending=False)
        else:
            df_top = df_cap.sort_values(by='시가총액', ascending=False).head(stock_count)

        # 3. 업종별 평균 PBR, PER 계산
        df_merged = pd.concat([df_fundamental, df_sector[['업종명']]], axis=1)
        df_valid = df_merged[(df_merged['PBR'] > 0) & (df_merged['PER'] > 0)]
        industry_avg = df_valid.groupby('업종명')[['PBR', 'PER']].mean()
        industry_avg_dict = industry_avg.to_dict('index')

        # 4. 종목별 상세 데이터 수집
        current_year = datetime.now().year - 1
        results = []

        end_date = latest_date
        start_date = (datetime.strptime(latest_date, "%Y%m%d") - timedelta(days=365)).strftime("%Y%m%d")

        total = len(df_top)
        for idx, ticker in enumerate(df_top.index, 1):
            name = stock.get_market_ticker_name(ticker)

            # 진행률 업데이트
            progress = int((idx / total) * 100)
            tasks[task_id]['progress'] = progress
            tasks[task_id]['message'] = f'진행률: [{idx}/{total}] {progress}% 완료'

            # pykrx 데이터
            pbr = df_fundamental.loc[ticker, 'PBR'] if ticker in df_fundamental.index else 0.0
            per = df_fundamental.loc[ticker, 'PER'] if ticker in df_fundamental.index else 0.0
            eps = df_fundamental.loc[ticker, 'EPS'] if ticker in df_fundamental.index else 0.0
            bps = df_fundamental.loc[ticker, 'BPS'] if ticker in df_fundamental.index else 0.0
            div_yield = df_fundamental.loc[ticker, 'DIV'] if ticker in df_fundamental.index else 0.0
            roe = (eps / bps * 100) if bps > 0 else 0.0

            # 52주 최고가/최저가
            high_52w = low_52w = 0
            try:
                df_ohlcv = stock.get_market_ohlcv_by_date(start_date, end_date, ticker)
                if not df_ohlcv.empty:
                    high_52w = int(df_ohlcv['고가'].max())
                    low_52w = int(df_ohlcv['저가'].min())
            except:
                pass

            # 업종 정보
            sector = get_sector_from_pykrx(ticker, df_sector)
            avg_pbr = avg_per = 0.0
            if sector in industry_avg_dict:
                avg_pbr = industry_avg_dict[sector]['PBR']
                avg_per = industry_avg_dict[sector]['PER']

            # DART 재무 데이터 (API 사용 가능한 경우만)
            if dart:
                revenue, op, re, cash, liabilities, equity, ocf, capex, da = get_dart_financials(dart, ticker, current_year)
            else:
                revenue = op = re = cash = liabilities = equity = ocf = capex = da = 0

            # 추가 지표 계산
            debt_ratio = (liabilities / equity * 100) if equity > 0 else 0.0
            fcf = ocf - capex
            ebitda = op + da

            results.append({
                '종목코드': ticker,
                '종목명': name,
                '업종': sector,
                'PBR': round(pbr, 2),
                '업종평균PBR': round(avg_pbr, 2),
                'PER': round(per, 2),
                '업종평균PER': round(avg_per, 2),
                'ROE': round(roe, 2),
                'EPS': int(eps),
                'BPS': int(bps),
                '배당수익률': round(div_yield, 2),
                '매출액': int(revenue),
                '영업이익': int(op),
                '이익잉여금': int(re),
                '현금및현금성자산': int(cash),
                '52주최고가': int(high_52w),
                '52주최저가': int(low_52w),
                '부채비율': round(debt_ratio, 2),
                'FCF': int(fcf),
                'EBITDA': int(ebitda)
            })

            time.sleep(0.05)

        df_result = pd.DataFrame(results)

        # 선택된 필드만 필터링
        if selected_fields:
            required_fields = ['종목코드', '종목명']
            fields_to_include = required_fields + [f for f in selected_fields if f not in required_fields]

            if 'PBR' in fields_to_include and '업종평균PBR' not in fields_to_include:
                idx_pbr = fields_to_include.index('PBR')
                fields_to_include.insert(idx_pbr + 1, '업종평균PBR')
            if 'PER' in fields_to_include and '업종평균PER' not in fields_to_include:
                idx_per = fields_to_include.index('PER')
                fields_to_include.insert(idx_per + 1, '업종평균PER')

            fields_to_include = [f for f in fields_to_include if f in df_result.columns]
            df_result = df_result[fields_to_include]

        # 엑셀 저장
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        count_label = 'all' if stock_count == 0 else f'top{stock_count}'
        result_filename = f'kospi_{count_label}_{timestamp}.xlsx'
        result_path = os.path.join(RESULTS_DIR, result_filename)

        try:
            # 엑셀 파일 생성 및 저장
            df_result.to_excel(result_path, index=False, engine='openpyxl')

            # 엑셀 포맷팅
            wb = load_workbook(result_path)
            ws = wb.active
            ws.auto_filter.ref = ws.dimensions

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            wb.save(result_path)

        except PermissionError as e:
            tasks[task_id]['status'] = 'error'
            tasks[task_id]['message'] = f'파일 저장 권한 오류: {RESULTS_DIR} 디렉토리에 쓰기 권한이 없습니다.'
            print(f"파일 저장 권한 오류: {e}")
            return
        except Exception as e:
            tasks[task_id]['status'] = 'error'
            tasks[task_id]['message'] = f'엑셀 파일 저장 실패: {str(e)[:100]}'
            print(f"엑셀 저장 오류: {e}")
            return

        tasks[task_id]['status'] = 'completed'
        tasks[task_id]['progress'] = 100
        tasks[task_id]['message'] = '데이터 수집 완료!'
        tasks[task_id]['result_file'] = result_filename

        cleanup_old_results()

    except Exception as e:
        tasks[task_id]['status'] = 'error'
        tasks[task_id]['message'] = f'오류 발생: {str(e)}'

# ============================================================================
# Flask Helper Functions
# ============================================================================

def is_logged_in():
    """로그인 상태 확인"""
    return 'user_id' in session

def get_user_info():
    """현재 로그인한 사용자 정보 반환 (세션 우선)"""
    if is_logged_in():
        # 세션에 저장된 user_info를 우선 사용
        if 'user_info' in session:
            return session['user_info']
        # 없으면 데이터베이스에서 조회
        from auth import get_current_user
        db_user_info = get_current_user()
        return db_user_info
    return None

def cleanup_old_results(max_files=20):
    """오래된 결과 파일 자동 삭제"""
    try:
        files = []
        for filename in os.listdir(RESULTS_DIR):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(RESULTS_DIR, filename)
                files.append((file_path, os.path.getctime(file_path)))

        # 생성 시간순 정렬 (오래된 순)
        files.sort(key=lambda x: x[1])

        # max_files를 초과하는 파일 삭제
        if len(files) > max_files:
            for i in range(len(files) - max_files):
                os.remove(files[i][0])
                print(f"자동 삭제됨: {files[i][0]}")
    except Exception as e:
        print(f"파일 정리 중 오류: {e}")

def run_data_collection(task_id, stock_count=100, fields=None):
    """백그라운드에서 데이터 수집 실행 (내부 함수 호출)"""
    tasks[task_id]['stock_count'] = stock_count
    collect_stock_data(task_id, stock_count, fields)

@bp_link10.route('/link10')
def index():
    """메인 페이지"""
    user_logged_in = is_logged_in()
    user_info = get_user_info()
    return render_template('link10.jsp',
                         is_logged_in=user_logged_in,
                         user_info=user_info)

@bp_link10.route('/link10/api/collect', methods=['POST'])
def start_collection():
    """데이터 수집 시작"""
    data = request.get_json() or {}
    stock_count = data.get('stock_count', 100)
    fields = data.get('fields', [])

    # 유효성 검사
    if not isinstance(stock_count, int) or stock_count < 0 or stock_count > 10000:
        return jsonify({
            'success': False,
            'message': '종목 수는 0(전체) 또는 1~10000 사이의 숫자여야 합니다.'
        }), 400

    task_id = str(uuid.uuid4())
    tasks[task_id] = {
        'status': 'pending',
        'progress': 0,
        'message': '대기 중...',
        'stock_count': stock_count,
        'created_at': datetime.now().isoformat()
    }

    # 백그라운드 스레드로 실행
    thread = threading.Thread(target=run_data_collection, args=(task_id, stock_count, fields))
    thread.start()

    message = '전체 종목 데이터 수집이 시작되었습니다.' if stock_count == 0 else f'{stock_count}개 종목 데이터 수집이 시작되었습니다.'

    return jsonify({
        'success': True,
        'task_id': task_id,
        'message': message
    })

@bp_link10.route('/link10/api/status/<task_id>', methods=['GET'])
def get_status(task_id):
    """작업 상태 조회"""
    if task_id not in tasks:
        return jsonify({'error': '작업을 찾을 수 없습니다.'}), 404

    return jsonify(tasks[task_id])

@bp_link10.route('/link10/api/download/<filename>', methods=['GET'])
def download_file(filename):
    """결과 파일 다운로드"""
    file_path = os.path.join(RESULTS_DIR, filename)

    if not os.path.exists(file_path):
        return jsonify({'error': '파일을 찾을 수 없습니다.'}), 404

    return send_file(file_path, as_attachment=True, download_name=filename)

@bp_link10.route('/link10/api/delete/<filename>', methods=['DELETE'])
def delete_file(filename):
    """결과 파일 삭제"""
    file_path = os.path.join(RESULTS_DIR, filename)

    if not os.path.exists(file_path):
        return jsonify({'success': False, 'message': '파일을 찾을 수 없습니다.'}), 404

    try:
        os.remove(file_path)
        return jsonify({'success': True, 'message': '파일이 삭제되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'삭제 실패: {str(e)}'}), 500

@bp_link10.route('/link10/api/results', methods=['GET'])
def list_results():
    """저장된 결과 파일 목록"""
    files = []
    if os.path.exists(RESULTS_DIR):
        for filename in os.listdir(RESULTS_DIR):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(RESULTS_DIR, filename)
                stat = os.stat(file_path)
                files.append({
                    'filename': filename,
                    'size': stat.st_size,
                    'created_at': datetime.fromtimestamp(stat.st_ctime).isoformat()
                })

    # 최신순 정렬
    files.sort(key=lambda x: x['created_at'], reverse=True)
    return jsonify(files)
