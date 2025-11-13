from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, session
from auth import get_db, get_current_user, login_required, get_user_activity_logs, get_activity_log_count, get_all_rcms, create_rcm, get_rcm_details, save_rcm_details, grant_rcm_access
import tempfile
import os
from datetime import date, timedelta

# Blueprint 생성
admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

def get_user_info():
    """현재 로그인한 사용자 정보 반환"""
    user_id = session.get('user_id')
    if not user_id:
        return None
    
    with get_db() as conn:
        user = conn.execute(
            'SELECT * FROM sb_user WHERE user_id = %s', (user_id,)
        ).fetchone()
        return dict(user) if user else None

def is_logged_in():
    """로그인 상태 확인"""
    return 'user_id' in session

def require_admin():
    """관리자 권한 확인 함수"""
    if not is_logged_in():
        flash('로그인이 필요합니다.')
        return redirect(url_for('login'))
    
    user_info = get_user_info()
    if not user_info or user_info.get('admin_flag') != 'Y':
        flash('관리자 권한이 필요합니다.')
        return redirect(url_for('index'))
    
    return None

def perform_auto_mapping(headers):
    """컬럼명 기반 자동 매핑"""
    # 시스템 필드와 가능한 컬럼명 패턴 정의
    field_patterns = {
        'control_code': [
            '통제코드', '통제 코드', 'control_code', 'control code', 
            '코드', 'code', '번호', 'no', 'id'
        ],
        'control_name': [
            '통제명', '통제 명', '통제이름', '통제 이름', 'control_name', 'control name',
            '통제활동명', '명칭', 'name', 'title', '제목'
        ],
        'control_description': [
            '통제활동설명', '통제 활동 설명', '설명', '상세설명', 'description', 'detail',
            '내용', 'content', '통제활동', '활동설명'
        ],
        'key_control': [
            '핵심통제여부', '핵심통제', '핵심 통제', 'key_control', 'key control',
            '중요통제', 'key', '핵심', '중요'
        ],
        'control_frequency': [
            '통제주기', '통제 주기', '주기', 'frequency', 'cycle',
            '빈도', '실행주기', '수행주기'
        ],
        'control_type': [
            '통제유형', '통제 유형', '유형', 'type', 'control_type',
            '예방적발', '예방/적발', '통제타입'
        ],
        'control_nature': [
            '통제구분', '통제 구분', '구분', 'nature', 'control_nature',
            '자동수동', '자동/수동', '수행방식'
        ],
        'population': [
            '모집단', '대상', 'population', '범위', 'scope',
            '테스트대상', '검토대상'
        ],
        'population_completeness_check': [
            '모집단완전성확인', '모집단 완전성 확인', '완전성확인', '완전성 확인',
            'completeness', '완전성', 'population_completeness'
        ],
        'population_count': [
            '모집단갯수', '모집단 갯수', '갯수', '건수', 'count',
            '수량', '개수', 'population_count'
        ],
        'test_procedure': [
            '테스트절차', '테스트 절차', '절차', 'procedure', 'test_procedure',
            '검토절차', '확인절차', '감사절차'
        ]
    }
    
    auto_mapping = {}
    
    # 각 헤더에 대해 패턴 매칭 수행
    for i, header in enumerate(headers):
        if not header:
            continue
            
        header_lower = str(header).lower().strip()
        
        # 각 시스템 필드별로 패턴 매칭
        for field_name, patterns in field_patterns.items():
            for pattern in patterns:
                if pattern.lower() in header_lower or header_lower in pattern.lower():
                    auto_mapping[field_name] = i
                    break
            if field_name in auto_mapping:
                break
    
    return auto_mapping

# 관리자 메인 페이지
@admin_bp.route('/')
def admin():
    """관리자 메인 페이지"""
    admin_check = require_admin()
    if admin_check:
        return admin_check
    
    return render_template('admin.jsp',
                         is_logged_in=is_logged_in(),
                         user_info=get_user_info(),
                         remote_addr=request.remote_addr)

# 사용자 관리
@admin_bp.route('/users')
def admin_users():
    """사용자 관리 페이지"""
    admin_check = require_admin()
    if admin_check:
        return admin_check
    
    with get_db() as conn:
        users = conn.execute('''
            SELECT user_id, company_name, user_name, user_email, phone_number, 
                   admin_flag, effective_start_date, effective_end_date, 
                   creation_date, last_login_date
            FROM sb_user 
            ORDER BY creation_date DESC
        ''').fetchall()
        
        users_list = [dict(user) for user in users]
    
    return render_template('admin_users.jsp',
                         users=users_list,
                         is_logged_in=is_logged_in(),
                         user_info=get_user_info(),
                         remote_addr=request.remote_addr)

@admin_bp.route('/users/add', methods=['POST'])
def admin_add_user():
    """사용자 추가"""
    admin_check = require_admin()
    if admin_check:
        return admin_check
    
    try:
        company_name = request.form.get('company_name')
        user_name = request.form.get('user_name')
        user_email = request.form.get('user_email')
        phone_number = request.form.get('phone_number')
        admin_flag = request.form.get('admin_flag', 'N')
        effective_start_date = request.form.get('effective_start_date')
        effective_end_date = request.form.get('effective_end_date')
        
        # effective_end_date가 빈 문자열이면 NULL로 처리
        if not effective_end_date:
            effective_end_date = None
        
        with get_db() as conn:
            conn.execute('''
                INSERT INTO sb_user (company_name, user_name, user_email, phone_number, 
                                   admin_flag, effective_start_date, effective_end_date)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''', (company_name, user_name, user_email, phone_number, 
                 admin_flag, effective_start_date, effective_end_date))
            conn.commit()
        
        flash('사용자가 성공적으로 추가되었습니다.')
    except Exception as e:
        flash(f'사용자 추가 중 오류가 발생했습니다: {str(e)}')
    
    return redirect(url_for('admin.admin_users'))

@admin_bp.route('/users/edit/<int:user_id>', methods=['POST'])
def admin_edit_user(user_id):
    """사용자 정보 수정"""
    admin_check = require_admin()
    if admin_check:
        return admin_check
    
    try:
        company_name = request.form.get('company_name')
        user_name = request.form.get('user_name')
        user_email = request.form.get('user_email')
        phone_number = request.form.get('phone_number')
        admin_flag = request.form.get('admin_flag', 'N')
        effective_start_date = request.form.get('effective_start_date')
        effective_end_date = request.form.get('effective_end_date')
        
        # effective_end_date가 빈 문자열이면 NULL로 처리
        if not effective_end_date:
            effective_end_date = None
        
        with get_db() as conn:
            conn.execute('''
                UPDATE sb_user 
                SET company_name = %s, user_name = %s, user_email = %s, phone_number = %s, 
                    admin_flag = %s, effective_start_date = %s, effective_end_date = %s
                WHERE user_id = %s
            ''', (company_name, user_name, user_email, phone_number, 
                 admin_flag, effective_start_date, effective_end_date, user_id))
            conn.commit()
        
        flash('사용자 정보가 성공적으로 수정되었습니다.')
    except Exception as e:
        flash(f'사용자 정보 수정 중 오류가 발생했습니다: {str(e)}')
    
    return redirect(url_for('admin.admin_users'))

@admin_bp.route('/users/delete/<int:user_id>', methods=['POST'])
def admin_delete_user(user_id):
    """사용자 삭제"""
    admin_check = require_admin()
    if admin_check:
        return admin_check
    
    try:
        with get_db() as conn:
            conn.execute('DELETE FROM sb_user WHERE user_id = %s', (user_id,))
            conn.commit()
        
        flash('사용자가 성공적으로 삭제되었습니다.')
    except Exception as e:
        flash(f'사용자 삭제 중 오류가 발생했습니다: {str(e)}')
    
    return redirect(url_for('admin.admin_users'))

@admin_bp.route('/users/extend/<int:user_id>', methods=['POST'])
def admin_extend_user(user_id):
    """사용자 1년 연장"""
    admin_check = require_admin()
    if admin_check:
        return admin_check
    
    try:
        today = date.today().strftime('%Y-%m-%d')
        next_year = (date.today() + timedelta(days=365)).strftime('%Y-%m-%d')
        
        with get_db() as conn:
            conn.execute('''
                UPDATE sb_user 
                SET effective_start_date = %s, effective_end_date = %s
                WHERE user_id = %s
            ''', (today, next_year, user_id))
            conn.commit()
        
        flash('사용자의 사용 기간이 1년 연장되었습니다.')
    except Exception as e:
        flash(f'사용자 기간 연장 중 오류가 발생했습니다: {str(e)}')
    
    return redirect(url_for('admin.admin_users'))

# 로그 관리
@admin_bp.route('/logs')
def admin_logs():
    """사용자 활동 로그 조회 페이지"""
    admin_check = require_admin()
    if admin_check:
        return admin_check
    
    user_info = get_user_info()
    # 로그 조회 페이지는 로그 기록하지 않음 (무한 루프 방지)
    
    page = int(request.args.get('page', 1))
    per_page = 10  # 10개씩 표시
    offset = (page - 1) * per_page
    
    # 필터링 옵션
    user_filter = request.args.get('user_id')
    action_filter = request.args.get('action_type')
    
    # 로그 조회
    if user_filter:
        logs = get_user_activity_logs(limit=per_page, offset=offset, user_id=user_filter)
        total_count = get_activity_log_count(user_id=user_filter)
    else:
        logs = get_user_activity_logs(limit=per_page, offset=offset)
        total_count = get_activity_log_count()
    
    # 페이지네이션 계산
    total_pages = (total_count + per_page - 1) // per_page
    
    # 사용자 목록 (필터용)
    with get_db() as conn:
        users = conn.execute('SELECT user_id, user_name, user_email FROM sb_user ORDER BY user_name').fetchall()
        users_list = [dict(user) for user in users]
    
    return render_template('admin_logs.jsp',
                         logs=logs,
                         users=users_list,
                         current_page=page,
                         total_pages=total_pages,
                         total_count=total_count,
                         user_filter=user_filter,
                         action_filter=action_filter,
                         is_logged_in=is_logged_in(),
                         user_info=get_user_info(),
                         remote_addr=request.remote_addr)

# RCM 관리
@admin_bp.route('/rcm')
def admin_rcm():
    """RCM 관리 메인 페이지"""
    user_info = get_user_info()
    if not user_info or user_info.get('admin_flag') != 'Y':
        flash('관리자 권한이 필요합니다.')
        return redirect(url_for('index'))
    
    # 모든 RCM 목록 조회
    rcms = get_all_rcms()
    
    return render_template('admin_rcm.jsp',
                         rcms=rcms,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr)

@admin_bp.route('/rcm/upload')
def admin_rcm_upload():
    """RCM 업로드 페이지"""
    user_info = get_user_info()
    if not user_info or user_info.get('admin_flag') != 'Y':
        flash('관리자 권한이 필요합니다.')
        return redirect(url_for('index'))
    
    # 모든 사용자 목록 조회 (회사별 RCM 업로드용)
    with get_db() as conn:
        users = conn.execute('''
            SELECT user_id, user_name, user_email, company_name 
            FROM sb_user 
            WHERE effective_end_date IS NULL OR effective_end_date > CURRENT_TIMESTAMP
            ORDER BY company_name, user_name
        ''').fetchall()
        users_list = [dict(user) for user in users]
    
    return render_template('admin_rcm_upload.jsp',
                         users=users_list,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr)

@admin_bp.route('/rcm/process_upload', methods=['POST'])
def admin_rcm_process_upload():
    """Excel 파일 업로드 처리"""
    user_info = get_user_info()
    if not user_info or user_info.get('admin_flag') != 'Y':
        return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'})
    
    try:
        rcm_name = request.form.get('rcm_name', '').strip()
        description = request.form.get('description', '').strip()
        target_user_id = request.form.get('target_user_id', '').strip()
        
        if not rcm_name:
            return jsonify({'success': False, 'message': 'RCM명은 필수입니다.'})
        
        if not target_user_id:
            return jsonify({'success': False, 'message': '대상 사용자를 선택해주세요.'})
        
        target_user_id = int(target_user_id)
        
        if 'excel_file' not in request.files:
            return jsonify({'success': False, 'message': 'Excel 파일을 선택해주세요.'})
        
        file = request.files['excel_file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Excel 파일을 선택해주세요.'})
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Excel 파일(.xlsx, .xls)만 업로드 가능합니다.'})
        
        # RCM 생성 (선택된 사용자를 소유자로 설정)
        rcm_id = create_rcm(rcm_name, description, target_user_id, file.filename)
        
        # Excel 파일 읽기
        from openpyxl import load_workbook
        
        # 임시 파일로 저장
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()
        
        try:
            # Excel 파일 읽기
            workbook = load_workbook(temp_file.name)
            
            # RCM 시트 선택 (RCM 시트가 있으면 사용, 없으면 활성 시트 사용)
            if 'RCM' in workbook.sheetnames:
                sheet = workbook['RCM']
            else:
                sheet = workbook.active
            
            # 헤더 추출 (첫 번째 행)
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value if cell.value else '')
            
            # 샘플 데이터 추출 (최대 5행)
            sample_data = []
            max_sample_rows = min(6, sheet.max_row)  # 헤더 + 5행
            for row_num in range(2, max_sample_rows + 1):
                row_data = []
                for col_num in range(1, len(headers) + 1):
                    cell_value = sheet.cell(row=row_num, column=col_num).value
                    row_data.append(str(cell_value) if cell_value is not None else '')
                sample_data.append(row_data)
            
            # 자동 매핑 수행
            auto_mapping = perform_auto_mapping(headers)
            
            # 세션에 업로드 정보 저장
            session[f'rcm_upload_{rcm_id}'] = {
                'headers': headers,
                'sample_data': sample_data,
                'file_path': temp_file.name,
                'total_rows': sheet.max_row - 1,  # 헤더 제외
                'auto_mapping': auto_mapping,
                'header_row': 1  # 기본 헤더 행
            }
            
            return jsonify({
                'success': True, 
                'rcm_id': rcm_id,
                'headers': headers,
                'sample_data': sample_data
            })
            
        finally:
            # Excel 파일은 매핑 완료 후 삭제하므로 여기서는 삭제하지 않음
            pass
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'업로드 처리 중 오류가 발생했습니다: {str(e)}'})

@admin_bp.route('/rcm/mapping/<int:rcm_id>')
def admin_rcm_mapping(rcm_id):
    """컬럼 매핑 페이지"""
    user_info = get_user_info()
    if not user_info or user_info.get('admin_flag') != 'Y':
        flash('관리자 권한이 필요합니다.')
        return redirect(url_for('index'))
    
    # 세션에서 업로드 정보 가져오기
    upload_key = f'rcm_upload_{rcm_id}'
    if upload_key not in session:
        flash('업로드 정보를 찾을 수 없습니다. 다시 업로드해주세요.')
        return redirect(url_for('admin.admin_rcm_upload'))
    
    upload_info = session[upload_key]
    
    # RCM 정보 조회
    with get_db() as conn:
        rcm_info = conn.execute(
            'SELECT * FROM sb_rcm WHERE rcm_id = %s', (rcm_id,)
        ).fetchone()
        
        if not rcm_info:
            flash('RCM 정보를 찾을 수 없습니다.')
            return redirect(url_for('admin.admin_rcm'))
    
    # 시스템 필드 정의
    system_fields = [
        {'key': 'control_code', 'name': '통제코드', 'required': True, 'description': ''},
        {'key': 'control_name', 'name': '통제명', 'required': True, 'description': ''},
        {'key': 'control_description', 'name': '통제활동설명', 'required': False, 'description': ''},
        {'key': 'key_control', 'name': '핵심통제여부', 'required': False, 'description': ''},
        {'key': 'control_frequency', 'name': '통제주기', 'required': False, 'description': ''},
        {'key': 'control_type', 'name': '통제유형', 'required': False, 'description': '예방 또는 적발'},
        {'key': 'control_nature', 'name': '통제구분', 'required': False, 'description': '수동 또는 자동'},
        {'key': 'population', 'name': '모집단', 'required': False, 'description': ''},
        {'key': 'population_completeness_check', 'name': '모집단완전성확인', 'required': False, 'description': ''},
        {'key': 'population_count', 'name': '모집단갯수', 'required': False, 'description': ''},
        {'key': 'test_procedure', 'name': '테스트절차', 'required': False, 'description': ''}
    ]
    
    return render_template('admin_rcm_mapping.jsp',
                         rcm_info=dict(rcm_info),
                         headers=upload_info['headers'],
                         sample_data=upload_info['sample_data'],
                         system_fields=system_fields,
                         total_rows=upload_info['total_rows'],
                         auto_mapping=upload_info.get('auto_mapping', {}),
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr)

@admin_bp.route('/rcm/save_mapping', methods=['POST'])
def admin_rcm_save_mapping():
    """매핑된 RCM 데이터 저장"""
    user_info = get_user_info()
    if not user_info or user_info.get('admin_flag') != 'Y':
        return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'})
    
    try:
        rcm_id = request.form.get('rcm_id')
        if not rcm_id:
            return jsonify({'success': False, 'message': 'RCM ID가 필요합니다.'})
        
        rcm_id = int(rcm_id)
        
        # 세션에서 업로드 정보 가져오기
        upload_key = f'rcm_upload_{rcm_id}'
        
        if upload_key not in session:
            return jsonify({'success': False, 'message': f'업로드 정보를 찾을 수 없습니다. 키: {upload_key}, 세션 키들: {list(session.keys())}'})
        
        upload_info = session[upload_key]
        file_path = upload_info['file_path']
        headers = upload_info['headers']
        
        # 매핑 정보 수집
        field_mapping = {}
        for key in request.form:
            if key.startswith('mapping_'):
                field_name = key.replace('mapping_', '')
                column_index = request.form.get(key)
                if column_index:
                    field_mapping[field_name] = int(column_index)
        
        
        # Excel 파일 다시 읽기
        from openpyxl import load_workbook
        
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'message': 'Excel 파일을 찾을 수 없습니다.'})
        
        try:
            workbook = load_workbook(file_path)
            # RCM 시트 선택 (RCM 시트가 있으면 사용, 없으면 활성 시트 사용)
            if 'RCM' in workbook.sheetnames:
                sheet = workbook['RCM']
            else:
                sheet = workbook.active
            
            # 헤더 행 번호 가져오기 (기본값: 1)
            header_row = int(request.form.get('header_row', 1))
            
            # 데이터 처리
            rcm_data = []
            for row_num in range(header_row + 1, sheet.max_row + 1):  # 헤더 제외
                row_data = {}
                
                # 각 필드별로 매핑된 컬럼에서 데이터 추출
                for field_name, column_index in field_mapping.items():
                    cell_value = sheet.cell(row=row_num, column=column_index + 1).value  # Excel은 1부터 시작
                    row_data[field_name] = str(cell_value) if cell_value is not None else ''
                
                rcm_data.append(row_data)
        finally:
            # 워크북 명시적으로 닫기
            if 'workbook' in locals():
                workbook.close()
        
        # DB에 데이터 저장
        save_rcm_details(rcm_id, rcm_data)
        
        # 회사별 폴더에 원본 파일 저장
        try:
            # RCM 정보에서 회사명 가져오기
            with get_db() as conn:
                rcm_info = conn.execute('''
                    SELECT r.original_filename, u.company_name 
                    FROM sb_rcm r
                    JOIN sb_user u ON r.upload_user_id = u.user_id
                    WHERE r.rcm_id = %s
                ''', (rcm_id,)).fetchone()
                
                if rcm_info and rcm_info['company_name'] and rcm_info['original_filename']:
                    company_name = rcm_info['company_name']
                    original_filename = rcm_info['original_filename']
                    
                    # 회사별 업로드 디렉토리 생성
                    company_upload_dir = os.path.join('uploads', company_name)
                    os.makedirs(company_upload_dir, exist_ok=True)
                    
                    # 회사별 폴더에 파일 저장
                    permanent_file_path = os.path.join(company_upload_dir, original_filename)
                    
                    # 임시 파일을 영구 경로로 복사
                    import shutil
                    shutil.copy2(file_path, permanent_file_path)
                    
                    
                    # DB에 새로운 파일 경로 업데이트 (상대 경로로 저장)
                    relative_path = os.path.join(company_name, original_filename)
                    conn.execute('''
                        UPDATE sb_rcm 
                        SET original_filename = %s
                        WHERE rcm_id = %s
                    ''', (relative_path, rcm_id))
                    conn.commit()
                    
        except Exception as e:
            # 실패해도 계속 진행
            pass
        
        # 세션에서 업로드 정보 삭제 및 임시 파일 삭제
        del session[upload_key]
        
        # 임시 파일 안전하게 삭제
        try:
            if os.path.exists(file_path):
                os.unlink(file_path)
        except PermissionError as e:
            # 파일 삭제 실패는 무시 (시스템이 나중에 정리함)
            pass
        
        return jsonify({
            'success': True,
            'message': f'{len(rcm_data)}개의 통제 데이터가 성공적으로 저장되었습니다.',
            'rcm_id': rcm_id
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'데이터 저장 중 오류가 발생했습니다: {str(e)}'})

@admin_bp.route('/rcm/<int:rcm_id>/view')
def admin_rcm_view(rcm_id):
    """RCM 상세 보기"""
    user_info = get_user_info()
    if not user_info or user_info.get('admin_flag') != 'Y':
        flash('관리자 권한이 필요합니다.')
        return redirect(url_for('index'))
    
    # RCM 기본 정보 조회
    with get_db() as conn:
        rcm_info = conn.execute('''
            SELECT r.*, u.user_name as upload_user_name, u.company_name
            FROM sb_rcm r
            LEFT JOIN sb_user u ON r.upload_user_id = u.user_id
            WHERE r.rcm_id = %s AND r.is_active = 'Y'
        ''', (rcm_id,)).fetchone()
        
        if not rcm_info:
            flash('RCM을 찾을 수 없습니다.')
            return redirect(url_for('admin.admin_rcm'))
    
    # RCM 상세 데이터 조회
    rcm_details = get_rcm_details(rcm_id)
    
    return render_template('admin_rcm_view.jsp',
                         rcm_info=dict(rcm_info),
                         rcm_details=rcm_details,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr)

@admin_bp.route('/rcm/<int:rcm_id>/users')
def admin_rcm_users(rcm_id):
    """RCM 사용자 관리"""
    user_info = get_user_info()
    if not user_info:
        flash('로그인이 필요합니다.')
        return redirect(url_for('login'))

    is_admin = user_info.get('admin_flag') == 'Y'

    # RCM 기본 정보 조회
    with get_db() as conn:
        rcm_info = conn.execute('''
            SELECT r.*, u.user_name as upload_user_name, u.company_name
            FROM sb_rcm r
            LEFT JOIN sb_user u ON r.upload_user_id = u.user_id
            WHERE r.rcm_id = %s AND r.is_active = 'Y'
        ''', (rcm_id,)).fetchone()

        if not rcm_info:
            flash('RCM을 찾을 수 없습니다.')
            return redirect(url_for('index'))

        # 시스템 관리자가 아닌 경우, RCM admin 권한 확인
        if not is_admin:
            user_permission = conn.execute('''
                SELECT permission_type FROM sb_user_rcm
                WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
            ''', (user_info['user_id'], rcm_id)).fetchone()

            if not user_permission or user_permission['permission_type'] != 'admin':
                flash('해당 RCM의 관리자 권한이 필요합니다.')
                return redirect(url_for('index'))

        # RCM 접근 권한이 있는 사용자 목록 (시스템 관리자 제외)
        rcm_users = conn.execute('''
            SELECT ur.*, u.user_name, u.user_email, u.company_name, u.admin_flag,
                   gb.user_name as granted_by_name
            FROM sb_user_rcm ur
            JOIN sb_user u ON ur.user_id = u.user_id
            LEFT JOIN sb_user gb ON ur.granted_by = gb.user_id
            WHERE ur.rcm_id = %s AND ur.is_active = 'Y'
            AND (u.admin_flag IS NULL OR u.admin_flag != 'Y')
            ORDER BY ur.granted_date DESC
        ''', (rcm_id,)).fetchall()
        
        # 활성 사용자 목록 (권한 부여용)
        # 시스템 관리자: 모든 사용자, 일반 사용자: 본인 회사 사용자만
        if is_admin:
            # 관리자는 모든 사용자 조회 (시스템 관리자 제외)
            all_users = conn.execute('''
                SELECT user_id, user_name, user_email, company_name
                FROM sb_user
                WHERE (effective_end_date IS NULL OR effective_end_date > CURRENT_TIMESTAMP)
                AND (admin_flag IS NULL OR admin_flag != 'Y')
                AND user_id NOT IN (
                    SELECT user_id FROM sb_user_rcm
                    WHERE rcm_id = %s AND is_active = 'Y'
                )
                ORDER BY company_name, user_name
            ''', (rcm_id,)).fetchall()
        else:
            # 일반 사용자는 본인 회사 사용자만 조회 (시스템 관리자 제외)
            user_company = user_info.get('company_name', '')
            all_users = conn.execute('''
                SELECT user_id, user_name, user_email, company_name
                FROM sb_user
                WHERE (effective_end_date IS NULL OR effective_end_date > CURRENT_TIMESTAMP)
                AND company_name = %s
                AND (admin_flag IS NULL OR admin_flag != 'Y')
                AND user_id NOT IN (
                    SELECT user_id FROM sb_user_rcm
                    WHERE rcm_id = %s AND is_active = 'Y'
                )
                ORDER BY user_name
            ''', (user_company, rcm_id)).fetchall()
        
        rcm_users_list = [dict(user) for user in rcm_users]
        all_users_list = [dict(user) for user in all_users]
    
    return render_template('admin_rcm_users.jsp',
                         rcm_info=dict(rcm_info),
                         rcm_users=rcm_users_list,
                         all_users=all_users_list,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr)

@admin_bp.route('/rcm/grant_access', methods=['POST'])
def admin_rcm_grant_access():
    """사용자에게 RCM 접근 권한 부여 (시스템 관리자 또는 RCM admin)"""
    user_info = get_user_info()
    if not user_info:
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'})

    try:
        # JSON 또는 form 데이터 처리
        data = request.get_json() if request.is_json else request.form

        rcm_id = data.get('rcm_id')
        # user_id 또는 target_user_id 둘 다 지원
        target_user_id = data.get('user_id') or data.get('target_user_id')
        permission_type = data.get('permission_type', 'read')

        # 필수 파라미터 검증
        if not rcm_id or not target_user_id:
            return jsonify({'success': False, 'message': 'RCM ID와 사용자 ID는 필수입니다.'})

        # 정수 변환
        rcm_id = int(rcm_id)
        target_user_id = int(target_user_id)

        # 권한 확인: 시스템 관리자 또는 RCM admin
        is_admin = user_info.get('admin_flag') == 'Y'
        with get_db() as conn:
            if not is_admin:
                user_permission = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()

                if not user_permission or user_permission['permission_type'] != 'admin':
                    return jsonify({'success': False, 'message': '해당 RCM의 관리자 권한이 필요합니다.'})

                # 일반 사용자는 본인 회사 사용자에게만 권한 부여 가능
                user_company = user_info.get('company_name', '')
                target_user = conn.execute('''
                    SELECT company_name FROM sb_user WHERE user_id = %s
                ''', (target_user_id,)).fetchone()

                if not target_user or target_user['company_name'] != user_company:
                    return jsonify({'success': False, 'message': '본인 회사 사용자에게만 권한을 부여할 수 있습니다.'})

        # 권한 부여
        grant_rcm_access(target_user_id, rcm_id, permission_type, user_info['user_id'])

        return jsonify({
            'success': True,
            'message': '사용자에게 RCM 접근 권한이 부여되었습니다.'
        })

    except ValueError as e:
        return jsonify({'success': False, 'message': f'잘못된 파라미터입니다: {str(e)}'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'권한 부여 중 오류가 발생했습니다: {str(e)}'})

@admin_bp.route('/rcm/change_permission', methods=['POST'])
def admin_rcm_change_permission():
    """사용자의 RCM 접근 권한 변경 (시스템 관리자 또는 RCM admin)"""
    user_info = get_user_info()
    if not user_info:
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'})

    try:
        # JSON 또는 form 데이터 처리
        data = request.get_json() if request.is_json else request.form

        rcm_id = data.get('rcm_id')
        # user_id 또는 target_user_id 둘 다 지원
        target_user_id = data.get('user_id') or data.get('target_user_id')
        permission_type = data.get('permission_type')

        # 필수 파라미터 검증
        if not rcm_id or not target_user_id or not permission_type:
            return jsonify({'success': False, 'message': 'RCM ID, 사용자 ID, 권한 유형은 필수입니다.'})

        # 정수 변환
        rcm_id = int(rcm_id)
        target_user_id = int(target_user_id)

        # 권한 확인: 시스템 관리자 또는 RCM admin
        is_admin = user_info.get('admin_flag') == 'Y'
        with get_db() as conn:
            if not is_admin:
                user_permission = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()

                if not user_permission or user_permission['permission_type'] != 'admin':
                    return jsonify({'success': False, 'message': '해당 RCM의 관리자 권한이 필요합니다.'})

            conn.execute('''
                UPDATE sb_user_rcm
                SET permission_type = %s
                WHERE rcm_id = %s AND user_id = %s AND is_active = 'Y'
            ''', (permission_type, rcm_id, target_user_id))
            conn.commit()

        return jsonify({
            'success': True,
            'message': '사용자 권한이 성공적으로 변경되었습니다.'
        })

    except ValueError as e:
        return jsonify({'success': False, 'message': f'잘못된 파라미터입니다: {str(e)}'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'권한 변경 중 오류가 발생했습니다: {str(e)}'})

@admin_bp.route('/rcm/revoke_access', methods=['POST'])
def admin_rcm_revoke_access():
    """사용자의 RCM 접근 권한 제거 (시스템 관리자 또는 RCM admin)"""
    user_info = get_user_info()
    if not user_info:
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'})

    try:
        # JSON 또는 form 데이터 처리
        data = request.get_json() if request.is_json else request.form

        rcm_id = data.get('rcm_id')
        # user_id 또는 target_user_id 둘 다 지원
        target_user_id = data.get('user_id') or data.get('target_user_id')

        # 필수 파라미터 검증
        if not rcm_id or not target_user_id:
            return jsonify({'success': False, 'message': 'RCM ID와 사용자 ID는 필수입니다.'})

        # 정수 변환
        rcm_id = int(rcm_id)
        target_user_id = int(target_user_id)

        # 권한 확인: 시스템 관리자 또는 RCM admin
        is_admin = user_info.get('admin_flag') == 'Y'
        with get_db() as conn:
            if not is_admin:
                user_permission = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()

                if not user_permission or user_permission['permission_type'] != 'admin':
                    return jsonify({'success': False, 'message': '해당 RCM의 관리자 권한이 필요합니다.'})

            conn.execute('''
                UPDATE sb_user_rcm
                SET is_active = 'N', last_updated = CURRENT_TIMESTAMP
                WHERE rcm_id = %s AND user_id = %s AND is_active = 'Y'
            ''', (rcm_id, target_user_id))
            conn.commit()
        
        return jsonify({
            'success': True,
            'message': '사용자의 RCM 접근 권한이 제거되었습니다.'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'권한 제거 중 오류가 발생했습니다: {str(e)}'})

@admin_bp.route('/rcm/delete', methods=['POST'])
def admin_rcm_delete():
    """RCM 삭제 (소프트 삭제)"""
    user_info = get_user_info()
    if not user_info or user_info.get('admin_flag') != 'Y':
        return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'})
    
    try:
        rcm_id = int(request.form.get('rcm_id'))
        
        with get_db() as conn:
            # RCM 비활성화 (소프트 삭제)
            conn.execute('''
                UPDATE sb_rcm 
                SET is_active = 'N'
                WHERE rcm_id = %s
            ''', (rcm_id,))
            
            # 관련 사용자 권한도 비활성화
            conn.execute('''
                UPDATE sb_user_rcm 
                SET is_active = 'N'
                WHERE rcm_id = %s
            ''', (rcm_id,))
            
            conn.commit()
        
        return jsonify({
            'success': True,
            'message': 'RCM이 성공적으로 삭제되었습니다.'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'RCM 삭제 중 오류가 발생했습니다: {str(e)}'})

# 사용자 전환 기능
@admin_bp.route('/switch_user', methods=['POST'])
def admin_switch_user():
    """관리자가 다른 사용자로 스위치"""
    user_info = get_user_info()
    if not user_info or user_info.get('admin_flag') != 'Y':
        return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'})
    
    try:
        target_user_id = int(request.form.get('target_user_id'))
        
        # 원래 관리자 정보 백업
        session['original_admin_id'] = session['user_id']
        
        # 대상 사용자로 세션 변경
        session['user_id'] = target_user_id
        
        # 세션에 캐시된 user_info 삭제 (새로운 사용자 정보로 다시 로드되도록)
        if 'user_info' in session:
            del session['user_info']
        
        return jsonify({
            'success': True,
            'message': '사용자 전환이 완료되었습니다.'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'사용자 전환 중 오류가 발생했습니다: {str(e)}'})

@admin_bp.route('/switch_back')
def admin_switch_back():
    """관리자 계정으로 돌아가기"""
    if 'original_admin_id' in session:
        session['user_id'] = session['original_admin_id']
        del session['original_admin_id']

        # 세션에 캐시된 user_info 삭제 (관리자 정보로 다시 로드되도록)
        if 'user_info' in session:
            del session['user_info']

        flash('관리자 계정으로 돌아왔습니다.')
    else:
        flash('원래 계정 정보를 찾을 수 없습니다.')

    # 이전 페이지로 돌아가기 (referrer가 없으면 홈으로)
    return redirect(request.referrer or url_for('index'))

# ============================================================================
# API 엔드포인트 - 사용자 전환 (네비게이션용)
# ============================================================================

@admin_bp.route('/api/admin/users', methods=['GET'])
@login_required
def api_get_users():
    """모든 사용자 목록 조회 API (관리자 전용)"""
    print("\n========== [API] /api/admin/users 호출 ==========")
    user_info = get_user_info()
    print(f"[API] user_info: {user_info}")

    # 관리자 권한 확인
    if user_info.get('admin_flag') != 'Y':
        print(f"[API] ❌ 관리자 권한 없음 - admin_flag: {user_info.get('admin_flag')}")
        return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'})

    print("[API] ✅ 관리자 권한 확인 완료")
    try:
        print("[API] DB 조회 시작...")
        with get_db() as conn:
            rows = conn.execute('''
                SELECT user_id, user_email, company_name, admin_flag
                FROM sb_user
                WHERE effective_end_date IS NULL OR effective_end_date > datetime('now')
                ORDER BY company_name
            ''').fetchall()

            print(f"[API] 조회된 사용자 수: {len(rows)}")

            users = []
            for row in rows:
                user_data = {
                    'user_id': row[0],
                    'user_email': row[1],
                    'company_name': row[2],
                    'admin_flag': row[3]
                }
                users.append(user_data)
                print(f"[API]  - {user_data}")

            print(f"[API] ✅ 성공: {len(users)}명의 사용자 반환")
            print("=" * 50)
            return jsonify({'success': True, 'users': users})
    except Exception as e:
        print(f"[API] ❌ 오류 발생: {str(e)}")
        import traceback
        traceback.print_exc()
        print("=" * 50)
        return jsonify({'success': False, 'message': f'사용자 목록 조회 중 오류: {str(e)}'})

@admin_bp.route('/api/admin/switch-user', methods=['POST'])
@login_required
def api_switch_user():
    """사용자 전환 API (관리자 전용)"""
    user_info = get_user_info()

    # 관리자 권한 확인
    if user_info.get('admin_flag') != 'Y':
        return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'})

    data = request.get_json()
    target_user_id = data.get('user_id')

    if not target_user_id:
        return jsonify({'success': False, 'message': '사용자 ID가 필요합니다.'})

    try:
        with get_db() as conn:
            # 대상 사용자 확인
            target_user = conn.execute('''
                SELECT user_id, company_name
                FROM sb_user
                WHERE user_id = %s AND (effective_end_date IS NULL OR effective_end_date > datetime('now'))
            ''', (target_user_id,)).fetchone()

            if not target_user:
                return jsonify({'success': False, 'message': '존재하지 않는 사용자입니다.'})

            # 관리자 ID 저장 (아직 저장되지 않은 경우에만)
            if 'original_admin_id' not in session:
                session['original_admin_id'] = session['user_id']

            # 사용자 전환
            session['user_id'] = target_user_id

            # 세션에 캐시된 user_info 삭제 (새 사용자 정보로 다시 로드되도록)
            if 'user_info' in session:
                del session['user_info']

            return jsonify({
                'success': True,
                'message': f"{target_user[1]} 계정으로 전환되었습니다."
            })
    except Exception as e:
        return jsonify({'success': False, 'message': f'사용자 전환 중 오류: {str(e)}'})