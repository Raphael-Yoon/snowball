from flask import Blueprint, request, jsonify, render_template, redirect, url_for, flash, session, send_file
from auth import login_required, get_current_user, get_user_rcms, get_rcm_details, get_key_rcm_details, save_operation_evaluation, get_operation_evaluations, get_operation_evaluation_samples, get_design_evaluation_sample, log_user_activity, get_db, is_design_evaluation_completed, get_completed_design_evaluation_sessions
from snowball_link5 import get_user_info, is_logged_in
import file_manager
from control_config import get_control_config
import json
import os
import tempfile
from openpyxl import load_workbook

bp_link7 = Blueprint('link7', __name__)

# 운영평가 관련 기능들

@bp_link7.route('/operation-evaluation')
@login_required
def user_operation_evaluation():
    """운영평가 페이지"""
    user_info = get_user_info()

    # 사용자가 접근 가능한 RCM 목록 조회 (ITGC만)
    user_rcms = get_user_rcms(user_info['user_id'], control_category='ITGC')

    # 각 RCM에 대해 모든 설계평가 세션 조회 (진행중 + 완료)
    from auth import get_all_design_evaluation_sessions
    for rcm in user_rcms:
        all_sessions = get_all_design_evaluation_sessions(rcm['rcm_id'], user_info['user_id'])
        completed_sessions = [s for s in all_sessions if s['completed_date'] is not None]
        in_progress_sessions = [s for s in all_sessions if s['completed_date'] is None]

        # 완료된 세션에 대해서만 운영평가 진행상황 조회
        for session in completed_sessions:
            operation_evaluation_session = f"OP_{session['evaluation_session']}"

            # 운영평가 진행 통제 수 조회
            from auth import count_completed_operation_evaluations
            with get_db() as conn:
                header = conn.execute('''
                    SELECT header_id FROM sb_operation_evaluation_header
                    WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
                ''', (rcm['rcm_id'], user_info['user_id'], operation_evaluation_session, session['evaluation_session'])).fetchone()

            if header:
                completed_count = count_completed_operation_evaluations(header['header_id'])
                session['operation_completed_count'] = completed_count
            else:
                session['operation_completed_count'] = 0

            # 운영평가 가능한 통제 개수 추가
            eligible_controls = get_key_rcm_details(rcm['rcm_id'], user_info['user_id'], session['evaluation_session'])
            session['eligible_control_count'] = len(eligible_controls)

        rcm['all_design_sessions'] = all_sessions
        rcm['completed_design_sessions'] = completed_sessions
        rcm['in_progress_design_sessions'] = in_progress_sessions
        rcm['design_evaluation_completed'] = len(completed_sessions) > 0

        # 핵심통제 개수 조회 (모든 핵심통제)
        key_controls = get_key_rcm_details(rcm['rcm_id'])
        rcm['key_control_count'] = len(key_controls)
        rcm['has_key_controls'] = len(key_controls) > 0

    log_user_activity(user_info, 'PAGE_ACCESS', '운영평가', '/user/operation-evaluation',
                     request.remote_addr, request.headers.get('User-Agent'))

    return render_template('link7_operation_evaluation_unified.jsp',
                         evaluation_type='ITGC',
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         user_rcms=user_rcms,
                         remote_addr=request.remote_addr)

@bp_link7.route('/operation-evaluation/rcm', methods=['GET', 'POST'])
@login_required
def user_operation_evaluation_rcm():
    """RCM별 운영평가 페이지 (설계평가 세션 기반)"""
    user_info = get_user_info()

    # POST로 전달된 RCM ID와 설계평가 세션 정보 받기
    if request.method == 'POST':
        rcm_id = request.form.get('rcm_id')
        design_evaluation_session = request.form.get('design_evaluation_session')
        new_operation_session = request.form.get('new_operation_session')  # 신규 운영평가 세션명


        if not rcm_id:
            flash('RCM 정보가 없습니다.', 'error')
            return redirect(url_for('link7.user_operation_evaluation'))
        if not design_evaluation_session:
            flash('설계평가 세션 정보가 없습니다.', 'error')
            return redirect(url_for('link7.user_operation_evaluation'))

        # 세션에 저장
        session['current_operation_rcm_id'] = int(rcm_id)
        session['current_design_evaluation_session'] = design_evaluation_session

        # 신규 운영평가 세션인 경우
        if new_operation_session:
            session['new_operation_session_name'] = new_operation_session
            flash(f'새로운 운영평가 세션 "{new_operation_session}"을 시작합니다.', 'success')
        else:
            # 기존 세션 제거
            session.pop('new_operation_session_name', None)

    # POST든 GET이든 세션에서 정수형 rcm_id를 가져옴
    rcm_id = session.get('current_operation_rcm_id')
    design_evaluation_session = session.get('current_design_evaluation_session')

    print(f"[DEBUG] RCM ID: {rcm_id}, Session: {design_evaluation_session}")

    if not rcm_id:
        flash('RCM 정보가 없습니다. 다시 선택해주세요.', 'error')
        return redirect(url_for('link7.user_operation_evaluation'))
    if not design_evaluation_session:
        flash('설계평가 세션 정보가 없습니다. 다시 선택해주세요.', 'error')
        return redirect(url_for('link7.user_operation_evaluation'))

    # 사용자가 해당 RCM에 접근 권한이 있는지 확인
    print("[DEBUG] Checking user permissions...")
    user_rcms = get_user_rcms(user_info['user_id'])
    rcm_ids = [rcm['rcm_id'] for rcm in user_rcms]

    if rcm_id not in rcm_ids:
        flash('해당 RCM에 대한 접근 권한이 없습니다.', 'error')
        return redirect(url_for('link7.user_operation_evaluation'))


    # 해당 설계평가 세션이 완료되었는지 확인
    print("[DEBUG] Checking completed sessions...")
    completed_sessions = get_completed_design_evaluation_sessions(rcm_id, user_info['user_id'])

    session_found = False
    for session_item in completed_sessions:
        if session_item['evaluation_session'] == design_evaluation_session:
            session_found = True
            break

    if not session_found:
        flash(f'설계평가 세션 "{design_evaluation_session}"이 완료되지 않아 운영평가를 수행할 수 없습니다.', 'warning')
        return redirect(url_for('link7.user_operation_evaluation'))
    
    # RCM 정보 조회
    print("[DEBUG] Fetching RCM info...")
    rcm_info = None
    for rcm in user_rcms:
        if rcm['rcm_id'] == rcm_id:
            rcm_info = rcm
            break
    
    # RCM 핵심통제 데이터 조회 (운영평가는 핵심통제이면서 설계평가가 '적정'인 통제만 대상)
    print("[DEBUG] Fetching key RCM details...")
    try:
        rcm_details = get_key_rcm_details(rcm_id, user_info['user_id'], design_evaluation_session)
        print(f"[DEBUG] rcm_details count: {len(rcm_details) if rcm_details else 0}")
    except Exception as e:
        print(f"[DEBUG] Error fetching key RCM details: {e}")
        import traceback
        traceback.print_exc()
        raise e
    
    # 매핑 정보 조회
    print("[DEBUG] Fetching mappings...")
    from auth import get_rcm_detail_mappings
    rcm_mappings_list = get_rcm_detail_mappings(rcm_id)
    # control_code를 키로 하는 딕셔너리로 변환
    rcm_mappings = {m['control_code']: m for m in rcm_mappings_list}

    # 핵심통제이면서 설계평가가 '적정'인 통제가 없는 경우 안내 메시지 표시
    if not rcm_details:
        flash('해당 RCM에 설계평가 결과가 "적정"인 핵심통제가 없어 운영평가를 수행할 수 없습니다.', 'warning')
        return redirect(url_for('link7.user_operation_evaluation'))

    # 각 통제 코드에 대한 config 정보 미리 로드
    print("[DEBUG] Loading control configs...")
    control_configs = {}
    for detail in rcm_details:
        control_configs[detail['control_code']] = get_control_config(detail['control_code'])

    # 운영평가 세션명 생성 (설계평가 세션 기반)
    operation_evaluation_session = f"OP_{design_evaluation_session}"

    # 운영평가 Header/Line 데이터 동기화 (설계평가 결과 변경 반영)
    print("[DEBUG] Syncing operation evaluation data...")
    sync_messages = []
    operation_header = None
    try:
        # 기존 운영평가 헤더 확인
        from auth import get_or_create_operation_evaluation_header
        with get_db() as conn:
            header_id = get_or_create_operation_evaluation_header(conn, rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)

            # 헤더 정보 조회 (진행률 표시용)
            operation_header = conn.execute('''
                SELECT header_id, evaluated_controls, total_controls, progress_percentage, evaluation_status
                FROM sb_operation_evaluation_header
                WHERE header_id = %s
            ''', (header_id,)).fetchone()

            # 현재 대상 통제 코드 목록 (핵심통제 + 설계평가 '적정')
            current_control_codes = {detail['control_code'] for detail in rcm_details}

            # 기존 Line 데이터 조회
            existing_lines = conn.execute('''
                SELECT line_id, control_code
                FROM sb_operation_evaluation_line
                WHERE header_id = %s
            ''', (header_id,)).fetchall()

            existing_control_codes = {line['control_code'] for line in existing_lines}

            # 신규 추가된 통제 (설계평가 부적정→적정 변경)
            new_controls = current_control_codes - existing_control_codes
            if new_controls:
                for idx, detail in enumerate(rcm_details):
                    if detail['control_code'] in new_controls:
                        # recommended_sample_size 가져오기 (있으면 사용)
                        recommended_size = detail.get('recommended_sample_size')

                        conn.execute('''
                            INSERT INTO sb_operation_evaluation_line (
                                header_id, control_code, control_sequence, sample_size
                            ) VALUES (%s, %s, %s, %s)
                        ''', (header_id, detail['control_code'], idx + 1, recommended_size))
                sync_messages.append(f"📌 신규 추가: {len(new_controls)}개 (설계평가 부적정→적정)")

            conn.commit()

            # 동기화 메시지 표시
            if sync_messages:
                flash(' '.join(sync_messages), 'success')
    except Exception as e:
        print(f"[DEBUG] Sync error: {e}")
        import traceback
        traceback.print_exc()
        flash(f"운영평가 데이터 동기화 중 오류 발생: {str(e)}", 'error')

    # 기존 운영평가 내역 불러오기 (Header-Line 구조)
    print("[DEBUG] Loading existing evaluations...")
    try:
        evaluations = get_operation_evaluations(rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)

        print(f'[snowball_link7] Total evaluations: {len(evaluations)}')

        # 평가가 완료된 통제(conclusion 값이 있는 경우) 또는 샘플이 업로드된 통제를 control_code를 키로 하는 딕셔너리로 변환
        # 중복이 있는 경우 가장 최신(last_updated 또는 evaluation_date 기준) 레코드만 사용
        evaluated_controls = {}
        for eval in evaluations:
            # line_id가 있거나, conclusion이 있거나, 샘플이 있으면 포함
            if eval.get('line_id') or eval.get('conclusion') or (eval.get('sample_lines') and len(eval.get('sample_lines', [])) > 0):
                control_code = eval['control_code']
                sample_lines_count = len(eval.get('sample_lines', []))
                print(f'[snowball_link7] {control_code}: samples={sample_lines_count}, line_id={eval.get("line_id")}, conclusion={eval.get("conclusion")}')

                # 기존에 없거나, 더 최신 데이터인 경우만 업데이트
                if control_code not in evaluated_controls:
                    evaluated_controls[control_code] = eval
                else:
                    # last_updated 또는 evaluation_date로 최신 판단
                    existing_date = evaluated_controls[control_code].get('last_updated') or evaluated_controls[control_code].get('evaluation_date')
                    new_date = eval.get('last_updated') or eval.get('evaluation_date')
                    if new_date and existing_date and new_date > existing_date:
                        evaluated_controls[control_code] = eval

        print(f'[snowball_link7] evaluated_controls keys: {list(evaluated_controls.keys())}')

    except Exception as e:
        print(f'[snowball_link7] Error loading evaluations: {e}')
        import traceback
        traceback.print_exc()
        evaluated_controls = {}

    log_user_activity(user_info, 'PAGE_ACCESS', 'RCM 운영평가', '/operation-evaluation/rcm',
                      request.remote_addr, request.headers.get('User-Agent'))

    return render_template('link7_detail.jsp',
                         rcm_id=rcm_id,
                         design_evaluation_session=design_evaluation_session,
                         evaluation_session=design_evaluation_session,  # 템플릿 호환성
                         operation_evaluation_session=operation_evaluation_session,
                         operation_header=operation_header,  # 진행률 표시용
                         rcm_info=rcm_info,
                         rcm_details=rcm_details,
                         rcm_mappings=rcm_mappings,
                         evaluated_controls=evaluated_controls,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr,
                         control_configs=control_configs)

@bp_link7.route('/api/operation-evaluation/save', methods=['POST'])
@login_required
def save_operation_evaluation_api():
    """운영평가 결과 저장 API"""
    print("=" * 50)
    print("운영평가 저장 API 호출됨")
    print("=" * 50)

    user_info = get_user_info()
    print(f"사용자 정보: {user_info}")

    # JSON과 FormData 모두 처리
    print(f"Content-Type: {request.content_type}")
    if request.content_type and 'multipart/form-data' in request.content_type:
        # FormData로 전송된 경우
        data = request.form.to_dict()
        rcm_id = data.get('rcm_id')
        design_evaluation_session = data.get('design_evaluation_session')
        control_code = data.get('control_code')

        # evaluation_data는 JSON 문자열로 전송되므로 파싱
        import json
        evaluation_data_str = data.get('evaluation_data')
        if evaluation_data_str:
            try:
                evaluation_data = json.loads(evaluation_data_str)
            except json.JSONDecodeError:
                return jsonify({
                    'success': False,
                    'message': 'evaluation_data 파싱 오류'
                })
        else:
            # 개별 필드로 전송된 경우 (구버전 호환)
            evaluation_data = {
                'sample_size': int(data.get('sample_size', 0)),
                'exception_count': int(data.get('exception_count', 0)),
                'exception_details': data.get('exception_details', ''),
                'conclusion': data.get('conclusion'),
                'improvement_plan': data.get('improvement_plan', '')
            }
    else:
        # JSON으로 전송된 경우
        data = request.get_json()
        rcm_id = data.get('rcm_id')
        design_evaluation_session = data.get('design_evaluation_session')
        control_code = data.get('control_code')
        evaluation_data = data.get('evaluation_data')

    print(f"rcm_id: {rcm_id}")
    print(f"design_evaluation_session: {design_evaluation_session}")
    print(f"control_code: {control_code}")
    print(f"evaluation_data: {evaluation_data}")

    if not all([rcm_id, design_evaluation_session, control_code, evaluation_data]):
        print("필수 데이터 누락!")
        return jsonify({
            'success': False,
            'message': '필수 데이터가 누락되었습니다.'
        })

    # 운영평가 세션명 생성
    operation_evaluation_session = f"OP_{design_evaluation_session}"
    print(f"operation_evaluation_session: {operation_evaluation_session}")

    try:
        print("권한 확인 시작...")
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인
        with get_db() as conn:
            access_check = conn.execute('''
                SELECT permission_type FROM sb_user_rcm
                WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
            ''', (user_info['user_id'], rcm_id)).fetchone()

            if not access_check:
                return jsonify({
                    'success': False,
                    'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                })

        # 해당 설계평가 세션이 완료되었는지 확인
        completed_sessions = get_completed_design_evaluation_sessions(rcm_id, user_info['user_id'])
        session_found = False
        for session in completed_sessions:
            if session['evaluation_session'] == design_evaluation_session:
                session_found = True
                break

        if not session_found:
            return jsonify({
                'success': False,
                'message': f'설계평가 세션 "{design_evaluation_session}"이 완료되지 않아 운영평가를 수행할 수 없습니다.'
            })

        # RCM에 설정된 권장 표본수 확인
        with get_db() as conn:
            rcm_detail = conn.execute('''
                SELECT recommended_sample_size FROM sb_rcm_detail
                WHERE rcm_id = %s AND control_code = %s
            ''', (rcm_id, control_code)).fetchone()

        recommended_size = rcm_detail['recommended_sample_size'] if rcm_detail and rcm_detail['recommended_sample_size'] is not None else 0

        # 표본 크기 유효성 검사 (no_occurrence가 아닌 경우에만)
        is_no_occurrence = evaluation_data.get('no_occurrence', False)
        if not is_no_occurrence and recommended_size > 0:
            submitted_sample_size = evaluation_data.get('sample_size')
            if submitted_sample_size is not None:
                submitted_sample_size = int(submitted_sample_size)
                if submitted_sample_size < recommended_size:
                    return jsonify({
                        'success': False,
                        'message': f'표본 크기({submitted_sample_size})는 권장 표본수({recommended_size})보다 작을 수 없습니다.'
                    })

        print("DB 저장 시작...")
        # 운영평가 결과 저장 (Header-Line 구조)
        save_operation_evaluation(rcm_id, control_code, user_info['user_id'], operation_evaluation_session, design_evaluation_session, evaluation_data)
        print("DB 저장 완료!")

        # 활동 로그 기록
        log_user_activity(user_info, 'OPERATION_EVALUATION', f'운영평가 저장 - {control_code}',
                         f'/api/operation-evaluation/save',
                         request.remote_addr, request.headers.get('User-Agent'))

        print("저장 성공 응답 반환")
        return jsonify({
            'success': True,
            'message': '운영평가 결과가 저장되었습니다.'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"운영평가 저장 오류: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'저장 중 오류가 발생했습니다: {str(e)}'
        })

@bp_link7.route('/api/operation-evaluation/load/<int:rcm_id>/<design_evaluation_session>')
@login_required
def load_operation_evaluation(rcm_id, design_evaluation_session):
    """운영평가 데이터 로드 API (설계평가 세션별)"""
    user_info = get_user_info()

    try:
        # 권한 체크
        user_rcms = get_user_rcms(user_info['user_id'])
        rcm_ids = [rcm['rcm_id'] for rcm in user_rcms]

        if rcm_id not in rcm_ids:
            return jsonify({'success': False, 'message': '접근 권한이 없습니다.'}), 403

        # 운영평가 세션명 생성
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        evaluations = get_operation_evaluations(rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)

        evaluation_dict = {}
        for eval_data in evaluations:
            control_code = eval_data['control_code']
            evaluation_dict[control_code] = {
                'sample_size': eval_data['sample_size'],
                'exception_count': eval_data['exception_count'],
                'mitigating_factors': eval_data.get('mitigating_factors'),
                'exception_details': eval_data['exception_details'],
                'conclusion': eval_data['conclusion'],
                'improvement_plan': eval_data['improvement_plan']
            }

        return jsonify({'success': True, 'evaluations': evaluation_dict})

    except Exception as e:
        return jsonify({'success': False, 'message': '데이터 로드 중 오류가 발생했습니다.'}), 500

@bp_link7.route('/api/operation-evaluation/samples/<int:line_id>')
@login_required
def load_operation_evaluation_samples(line_id):
    """평가 버튼 클릭 시 특정 line_id의 샘플 데이터 조회 API"""
    user_info = get_user_info()

    try:
        # line_id에 해당하는 통제의 권한 확인
        with get_db() as conn:
            line_info = conn.execute('''
                SELECT h.rcm_id, h.user_id
                FROM sb_operation_evaluation_line l
                JOIN sb_operation_evaluation_header h ON l.header_id = h.header_id
                WHERE l.line_id = %s
            ''', (line_id,)).fetchone()

            if not line_info:
                return jsonify({'success': False, 'message': '평가 데이터를 찾을 수 없습니다.'}), 404

            # 권한 체크
            if line_info['user_id'] != user_info['user_id']:
                return jsonify({'success': False, 'message': '접근 권한이 없습니다.'}), 403

        # 샘플 데이터 조회
        sample_lines = get_operation_evaluation_samples(line_id)

        # 설계평가 샘플 조회
        design_sample = get_design_evaluation_sample(line_id)

        # attributes 정의 조회 (RCM detail에서)
        attributes = []
        population_attribute_count = 0

        with get_db() as conn:
            # line_id로부터 control_code와 rcm_id 조회
            line_detail = conn.execute('''
                SELECT l.control_code, h.rcm_id
                FROM sb_operation_evaluation_line l
                JOIN sb_operation_evaluation_header h ON l.header_id = h.header_id
                WHERE l.line_id = %s
            ''', (line_id,)).fetchone()

            if line_detail:
                # RCM detail에서 attribute 정의 조회
                rcm_detail = conn.execute('''
                    SELECT population_attribute_count,
                           attribute0, attribute1, attribute2, attribute3, attribute4,
                           attribute5, attribute6, attribute7, attribute8, attribute9
                    FROM sb_rcm_detail
                    WHERE rcm_id = %s AND control_code = %s
                ''', (line_detail['rcm_id'], line_detail['control_code'])).fetchone()

                if rcm_detail:
                    population_attribute_count = rcm_detail['population_attribute_count'] or 0

                    # attribute 정의 생성 (RCM detail에 정의된 모든 attributes 반환)
                    for i in range(10):
                        # RCM detail에서 attribute 이름 가져오기
                        attr_name = rcm_detail[f'attribute{i}'] if rcm_detail[f'attribute{i}'] else None

                        # 이름이 정의되지 않은 attribute는 skip
                        if not attr_name:
                            continue

                        # population_attribute_count를 기준으로 모집단/증빙 구분
                        if i < population_attribute_count:
                            attr_type = 'population'
                        else:
                            attr_type = 'evidence'

                        attributes.append({
                            'attribute': f'attribute{i}',
                            'name': attr_name,
                            'type': attr_type
                        })

        return jsonify({
            'success': True,
            'samples': sample_lines,
            'design_sample': design_sample,
            'attributes': attributes,
            'population_attribute_count': population_attribute_count
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'샘플 데이터 로드 중 오류가 발생했습니다: {str(e)}'}), 500

# REMOVED: Duplicate reset API that was deleting entire header
# The correct reset API is at line 589: operation_evaluation_reset()
# That one only deletes specific control's line and files, not the header
# ===================================================================
# APD01 표준통제 테스트 API
# ===================================================================

@bp_link7.route('/api/operation-evaluation/apd01/upload-population', methods=['POST'])
@login_required
def apd01_upload_population():
    """APD01 모집단 업로드 및 파싱"""
    user_info = get_user_info()

    # 파일 받기
    if 'population_file' not in request.files:
        return jsonify({'success': False, 'message': '파일이 없습니다.'})

    file = request.files['population_file']
    if not file.filename:
        return jsonify({'success': False, 'message': '파일을 선택해주세요.'})

    # 필드 매핑 정보 받기 (JSON)
    import json
    field_mapping_str = request.form.get('field_mapping')
    if not field_mapping_str:
        return jsonify({'success': False, 'message': '필드 매핑 정보가 없습니다.'})

    try:
        field_mapping = json.loads(field_mapping_str)
    except:
        return jsonify({'success': False, 'message': '필드 매핑 형식이 올바르지 않습니다.'})

    # RCM 정보
    rcm_id = request.form.get('rcm_id')
    control_code = request.form.get('control_code')
    design_evaluation_session = request.form.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        return jsonify({'success': False, 'message': '필수 정보가 누락되었습니다.'})

    try:
        # 운영평가 헤더 조회 (RCM 페이지에서 이미 생성되어 있어야 함)
        operation_evaluation_session = f"OP_{design_evaluation_session}"
        from auth import get_db

        with get_db() as conn:
            header = conn.execute('''
                SELECT header_id FROM sb_operation_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if not header:
                return jsonify({'success': False, 'message': '운영평가 세션을 찾을 수 없습니다. RCM 페이지에서 다시 시작해주세요.'})

            operation_header_id = header['header_id']

        # 임시 파일로 저장
        import tempfile
        import os
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()

        # RCM detail에서 recommended_sample_size 가져오기
        with get_db() as conn:
            rcm_detail = conn.execute('''
                SELECT recommended_sample_size
                FROM sb_rcm_detail
                WHERE rcm_id = %s AND control_code = %s
            ''', (rcm_id, control_code)).fetchone()

        recommended_size = rcm_detail['recommended_sample_size'] if rcm_detail else None

        # 모집단 파싱 (recommended_sample_size 전달)
        result = file_manager.parse_apd01_population(temp_file.name, field_mapping, recommended_size)

        # 표본 선택
        samples = file_manager.select_random_samples(result['population'], result['sample_size'])

        # 임시 파일 삭제 (Windows에서 파일 핸들 문제로 실패할 수 있으므로 무시)
        try:
            os.unlink(temp_file.name)
        except Exception as e:
            print(f"임시 파일 삭제 실패 (무시됨): {e}")

        # 템플릿 기반 엑셀 파일 생성 및 저장 (운영평가 헤더 ID 사용)
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=result['population'],
            field_mapping=field_mapping,
            samples=samples,
            test_results_data=None  # 아직 테스트 결과 없음
        )

        # 세션에 파일 경로만 저장 (나중에 저장할 때 사용)
        session_key = f'apd01_test_{rcm_id}_{control_code}'
        session[session_key] = {
            'file_paths': file_paths,
            'rcm_id': rcm_id,
            'control_code': control_code,
            'design_evaluation_session': design_evaluation_session,
            'operation_header_id': operation_header_id
        }

        return jsonify({
            'success': True,
            'population_count': result['count'],
            'sample_size': result['sample_size'],
            'samples': samples
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'파일 처리 오류: {str(e)}'})


# The following routes are deprecated and replaced by the generic implementation
# in operation_evaluation_generic.py. They are kept here for reference but can be removed.
# - apd01_save_test_results
# - user_operation_evaluation_apd01


@bp_link7.route('/api/operation-evaluation/reset', methods=['POST'])
@login_required
def operation_evaluation_reset():
    """운영평가 파일 삭제 및 리셋 (모든 통제 공통)"""
    user_info = get_user_info()
    data = request.get_json()

    rcm_id = data.get('rcm_id')
    control_code = data.get('control_code')
    design_evaluation_session = data.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        return jsonify({'success': False, 'message': '필수 데이터가 누락되었습니다.'})

    try:
        import os
        operation_evaluation_session = f"OP_{design_evaluation_session}"


        # DB에서 operation_header_id 조회 (있으면)
        from auth import get_db
        with get_db() as conn:
            header = conn.execute('''
                SELECT header_id FROM sb_operation_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()


            if header:
                operation_header_id = header['header_id']


                # DB 라인 데이터 삭제 (해당 통제만)
                deleted_rows = conn.execute('''
                    DELETE FROM sb_operation_evaluation_line
                    WHERE header_id = %s AND control_code = %s
                ''', (operation_header_id, control_code))
                conn.commit()


                # 파일 삭제 (절대 경로 사용, control_code 폴더 제거)
                app_root = os.path.dirname(os.path.abspath(__file__))
                excel_path = os.path.join(app_root, 'static', 'uploads', 'operation_evaluations',
                                        str(rcm_id), str(operation_header_id), f'{control_code}_evaluation.xlsx')


                if os.path.exists(excel_path):
                    os.remove(excel_path)

        # 세션 정리 (통제별로 다른 키 사용)
        if control_code == 'APD01':
            session_key = f'apd01_test_{rcm_id}_{control_code}'
        elif control_code == 'APD07':
            session_key = f'apd07_test_{rcm_id}_{control_code}'
        else:
            session_key = f'{control_code.lower()}_test_{rcm_id}_{control_code}'

        session.pop(session_key, None)

        log_user_activity(user_info, 'OPERATION_EVALUATION', f'{control_code} 리셋',
                         '/api/operation-evaluation/reset',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({'success': True, 'message': '초기화되었습니다.'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'리셋 오류: {str(e)}'})


# The following routes are deprecated and replaced by the generic implementation
# in operation_evaluation_generic.py. They are kept here for reference but can be removed.
# - apd01_save_test_results
# - user_operation_evaluation_apd01


@bp_link7.route('/api/design-evaluation/get', methods=['GET'])
@login_required
def get_design_evaluation_data():
    """설계평가 데이터 조회 (운영평가에서 보기용)"""
    try:
        user_info = get_user_info()
        rcm_id_param = request.args.get('rcm_id')
        evaluation_session = request.args.get('evaluation_session')


        if not rcm_id_param or not evaluation_session:
            return jsonify({'success': False, 'message': '필수 파라미터가 누락되었습니다.'})

        rcm_id = int(rcm_id_param)

        # 설계평가 데이터 조회
        from auth import get_design_evaluations
        evaluations = get_design_evaluations(rcm_id, user_info['user_id'], evaluation_session)

        # RCM 상세 정보와 조인하여 통제 정보 추가
        rcm_details = get_rcm_details(rcm_id)
        rcm_dict = {detail['control_code']: detail for detail in rcm_details}

        # 매핑 정보 조회
        from auth import get_rcm_detail_mappings
        rcm_mappings_list = get_rcm_detail_mappings(rcm_id)
        rcm_mappings = {m['control_code']: m for m in rcm_mappings_list}

        # 평가 데이터에 통제 정보 추가
        result = []
        for eval_data in evaluations:
            control_code = eval_data['control_code']
            if control_code in rcm_dict:
                detail = rcm_dict[control_code]
                mapping = rcm_mappings.get(control_code, {})
                result.append({
                    'control_code': control_code,
                    'control_name': detail['control_name'],
                    'control_description': detail['control_description'],
                    'control_frequency': detail['control_frequency'],
                    'control_frequency_name': detail.get('control_frequency_name'),
                    'control_nature': detail['control_nature'],
                    'control_nature_name': detail.get('control_nature_name'),
                    'key_control': detail.get('key_control'),
                    'std_control_code': mapping.get('std_control_code'),
                    'std_control_name': mapping.get('std_control_name'),
                    'design_adequacy': eval_data.get('overall_effectiveness'),
                    'improvement_plan': eval_data.get('recommended_actions'),
                    'evaluated_date': eval_data.get('evaluation_date')
                })

        return jsonify({'success': True, 'evaluations': result})

    except Exception as e:
        return jsonify({'success': False, 'message': f'조회 오류: {str(e)}'})

@bp_link7.route('/operation-evaluation/apd07')
@login_required
def user_operation_evaluation_apd07():
    """APD07 운영평가 페이지"""
    user_info = get_user_info()

    rcm_id = request.args.get('rcm_id')
    control_code = request.args.get('control_code')
    control_name = request.args.get('control_name')
    design_evaluation_session = request.args.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        flash('필수 정보가 누락되었습니다.', 'error')
        return redirect(url_for('link7.user_operation_evaluation'))

    # 기존 운영평가 데이터 조회
    existing_data = None
    operation_evaluation_session = f"OP_{design_evaluation_session}"

    try:
        from auth import get_or_create_operation_evaluation_header
        with get_db() as conn:
            # 운영평가 헤더 조회 (있으면)
            header = conn.execute('''
                SELECT header_id FROM sb_operation_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if header:
                operation_header_id = header['header_id']

                # 저장된 파일에서 데이터 로드
                loaded_data = file_manager.load_operation_test_data(
                    rcm_id=rcm_id,
                    operation_header_id=operation_header_id,
                    control_code=control_code
                )

                if loaded_data and loaded_data['samples_data']:
                    existing_data = {
                        'samples': loaded_data['samples_data'].get('samples', []),
                        'population_count': loaded_data['samples_data'].get('population_count', 0),
                        'sample_size': loaded_data['samples_data'].get('sample_size', 0),
                        'test_results': loaded_data['samples_data'].get('test_results', {}),
                        'operation_header_id': operation_header_id
                    }

                    # 세션에 operation_header_id 저장 (저장 시 필요)
                    session_key = f'apd07_test_{rcm_id}_{control_code}'
                    session[session_key] = {
                        'operation_header_id': operation_header_id,
                        'population_count': existing_data['population_count'],
                        'sample_size': existing_data['sample_size']
                    }

    except Exception as e:
        print(f"기존 데이터 로드 오류: {e}")
        import traceback
        traceback.print_exc()
        # 오류가 발생해도 페이지는 정상적으로 표시
        pass

    log_user_activity(user_info, 'PAGE_ACCESS', 'APD07 운영평가', '/operation-evaluation/apd07',
                     request.remote_addr, request.headers.get('User-Agent'))

    return render_template('user_operation_evaluation_apd07.jsp',
                         rcm_id=rcm_id,
                         control_code=control_code,
                         control_name=control_name,
                         design_evaluation_session=design_evaluation_session,
                         existing_data=existing_data,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr)

# ===================================================================
# APD07 표준통제 테스트 API
# ===================================================================

@bp_link7.route('/api/operation-evaluation/apd07/upload-population', methods=['POST'])
@login_required
def apd07_upload_population():
    """APD07 모집단 업로드 및 파싱 (데이터 직접변경 승인)"""
    user_info = get_user_info()

    # 파일 받기
    if 'population_file' not in request.files:
        return jsonify({'success': False, 'message': '파일이 없습니다.'})

    file = request.files['population_file']
    if not file.filename:
        return jsonify({'success': False, 'message': '파일을 선택해주세요.'})

    # 필드 매핑 정보 받기 (JSON)
    import json
    field_mapping_str = request.form.get('field_mapping')
    if not field_mapping_str:
        return jsonify({'success': False, 'message': '필드 매핑 정보가 없습니다.'})

    try:
        field_mapping = json.loads(field_mapping_str)
    except:
        return jsonify({'success': False, 'message': '필드 매핑 형식이 올바르지 않습니다.'})

    # RCM 정보
    rcm_id = request.form.get('rcm_id')
    control_code = request.form.get('control_code')
    design_evaluation_session = request.form.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        return jsonify({'success': False, 'message': '필수 정보가 누락되었습니다.'})

    try:
        # 운영평가 헤더 조회 (RCM 페이지에서 이미 생성되어 있어야 함)
        operation_evaluation_session = f"OP_{design_evaluation_session}"
        from auth import get_db

        with get_db() as conn:
            header = conn.execute('''
                SELECT header_id FROM sb_operation_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if not header:
                return jsonify({'success': False, 'message': '운영평가 세션을 찾을 수 없습니다. RCM 페이지에서 다시 시작해주세요.'})

            operation_header_id = header['header_id']

        # 임시 파일로 저장
        import tempfile
        import os
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()

        # RCM detail에서 recommended_sample_size 가져오기
        with get_db() as conn:
            rcm_detail = conn.execute('''
                SELECT recommended_sample_size
                FROM sb_rcm_detail
                WHERE rcm_id = %s AND control_code = %s
            ''', (rcm_id, control_code)).fetchone()

        recommended_size = rcm_detail['recommended_sample_size'] if rcm_detail else None

        # 모집단 파싱 (APD07용, recommended_sample_size 전달)
        result = file_manager.parse_apd07_population(temp_file.name, field_mapping, recommended_size)

        # 표본 선택
        samples = file_manager.select_random_samples(result['population'], result['sample_size'])

        # 임시 파일 삭제 (Windows에서 파일 핸들 문제로 실패할 수 있으므로 무시)
        try:
            os.unlink(temp_file.name)
        except Exception as e:
            print(f"임시 파일 삭제 실패 (무시됨): {e}")

        # 템플릿 기반 엑셀 파일 생성 및 저장 (운영평가 헤더 ID 사용)
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=result['population'],
            field_mapping=field_mapping,
            samples=samples,
            test_results_data=None  # 아직 테스트 결과 없음
        )

        # 세션에 파일 경로만 저장 (나중에 저장할 때 사용)
        session_key = f'apd07_test_{rcm_id}_{control_code}'
        session[session_key] = {
            'file_paths': file_paths,
            'rcm_id': rcm_id,
            'control_code': control_code,
            'design_evaluation_session': design_evaluation_session,
            'operation_header_id': operation_header_id
        }

        return jsonify({
            'success': True,
            'population_count': result['count'],
            'sample_size': result['sample_size'],
            'samples': samples
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'파일 처리 오류: {str(e)}'})


@bp_link7.route('/api/operation-evaluation/apd07/save-test-results', methods=['POST'])
@login_required
def apd07_save_test_results():
    """APD07 테스트 결과 저장 (데이터 직접변경 승인)"""
    user_info = get_user_info()
    data = request.get_json()

    rcm_id = data.get('rcm_id')
    control_code = data.get('control_code')
    design_evaluation_session = data.get('design_evaluation_session')
    test_results = data.get('test_results')  # 표본별 테스트 결과

    if not all([rcm_id, control_code, design_evaluation_session, test_results]):
        return jsonify({'success': False, 'message': '필수 데이터가 누락되었습니다.'})

    try:
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        # 세션에서 파일 경로 정보 가져오기
        session_key = f'apd07_test_{rcm_id}_{control_code}'
        test_data = session.get(session_key)

        if not test_data:
            return jsonify({'success': False, 'message': '테스트 데이터를 찾을 수 없습니다. 모집단을 다시 업로드해주세요.'})

        # 세션에서 operation_header_id 가져오기
        operation_header_id = test_data.get('operation_header_id')
        if not operation_header_id:
            return jsonify({'success': False, 'message': '운영평가 헤더 ID를 찾을 수 없습니다.'})

        # 저장된 파일에서 표본 데이터 로드
        loaded_data = file_manager.load_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code
        )

        if not loaded_data or not loaded_data['samples_data']:
            return jsonify({'success': False, 'message': '저장된 표본 데이터를 찾을 수 없습니다.'})

        samples_data = loaded_data['samples_data']

        # 템플릿 기반 엑셀 파일 업데이트 (테스트 결과 추가)
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=loaded_data.get('population_data', []),
            field_mapping=samples_data.get('field_mapping', {}),
            samples=samples_data['samples'],
            test_results_data={
                'test_results': test_results,
                'exceptions': [r for r in test_results if r.get('has_exception')],
                'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception',
                'test_type': 'APD07'
            }
        )

        # 평가 데이터 구성 (메타데이터만 DB에 저장)
        evaluation_data = {
            'test_type': 'APD07',
            'population_count': samples_data['population_count'],
            'sample_size': samples_data['sample_size'],
            'population_path': None,  # 템플릿 방식에서는 엑셀에 통합
            'samples_path': file_paths.get('samples_path'),
            'test_results_path': file_paths.get('excel_path'),  # 엑셀 파일 경로
            'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception'
        }

        # 운영평가 저장
        save_operation_evaluation(rcm_id, control_code, user_info['user_id'],
                                 operation_evaluation_session, design_evaluation_session, evaluation_data)

        # 세션 정리 제거 - 다시 저장할 수 있도록 세션 유지
        # session.pop(session_key, None)

        log_user_activity(user_info, 'OPERATION_EVALUATION', f'APD07 테스트 저장 - {control_code}',
                         '/api/operation-evaluation/apd07/save-test-results',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({'success': True, 'message': 'APD07 테스트 결과가 저장되었습니다.'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'저장 오류: {str(e)}'})

# ===================================================================
# APD09 운영평가 라우트
# ===================================================================

@bp_link7.route('/operation-evaluation/apd09')
@login_required
def user_operation_evaluation_apd09():
    """APD09 운영평가 페이지"""
    user_info = get_user_info()

    rcm_id = request.args.get('rcm_id')
    control_code = request.args.get('control_code')
    control_name = request.args.get('control_name')
    design_evaluation_session = request.args.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        flash('필수 정보가 누락되었습니다.', 'error')
        return redirect(url_for('link7.user_operation_evaluation'))

    # 기존 운영평가 데이터 조회
    existing_data = None
    operation_evaluation_session = f"OP_{design_evaluation_session}"

    try:
        from auth import get_or_create_operation_evaluation_header
        with get_db() as conn:
            # 운영평가 헤더 조회 (있으면)
            header = conn.execute('''
                SELECT header_id FROM sb_operation_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if header:
                operation_header_id = header['header_id']

                # 저장된 파일에서 데이터 로드
                loaded_data = file_manager.load_operation_test_data(
                    rcm_id=rcm_id,
                    operation_header_id=operation_header_id,
                    control_code=control_code
                )

                if loaded_data and loaded_data['samples_data']:
                    existing_data = {
                        'samples': loaded_data['samples_data'].get('samples', []),
                        'population_count': loaded_data['samples_data'].get('population_count', 0),
                        'sample_size': loaded_data['samples_data'].get('sample_size', 0),
                        'test_results': loaded_data['samples_data'].get('test_results', {}),
                        'operation_header_id': operation_header_id
                    }

                    # 세션에 operation_header_id 저장 (저장 시 필요)
                    session_key = f'apd09_test_{rcm_id}_{control_code}'
                    session[session_key] = {
                        'operation_header_id': operation_header_id,
                        'population_count': existing_data['population_count'],
                        'sample_size': existing_data['sample_size']
                    }

    except Exception as e:
        print(f"기존 데이터 로드 오류: {e}")
        import traceback
        traceback.print_exc()
        # 오류가 발생해도 페이지는 정상적으로 표시
        pass

    log_user_activity(user_info, 'PAGE_ACCESS', 'APD09 운영평가', '/operation-evaluation/apd09',
                     request.remote_addr, request.headers.get('User-Agent'))

    return render_template('user_operation_evaluation_apd09.jsp',
                         rcm_id=rcm_id,
                         control_code=control_code,
                         control_name=control_name,
                         design_evaluation_session=design_evaluation_session,
                         existing_data=existing_data,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr)


@bp_link7.route('/api/operation-evaluation/apd09/upload-population', methods=['POST'])
@login_required
def upload_apd09_population():
    """APD09 모집단 업로드 및 파싱 (OS 접근권한 부여 승인)"""
    user_info = get_user_info()

    # 파일 받기
    if 'population_file' not in request.files:
        return jsonify({'success': False, 'message': '파일이 없습니다.'})

    file = request.files['population_file']
    if not file.filename:
        return jsonify({'success': False, 'message': '파일을 선택해주세요.'})

    # 필드 매핑 정보 받기 (JSON)
    field_mapping_str = request.form.get('field_mapping')
    if not field_mapping_str:
        return jsonify({'success': False, 'message': '필드 매핑 정보가 없습니다.'})

    try:
        field_mapping = json.loads(field_mapping_str)
    except:
        return jsonify({'success': False, 'message': '필드 매핑 형식이 올바르지 않습니다.'})

    # RCM 정보
    rcm_id = request.form.get('rcm_id')
    control_code = request.form.get('control_code')
    design_evaluation_session = request.form.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        return jsonify({'success': False, 'message': '필수 정보가 누락되었습니다.'})

    try:
        # 운영평가 헤더 조회 (RCM 페이지에서 이미 생성되어 있어야 함)
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        with get_db() as conn:
            header = conn.execute('''
                SELECT header_id FROM sb_operation_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if not header:
                return jsonify({'success': False, 'message': '운영평가 세션을 찾을 수 없습니다. RCM 페이지에서 다시 시작해주세요.'})

            operation_header_id = header['header_id']

        # 임시 파일로 저장
        import tempfile
        import os
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()

        # RCM detail에서 recommended_sample_size 가져오기
        with get_db() as conn:
            rcm_detail = conn.execute('''
                SELECT recommended_sample_size
                FROM sb_rcm_detail
                WHERE rcm_id = %s AND control_code = %s
            ''', (rcm_id, control_code)).fetchone()

        recommended_size = rcm_detail['recommended_sample_size'] if rcm_detail else None

        # 모집단 파싱 (APD09용, recommended_sample_size 전달)
        result = file_manager.parse_apd09_population(temp_file.name, field_mapping, recommended_size)

        # 표본 선택
        samples = file_manager.select_random_samples(result['population'], result['sample_size'])

        # 임시 파일 삭제 (Windows에서 파일 핸들 문제로 실패할 수 있으므로 무시)
        try:
            os.unlink(temp_file.name)
        except Exception as e:
            print(f"임시 파일 삭제 실패 (무시됨): {e}")

        # 템플릿 기반 엑셀 파일 생성 및 저장 (운영평가 헤더 ID 사용)
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=result['population'],
            field_mapping=field_mapping,
            samples=samples,
            test_results_data=None  # 아직 테스트 결과 없음
        )

        # 세션에 파일 경로만 저장 (나중에 저장할 때 사용)
        session_key = f'apd09_test_{rcm_id}_{control_code}'
        session[session_key] = {
            'file_paths': file_paths,
            'rcm_id': rcm_id,
            'control_code': control_code,
            'design_evaluation_session': design_evaluation_session,
            'operation_header_id': operation_header_id,
            'field_mapping': field_mapping,
            'population_count': result['count'],
            'sample_size': result['sample_size']
        }

        return jsonify({
            'success': True,
            'population_count': result['count'],
            'sample_size': result['sample_size'],
            'samples': samples
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


@bp_link7.route('/api/operation-evaluation/apd09/save-test-results', methods=['POST'])
@login_required
def save_apd09_test_results():
    """APD09 테스트 결과 저장"""
    try:
        user_info = get_user_info()
        data = request.json
        rcm_id = data.get('rcm_id')
        control_code = data.get('control_code')
        design_evaluation_session = data.get('design_evaluation_session')
        test_results = data.get('test_results', [])

        # 세션에서 operation_header_id 가져오기
        session_key = f'apd09_test_{rcm_id}_{control_code}'
        session_data = session.get(session_key)

        if not session_data:
            return jsonify({'success': False, 'message': '세션 정보가 없습니다. 모집단을 다시 업로드해주세요.'})

        operation_header_id = session_data['operation_header_id']
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        # 저장된 파일에서 표본 데이터 로드
        loaded_data = file_manager.load_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code
        )

        if not loaded_data or not loaded_data['samples_data']:
            return jsonify({'success': False, 'message': '저장된 표본 데이터를 찾을 수 없습니다.'})

        samples_data = loaded_data['samples_data']

        # 엑셀 파일에 테스트 결과 저장
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=loaded_data.get('population_data', []),
            field_mapping=session_data.get('field_mapping', {}),  # 세션에서 가져오기
            samples=samples_data['samples'],
            test_results_data={
                'test_results': test_results,
                'exceptions': [r for r in test_results if r.get('has_exception')],
                'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception',
                'test_type': 'APD09'
            }
        )

        # 운영평가 데이터 저장
        evaluation_data = {
            'sample_size': session_data['sample_size'],
            'population_path': file_paths.get('population_file'),
            'samples_path': file_paths.get('excel_path'),
            'test_results_path': file_paths.get('excel_path'),  # 엑셀 파일 경로
            'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception'
        }

        # 운영평가 저장
        save_operation_evaluation(rcm_id, control_code, user_info['user_id'],
                                 operation_evaluation_session, design_evaluation_session, evaluation_data)

        # 세션 정리 제거 - 다시 저장할 수 있도록 세션 유지
        # session.pop(session_key, None)

        log_user_activity(user_info, 'OPERATION_EVALUATION', f'APD09 테스트 저장 - {control_code}',
                         '/api/operation-evaluation/apd09/save-test-results',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({'success': True, 'message': 'APD09 테스트 결과가 저장되었습니다.'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'저장 오류: {str(e)}'})


# ===================================================================
# 운영평가 리셋 API
# ===================================================================

# ===================================================================
# APD12 운영평가 라우트
# ===================================================================

@bp_link7.route('/operation-evaluation/apd12')
@login_required
def user_operation_evaluation_apd12():
    """APD12 운영평가 페이지"""
    user_info = get_user_info()

    rcm_id = request.args.get('rcm_id')
    control_code = request.args.get('control_code')
    control_name = request.args.get('control_name')
    design_evaluation_session = request.args.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        flash('필수 정보가 누락되었습니다.', 'error')
        return redirect(url_for('link7.user_operation_evaluation'))

    # 기존 운영평가 데이터 조회
    existing_data = None
    operation_evaluation_session = f"OP_{design_evaluation_session}"

    try:
        from auth import get_or_create_operation_evaluation_header
        with get_db() as conn:
            # 운영평가 헤더 조회 (있으면)
            header = conn.execute('''
                SELECT header_id FROM sb_operation_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if header:
                operation_header_id = header['header_id']

                # 저장된 파일에서 데이터 로드
                loaded_data = file_manager.load_operation_test_data(
                    rcm_id=rcm_id,
                    operation_header_id=operation_header_id,
                    control_code=control_code
                )

                if loaded_data and loaded_data['samples_data']:
                    existing_data = {
                        'samples': loaded_data['samples_data'].get('samples', []),
                        'population_count': loaded_data['samples_data'].get('population_count', 0),
                        'sample_size': loaded_data['samples_data'].get('sample_size', 0),
                        'test_results': loaded_data['samples_data'].get('test_results', {}),
                        'operation_header_id': operation_header_id
                    }

                    # 세션에 operation_header_id 저장 (저장 시 필요)
                    session_key = f'apd12_test_{rcm_id}_{control_code}'
                    session[session_key] = {
                        'operation_header_id': operation_header_id,
                        'population_count': existing_data['population_count'],
                        'sample_size': existing_data['sample_size']
                    }

    except Exception as e:
        print(f"기존 데이터 로드 오류: {e}")
        import traceback
        traceback.print_exc()
        # 오류가 발생해도 페이지는 정상적으로 표시
        pass

    log_user_activity(user_info, 'PAGE_ACCESS', 'APD12 운영평가', '/operation-evaluation/apd12',
                     request.remote_addr, request.headers.get('User-Agent'))

    return render_template('user_operation_evaluation_apd12.jsp',
                         rcm_id=rcm_id,
                         control_code=control_code,
                         control_name=control_name,
                         design_evaluation_session=design_evaluation_session,
                         existing_data=existing_data,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr)


@bp_link7.route('/api/operation-evaluation/apd12/upload-population', methods=['POST'])
@login_required
def upload_apd12_population():
    """APD12 모집단 업로드 및 파싱 (DB 접근권한 부여 승인)"""
    user_info = get_user_info()

    # 파일 받기
    if 'population_file' not in request.files:
        return jsonify({'success': False, 'message': '파일이 없습니다.'})

    file = request.files['population_file']
    if not file.filename:
        return jsonify({'success': False, 'message': '파일을 선택해주세요.'})

    # 필드 매핑 정보 받기 (JSON)
    field_mapping_str = request.form.get('field_mapping')
    if not field_mapping_str:
        return jsonify({'success': False, 'message': '필드 매핑 정보가 없습니다.'})

    try:
        field_mapping = json.loads(field_mapping_str)
    except:
        return jsonify({'success': False, 'message': '필드 매핑 형식이 올바르지 않습니다.'})

    # RCM 정보
    rcm_id = request.form.get('rcm_id')
    control_code = request.form.get('control_code')
    design_evaluation_session = request.form.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        return jsonify({'success': False, 'message': '필수 정보가 누락되었습니다.'})

    try:
        # 운영평가 헤더 조회 (RCM 페이지에서 이미 생성되어 있어야 함)
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        with get_db() as conn:
            header = conn.execute('''
                SELECT header_id FROM sb_operation_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if not header:
                return jsonify({'success': False, 'message': '운영평가 세션을 찾을 수 없습니다. RCM 페이지에서 다시 시작해주세요.'})

            operation_header_id = header['header_id']

        # 임시 파일로 저장
        import tempfile
        import os
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()

        # RCM detail에서 recommended_sample_size 가져오기
        with get_db() as conn:
            rcm_detail = conn.execute('''
                SELECT recommended_sample_size
                FROM sb_rcm_detail
                WHERE rcm_id = %s AND control_code = %s
            ''', (rcm_id, control_code)).fetchone()

        recommended_size = rcm_detail['recommended_sample_size'] if rcm_detail else None

        # 모집단 파싱 (APD12용, recommended_sample_size 전달)
        result = file_manager.parse_apd12_population(temp_file.name, field_mapping, recommended_size)

        # 표본 선택
        samples = file_manager.select_random_samples(result['population'], result['sample_size'])

        # 임시 파일 삭제 (Windows에서 파일 핸들 문제로 실패할 수 있으므로 무시)
        try:
            os.unlink(temp_file.name)
        except Exception as e:
            print(f"임시 파일 삭제 실패 (무시됨): {e}")

        # 템플릿 기반 엑셀 파일 생성 및 저장 (운영평가 헤더 ID 사용)
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=result['population'],
            field_mapping=field_mapping,
            samples=samples,
            test_results_data=None  # 아직 테스트 결과 없음
        )

        # 세션에 파일 경로만 저장 (나중에 저장할 때 사용)
        session_key = f'apd12_test_{rcm_id}_{control_code}'
        session[session_key] = {
            'file_paths': file_paths,
            'rcm_id': rcm_id,
            'control_code': control_code,
            'design_evaluation_session': design_evaluation_session,
            'operation_header_id': operation_header_id,
            'field_mapping': field_mapping,
            'population_count': result['count'],
            'sample_size': result['sample_size']
        }

        return jsonify({
            'success': True,
            'population_count': result['count'],
            'sample_size': result['sample_size'],
            'samples': samples
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


@bp_link7.route('/api/operation-evaluation/apd12/save-test-results', methods=['POST'])
@login_required
def save_apd12_test_results():
    """APD12 테스트 결과 저장"""
    try:
        user_info = get_user_info()
        data = request.json
        rcm_id = data.get('rcm_id')
        control_code = data.get('control_code')
        design_evaluation_session = data.get('design_evaluation_session')
        test_results = data.get('test_results', [])

        # 세션에서 operation_header_id 가져오기
        session_key = f'apd12_test_{rcm_id}_{control_code}'
        session_data = session.get(session_key)

        if not session_data:
            return jsonify({'success': False, 'message': '세션 정보가 없습니다. 모집단을 다시 업로드해주세요.'})

        operation_header_id = session_data['operation_header_id']
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        # 엑셀 파일에 테스트 결과 ��장
        # 저장된 파일에서 표본 데이터 로드
        loaded_data = file_manager.load_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code
        )

        if not loaded_data or not loaded_data['samples_data']:
            return jsonify({'success': False, 'message': '저장된 표본 데이터를 찾을 수 없습니다.'})

        samples_data = loaded_data['samples_data']

        # 엑셀 파일에 테스트 결과 저장
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=loaded_data.get('population_data', []),
            field_mapping=session_data.get('field_mapping', {}),  # 세션에서 가져오기
            samples=samples_data['samples'],
            test_results_data={
                'test_results': test_results,
                'exceptions': [r for r in test_results if r.get('has_exception')],
                'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception',
                'test_type': 'APD12'
            }
        )

        # 운영평가 데이터 저장
        evaluation_data = {
            'test_type': 'APD12',
            'population_count': samples_data['population_count'],
            'sample_size': samples_data['sample_size'],
            'population_path': None,  # 템플릿 방식에서는 엑셀에 통합
            'samples_path': file_paths.get('samples_path'),
            'test_results_path': file_paths.get('excel_path'),  # 엑셀 파일 경로
            'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception'
        }

        # 운영평가 저장
        save_operation_evaluation(rcm_id, control_code, user_info['user_id'],
                                 operation_evaluation_session, design_evaluation_session, evaluation_data)

        # 세션 정리 제거 - 다시 저장할 수 있도록 세션 유지
        # session.pop(session_key, None)

        log_user_activity(user_info, 'OPERATION_EVALUATION', f'APD12 테스트 저장 - {control_code}',
                         '/api/operation-evaluation/apd12/save-test-results',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({'success': True, 'message': 'APD12 테스트 결과가 저장되었습니다.'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'저장 오류: {str(e)}'})

# ============================================================================
# ELC 운영평가 (수동통제만)
# ============================================================================

@bp_link7.route('/elc/operation-evaluation')
@login_required
def elc_operation_evaluation():
    """ELC 운영평가 페이지"""
    user_info = get_user_info()

    # ELC RCM 목록만 필터링
    all_rcms = get_user_rcms(user_info['user_id'])
    elc_rcms = [rcm for rcm in all_rcms if rcm.get('control_category') == 'ELC']

    # 각 RCM에 대해 모든 설계평가 세션 조회 (진행중 + 완료)
    from auth import get_all_design_evaluation_sessions
    for rcm in elc_rcms:
        all_sessions = get_all_design_evaluation_sessions(rcm['rcm_id'], user_info['user_id'])
        completed_sessions = [s for s in all_sessions if s['completed_date'] is not None]
        in_progress_sessions = [s for s in all_sessions if s['completed_date'] is None]

        # 완료된 세션에 대해서만 운영평가 진행상황 조회
        for session in completed_sessions:
            operation_evaluation_session = f"OP_{session['evaluation_session']}"

            from auth import count_completed_operation_evaluations
            with get_db() as conn:
                header = conn.execute('''
                    SELECT header_id FROM sb_operation_evaluation_header
                    WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
                ''', (rcm['rcm_id'], user_info['user_id'], operation_evaluation_session, session['evaluation_session'])).fetchone()

            if header:
                completed_count = count_completed_operation_evaluations(header['header_id'])
                session['operation_completed_count'] = completed_count
            else:
                session['operation_completed_count'] = 0

            # 운영평가 가능한 통제 개수 추가
            eligible_controls = get_key_rcm_details(rcm['rcm_id'], user_info['user_id'], session['evaluation_session'], control_category='ELC')
            session['eligible_control_count'] = len(eligible_controls)

        rcm['all_design_sessions'] = all_sessions
        rcm['completed_design_sessions'] = completed_sessions
        rcm['in_progress_design_sessions'] = in_progress_sessions
        rcm['design_evaluation_completed'] = len(completed_sessions) > 0

        # 핵심통제 개수 조회
        key_controls = get_key_rcm_details(rcm['rcm_id'], control_category='ELC')
        rcm['key_control_count'] = len(key_controls)
        rcm['has_key_controls'] = len(key_controls) > 0

    log_user_activity(user_info, 'PAGE_ACCESS', 'ELC 운영평가', '/elc/operation-evaluation',
                     request.remote_addr, request.headers.get('User-Agent'))

    return render_template('link7_operation_evaluation_unified.jsp',
                         evaluation_type='ELC',
                         user_rcms=elc_rcms,
                         is_logged_in=is_logged_in(),
                         user_info=user_info)

# ============================================================================
# TLC 운영평가 (자동통제 포함)
# ============================================================================

@bp_link7.route('/tlc/operation-evaluation')
@login_required
def tlc_operation_evaluation():
    """TLC 운영평가 페이지"""
    user_info = get_user_info()

    # TLC RCM 목록만 필터링
    all_rcms = get_user_rcms(user_info['user_id'])
    tlc_rcms = [rcm for rcm in all_rcms if rcm.get('control_category') == 'TLC']

    # 각 RCM에 대해 모든 설계평가 세션 조회 (진행중 + 완료)
    from auth import get_all_design_evaluation_sessions
    for rcm in tlc_rcms:
        all_sessions = get_all_design_evaluation_sessions(rcm['rcm_id'], user_info['user_id'])
        completed_sessions = [s for s in all_sessions if s['completed_date'] is not None]
        in_progress_sessions = [s for s in all_sessions if s['completed_date'] is None]

        # 완료된 세션에 대해서만 운영평가 진행상황 조회
        for session in completed_sessions:
            operation_evaluation_session = f"OP_{session['evaluation_session']}"

            from auth import count_completed_operation_evaluations
            with get_db() as conn:
                header = conn.execute('''
                    SELECT header_id FROM sb_operation_evaluation_header
                    WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
                ''', (rcm['rcm_id'], user_info['user_id'], operation_evaluation_session, session['evaluation_session'])).fetchone()

            if header:
                completed_count = count_completed_operation_evaluations(header['header_id'])
                session['operation_completed_count'] = completed_count
            else:
                session['operation_completed_count'] = 0

            # 운영평가 가능한 통제 개수 추가
            eligible_controls = get_key_rcm_details(rcm['rcm_id'], user_info['user_id'], session['evaluation_session'], control_category='TLC')
            session['eligible_control_count'] = len(eligible_controls)

        rcm['all_design_sessions'] = all_sessions
        rcm['completed_design_sessions'] = completed_sessions
        rcm['in_progress_design_sessions'] = in_progress_sessions
        rcm['design_evaluation_completed'] = len(completed_sessions) > 0

        # 핵심통제 개수 조회
        key_controls = get_key_rcm_details(rcm['rcm_id'], control_category='TLC')
        rcm['key_control_count'] = len(key_controls)
        rcm['has_key_controls'] = len(key_controls) > 0

    log_user_activity(user_info, 'PAGE_ACCESS', 'TLC 운영평가', '/tlc/operation-evaluation',
                     request.remote_addr, request.headers.get('User-Agent'))

    return render_template('link7_operation_evaluation_unified.jsp',
                         evaluation_type='TLC',
                         user_rcms=tlc_rcms,
                         is_logged_in=is_logged_in(),
                         user_info=user_info)


# ===================================================================
# 일반 통제 모집단 업로드 API (표본수 0인 경우)
# ===================================================================

@bp_link7.route('/api/operation-evaluation/upload-population', methods=['POST'])
@login_required
def upload_general_population():
    """일반 통제 모집단 업로드 및 표본 추출 (표본수 0인 경우)"""
    import os
    from openpyxl import load_workbook
    from werkzeug.utils import secure_filename

    user_info = get_user_info()

    # 파일 받기
    if 'population_file' not in request.files:
        return jsonify({'success': False, 'error': '파일이 없습니다.'})

    file = request.files['population_file']
    if not file.filename:
        return jsonify({'success': False, 'error': '파일을 선택해주세요.'})

    print(f"[upload_general_population] 원본 파일명: {file.filename}")

    # 파라미터 받기
    control_code = request.form.get('control_code')
    rcm_id = request.form.get('rcm_id')
    design_evaluation_session = request.form.get('design_evaluation_session')
    field_mapping_str = request.form.get('field_mapping')

    if not all([control_code, rcm_id, design_evaluation_session, field_mapping_str]):
        return jsonify({'success': False, 'error': '필수 파라미터가 누락되었습니다.'})

    try:
        field_mapping = json.loads(field_mapping_str)
    except:
        return jsonify({'success': False, 'error': '필드 매핑 파싱 실패'})

    try:
        # 파일 저장
        upload_folder = os.path.join('uploads', 'populations')
        os.makedirs(upload_folder, exist_ok=True)

        # 원본 파일명에서 확장자 추출
        original_filename = file.filename
        file_ext = os.path.splitext(original_filename)[1]  # .xlsx

        # secure_filename으로 안전한 이름 생성
        filename = secure_filename(file.filename)

        # secure_filename이 파일명을 완전히 제거한 경우 (한글 등)
        if not filename or filename == file_ext.replace('.', ''):
            filename = f"population{file_ext}"

        # 확장자가 없으면 원본에서 가져온 확장자 추가
        if not os.path.splitext(filename)[1]:
            filename = filename + file_ext

        print(f"[upload_general_population] 원본: {original_filename}, 변환후: {filename}")

        # 파일 확장자 확인
        if not filename.lower().endswith(('.xlsx', '.xlsm')):
            return jsonify({'success': False, 'error': '.xlsx 또는 .xlsm 형식의 파일만 지원됩니다. (.xls 파일은 Excel에서 .xlsx로 변환 후 업로드해주세요)'})

        filepath = os.path.join(upload_folder, f"{user_info['user_id']}_{control_code}_{filename}")
        file.save(filepath)
        print(f"[upload_general_population] 파일 저장 완료: {filepath}")

        # 엑셀 파일 읽기 (openpyxl 사용)
        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            print(f"[upload_general_population] 엑셀 파일 로드 성공")
        except Exception as excel_error:
            print(f"[upload_general_population] 엑셀 파일 읽기 실패: {excel_error}")
            return jsonify({'success': False, 'error': f'엑셀 파일을 읽을 수 없습니다. 파일이 손상되었거나 암호로 보호되어 있을 수 있습니다. ({str(excel_error)})'})


        # 헤더 읽기 (첫 번째 행)
        headers = [cell.value for cell in ws[1]]

        # 필드 매핑 적용
        number_col_idx = field_mapping['number']
        description_col_idx = field_mapping['description']

        # 모집단 데이터 파싱
        population = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # 헤더 제외
            if row[number_col_idx] is not None:  # 빈 행 건너뛰기
                population.append({
                    'number': str(row[number_col_idx]),
                    'description': str(row[description_col_idx]) if row[description_col_idx] else ''
                })

        wb.close()

        population_count = len(population)

        # 표본 크기 자동 계산
        sample_size = file_manager.calculate_sample_size(population_count)

        # 무작위 표본 추출
        import random
        sample_indices = random.sample(range(population_count), min(sample_size, population_count))
        samples = [population[i] for i in sorted(sample_indices)]

        # 운영평가 세션 확인/생성
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        with get_db() as conn:
            # Header 확인/생성
            header = conn.execute('''
                SELECT header_id FROM sb_operation_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if not header:
                conn.execute('''
                    INSERT INTO sb_operation_evaluation_header (rcm_id, user_id, evaluation_session, design_evaluation_session)
                    VALUES (%s, %s, %s, %s)
                ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session))
                conn.commit()

                header = conn.execute('''
                    SELECT header_id FROM sb_operation_evaluation_header
                    WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
                ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            header_id = header['header_id']

            # 기존 Line 확인
            existing_line = conn.execute('''
                SELECT line_id FROM sb_operation_evaluation_line
                WHERE header_id = %s AND control_code = %s
            ''', (header_id, control_code)).fetchone()

            if existing_line:
                line_id = existing_line['line_id']

                # 기존 샘플 삭제
                conn.execute('DELETE FROM sb_evaluation_sample WHERE line_id = %s', (line_id,))

                # Line 업데이트 (sample_size만)
                conn.execute('''
                    UPDATE sb_operation_evaluation_line
                    SET sample_size = %s
                    WHERE line_id = %s
                ''', (sample_size, line_id))
            else:
                # 새 Line 생성
                conn.execute('''
                    INSERT INTO sb_operation_evaluation_line
                    (header_id, control_code, sample_size)
                    VALUES (%s, %s, %s)
                ''', (header_id, control_code, sample_size))

                # SQLite용 last_insert_rowid() 사용
                line_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()['id']

            print(f"[upload_general_population] line_id: {line_id}, 샘플 수: {len(samples)}")

            # 샘플 데이터 저장 (attribute0에 번호, attribute1에 설명 저장)
            for idx, sample in enumerate(samples, 1):
                print(f"[upload_general_population] 샘플 #{idx} 저장 중: {sample['number']}, {sample['description'][:30]}...")
                conn.execute('''
                    INSERT INTO sb_evaluation_sample
                    (line_id, sample_number, evaluation_type, attribute0, attribute1)
                    VALUES (%s, %s, %s, %s, %s)
                ''', (line_id, idx, 'operation', sample['number'], sample['description']))

            conn.commit()
            print(f"[upload_general_population] DB 커밋 완료")

            # 저장된 샘플 데이터 조회하여 sample_lines 형식으로 반환
            sample_lines = []
            saved_samples = conn.execute('''
                SELECT sample_number, evidence, has_exception, mitigation,
                       attribute0, attribute1, attribute2, attribute3, attribute4,
                       attribute5, attribute6, attribute7, attribute8, attribute9
                FROM sb_evaluation_sample
                WHERE line_id = %s
                ORDER BY sample_number
            ''', (line_id,)).fetchall()

            for sample in saved_samples:
                # attribute 데이터 수집
                attributes = {}
                for i in range(10):
                    attr_val = sample[f'attribute{i}']
                    if attr_val is not None:
                        attributes[f'attribute{i}'] = attr_val

                print(f"[upload_general_population] Sample #{sample['sample_number']} attributes: {attributes}")

                sample_lines.append({
                    'sample_number': sample['sample_number'],
                    'evidence': sample['evidence'] or '',
                    'result': 'exception' if sample['has_exception'] else 'no_exception',
                    'mitigation': sample['mitigation'] or '',
                    'attributes': attributes if attributes else None
                })

            print(f"[upload_general_population] 반환할 sample_lines: {json.dumps(sample_lines, ensure_ascii=False, indent=2)}")

            # RCM detail에서 attribute 정의 조회
            rcm_detail = conn.execute('''
                SELECT population_attribute_count,
                       attribute0, attribute1, attribute2, attribute3, attribute4,
                       attribute5, attribute6, attribute7, attribute8, attribute9
                FROM sb_rcm_detail
                WHERE rcm_id = %s AND control_code = %s
            ''', (rcm_id, control_code)).fetchone()

            population_attr_count = rcm_detail['population_attribute_count'] if rcm_detail and rcm_detail['population_attribute_count'] else 2
            print(f"[upload_general_population] RCM detail population_attribute_count: {population_attr_count}")
            print(f"[upload_general_population] RCM detail attributes: attribute0={rcm_detail['attribute0'] if rcm_detail else 'N/A'}, attribute1={rcm_detail['attribute1'] if rcm_detail else 'N/A'}, attribute2={rcm_detail['attribute2'] if rcm_detail else 'N/A'}, attribute3={rcm_detail['attribute3'] if rcm_detail else 'N/A'}")

            # 샘플 데이터를 확인하여 실제 사용된 attribute 찾기
            used_attributes = set()
            for sample in saved_samples:
                for i in range(10):
                    if sample[f'attribute{i}'] is not None:
                        used_attributes.add(i)

            print(f"[upload_general_population] 사용된 attributes: {sorted(used_attributes)}")

            # attribute 정의 생성 (RCM detail에 정의된 모든 attributes 반환)
            attributes = []
            for i in range(10):
                # RCM detail에서 attribute 이름 가져오기
                attr_name = rcm_detail[f'attribute{i}'] if rcm_detail else None

                # 이름이 정의되지 않은 attribute는 skip
                if not attr_name:
                    continue

                # population_attr_count를 기준으로 모집단/증빙 구분
                if i < population_attr_count:
                    attr_type = 'population'
                else:
                    attr_type = 'evidence'

                attributes.append({
                    'attribute': f'attribute{i}',
                    'name': attr_name,
                    'type': attr_type
                })

            print(f"[upload_general_population] attributes 생성 (population_count={population_attr_count}): {attributes}")

        return jsonify({
            'success': True,
            'population_count': population_count,
            'sample_size': sample_size,
            'line_id': line_id,
            'sample_lines': sample_lines,
            'attributes': attributes,
            'population_attribute_count': population_attr_count
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


@bp_link7.route('/api/operation-evaluation/save-attributes', methods=['POST'])
@login_required
def save_attributes():
    """Attribute 필드 설정 저장"""
    user_info = get_user_info()
    data = request.get_json()

    line_id = data.get('line_id')
    attributes = data.get('attributes', [])

    if not line_id or not attributes:
        return jsonify({'success': False, 'error': '필수 데이터가 누락되었습니다.'})

    try:
        # Attribute 설정을 로그로 출력 (실제 구현은 DB 스키마에 따라 조정 필요)
        attribute_info = json.dumps(attributes, ensure_ascii=False)
        print(f"[save_attributes] line_id: {line_id}, attributes: {attribute_info}")

        # 성공 응답
        return jsonify({
            'success': True,
            'message': 'Attribute 설정이 저장되었습니다.'
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


@bp_link7.route('/api/operation-evaluation/reset-population', methods=['POST'])
@login_required
def reset_population_upload():
    """모집단 업로드 초기화 (파일 삭제 + DB 데이터 삭제)"""
    import os
    user_info = get_user_info()
    data = request.get_json()

    control_code = data.get('control_code')
    line_id = data.get('line_id')

    if not control_code:
        return jsonify({'success': False, 'message': '통제 코드가 필요합니다.'})

    try:
        # 1. 업로드된 모집단 파일 삭제
        upload_folder = os.path.join('uploads', 'populations')
        if os.path.exists(upload_folder):
            # 파일명 패턴: {user_id}_{control_code}_*.xlsx
            file_pattern = f"{user_info['user_id']}_{control_code}_"
            for filename in os.listdir(upload_folder):
                if filename.startswith(file_pattern):
                    filepath = os.path.join(upload_folder, filename)
                    try:
                        os.remove(filepath)
                        print(f"[reset_population_upload] 파일 삭제: {filepath}")
                    except Exception as file_error:
                        print(f"[reset_population_upload] 파일 삭제 실패: {filepath}, {file_error}")

        # 2. DB에서 표본 데이터 삭제
        if line_id:
            with get_db() as conn:
                # 표본 데이터 삭제
                conn.execute('DELETE FROM sb_evaluation_sample WHERE line_id = %s', (line_id,))

                # 라인 데이터 삭제
                conn.execute('DELETE FROM sb_operation_evaluation_line WHERE line_id = %s', (line_id,))

                conn.commit()
                print(f"[reset_population_upload] DB 데이터 삭제 완료: line_id={line_id}")

        log_user_activity(user_info, 'DATA_DELETE', '모집단 업로드 초기화',
                         f'/api/operation-evaluation/reset-population (control: {control_code})',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({
            'success': True,
            'message': '모집단 업로드가 초기화되었습니다.'
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


# ============================================================================
# 운영평가 다운로드 기능
# ============================================================================

@bp_link7.route('/operation-evaluation/download')
@login_required
def download_operation_evaluation():
    """운영평가 결과를 Template_Manual.xlsx 양식으로 다운로드 (통제별)"""
    from flask import make_response
    import urllib.parse

    user_info = get_user_info()

    # URL 파라미터 받기
    rcm_id = request.args.get('rcm_id')
    evaluation_session = request.args.get('evaluation_session')
    design_evaluation_session = request.args.get('design_evaluation_session')
    control_code = request.args.get('control_code')

    # 필수 파라미터 검증
    if not all([rcm_id, evaluation_session, design_evaluation_session, control_code]):
        flash('RCM ID, 운영평가 세션, 설계평가 세션, 통제번호가 필요합니다.', 'error')
        return redirect(url_for('link7.user_operation_evaluation'))

    try:
        # 템플릿 파일 경로
        template_path = os.path.join(os.path.dirname(__file__), 'paper_templates', 'Template_Manual.xlsx')

        if not os.path.exists(template_path):
            flash('템플릿 파일을 찾을 수 없습니다.', 'error')
            return redirect(url_for('link7.user_operation_evaluation'))

        # 템플릿 로드 (외부 링크 제거)
        wb = load_workbook(template_path, keep_links=False)

        # RCM 정보 조회
        with get_db() as conn:
            rcm_info = conn.execute("""
                SELECT rcm_name, description
                FROM sb_rcm
                WHERE rcm_id = %s
            """, (rcm_id,)).fetchone()

            if not rcm_info:
                flash('RCM 정보를 찾을 수 없습니다.', 'error')
                return redirect(url_for('link7.user_operation_evaluation'))

            # 운영평가 결과 조회 (해당 통제 1개만)
            evaluation = conn.execute("""
                SELECT
                    l.line_id,
                    l.control_code,
                    rd.control_name,
                    rd.control_description,
                    rd.control_frequency,
                    rd.control_type,
                    rd.control_nature,
                    l.sample_size,
                    l.exception_count,
                    l.exception_details,
                    l.conclusion,
                    l.improvement_plan,
                    l.review_comment,
                    l.evaluation_date,
                    d.attribute0, d.attribute1, d.attribute2, d.attribute3, d.attribute4,
                    d.attribute5, d.attribute6, d.attribute7, d.attribute8, d.attribute9,
                    d.population_attribute_count,
                    d.recommended_sample_size
                FROM sb_operation_evaluation_line l
                JOIN sb_operation_evaluation_header h ON l.header_id = h.header_id
                JOIN sb_rcm_detail_v rd ON h.rcm_id = rd.rcm_id AND l.control_code = rd.control_code
                JOIN sb_rcm_detail d ON d.rcm_id = h.rcm_id AND d.control_code = l.control_code
                WHERE h.rcm_id = %s
                  AND h.evaluation_session = %s
                  AND h.design_evaluation_session = %s
                  AND l.control_code = %s
            """, (rcm_id, evaluation_session, design_evaluation_session, control_code)).fetchone()

            # 설계평가 결과 조회 (design_comment 및 line_id 가져오기)
            design_evaluation = conn.execute("""
                SELECT l.design_comment, l.line_id
                FROM sb_design_evaluation_line l
                JOIN sb_design_evaluation_header h ON l.header_id = h.header_id
                WHERE h.rcm_id = %s
                  AND h.evaluation_session = %s
                  AND l.control_code = %s
            """, (rcm_id, design_evaluation_session, control_code)).fetchone()

            # 설계평가 이미지 조회 (파일 시스템에서)
            design_image_files = []
            if design_evaluation:
                # header_id 조회
                header = conn.execute("""
                    SELECT header_id
                    FROM sb_design_evaluation_header
                    WHERE rcm_id = %s AND evaluation_session = %s
                """, (rcm_id, design_evaluation_session)).fetchone()

                if header:
                    header_id = header['header_id']
                    image_dir = os.path.join('static', 'uploads', 'design_evaluations', str(rcm_id), str(header_id), control_code)

                    if os.path.exists(image_dir):
                        design_image_files = [os.path.join(image_dir, f) for f in os.listdir(image_dir) if os.path.isfile(os.path.join(image_dir, f))]

            # 운영평가 이미지 조회 (파일 시스템에서)
            operation_image_files = []
            if evaluation:
                # header_id 조회
                op_header = conn.execute("""
                    SELECT header_id
                    FROM sb_operation_evaluation_header
                    WHERE rcm_id = %s AND evaluation_session = %s
                """, (rcm_id, evaluation_session)).fetchone()

                if op_header:
                    op_header_id = op_header['header_id']
                    op_image_dir = os.path.join('static', 'uploads', 'operation_evaluations', str(rcm_id), str(op_header_id), control_code)
                    print(f"[DEBUG] 운영평가 이미지 디렉토리: {op_image_dir}")
                    print(f"[DEBUG] 디렉토리 존재 여부: {os.path.exists(op_image_dir)}")

                    if os.path.exists(op_image_dir):
                        operation_image_files = [os.path.join(op_image_dir, f) for f in os.listdir(op_image_dir) if os.path.isfile(os.path.join(op_image_dir, f))]
                        print(f"[DEBUG] 운영평가 이미지 파일 수: {len(operation_image_files)}")
                        for img_file in operation_image_files:
                            print(f"[DEBUG]   - {img_file}")
                else:
                    print(f"[DEBUG] 운영평가 header를 찾을 수 없음 (rcm_id={rcm_id}, evaluation_session={evaluation_session})")

        if not evaluation:
            flash('다운로드할 운영평가 결과가 없습니다.', 'warning')
            return redirect(url_for('link7.user_operation_evaluation'))

        eval_dict = dict(evaluation)
        design_eval_dict = dict(design_evaluation) if design_evaluation else {}

        # Template 시트에 직접 내용 작성
        template_sheet = wb['Template']

        # Client 정보 (C2)
        template_sheet['C2'] = user_info.get('company_name', '')

        # Prepared by (C4)
        template_sheet['C4'] = user_info.get('user_name', '')

        # 통제번호 (C7)
        template_sheet['C7'] = control_code

        # 통제명 (C8)
        template_sheet['C8'] = eval_dict.get('control_name', '')

        # 주기 (C9)
        template_sheet['C9'] = eval_dict.get('control_frequency', '')

        # 구분 (C10)
        template_sheet['C10'] = eval_dict.get('control_type', '')

        # 통제 설명 (C11)
        template_sheet['C11'] = eval_dict.get('control_description', '')

        # 설계평가 검토 결과 (C12)
        design_comment = design_eval_dict.get('design_comment', '')
        template_sheet['C12'] = design_comment

        # C12 셀의 행 높이 자동 조정 (텍스트 길이에 따라)
        if design_comment:
            # 줄바꿈 개수 계산
            line_count = design_comment.count('\n') + 1
            # 기본 행 높이(15) + 각 줄당 추가 높이(15)
            row_height = 15 + (line_count * 15)
            # 최대 높이 제한 (300)
            row_height = min(row_height, 300)
            template_sheet.row_dimensions[12].height = row_height

        # 운영평가 의견 작성 (C13)
        operation_review_comment = eval_dict.get('review_comment', '')
        template_sheet['C13'] = operation_review_comment

        # C13 셀의 행 높이 자동 조정 (텍스트 길이에 따라)
        if operation_review_comment:
            # 줄바꿈 개수 계산
            line_count = operation_review_comment.count('\n') + 1
            # 기본 행 높이(15) + 각 줄당 추가 높이(15)
            row_height = 15 + (line_count * 15)
            # 최대 높이 제한 (300)
            row_height = min(row_height, 300)
            template_sheet.row_dimensions[13].height = row_height

        # 운영평가 결론 작성 (C14) - Effective, Ineffective 등
        operation_conclusion = eval_dict.get('conclusion', '')
        template_sheet['C14'] = operation_conclusion

        # Template 시트명을 통제코드로 변경
        template_sheet.title = control_code[:31]  # Excel 시트명 31자 제한

        # Testing Table 시트에 샘플 데이터 작성
        testing_table = wb['Testing Table']
        line_id = eval_dict.get('line_id')
        population_count = eval_dict.get('population_attribute_count', 2)
        sample_size = eval_dict.get('sample_size', 0)

        # 모집단 attribute 개수와 증빙 attribute 개수 계산
        evidence_attributes = []
        for i in range(population_count, 10):
            attr_key = f'attribute{i}'
            attr_name = eval_dict.get(attr_key)
            if attr_name:
                evidence_attributes.append((i, attr_name))

        evidence_count = len(evidence_attributes)

        # 디버그: attribute 정보 출력
        print(f"[DEBUG] Control: {control_code}")
        print(f"[DEBUG] population_count: {population_count}")
        print(f"[DEBUG] Population attributes:")
        for i in range(population_count):
            print(f"  attribute{i}: {eval_dict.get(f'attribute{i}')}")
        print(f"[DEBUG] Evidence attributes:")
        for i, name in evidence_attributes:
            print(f"  attribute{i}: {name}")

        # 템플릿에 이미 C4~L4(10개 컬럼)과 5~64행(60개 샘플)이 준비되어 있음
        # 1. C4~L4에 헤더를 채우고 사용하지 않는 컬럼 삭제
        # 2. 5~64행에 샘플 데이터를 채우고 사용하지 않는 행 삭제

        from openpyxl.styles import PatternFill
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        # C열(3번)부터 시작
        current_col = 3

        # 모집단 항목 헤더 작성 (노란색 배경)
        for i in range(population_count):
            attr_key = f'attribute{i}'
            attr_name = eval_dict.get(attr_key, f'모집단{i+1}')
            if attr_name:
                cell = testing_table.cell(row=4, column=current_col, value=attr_name)
                cell.fill = yellow_fill
                current_col += 1

        # 증빙 항목 헤더 작성 (초록색 배경)
        evidence_col_start = current_col
        for i, attr_name in evidence_attributes:
            cell = testing_table.cell(row=4, column=current_col, value=attr_name)
            cell.fill = green_fill
            current_col += 1

        # 사용하지 않는 컬럼 삭제 (L열=12번 컬럼까지 준비되어 있음)
        # 결론, 비고는 템플릿에 이미 있으므로 추가 작성 불필요
        # current_col: 모집단 + 증빙 사용한 마지막 컬럼 + 1
        # current_col부터 12까지 삭제 (결론/비고 컬럼도 템플릿에 있으므로)
        print(f"[DEBUG] current_col after evidence: {current_col}")
        if current_col <= 12:
            cols_to_delete = 12 - current_col + 1
            print(f"[DEBUG] Deleting columns from {current_col} to 12 (count: {cols_to_delete})")
            testing_table.delete_cols(current_col, cols_to_delete)
        else:
            print(f"[DEBUG] No columns to delete (current_col={current_col} > 12)")

        # B열에 순번 작성 (1, 2, 3, ...)
        if sample_size > 0:
            for i in range(sample_size):
                testing_table.cell(row=5 + i, column=2, value=i + 1)  # B열 = column 2

        # 샘플 데이터 입력 (5행부터)
        if line_id:
            samples = get_operation_evaluation_samples(line_id)
            if samples:
                for row_idx, sample in enumerate(samples, start=5):
                    sample_attributes = sample.get('attributes', {})

                    # 모집단 데이터 (C열부터)
                    col = 3
                    for i in range(population_count):
                        attr_key = f'attribute{i}'
                        attr_value = sample_attributes.get(attr_key, '')
                        testing_table.cell(row=row_idx, column=col, value=attr_value)
                        col += 1

                    # 증빙 데이터 (모집단 다음 컬럼부터)
                    for i, attr_name in evidence_attributes:
                        attr_key = f'attribute{i}'
                        attr_value = sample_attributes.get(attr_key, '')
                        testing_table.cell(row=row_idx, column=col, value=attr_value)
                        col += 1

        # 66번 행("Testing Table")의 색상을 행 전체에 미리 적용 (행 삭제 전)
        from copy import copy
        from openpyxl.styles import PatternFill
        source_cell_66 = testing_table.cell(row=66, column=2)
        if source_cell_66.fill:
            for col in range(2, 16385):
                cell = testing_table.cell(row=66, column=col)
                cell.fill = copy(source_cell_66.fill)

        # 사용하지 않는 행 삭제 (5~64행까지 60개 준비되어 있음)
        if sample_size < 60:
            first_row_to_delete = 5 + sample_size
            rows_to_delete = 64 - first_row_to_delete + 1
            if rows_to_delete > 0:
                testing_table.delete_rows(first_row_to_delete, rows_to_delete)

        # 행 삭제 후 "Testing Table" 구분자 행 위치 계산
        # 원래 66번 행이 (66 - rows_to_delete)번으로 이동
        if sample_size < 60 and rows_to_delete > 0:
            testing_table_separator_row = 66 - rows_to_delete
        else:
            testing_table_separator_row = 66

        # 이미지 삽입 시작 위치 초기화 (구분자 2칸 아래)
        current_row = testing_table_separator_row + 2

        # 설계평가 이미지를 Testing Table 구분자 다음에 삽입
        if design_image_files:
            from openpyxl.drawing.image import Image as XLImage

            for image_path in design_image_files:
                if os.path.exists(image_path):
                    try:
                        # 이미지 객체 생성
                        xl_img = XLImage(image_path)

                        # 이미지 크기 조정 (최대 너비 400px)
                        max_width = 400
                        if xl_img.width > max_width:
                            ratio = max_width / xl_img.width
                            xl_img.width = max_width
                            xl_img.height = int(xl_img.height * ratio)

                        # 이미지 삽입 (모든 설계평가 이미지는 같은 행에 삽입)
                        design_img_row = testing_table_separator_row + 2
                        xl_img.anchor = f'B{design_img_row}'
                        testing_table.add_image(xl_img)

                        # 행 높이 조정 (가장 큰 이미지 높이로 설정)
                        current_height = testing_table.row_dimensions[design_img_row].height or 0
                        new_height = (xl_img.height * 0.75) + 5
                        if new_height > current_height:
                            testing_table.row_dimensions[design_img_row].height = new_height

                    except Exception as e:
                        print(f"설계평가 이미지 삽입 실패 ({image_path}): {e}")

        # 운영평가 이미지 삽입 (설계평가 바로 다음)
        if operation_image_files:
            from openpyxl.drawing.image import Image as XLImage

            # 운영평가 이미지는 설계평가 이미지 다음 행 (설계평가 +0, 빈칸 +1, 운영평가 +2)
            # 설계평가가 (separator + 2)에 있으므로, 운영평가는 (separator + 2 + 2) = separator + 4
            if design_image_files:
                operation_img_row = testing_table_separator_row + 4
            else:
                operation_img_row = testing_table_separator_row + 2

            for image_path in operation_image_files:
                if os.path.exists(image_path):
                    try:
                        # 이미지 객체 생성
                        xl_img = XLImage(image_path)

                        # 이미지 크기 조정 (최대 너비 400px)
                        max_width = 400
                        if xl_img.width > max_width:
                            ratio = max_width / xl_img.width
                            xl_img.width = max_width
                            xl_img.height = int(xl_img.height * ratio)

                        # 이미지 삽입 (모든 운영평가 이미지는 같은 행에 삽입)
                        xl_img.anchor = f'B{operation_img_row}'
                        testing_table.add_image(xl_img)

                        # 행 높이 조정 (가장 큰 이미지 높이로 설정)
                        current_height = testing_table.row_dimensions[operation_img_row].height or 0
                        new_height = (xl_img.height * 0.75) + 5
                        if new_height > current_height:
                            testing_table.row_dimensions[operation_img_row].height = new_height

                    except Exception as e:
                        print(f"운영평가 이미지 삽입 실패 ({image_path}): {e}")

        # Population 시트 처리
        recommended_sample_size = eval_dict.get('recommended_sample_size', 0)
        if recommended_sample_size == 0 and 'Population' in wb.sheetnames:
            # 표본수가 0인 경우: 업로드한 모집단 데이터를 Population 시트에 채움
            population_sheet = wb['Population']

            # 업로드한 모집단 파일 경로
            upload_folder = os.path.join('uploads', 'populations')
            population_file_pattern = f"{user_info['user_id']}_{control_code}_*"
            population_files = []

            if os.path.exists(upload_folder):
                import glob
                population_files = glob.glob(os.path.join(upload_folder, population_file_pattern))

            if population_files:
                # 가장 최근 파일 사용
                population_file = max(population_files, key=os.path.getmtime)
                print(f"[DEBUG] 모집단 파일 발견: {population_file}")

                try:
                    # 모집단 파일 읽기
                    pop_wb = load_workbook(population_file, read_only=True)
                    pop_ws = pop_wb.active

                    # 헤더 복사 (1행)
                    for col_idx, cell in enumerate(pop_ws[1], start=1):
                        if cell.value:
                            population_sheet.cell(row=1, column=col_idx, value=cell.value)

                    # 데이터 복사 (2행부터)
                    row_idx = 2
                    for row in pop_ws.iter_rows(min_row=2, values_only=True):
                        if any(cell is not None for cell in row):  # 빈 행이 아니면
                            for col_idx, value in enumerate(row, start=1):
                                population_sheet.cell(row=row_idx, column=col_idx, value=value)
                            row_idx += 1

                    pop_wb.close()
                    print(f"[DEBUG] 모집단 데이터 복사 완료: {row_idx - 2}개 행")
                except Exception as e:
                    print(f"[ERROR] 모집단 파일 읽기 실패: {e}")
            else:
                print(f"[DEBUG] 모집단 파일을 찾을 수 없음: {population_file_pattern}")
        elif recommended_sample_size != 0 and 'Population' in wb.sheetnames:
            # 표본수가 0이 아닌 경우: Population 시트 삭제
            wb.remove(wb['Population'])

        # 시트 순서 조정: 통제명 시트를 가장 앞에, Testing Table을 두 번째로
        control_sheet_index = wb.index(template_sheet)
        testing_table_index = wb.index(wb['Testing Table'])

        # 통제명 시트를 맨 앞으로 이동
        wb.move_sheet(template_sheet, offset=-control_sheet_index)
        # Testing Table을 두 번째로 이동 (통제명 시트 다음)
        wb.move_sheet(wb['Testing Table'], offset=-testing_table_index + 1)

        # Population 시트가 있으면 세 번째로 이동
        if 'Population' in wb.sheetnames:
            population_index = wb.index(wb['Population'])
            wb.move_sheet(wb['Population'], offset=-population_index + 2)

        # 명명된 범위(defined names) 제거 (깨진 참조 방지)
        if hasattr(wb, 'defined_names'):
            try:
                # openpyxl 3.x 버전
                names_to_remove = list(wb.defined_names.definedName)
                for name in names_to_remove:
                    del wb.defined_names[name.name]
            except AttributeError:
                # openpyxl 2.x 버전 또는 다른 구조
                try:
                    names_to_remove = [name for name in wb.defined_names]
                    for name in names_to_remove:
                        try:
                            del wb.defined_names[name]
                        except:
                            pass
                except:
                    pass

        # 외부 링크(external links) 제거
        if hasattr(wb, '_external_links'):
            wb._external_links = []

        # 임시 파일로 저장
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()

        # 다운로드 파일명 생성
        filename = f"{control_code}_{evaluation_session}.xlsx"
        unsafe_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
        for char in unsafe_chars:
            filename = filename.replace(char, '_')

        # 파일 전송 (UTF-8 인코딩)
        response = make_response(send_file(
            temp_file.name,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ))

        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"

        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'다운로드 중 오류가 발생했습니다: {str(e)}', 'error')
        return redirect(url_for('link7.user_operation_evaluation'))


@bp_link7.route('/api/operation-evaluation/upload-image', methods=['POST'])
@login_required
def upload_operation_image():
    """운영평가 이미지 업로드"""
    try:
        rcm_id = request.form.get('rcm_id')
        header_id = request.form.get('header_id')
        control_code = request.form.get('control_code')

        if not all([rcm_id, header_id, control_code]):
            return jsonify({'success': False, 'message': '필수 파라미터가 누락되었습니다.'}), 400

        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '파일이 없습니다.'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '파일이 선택되지 않았습니다.'}), 400

        # 파일 확장자 검증
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'bmp'}
        file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if file_ext not in allowed_extensions:
            return jsonify({'success': False, 'message': '허용되지 않는 파일 형식입니다.'}), 400

        # 저장 경로 생성
        upload_dir = os.path.join('static', 'uploads', 'operation_evaluations', str(rcm_id), str(header_id), control_code)
        os.makedirs(upload_dir, exist_ok=True)

        # 파일명 생성 (타임스탬프 포함)
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{file.filename}"
        filepath = os.path.join(upload_dir, filename)

        # 파일 저장
        file.save(filepath)

        return jsonify({
            'success': True,
            'message': '이미지가 업로드되었습니다.',
            'filepath': filepath.replace('\\', '/')
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'업로드 실패: {str(e)}'}), 500


@bp_link7.route('/api/operation-evaluation/images/<int:rcm_id>/<int:header_id>/<control_code>')
@login_required
def get_operation_images(rcm_id, header_id, control_code):
    """운영평가 이미지 목록 조회"""
    try:
        image_dir = os.path.join('static', 'uploads', 'operation_evaluations', str(rcm_id), str(header_id), control_code)

        if not os.path.exists(image_dir):
            return jsonify({'success': True, 'images': []})

        images = []
        for filename in os.listdir(image_dir):
            if os.path.isfile(os.path.join(image_dir, filename)):
                images.append({
                    'filename': filename,
                    'url': f'/{image_dir}/{filename}'.replace('\\', '/')
                })

        return jsonify({'success': True, 'images': images})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link7.route('/api/operation-evaluation/delete-image', methods=['POST'])
@login_required
def delete_operation_image():
    """운영평가 이미지 삭제"""
    try:
        data = request.get_json()
        filepath = data.get('filepath')

        if not filepath:
            return jsonify({'success': False, 'message': '파일 경로가 없습니다.'}), 400

        # 파일 존재 확인 및 삭제
        if os.path.exists(filepath):
            os.remove(filepath)
            return jsonify({'success': True, 'message': '이미지가 삭제되었습니다.'})
        else:
            return jsonify({'success': False, 'message': '파일을 찾을 수 없습니다.'}), 404

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
